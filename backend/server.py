from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import io
import re
import csv
import json
import time
import html
import hmac
import hashlib
import uuid
import logging
import secrets
import asyncio
import ipaddress
import sys
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal, Any
from urllib.parse import quote


import bcrypt
import jwt
import httpx
import resend
from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, Query, UploadFile, File, Body
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.cors import CORSMiddleware
from PIL import Image as PILImage
from motor.motor_asyncio import AsyncIOMotorClient
from pymongo import ReturnDocument
from pymongo.errors import DuplicateKeyError
from pydantic import BaseModel, Field, EmailStr, ConfigDict, field_validator
from reportlab.lib.pagesizes import LETTER
from reportlab.lib import colors as rl_colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas

sys.modules.setdefault("server", sys.modules[__name__])
sys.modules.setdefault("backend.server", sys.modules[__name__])


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]
JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_MINUTES = max(5, int(os.environ.get("ACCESS_TOKEN_MINUTES", "15")))
REFRESH_TOKEN_DAYS = max(1, int(os.environ.get("REFRESH_TOKEN_DAYS", "30")))


def _private_ref(value: Any) -> str:
    """Return a stable correlation token without logging the source value."""
    digest = hmac.new(JWT_SECRET.encode(), str(value).encode(), hashlib.sha256).hexdigest()
    return digest[:12]


ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "admin@fironova.com")
# Aucun défaut : un mot de passe admin en dur dans le repo est un mot de passe public.
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD")
if not ADMIN_PASSWORD:
    raise RuntimeError(
        "ADMIN_PASSWORD est obligatoire. Définissez-le dans l'environnement "
        "avant de démarrer le serveur."
    )
INTERAC_EMAIL = os.environ.get("INTERAC_EMAIL", "orders@fironova.com")
# Confirmation automatique des virements Interac (Autodeposit) via Microsoft
# Graph API (app-only / client credentials). Boîte surveillée = INTERAC_EMAIL.
INTERAC_GRAPH_TENANT_ID = os.environ.get("INTERAC_GRAPH_TENANT_ID", "")
INTERAC_GRAPH_CLIENT_ID = os.environ.get("INTERAC_GRAPH_CLIENT_ID", "")
INTERAC_GRAPH_CLIENT_SECRET = os.environ.get("INTERAC_GRAPH_CLIENT_SECRET", "")
INTERAC_GRAPH_USER = os.environ.get("INTERAC_GRAPH_USER", "").strip().lower() or INTERAC_EMAIL.lower()
INTERAC_AUTOCONFIRM_MODE = os.environ.get("INTERAC_AUTOCONFIRM_MODE", "off").strip().lower()
if INTERAC_AUTOCONFIRM_MODE not in {"off", "strict"}:
    INTERAC_AUTOCONFIRM_MODE = "off"
INTERAC_TRUSTED_SENDER = os.environ.get("INTERAC_TRUSTED_SENDER", "").strip().lower()

# Strip placeholder values like <TENANT_ID> that were never replaced.
def _strip_placeholder(v: str) -> str:
    return "" if (v.startswith("<") and v.endswith(">")) else v

INTERAC_GRAPH_TENANT_ID = _strip_placeholder(INTERAC_GRAPH_TENANT_ID)
INTERAC_GRAPH_CLIENT_ID = _strip_placeholder(INTERAC_GRAPH_CLIENT_ID)
INTERAC_GRAPH_CLIENT_SECRET = _strip_placeholder(INTERAC_GRAPH_CLIENT_SECRET)

# Code d'accès facultatif demandé AVANT même l'écran de login admin, côté SPA
# publique. Ne remplace pas l'auth (JWT + rôle restent la vraie barrière sur
# /api/admin/*) — sert seulement à ce qu'un visiteur qui tombe sur l'URL admin
# obscure par hasard ne voie même pas d'écran de login. Si non défini, la
# passerelle est désactivée (aucun blocage).
ADMIN_GATE_CODE = os.environ.get("ADMIN_GATE_CODE")
TRUST_PROXY_IPS = {
    ip.strip()
    for ip in os.environ.get("TRUST_PROXY_IPS", "").split(",")
    if ip.strip()
}
INTERAC_PASSWORD_HINT = os.environ.get("INTERAC_PASSWORD_HINT", "FIRONOVA")
NOWPAYMENTS_API_KEY = os.environ.get("NOWPAYMENTS_API_KEY", "")
NOWPAYMENTS_IPN_SECRET = os.environ.get("NOWPAYMENTS_IPN_SECRET", "")
PUBLIC_BASE_URL = os.environ.get("PUBLIC_BASE_URL", "").rstrip("/")
APP_ENV = os.environ.get("APP_ENV", os.environ.get("ENV", os.environ.get("ENVIRONMENT", ""))).strip().lower()
IS_PRODUCTION = APP_ENV in {"prod", "production"}
if IS_PRODUCTION and not PUBLIC_BASE_URL:
    raise RuntimeError(
        "PUBLIC_BASE_URL est obligatoire en production. Définissez-la avant de démarrer le serveur."
    )
NOWPAYMENTS_BASE_URL = "https://api.nowpayments.io/v1"
RESEND_API_KEY = os.environ.get("RESEND_API_KEY", "")
SENDER_EMAIL = os.environ.get("SENDER_EMAIL", "orders@fironova.com")
MAGIC_SENDER_EMAIL = os.environ.get("MAGIC_SENDER_EMAIL", "")
# Expéditeur du programme d'affiliation (invitations, relances). Séparé de
# SENDER_EMAIL parce qu'une invitation à un partenariat commercial n'a pas à
# partir de l'adresse des commandes — et parce que si cette adresse n'est pas
# vérifiée chez le fournisseur d'envoi, seules les invitations échouent au lieu
# de faire tomber toute la messagerie transactionnelle.
AFFILIATE_SENDER_EMAIL = os.environ.get("AFFILIATE_SENDER_EMAIL", "")
ADMIN_NOTIFICATION_EMAIL = os.environ.get("ADMIN_NOTIFICATION_EMAIL", "admin@fironova.com")
SHIPPING_FLAT_CAD = float(os.environ.get("SHIPPING_FLAT_CAD", "20.00"))
FREE_SHIPPING_THRESHOLD_CAD = float(os.environ.get("FREE_SHIPPING_THRESHOLD_CAD", "200.00"))
UNPAID_ORDER_TTL_HOURS = float(os.environ.get("UNPAID_ORDER_TTL_HOURS", "24"))
# Seuil minimum en CAD sous lequel un payout affilié est reporté au cycle
# suivant (évite de gaspiller des frais gas crypto sur des micro-payouts).
# Les commissions restent en `approved` (payout_id=None) et roulent au mois suivant.
# Une notification email bilingue est envoyée UNE seule fois par (affilié, période).
AFFILIATE_PAYOUT_MIN_CAD = float(os.environ.get("AFFILIATE_PAYOUT_MIN_CAD", "25.00"))
PREORDER_RELEASE_INTERVAL_SECONDS = int(os.environ.get("PREORDER_RELEASE_INTERVAL_SECONDS", "300"))
# Rabais % du coupon auto-lié à chaque affilié (0 = pas de coupon auto).
AFFILIATE_COUPON_PERCENT = float(os.environ.get("AFFILIATE_COUPON_PERCENT", "10"))
_STANDARD_COUPON_MAX_PERCENT_RAW = os.environ.get("STANDARD_COUPON_MAX_PERCENT", "").strip()
STANDARD_COUPON_MAX_PERCENT = (
    float(_STANDARD_COUPON_MAX_PERCENT_RAW) if _STANDARD_COUPON_MAX_PERCENT_RAW else None
)
COOKIE_SECURE = os.environ.get("COOKIE_SECURE", "true").strip().lower() == "true"
COOKIE_SAMESITE = os.environ.get("COOKIE_SAMESITE", "none").strip().lower()
if COOKIE_SAMESITE not in {"lax", "strict", "none"}:
    COOKIE_SAMESITE = "none"
# Browsers reject SameSite=None when Secure=false. Fallback keeps local HTTP usable.
if not COOKIE_SECURE and COOKIE_SAMESITE == "none":
    COOKIE_SAMESITE = "lax"

# Si l'API est derrière un reverse-proxy, renseigner les CIDR de confiance ; sinon X-Forwarded-For est ignoré et le rate-limiting se basera sur l'IP partagée du proxy.
TRUSTED_PROXIES: list[str] = [
    e for e in (os.environ.get("TRUSTED_PROXIES", "") or "").split(",") if e.strip()
]

# Feature flags — toggle without deploying new code, just flip the env var and restart.
COA_PAGE_ENABLED = os.environ.get("COA_PAGE_ENABLED", "false").strip().lower() == "true"

# --- Prélancement (BLOC 3) --------------------------------------------------
# Défaut sûr : la boutique reste OUVERTE si la variable est absente.
PRELAUNCH_ENABLED = os.environ.get("PRELAUNCH_ENABLED", "false").strip().lower() == "true"
PRELAUNCH_PREVIEW_TOKEN = os.environ.get("PRELAUNCH_PREVIEW_TOKEN", "")  # ?preview=<token> contourne la porte
LAUNCH_COUPON_CODE = os.environ.get("LAUNCH_COUPON_CODE", "LAUNCH15").strip().upper()
LAUNCH_COUPON_ENABLED = os.environ.get("LAUNCH_COUPON_ENABLED", "false").strip().lower() == "true"

# File uploads (COA PDFs). Served statically at /uploads/coa/<file>.
UPLOAD_DIR = ROOT_DIR / "uploads"
COA_UPLOAD_DIR = UPLOAD_DIR / "coa"
COA_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
MAX_COA_UPLOAD_MB = float(os.environ.get("MAX_COA_UPLOAD_MB", "10"))

# Product images. Served statically at /uploads/images/<file>.
IMAGE_UPLOAD_DIR = UPLOAD_DIR / "images"
IMAGE_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
MAX_IMAGE_UPLOAD_MB = float(os.environ.get("MAX_IMAGE_UPLOAD_MB", "5"))
ALLOWED_IMAGE_TYPES = {"image/png", "image/jpeg", "image/webp", "image/gif"}
_IMAGE_FORMAT_EXT = {"PNG": "png", "JPEG": "jpg", "WEBP": "webp", "GIF": "gif"}

# --- Address validation (Google Maps Address Validation API) --------------
# Utilisée dans POST /checkout (blocking) et POST /checkout/validate-address
# (preview côté client). Cache 24h TTL pour économiser le quota gratuit (10 K/mois).
GOOGLE_MAPS_API_KEY = os.environ.get("GOOGLE_MAPS_API_KEY", "").strip()
_ADDRESS_CACHE: dict = {}
_ADDRESS_CACHE_TTL_SEC = 60 * 60 * 24  # 24h

# Étiquettes d'expédition Postes Canada — servies à /uploads/labels/<file>.
LABEL_UPLOAD_DIR = UPLOAD_DIR / "labels"
LABEL_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# Postes Canada (Canada Post) rating/tracking API — leave blank to keep using the flat-rate
# shipping_zones/shipping_methods system already in place.
CANADA_POST_API_KEY = os.environ.get("CANADA_POST_API_KEY", "")
CANADA_POST_CUSTOMER_NUMBER = os.environ.get("CANADA_POST_CUSTOMER_NUMBER", "")
CANADA_POST_CONTRACT_ID = os.environ.get("CANADA_POST_CONTRACT_ID", "")
CANADA_POST_ORIGIN_POSTAL_CODE = os.environ.get("CANADA_POST_ORIGIN_POSTAL_CODE", "")
# Expéditeur pour la création d'étiquettes. Emballage neutre : aucune mention
# du contenu ne figure sur l'étiquette (politique d'emballage discret).
CANADA_POST_SENDER_NAME = os.environ.get("CANADA_POST_SENDER_NAME", "FIRONOVA")
CANADA_POST_SENDER_ADDRESS = os.environ.get("CANADA_POST_SENDER_ADDRESS", "")
CANADA_POST_SENDER_CITY = os.environ.get("CANADA_POST_SENDER_CITY", "Montreal")
CANADA_POST_SENDER_PROVINCE = os.environ.get("CANADA_POST_SENDER_PROVINCE", "QC")
CANADA_POST_SENDER_PHONE = os.environ.get("CANADA_POST_SENDER_PHONE", "")
CANADA_POST_ENVIRONMENT = os.environ.get("CANADA_POST_ENVIRONMENT", "dev")  # "dev" (soa-gw) or "prod" (soa-gw prod host)
CANADA_POST_BASE_URL = (
    "https://ct.soa-gw.canadapost.ca" if CANADA_POST_ENVIRONMENT == "dev"
    else "https://soa-gw.canadapost.ca"
)
# New Shipping OpenAPI (OAuth2) support. Modes:
# - legacy: existing XML/basic-auth flow
# - openapi: OAuth2 + JSON flow from the shipping v1 OpenAPI spec
# - auto: openapi if OAuth creds are present, otherwise legacy
CANADA_POST_API_MODE = os.environ.get("CANADA_POST_API_MODE", "auto").strip().lower()
CANADA_POST_OPENAPI_BASE_URL = os.environ.get(
    "CANADA_POST_OPENAPI_BASE_URL",
    "https://api.canadapost-postescanada.ca/prod/devportal-portaildesdeveloppeurs/shipping/v1",
).rstrip("/")
CANADA_POST_OAUTH_TOKEN_URL = os.environ.get(
    "CANADA_POST_OAUTH_TOKEN_URL",
    "https://api.canadapost-postescanada.ca/prod/devportal-portaildesdeveloppeurs/cpc-api-native-oauth-provider/oauth2/token",
).strip()
CANADA_POST_OAUTH_CLIENT_ID = os.environ.get("CANADA_POST_OAUTH_CLIENT_ID", "").strip()
CANADA_POST_OAUTH_CLIENT_SECRET = os.environ.get("CANADA_POST_OAUTH_CLIENT_SECRET", "").strip()
CANADA_POST_PLATFORM_ID = os.environ.get("CANADA_POST_PLATFORM_ID", "").strip()
CANADA_POST_MAILED_BY = os.environ.get("CANADA_POST_MAILED_BY", CANADA_POST_CUSTOMER_NUMBER).strip()
CANADA_POST_MOBO = os.environ.get("CANADA_POST_MOBO", CANADA_POST_CUSTOMER_NUMBER).strip()
CANADA_POST_INTENDED_METHOD = os.environ.get("CANADA_POST_INTENDED_METHOD", "").strip()
CANADA_POST_DEFAULT_SERVICE_CODE = os.environ.get("CANADA_POST_DEFAULT_SERVICE_CODE", "DOM.EP").strip() or "DOM.EP"
CANADA_POST_AUTO_LABEL_INTERVAL_SECONDS = int(os.environ.get("CANADA_POST_AUTO_LABEL_INTERVAL_SECONDS", "60"))
CANADA_POST_AUTO_DELIVERY_SYNC_SECONDS = int(os.environ.get("CANADA_POST_AUTO_DELIVERY_SYNC_SECONDS", "900"))
CANADA_POST_SANDBOX_DELIVERY_FALLBACK = os.environ.get("CANADA_POST_SANDBOX_DELIVERY_FALLBACK", "false").strip().lower() == "true"
CANADA_POST_SANDBOX_DELIVER_AFTER_HOURS = int(os.environ.get("CANADA_POST_SANDBOX_DELIVER_AFTER_HOURS", "24"))

# Le repli sandbox fabrique de FAUSSES confirmations de livraison : une
# commande jamais expédiée apparaîtrait livrée. Il est déjà conditionné à
# CANADA_POST_ENVIRONMENT == "dev" côté service, mais cette variable vaut
# "dev" PAR DÉFAUT — un simple oubli en production suffisait à l'activer.
# On refuse de démarrer plutôt que de risquer de mentir à un client.
if IS_PRODUCTION and CANADA_POST_SANDBOX_DELIVERY_FALLBACK:
    raise RuntimeError(
        "CANADA_POST_SANDBOX_DELIVERY_FALLBACK=true est interdit en production : "
        "ce repli simule des livraisons et marquerait des commandes livrées à tort."
    )
if IS_PRODUCTION and CANADA_POST_ENVIRONMENT != "prod":
    logging.error(
        "[config] CANADA_POST_ENVIRONMENT=%r en production : tarifs et étiquettes "
        "sortiraient de l'environnement de test Postes Canada. Attendu : 'prod'.",
        CANADA_POST_ENVIRONMENT,
    )

# ---------------------------------------------------------------------------
# Dispatch batch (fenêtre de traitement des commandes)
# Cutoff 13h heure de l'Est, jours ouvrables seulement (week-ends + fériés exclus).
# ---------------------------------------------------------------------------
from zoneinfo import ZoneInfo
try:
    import holidays as _holidays_lib
    _CA_HOLIDAYS = _holidays_lib.Canada()  # fériés fédéraux canadiens
except Exception:  # pragma: no cover - lib absente => aucun férié bloqué
    _CA_HOLIDAYS = set()

ORDER_CUTOFF_HOUR = int(os.environ.get("ORDER_CUTOFF_HOUR", "13"))
ORDER_CUTOFF_TZ = os.environ.get("ORDER_CUTOFF_TZ", "America/Toronto")


def _is_business_day(d) -> bool:
    """False les samedis, dimanches et jours fériés fédéraux canadiens."""
    if d.weekday() >= 5:  # 5 = samedi, 6 = dimanche
        return False
    return d not in _CA_HOLIDAYS


def compute_dispatch_batch(paid_at) -> str:
    """
    Retourne la date du lot d'expédition (prochain jour ouvrable) au format
    'YYYY-MM-DD' en date locale (ORDER_CUTOFF_TZ), à partir d'un horodatage de
    paiement (datetime aware UTC, ou string ISO).
    Règle : payé avant le cutoff un jour ouvrable => jour même, sinon jour
    suivant ; puis on avance jusqu'au prochain jour ouvrable.
    """
    if isinstance(paid_at, str):
        paid_at = datetime.fromisoformat(paid_at.replace("Z", "+00:00"))
    if paid_at.tzinfo is None:
        paid_at = paid_at.replace(tzinfo=timezone.utc)
    local = paid_at.astimezone(ZoneInfo(ORDER_CUTOFF_TZ))
    candidate = local.date()
    if not (local.hour < ORDER_CUTOFF_HOUR and _is_business_day(candidate)):
        candidate = candidate + timedelta(days=1)
    while not _is_business_day(candidate):
        candidate = candidate + timedelta(days=1)
    return candidate.isoformat()

if RESEND_API_KEY:
    resend.api_key = RESEND_API_KEY

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI(
    title="FIRONOVA API",
    version="1.0.0",
    docs_url=None if IS_PRODUCTION else "/docs",
    redoc_url=None if IS_PRODUCTION else "/redoc",
    openapi_url=None if IS_PRODUCTION else "/openapi.json",
)
api = APIRouter(prefix="/api")


class ImmutableStaticFiles(StaticFiles):
    async def get_response(self, path: str, scope: dict) -> Response:
        response = await super().get_response(path, scope)
        if response.status_code == 200:
            response.headers["Cache-Control"] = "public, max-age=31536000, immutable"
        return response


# Only non-sensitive catalog assets are public. Shipping labels contain customer
# PII and are served through an authenticated admin route below.
app.mount("/uploads/coa", ImmutableStaticFiles(directory=str(COA_UPLOAD_DIR)), name="uploads-coa")
app.mount("/uploads/images", ImmutableStaticFiles(directory=str(IMAGE_UPLOAD_DIR)), name="uploads-images")
app.mount("/api/uploads/coa", ImmutableStaticFiles(directory=str(COA_UPLOAD_DIR)), name="uploads-api-coa")
app.mount("/api/uploads/images", ImmutableStaticFiles(directory=str(IMAGE_UPLOAD_DIR)), name="uploads-api-images")


# ---------------------------------------------------------------------------
# Helpers: password & JWT
# ---------------------------------------------------------------------------
PASSWORD_COMPLEXITY_RE = re.compile(r"^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[^A-Za-z0-9]).{8,}$")


def _validate_password_strength(password: str) -> str:
    if not PASSWORD_COMPLEXITY_RE.match(password):
        raise ValueError("Password must be at least 8 characters and include uppercase, lowercase, a number, and a special character")
    return password


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def create_access_token(user_id: str, email: str, role: str, token_version: int = 0) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "role": role,
        "tv": token_version,  # doit correspondre à users.token_version — sinon token révoqué
        "exp": datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_MINUTES),
        "type": "access",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def _is_trusted_proxy_peer(request: Request) -> bool:
    peer = request.client.host if request.client else ""
    try:
        peer_addr = ipaddress.ip_address(peer)
    except ValueError:
        return False
    for entry in TRUSTED_PROXIES or sorted(TRUST_PROXY_IPS):
        try:
            if "/" in entry:
                if peer_addr in ipaddress.ip_network(entry.strip(), strict=False):
                    return True
            elif peer.lower() == entry.strip().lower():
                return True
        except ValueError:
            continue
    return False


def _trusted_forwarded_header(request: Request, name: str) -> str:
    return request.headers.get(name, "") if _is_trusted_proxy_peer(request) else ""


def _cookie_secure_for_request(request: Optional[Request]) -> bool:
    """Keep secure cookies in production, but allow local HTTP development.

    Browsers reject `Secure` cookies on plain HTTP, which breaks local auth
    flows (admin upload endpoints rely on that auth cookie).
    """
    if not COOKIE_SECURE:
        return False
    if request is None:
        return COOKIE_SECURE
    host = (request.headers.get("host") or "").split(":", 1)[0].strip().lower()
    proto = (_trusted_forwarded_header(request, "x-forwarded-proto") or request.url.scheme or "").lower()
    if proto != "https" and host in {"localhost", "127.0.0.1"}:
        return False
    return COOKIE_SECURE


def set_auth_cookie(response: Response, token: str, request: Optional[Request] = None) -> None:
    secure_flag = _cookie_secure_for_request(request)
    samesite_flag = COOKIE_SAMESITE
    if not secure_flag and samesite_flag == "none":
        samesite_flag = "lax"
    response.set_cookie(
        key="access_token",
        value=token,
        httponly=True,
        secure=secure_flag,
        samesite=samesite_flag,
        max_age=ACCESS_TOKEN_MINUTES * 60,
        path="/",
    )


def _hash_refresh_token(token: str) -> str:
    return hmac.new(JWT_SECRET.encode(), token.encode(), hashlib.sha256).hexdigest()


def _set_refresh_cookie(response: Response, token: str, request: Optional[Request] = None) -> None:
    secure_flag = _cookie_secure_for_request(request)
    samesite_flag = COOKIE_SAMESITE
    if not secure_flag and samesite_flag == "none":
        samesite_flag = "lax"
    response.set_cookie(
        key="refresh_token",
        value=token,
        httponly=True,
        secure=secure_flag,
        samesite=samesite_flag,
        max_age=REFRESH_TOKEN_DAYS * 86400,
        path="/api/auth",
    )


def clear_auth_cookie(response: Response) -> None:
    response.delete_cookie(key="access_token", path="/")
    response.delete_cookie(key="refresh_token", path="/api/auth")


async def _start_session(response: Response, request: Request, user: dict, family_id: Optional[str] = None) -> None:
    now = datetime.now(timezone.utc)
    raw_refresh = secrets.token_urlsafe(48)
    await db.refresh_sessions.insert_one({
        "id": str(uuid.uuid4()),
        "family_id": family_id or str(uuid.uuid4()),
        "user_id": user["id"],
        "token_hash": _hash_refresh_token(raw_refresh),
        "token_version": user.get("token_version", 0),
        "created_at": now,
        "expires_at": now + timedelta(days=REFRESH_TOKEN_DAYS),
        "revoked_at": None,
    })
    access_token = create_access_token(
        user["id"], user["email"], user["role"], user.get("token_version", 0),
    )
    set_auth_cookie(response, access_token, request)
    _set_refresh_cookie(response, raw_refresh, request)


def _reject_refresh(detail: str) -> JSONResponse:
    response = JSONResponse(status_code=401, content={"detail": detail})
    clear_auth_cookie(response)
    return response


async def refresh_session(response: Response, request: Request):
    raw_refresh = request.cookies.get("refresh_token")
    if not raw_refresh:
        return _reject_refresh("Refresh session required")

    token_hash = _hash_refresh_token(raw_refresh)
    now = datetime.now(timezone.utc)
    existing = await db.refresh_sessions.find_one({"token_hash": token_hash})
    if not existing:
        return _reject_refresh("Invalid refresh session")
    if existing.get("revoked_at") is not None:
        await db.refresh_sessions.update_many(
            {"family_id": existing["family_id"], "revoked_at": None},
            {"$set": {"revoked_at": now, "revoke_reason": "reuse_detected"}},
        )
        return _reject_refresh("Refresh session reuse detected")

    session = await db.refresh_sessions.find_one_and_update(
        {"token_hash": token_hash, "revoked_at": None, "expires_at": {"$gt": now}},
        {"$set": {"revoked_at": now, "revoke_reason": "rotated"}},
        return_document=ReturnDocument.BEFORE,
    )
    if not session:
        return _reject_refresh("Refresh session expired")

    user = await db.users.find_one({"id": session["user_id"]})
    if not user or session.get("token_version", 0) != user.get("token_version", 0):
        return _reject_refresh("Refresh session revoked")
    await _start_session(response, request, user, family_id=session["family_id"])
    return {"ok": True}


async def _resolve_user(request: Request) -> Optional[dict]:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        return None
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            return None
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            return None
        # Révocation : si le token_version du JWT ne correspond plus à celui
        # stocké sur l'utilisateur (changement de mot de passe, changement
        # d'email confirmé, déconnexion globale, suppression de compte),
        # le token est considéré invalide même s'il n'a pas expiré.
        if payload.get("tv", 0) != user.get("token_version", 0):
            return None
        user.pop("token_version", None)  # détail interne, pas exposé à l'API
        return user
    except jwt.PyJWTError:
        return None


async def get_current_user(request: Request) -> dict:
    user = await _resolve_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return user


def _public_user_payload(user: Optional[dict]) -> dict:
    if not user:
        return {}
    payload = {
        k: v for k, v in user.items()
        if k not in {"password_hash", "token_version", "passwordless"}
    }
    # Ne pas exposer les permissions internes pour les comptes clients.
    if payload.get("role") not in {"staff", "admin"}:
        payload.pop("permissions", None)
    return payload


async def get_admin_user(request: Request) -> dict:
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    return user


# ---------------------------------------------------------------------------
# Rôles & permissions granulaires — inspiré du modèle de règles par entité
# vu sur Base44 (Créateur uniquement / Tous les utilisateurs / etc.), adapté
# à un admin e-commerce : trois rôles (user, staff, admin), et pour les
# comptes "staff", un niveau d'accès par zone fonctionnelle.
#   - "admin"  : accès complet à tout, y compris la gestion des autres
#                membres staff (équivalent "Owner" chez Base44).
#   - "staff"  : accès limité aux zones et niveaux qui lui sont accordés.
#   - "user"   : client normal, aucun accès admin.
# Niveaux par zone : none < view < manage (manage inclut view).
# ---------------------------------------------------------------------------
STAFF_AREAS = [
    "orders", "orders_reopen", "products", "coupons", "customers", "subscribers",
    "shipping", "dashboard", "settings", "categories", "menus", "affiliates",
    "seo", "emails", "audit", "trash", "staff",
]
_PERMISSION_ORDER = {"none": 0, "view": 1, "manage": 2}


def _has_area_permission(user: dict, area: str, level: str = "view") -> bool:
    if user.get("role") == "admin":
        return True
    if user.get("role") != "staff":
        return False
    perms = user.get("permissions") or {}
    have = _PERMISSION_ORDER.get(perms.get(area, "none"), 0)
    need = _PERMISSION_ORDER.get(level, 1)
    return have >= need


def require_area(area: str, level: str = "view"):
    """Dépendance FastAPI paramétrée : autorise si role == admin (accès total),
    ou si role == staff ET permissions[area] >= level demandé.
    Journalise automatiquement toute action de niveau "manage" (= mutation) —
    couvre les 35 endpoints admin existants sans qu'il faille les modifier
    un par un."""
    async def _dep(request: Request, user: dict = Depends(get_current_user)) -> dict:
        allowed = _has_area_permission(user, area, level)
        if not allowed:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        if level == "manage":
            asyncio.create_task(_log_action(
                user, action=f"{request.method} {request.url.path}", area=area,
            ))
        return user
    return _dep


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
class RegisterIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=8)
    # Prénom et nom séparés et OBLIGATOIRES. Le `required` du formulaire est une
    # commodité d'interface, pas une garantie : la contrainte doit tenir ici,
    # où elle ne peut pas être contournée. min_length=1 seul laisserait passer
    # une chaîne d'espaces, d'où le validateur.
    first_name: str = Field(min_length=1, max_length=60)
    last_name: str = Field(min_length=1, max_length=60)
    website: str = ""  # honeypot anti-bot — doit rester vide

    @field_validator("first_name", "last_name")
    @classmethod
    def _non_vide(cls, v: str) -> str:
        v = (v or "").strip()
        if not v:
            raise ValueError("Prénom et nom sont obligatoires.")
        return v

    @property
    def name(self) -> str:
        """Nom complet, recomposé. Le reste du code (courriels, en-têtes de
        commande, affichage admin) attend encore ce champ unique."""
        return f"{self.first_name} {self.last_name}".strip()

    @field_validator("password")
    @classmethod
    def _validate_password(cls, v: str) -> str:
        return _validate_password_strength(v)


class LoginIn(BaseModel):
    email: EmailStr
    password: str


class UserOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    email: str
    name: str
    role: str
    created_at: str


class ProductVariant(BaseModel):
    id: Optional[str] = None  # generated server-side if missing
    name: str  # "5mg", "10mg", "500mcg"
    price: float
    stock: int = 0
    sku: str = ""
    # COA status — single source of truth per variant (replaces the two legacy
    # booleans badge_coa_available / badge_coa_pending which could contradict).
    #   "available" : a COA exists for this lot (coa_url should be set).
    #   "pending"   : COA not ready yet but coming — shown for transparency.
    #                 The variant stays purchasable, with a visible warning.
    #   "none"      : nothing shown.
    coa_status: str = "none"  # "available" | "pending" | "none"
    # Legacy booleans kept for backward-compat / rollback; no longer read for display.
    badge_coa_available: bool = False
    badge_coa_pending: bool = False
    badge_coming_soon: bool = False
    coa_url: str = ""
    sale_price: Optional[float] = None  # special/discount price (must be < price to apply)
    preorder_enabled: bool = False
    preorder_delay_message: str = ""
    preorder_price: Optional[float] = None
    preorder_note: str = ""
    weight_grams: float = 50.0  # used to estimate parcel weight for live Canada Post rating


class ProductIn(BaseModel):
    slug: str
    name_en: str
    name_fr: str
    category: str  # healing | gh-secretagogues | weight-loss | cognitive | longevity
    sequence: Optional[str] = ""
    purity: str = "≥ 99%"
    dosage_mg: float = 0.0  # informational only — variants drive actual pricing/stock
    description_en: str
    description_fr: str
    price_cad: float = 0.0  # legacy/fallback (= price of first variant)
    stock: int = 0  # legacy/fallback (= total across variants)
    low_stock_threshold: int = 10
    image_url: str = ""
    lab_tested: bool = True
    active: bool = True
    featured: bool = False
    preorder_allowed: bool = False  # legacy product-level (variants override)
    coa_url: Optional[str] = ""
    coa_lot: Optional[str] = ""
    coa_date: Optional[str] = ""
    # --- SCIENTIFIC DATA (10 new fields) ---
    molecular_formula: str = ""        # ex. "C62H98N16O22"
    molecular_weight: Optional[float] = None   # ex. 1419.5 (Da)
    cas_number: str = ""               # ex. "137525-51-0"
    sequence_length: Optional[int] = None # ex. 31 (derivable from sequence but explicit here)
    storage: str = ""                  # ex. "-20°C, à l'abri de la lumière"
    solubility: str = ""               # ex. "Soluble dans l'eau, acide acétique 0,1%"
    appearance: str = ""               # ex. "Poudre lyophilisée blanche"
    mechanism: str = ""                # ex. "Potentialise la libération d'hormone de croissance"
    research_areas: List[str] = []     # ex. ["Réparation tissulaire", "Anti-inflammatoire"]
    synonyms: List[str] = []           # ex. ["Sermorelin", "GRF 1-29"]
    meta_title_en: str = ""
    meta_title_fr: str = ""
    meta_description_en: str = ""
    meta_description_fr: str = ""
    og_image_url: str = ""
    images: List[str] = []  # galerie : image_url reste la couverture (images[0] ou saisie manuelle)
    variants: List[ProductVariant] = []


class ProductOut(ProductIn):
    id: str
    created_at: str


class CartItem(BaseModel):
    product_id: str
    variant_id: Optional[str] = None
    qty: int = Field(ge=1, le=1000)


class ShippingAddress(BaseModel):
    full_name: str = Field(min_length=1, max_length=120)
    address1: str = Field(min_length=1, max_length=160)
    address2: Optional[str] = Field(default="", max_length=160)
    city: str = Field(min_length=1, max_length=100)
    province: str = Field(min_length=2, max_length=64)  # QC, ON, BC, AB, ...
    postal_code: str = Field(min_length=3, max_length=20)
    country: str = "CA"
    phone: Optional[str] = Field(default="", max_length=32)


class CheckoutIn(BaseModel):
    items: List[CartItem]
    shipping: ShippingAddress
    email: Optional[EmailStr] = None  # guest email; auth user's email is used if logged in
    payment_method: Literal["interac", "nowpayments"]
    pay_currency: Optional[str] = "btc"  # used only for nowpayments
    coupon_code: Optional[str] = None
    idempotency_key: Optional[str] = None  # optionally passed in body to avoid CORS preflight from a custom header
    accept_terms: bool
    confirm_age: bool
    confirm_research_use: bool


class CouponIn(BaseModel):
    code: str
    discount_type: Literal["percent", "fixed"]
    value: float = Field(gt=0)  # percent 1-100 or absolute CAD
    min_subtotal: float = 0.0
    usage_limit: Optional[int] = None  # None = unlimited
    active: bool = True
    expires_at: Optional[str] = None  # ISO string
    # --- Avance (tous optionnels ; vides/desactives = comportement de base) ---
    start_at: Optional[str] = None
    allowed_emails: List[str] = []
    per_customer_limit: Optional[int] = None
    first_order_only: bool = False
    max_discount_cad: Optional[float] = None
    restrict_products: List[str] = []
    restrict_categories: List[str] = []


class OrderNoteIn(BaseModel):
    text: str = Field(min_length=1)
    visible_to_customer: bool = False


class RefundIn(BaseModel):
    amount: float = Field(gt=0)


class ShippingInfoIn(BaseModel):
    carrier: Optional[str] = ""
    tracking_number: Optional[str] = ""
    shipped_at: Optional[str] = None  # ISO; defaults to now() if not provided


class CreateLabelIn(BaseModel):
    service_code: str = Field(min_length=1, max_length=40)  # ex. "DOM.EP"


class StockAdjustIn(BaseModel):
    delta: int  # positive to add, negative to subtract


class StockRestockDeltaIn(BaseModel):
    variant_id: Optional[str] = None   # None = ligne produit legacy sans variantes
    # Positif : restock. Négatif : ajustement inventaire (perte, casse, retour).
    # 0 est refusé pour éviter les no-ops silencieux.
    quantity: int = Field(ge=-100000, le=100000)
    note: Optional[str] = None

    @field_validator("quantity")
    @classmethod
    def _not_zero(cls, v: int) -> int:
        if v == 0:
            raise ValueError("quantity must be non-zero (positive to add, negative to remove)")
        return v


class StockRestockIn(BaseModel):
    deltas: List[StockRestockDeltaIn] = Field(min_length=1, max_length=200)
    reason: Optional[str] = None   # optionnel : "réception fournisseur", "casse", "retour client", etc.


class StockBulkRestockRowIn(BaseModel):
    # Identifie la variante par SKU (préféré) OU par (product_slug + variant_name).
    sku: Optional[str] = None
    product_slug: Optional[str] = None
    variant_name: Optional[str] = None
    quantity: int = Field(ge=-100000, le=100000)
    note: Optional[str] = None

    @field_validator("quantity")
    @classmethod
    def _not_zero_bulk(cls, v: int) -> int:
        if v == 0:
            raise ValueError("quantity must be non-zero")
        return v


class StockBulkRestockIn(BaseModel):
    rows: List[StockBulkRestockRowIn] = Field(min_length=1, max_length=500)
    reason: Optional[str] = None


class StockNotifyIn(BaseModel):
    email: EmailStr
    product_id: str
    variant_id: Optional[str] = None
    website: str = ""  # honeypot anti-bot — doit rester vide


class GuestOrderAccessIn(BaseModel):
    email: EmailStr


# --- Catégories (BLOC 1) ----------------------------------------------------
# Les produits référencent une catégorie par SLUG (string), pas par id : aucune
# migration de données sur les produits existants.
_SLUG_RE = re.compile(r"^[a-z0-9-]+$")


class CategoryIn(BaseModel):
    slug: str = Field(min_length=1, max_length=60)
    name_en: str = Field(min_length=1, max_length=120)
    name_fr: str = Field(min_length=1, max_length=120)
    published: bool = True
    display_order: int = 0

    @field_validator("slug")
    @classmethod
    def _check_slug(cls, v: str) -> str:
        v = v.strip().lower()
        if not _SLUG_RE.match(v):
            raise ValueError("slug must match ^[a-z0-9-]+$")
        return v


class CategoryOut(CategoryIn):
    id: str
    created_at: str


# --- Menus (BLOC 2) ---------------------------------------------------------
class MenuItemIn(BaseModel):
    id: Optional[str] = None  # généré côté serveur si absent
    label_en: str = Field(min_length=1, max_length=120)
    label_fr: str = Field(min_length=1, max_length=120)
    url: str = Field(min_length=1, max_length=500)
    published: bool = True
    display_order: int = 0
    open_new_tab: bool = False


class MenuIn(BaseModel):
    slug: str = Field(min_length=1, max_length=60)
    name_en: str = Field(min_length=1, max_length=120)
    name_fr: str = Field(min_length=1, max_length=120)
    location: Literal["header", "footer"]
    published: bool = True
    display_order: int = 0
    items: List[MenuItemIn] = []

    @field_validator("slug")
    @classmethod
    def _check_slug(cls, v: str) -> str:
        v = v.strip().lower()
        if not _SLUG_RE.match(v):
            raise ValueError("slug must match ^[a-z0-9-]+$")
        return v


class MenuOut(MenuIn):
    id: str
    created_at: str


class NewsletterSubscribeIn(BaseModel):
    email: EmailStr
    consent: Literal[True]
    lang: Optional[Literal["en", "fr"]] = "en"
    source: Optional[str] = "footer"  # d'où vient l'inscription (footer, popup, checkout…)
    website: str = ""  # honeypot anti-bot — doit rester vide


# --- Mon Compte -------------------------------------------------------------
class ProfileUpdateIn(BaseModel):
    model_config = ConfigDict(extra="forbid")

    name: str = Field(min_length=1, max_length=120)


class PasswordChangeIn(BaseModel):
    current_password: Optional[str] = None
    new_password: str = Field(min_length=8)

    @field_validator("new_password")
    @classmethod
    def _validate_new_password(cls, v: str) -> str:
        return _validate_password_strength(v)


class EmailChangeRequestIn(BaseModel):
    new_email: EmailStr
    current_password: Optional[str] = None


class SavedAddressIn(BaseModel):
    label: Optional[str] = ""  # "Maison", "Labo", …
    full_name: str
    address1: str
    address2: Optional[str] = ""
    city: str
    province: str
    postal_code: str
    country: str = "CA"
    phone: Optional[str] = ""
    is_default: bool = False


class AccountDeleteIn(BaseModel):
    current_password: Optional[str] = None


class ShippingZoneIn(BaseModel):
    name: str
    countries: List[str] = []  # e.g., ["CA"], ["US"], ["INTL"]
    provinces: List[str] = []  # optional sub-region restriction (Canadian provinces)


class ShippingMethodIn(BaseModel):
    zone_id: str
    name: str  # e.g., "Canada Post Xpresspost", "Expedited", "International Tracked"
    cost_cad: float
    eta_days: str = ""  # e.g., "2-3 business days"
    active: bool = True


class ShippingRateRequest(BaseModel):
    postal_code: str
    country: str = "CA"
    items: Optional[List[CartItem]] = None  # used to estimate parcel weight; omit for a default estimate


# ---------------------------------------------------------------------------
# Tax & shipping (no tax — shipping flat-rate)
# ---------------------------------------------------------------------------
PROVINCES_CA = ["AB", "BC", "MB", "NB", "NL", "NS", "NT", "NU", "ON", "PE", "QC", "SK", "YT"]


# ---------------------------------------------------------------------------
# Auth endpoints
# ---------------------------------------------------------------------------
async def register(payload: RegisterIn, response: Response, request: Request):
    await _rate_limit("register", _client_ip(request), 5, 3600, "Too many registration attempts. Try again later.")
    email = payload.email.lower().strip()
    name = payload.name.strip()

    # Honeypot : les robots remplissent ce champ, les humains non.
    if (payload.website or "").strip():
        return {"ok": True, "verification_sent": True}

    if not name:
        raise HTTPException(status_code=400, detail="Le nom est requis.")
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=409, detail="Email already registered")

    user_doc = {
        "id": str(uuid.uuid4()),
        "email": email,
        # Les deux formes coexistent volontairement : « name » reste la source
        # d'affichage pour tout le code existant (courriels, commandes, admin),
        # tandis que first_name permet de s'adresser à la personne par son
        # prénom sans deviner où couper un nom complet.
        "name": name,
        "first_name": payload.first_name,
        "last_name": payload.last_name,
        "password_hash": hash_password(payload.password),
        "role": "user",
        "token_version": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "email_verified": False,
    }
    try:
        await db.users.insert_one(user_doc)
    except DuplicateKeyError as exc:
        raise HTTPException(status_code=409, detail="Email already registered") from exc

    try:
        await db.subscribers.update_one({"email": email}, {"$set": {"converted": True}})
    except Exception as e:  # pragma: no cover
        logging.warning("subscriber conversion flag failed ref=%s error_type=%s", _private_ref(email), type(e).__name__)

    raw = await _issue_magic_token(email, name, is_signup=True, lang="fr", ip=_client_ip(request))
    base = _trusted_public_base_url()
    link = f"{base}/auth/callback?token={raw}"
    await _send_magic_email(email, link, "fr", is_signup=True)

    return {
        "id": user_doc["id"],
        "email": email,
        "name": user_doc["name"],
        "role": "user",
        "created_at": user_doc["created_at"],
        "email_verified": False,
        "verification_sent": True,
    }


_LOGIN_MAX_ATTEMPTS = 5
_LOGIN_WINDOW_SECONDS = 900


def _client_ip(request: Request) -> str:
    peer = request.client.host if request.client else "unknown"
    if not _is_trusted_proxy_peer(request):
        return peer

    # Cloudflare : ce header n'est fiable que si le pair immédiat est un
    # reverse-proxy explicitement approuvé; sinon un client direct peut le forger.
    cf_ip = (request.headers.get("cf-connecting-ip") or "").strip()
    if cf_ip:
        try:
            ipaddress.ip_address(cf_ip)
            return cf_ip
        except ValueError:
            pass

    xff = request.headers.get("x-forwarded-for", "").strip()
    if not xff:
        return peer
    first = xff.split(",")[0].strip()
    if not first:
        return peer
    try:
        ipaddress.ip_address(first)
    except ValueError:
        return peer
    return first


# Public abuse protections for launch-day stability.
CHECKOUT_MAX_PER_MINUTE = int(os.environ.get("CHECKOUT_MAX_PER_MINUTE", "20"))
WEBHOOK_MAX_PER_MINUTE = int(os.environ.get("WEBHOOK_MAX_PER_MINUTE", "120"))
PUBLIC_MUTATION_MAX_PER_MINUTE = int(os.environ.get("PUBLIC_MUTATION_MAX_PER_MINUTE", "120"))


async def _rate_limit_distributed(bucket: str, key: str, max_hits: int, window_seconds: int, detail: str):
    now_ts = datetime.now(timezone.utc).timestamp()
    slot = int(now_ts // window_seconds)
    doc_id = f"{bucket}:{key}:{window_seconds}:{slot}"
    expires_at = datetime.now(timezone.utc) + timedelta(seconds=max(window_seconds * 2, 120))
    try:
        doc = await db.rate_limit_counters.find_one_and_update(
            {"_id": doc_id},
            {
                "$inc": {"count": 1},
                "$setOnInsert": {
                    "bucket": bucket,
                    "key": key,
                    "window_seconds": window_seconds,
                    "slot": slot,
                    "expires_at": expires_at,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                },
            },
            upsert=True,
            return_document=ReturnDocument.AFTER,
        )
    except Exception as e:
        logging.error("[rate-limit] shared backend unavailable: %s", e)
        raise HTTPException(status_code=503, detail="Rate limiting service unavailable")

    if int((doc or {}).get("count", 0)) > max_hits:
        raise HTTPException(status_code=429, detail=detail)


async def _rate_limit(bucket: str, key: str, max_hits: int, window_seconds: int, detail: str):
    await _rate_limit_distributed(bucket, key, max_hits, window_seconds, detail)


async def _rate_limit_email(bucket: str, email: str, max_hits: int, window_seconds: int, detail: str):
    await _rate_limit(bucket, f"email:{email.lower().strip()}", max_hits, window_seconds, detail)


async def _cursor_all(cursor) -> list:
    """Materialize a cursor only when the endpoint contract requires all rows."""
    return [document async for document in cursor]


async def _register_webhook_event(provider: str, signature: str, raw_body: bytes) -> bool:
    """Deduplicate webhook deliveries to prevent replay side-effects.

    Returns True when this event is new, False when already seen.
    """
    body_hash = hashlib.sha256(raw_body).hexdigest()
    event_key = f"{provider}:{signature}:{body_hash}"
    now = datetime.now(timezone.utc)
    try:
        await db.webhook_events.insert_one(
            {
                "id": str(uuid.uuid4()),
                "event_key": event_key,
                "provider": provider,
                "created_at": now.isoformat(),
                "created_at_dt": now,
            }
        )
        return True
    except DuplicateKeyError:
        return False


# Implementation lives in services/nowpayments.py; re-exported so existing call
# sites (routers/, other server helpers) keep resolving these names here.
try:
    from services.nowpayments import (  # noqa: F401
        _verify_nowpayments_signature, _nowpayments_create, nowpayments_ipn, crypto_status,
        _refresh_np_jwt, NowPaymentsPayoutError, _np_auth_token, _np_create_payout,
        _np_verify_payout, _np_payout_status, nowpayments_payout_ipn,
    )
except ImportError:  # package-relative import (uvicorn backend.server:app)
    from backend.services.nowpayments import (  # noqa: F401
        _verify_nowpayments_signature, _nowpayments_create, nowpayments_ipn, crypto_status,
        _refresh_np_jwt, NowPaymentsPayoutError, _np_auth_token, _np_create_payout,
        _np_verify_payout, _np_payout_status, nowpayments_payout_ipn,
    )


class AdminGateIn(BaseModel):
    code: str


class TrashIdsIn(BaseModel):
    ids: List[str] = Field(min_length=1, max_length=200)


# ---------------------------------------------------------------------------
# Corbeille générique — même principe pour toutes les données supprimables :
# suppression douce (deleted_at posé, rien n'est perdu), restauration en un
# clic, purge automatique après 30 jours (SAUF les commandes — voir note).
#
# ⚠️ Les commandes ne sont PAS purgées automatiquement : une commande payée
# est une pièce comptable. Elle reste en corbeille indéfiniment jusqu'à une
# purge manuelle explicite par un owner. Tous les autres types (produits,
# coupons, zones/méthodes de livraison) suivent la règle des 30 jours.
# ---------------------------------------------------------------------------
TRASH_RESOURCES = {
    "products":         {"collection": "products",         "area": "products",  "has_active": True,  "auto_purge_days": 30},
    "coupons":          {"collection": "coupons",           "area": "coupons",   "has_active": True,  "auto_purge_days": 30},
    "shipping_zones":   {"collection": "shipping_zones",    "area": "shipping",  "has_active": False, "auto_purge_days": 30},
    "shipping_methods": {"collection": "shipping_methods",  "area": "shipping",  "has_active": True,  "auto_purge_days": 30},
    "shipping_boxes":   {"collection": "shipping_boxes",    "area": "shipping",  "has_active": True,  "auto_purge_days": 30},
    "orders":           {"collection": "orders",            "area": "orders",    "has_active": False, "auto_purge_days": None},
}


async def _soft_delete(resource: str, item_id: str, admin: dict) -> dict:
    cfg = TRASH_RESOURCES[resource]
    coll = getattr(db, cfg["collection"])
    doc = await coll.find_one({"id": item_id})
    if not doc:
        raise HTTPException(status_code=404, detail=f"{resource[:-1].capitalize()} not found")
    update = {"deleted_at": datetime.now(timezone.utc).isoformat(), "deleted_by": admin.get("id")}
    if cfg["has_active"]:
        update["_pre_delete_active"] = doc.get("active", True)
        update["active"] = False
    await coll.update_one({"id": item_id}, {"$set": update})
    await _log_action(admin, f"{resource}.delete", area=cfg["area"], detail=f"Moved {resource[:-1]} {item_id} to trash")
    return {"ok": True}


async def admin_list_trash(resource: str, admin: dict = Depends(get_admin_user)):
    if resource not in TRASH_RESOURCES:
        raise HTTPException(status_code=404, detail="Unknown resource")
    coll = getattr(db, TRASH_RESOURCES[resource]["collection"])
    items = await coll.find({"deleted_at": {"$ne": None}}, {"_id": 0}).sort("deleted_at", -1).to_list(500)
    return items


async def admin_restore_trash(resource: str, payload: TrashIdsIn, admin: dict = Depends(get_admin_user)):
    if resource not in TRASH_RESOURCES:
        raise HTTPException(status_code=404, detail="Unknown resource")
    cfg = TRASH_RESOURCES[resource]
    coll = getattr(db, cfg["collection"])
    restored = 0
    for item_id in payload.ids:
        doc = await coll.find_one({"id": item_id, "deleted_at": {"$ne": None}})
        if not doc:
            continue
        update = {}
        if cfg["has_active"]:
            update["active"] = doc.get("_pre_delete_active", True)
        unset = {"deleted_at": "", "deleted_by": "", "_pre_delete_active": ""}
        await coll.update_one({"id": item_id}, {"$set": update, "$unset": unset} if update else {"$unset": unset})
        restored += 1
    await _log_action(admin, f"{resource}.restore", area=cfg["area"], detail=f"Restored {restored} {resource}")
    return {"ok": True, "restored": restored}


async def admin_purge_trash(resource: str, payload: TrashIdsIn, admin: dict = Depends(get_admin_user)):
    """Suppression DÉFINITIVE et immédiate des éléments sélectionnés — action
    irréversible, y compris pour les commandes (purge manuelle uniquement)."""
    if resource not in TRASH_RESOURCES:
        raise HTTPException(status_code=404, detail="Unknown resource")
    cfg = TRASH_RESOURCES[resource]
    coll = getattr(db, cfg["collection"])
    res = await coll.delete_many({"id": {"$in": payload.ids}, "deleted_at": {"$ne": None}})
    await _log_action(admin, f"{resource}.purge", area=cfg["area"],
                      detail=f"Permanently deleted {res.deleted_count} {resource}")
    return {"ok": True, "purged": res.deleted_count}


async def _trash_auto_purge_watchdog():
    """Purge automatiquement après 30 jours — sauf les commandes (auto_purge_days=None).
    Tourne dans le même worker verrouillé que les autres tâches de fond."""
    while True:
        try:
            now = datetime.now(timezone.utc)
            for resource, cfg in TRASH_RESOURCES.items():
                if cfg["auto_purge_days"] is None:
                    continue
                cutoff = (now - timedelta(days=cfg["auto_purge_days"])).isoformat()
                coll = getattr(db, cfg["collection"])
                res = await coll.delete_many({"deleted_at": {"$ne": None, "$lt": cutoff}})
                if res.deleted_count:
                    logging.info("[trash] auto-purged %d %s older than %d days",
                                res.deleted_count, resource, cfg["auto_purge_days"])
        except Exception as e:
            logging.error("[trash] auto-purge watchdog error: %s", e)
        await asyncio.sleep(6 * 3600)  # toutes les 6h


class StaffPermissionsIn(BaseModel):
    orders: Literal["none", "view", "manage"] = "none"
    # Permission dédiée: autorise l'action sensible "Réouvrir" après paiement tardif.
    orders_reopen: Literal["none", "view", "manage"] = "none"
    products: Literal["none", "view", "manage"] = "none"
    coupons: Literal["none", "view", "manage"] = "none"
    customers: Literal["none", "view", "manage"] = "none"
    subscribers: Literal["none", "view", "manage"] = "none"
    shipping: Literal["none", "view", "manage"] = "none"
    dashboard: Literal["none", "view", "manage"] = "none"
    # Zones additionnelles — alignées sur la nav admin frontend pour qu'un
    # compte staff puisse réellement se voir accorder l'accès à chacune.
    categories: Literal["none", "view", "manage"] = "none"
    menus: Literal["none", "view", "manage"] = "none"
    affiliates: Literal["none", "view", "manage"] = "none"
    seo: Literal["none", "view", "manage"] = "none"
    emails: Literal["none", "view", "manage"] = "none"
    audit: Literal["none", "view", "manage"] = "none"
    trash: Literal["none", "view", "manage"] = "none"
    staff: Literal["none", "view", "manage"] = "none"


class StaffInviteIn(BaseModel):
    email: EmailStr
    name: str = Field(min_length=1, max_length=120)
    permissions: StaffPermissionsIn
    as_owner: bool = False  # invite directement comme admin (owner) — accès total


class StaffAcceptIn(BaseModel):
    token: str
    password: str = Field(min_length=8)

    @field_validator("password")
    @classmethod
    def _validate_password(cls, v: str) -> str:
        return _validate_password_strength(v)


async def admin_gate_verify(payload: AdminGateIn, request: Request):
    """Passerelle avant le login admin — voir ADMIN_GATE_CODE plus haut.
    Rate-limit distinct et plus strict que le login (pas de compte à protéger
    ici, juste dissuader le brute-force du code lui-même)."""
    if not ADMIN_GATE_CODE:
        return {"ok": True}  # passerelle désactivée si aucun code configuré
    await _rate_limit("admin_gate", _client_ip(request), 5, 3600, "Too many attempts. Try again later.")
    if not secrets.compare_digest(payload.code, ADMIN_GATE_CODE):
        raise HTTPException(status_code=403, detail="Invalid code")
    return {"ok": True}


async def admin_autologin(_admin: dict = Depends(get_admin_user)):
    """Autologin de la porte admin : si une session admin valide existe déjà
    (cookie httpOnly access_token), la porte est levée côté frontend sans
    ressaisir le code — AdminGate.jsx appelle cette route à l'ouverture."""
    return {"ok": True}


async def login(payload: LoginIn, response: Response, request: Request):
    email = payload.email.lower().strip()
    throttle_key = f"{_client_ip(request)}:{email}"
    await _rate_limit(
        "login", throttle_key, _LOGIN_MAX_ATTEMPTS, _LOGIN_WINDOW_SECONDS,
        "Too many login attempts. Try again in 15 minutes.",
    )
    await _rate_limit(
        "login_account", email, 20, _LOGIN_WINDOW_SECONDS,
        "Too many login attempts. Try again in 15 minutes.",
    )
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    # Comptes sans le champ (antérieurs) sont considérés vérifiés.
    if user.get("email_verified", True) is False:
        raise HTTPException(status_code=403, detail="Veuillez vérifier votre adresse email pour activer votre compte. / Please verify your email address to activate your account.")
    await _start_session(response, request, user)
    return {
        "id": user["id"],
        "email": user["email"],
        "name": user["name"],
        "role": user["role"],
        "created_at": user["created_at"],
    }


async def logout(response: Response, request: Request):
    raw_refresh = request.cookies.get("refresh_token")
    if raw_refresh:
        await db.refresh_sessions.update_one(
            {"token_hash": _hash_refresh_token(raw_refresh), "revoked_at": None},
            {"$set": {"revoked_at": datetime.now(timezone.utc), "revoke_reason": "logout"}},
        )
    clear_auth_cookie(response)
    return {"ok": True}


async def logout_all_devices(response: Response, user: dict = Depends(get_current_user)):
    """Révoque tous les tokens émis avant cet appel, sur tous les appareils.
    Utile après un changement de mot de passe suspect ou une perte d'appareil."""
    await db.users.update_one({"id": user["id"]}, {"$inc": {"token_version": 1}})
    await db.refresh_sessions.update_many(
        {"user_id": user["id"], "revoked_at": None},
        {"$set": {"revoked_at": datetime.now(timezone.utc), "revoke_reason": "logout_all"}},
    )
    clear_auth_cookie(response)
    return {"ok": True}


async def me(request: Request):
    """Returns the current user or null (200) for guests, instead of 401.
    Cette version évite les erreurs 401 bruyantes dans la console browser
    lorsque le frontend interroge la session au chargement pour un visiteur
    anonyme — c'est un pattern standard pour un endpoint 'who am I'."""
    if isinstance(request, dict) and request.get("id") and request.get("email"):
        return _public_user_payload(request)
    user = await _resolve_user(request)
    if not user:
        return None
    return _public_user_payload(user)


# ---------------------------------------------------------------------------
# Magic link natif — auth sans mot de passe sur NOTRE backend (pas Supabase).
# Cohabite avec le login/register classique. Cookie httpOnly identique.
# ---------------------------------------------------------------------------
MAGIC_TOKEN_TTL_MINUTES = 15
# Reglable par l'environnement : cinq demandes par heure protegent bien en
# production, mais rendent une seance de test impraticable — on epuise le quota
# avant d'avoir pu observer quoi que ce soit, et le blocage ressemble alors a la
# panne qu'on cherchait justement a diagnostiquer.
MAGIC_REQUEST_MAX = int(os.environ.get("MAGIC_REQUEST_MAX", "5"))
MAGIC_REQUEST_WINDOW = int(os.environ.get("MAGIC_REQUEST_WINDOW", "3600"))
MAGIC_VERIFY_MAX = 300
MAGIC_VERIFY_WINDOW = 3600


class MagicRequestIn(BaseModel):
    email: EmailStr
    name: Optional[str] = None
    # Renseignés à l'INSCRIPTION seulement (create=true). Une demande de
    # connexion n'a pas à les fournir : le compte existe déjà et les porte.
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    create: bool = False
    lang: str = "fr"
    website: str = ""  # honeypot anti-bot — doit rester vide


class ForgotPasswordIn(BaseModel):
    email: EmailStr
    lang: str = "fr"


class ResetPasswordIn(BaseModel):
    token: str = Field(min_length=10, max_length=200)
    password: str

    @field_validator("password")
    @classmethod
    def _pw(cls, v):
        return _validate_password_strength(v)


def _hash_magic_token(raw: str) -> str:
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


async def _issue_magic_token(email: str, name: str, is_signup: bool,
                             lang: str, ip: str,
                             first_name: str = "", last_name: str = "") -> str:
    raw = secrets.token_urlsafe(32)
    now = datetime.now(timezone.utc)
    await db.magic_tokens.insert_one({
        "id": str(uuid.uuid4()),
        "email": email,
        "token_hash": _hash_magic_token(raw),
        "is_signup": is_signup,
        "name": (name or "").strip()[:120],
        # Conservés séparément : le compte n'est créé qu'au clic sur le lien,
        # parfois bien après la saisie. Recomposer le prénom en coupant le nom
        # complet à cet instant-là se tromperait sur « Marie-Claude Saint-Jean ».
        "first_name": (first_name or "").strip()[:60],
        "last_name": (last_name or "").strip()[:60],
        "created_at": now.isoformat(),
        "expires_at": (now + timedelta(minutes=MAGIC_TOKEN_TTL_MINUTES)).isoformat(),
        "used_at": None,
        "ip": ip,
        "lang": lang,
    })
    return raw


def _hash_token(raw: str) -> str:
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _trusted_public_base_url() -> str:
    """Source unique des liens email; ne jamais utiliser request.base_url (Host header)."""
    base = (PUBLIC_BASE_URL or "").rstrip("/")
    if not base:
        raise HTTPException(503, "PUBLIC_BASE_URL is required for email links")
    return base


async def _send_magic_email(email: str, link: str, lang: str, is_signup: bool) -> None:
    fr = lang.startswith("fr")
    if fr:
        subject = "Votre lien de connexion Fironova"
        heading = "Bienvenue chez Fironova" if is_signup else "Connexion à Fironova"
        intro = ("Confirmez votre adresse pour activer votre compte."
                 if is_signup else "Cliquez pour vous connecter sans mot de passe.")
        cta = "Activer mon compte" if is_signup else "Me connecter"
        expiry = f"Ce lien expire dans {MAGIC_TOKEN_TTL_MINUTES} minutes et ne sert qu'une fois."
        ignore = "Si vous n'êtes pas à l'origine de cette demande, ignorez cet email."
    else:
        subject = "Your Fironova sign-in link"
        heading = "Welcome to Fironova" if is_signup else "Sign in to Fironova"
        intro = ("Confirm your address to activate your account."
                 if is_signup else "Click to sign in without a password.")
        cta = "Activate my account" if is_signup else "Sign me in"
        expiry = f"This link expires in {MAGIC_TOKEN_TTL_MINUTES} minutes and works only once."
        ignore = "If you didn't request this, you can safely ignore this email."
    html = f"""\
<div style="font-family:Inter,-apple-system,Segoe UI,sans-serif;max-width:520px;margin:0 auto;background:#F7FAFC;padding:40px 24px;">
  <div style="background:#0B2E4F;border-radius:20px 20px 0 0;padding:28px 32px;">
    <span style="font-family:'Space Grotesk',sans-serif;color:#F7FAFC;font-size:20px;font-weight:700;letter-spacing:-0.02em;">FIRONOVA</span>
    <span style="color:#00B8D4;font-size:20px;font-weight:700;"> ·</span>
  </div>
  <div style="background:#ffffff;border-radius:0 0 20px 20px;padding:36px 32px;border:1px solid #E2E8F0;border-top:none;">
    <h1 style="font-family:'Space Grotesk',sans-serif;color:#0B2E4F;font-size:24px;font-weight:700;margin:0 0 12px;">{heading}</h1>
    <p style="color:#334155;font-size:15px;line-height:1.6;margin:0 0 28px;">{intro}</p>
    <a href="{link}" style="display:inline-block;background:#00B8D4;color:#0B2E4F;font-weight:700;text-decoration:none;padding:14px 32px;border-radius:999px;font-size:15px;">{cta} &rarr;</a>
    <p style="color:#64748B;font-size:12px;line-height:1.6;margin:28px 0 0;font-family:'JetBrains Mono',monospace;">{expiry}</p>
    <p style="color:#94A3B8;font-size:12px;line-height:1.6;margin:8px 0 0;">{ignore}</p>
    <hr style="border:none;border-top:1px solid #E2E8F0;margin:24px 0 12px;">
    <p style="color:#94A3B8;font-size:11px;line-height:1.5;margin:0;">Produits destin&eacute;s &agrave; la recherche uniquement (RUO). R&eacute;serv&eacute; aux 18 ans et plus.<br>For Research Use Only. 18+ only.</p>
  </div>
</div>"""
    from_addr = MAGIC_SENDER_EMAIL or SENDER_EMAIL
    await _send_email(email, subject, html, from_email=from_addr)


async def magic_request(payload: MagicRequestIn, request: Request):
    """Émet un lien magique. Réponse uniforme : ne révèle jamais si l'email existe."""
    email = payload.email.lower().strip()
    await _rate_limit_email("magic_request", email, MAGIC_REQUEST_MAX,
                             MAGIC_REQUEST_WINDOW, "Trop de demandes. Réessayez plus tard.")

    # Honeypot : réponse neutre sans effet.
    if (payload.website or "").strip():
        return {"ok": True}

    existing = await db.users.find_one({"email": email})
    is_signup = payload.create and not existing
    # Login demandé sur un email inconnu -> réponse neutre, aucun email (anti-énumération).
    if not payload.create and not existing:
        # La réponse reste volontairement identique à un succès : la révéler
        # permettrait d'énumérer les comptes. Mais le SERVEUR doit le dire,
        # sinon ce cas est indiscernable d'une panne d'envoi — on a cherché
        # pendant des jours du côté du fournisseur un courriel qui n'avait
        # jamais été demandé.
        logging.warning(
            "[magic] aucune demande émise pour %s — aucun compte à cette "
            "adresse et create=false (réponse neutre côté client)",
            _private_ref(email),
        )
        return {"ok": True}

    if payload.create and existing:
        # Inscription demandée sur une adresse qui a déjà un compte. On
        # n'envoie RIEN : il n'y a pas de compte à activer, et un courriel
        # parlant d'activation ferait croire à la création d'un second compte
        # — impossible, l'index unique sur users.email l'interdit.
        #
        # Contrairement au cas « connexion sur adresse inconnue » juste
        # au-dessus, la réponse est ici explicite. Un formulaire d'INSCRIPTION
        # qui reste muet oblige la personne à attendre un courriel qui ne
        # viendra pas. Le compromis est assumé : cela confirme l'existence du
        # compte à qui saisit l'adresse, ce que fait la majorité des sites.
        logging.info("[magic] inscription refusée pour %s — compte déjà existant",
                     _private_ref(email))
        return {"ok": True, "existing": True}

    prenom = (payload.first_name or "").strip()
    nom = (payload.last_name or "").strip()
    if is_signup and (not prenom or not nom):
        # Obligation appliquée ici et non seulement dans le formulaire : le
        # `required` du navigateur se contourne, une requête directe non.
        raise HTTPException(400, "Prénom et nom sont obligatoires.")

    complet = f"{prenom} {nom}".strip() or (payload.name or "")
    raw = await _issue_magic_token(email, complet, is_signup,
                                   payload.lang or "fr", _client_ip(request),
                                   first_name=prenom, last_name=nom)
    base = _trusted_public_base_url()
    link = f"{base}/auth/callback?token={raw}"
    await _send_magic_email(email, link, payload.lang or "fr", is_signup)
    return {"ok": True}


async def magic_verify(response: Response, request: Request, token: str = Body(..., embed=True)):
    """Vérifie le token (usage unique + TTL), crée le compte si signup, pose le cookie."""
    await _rate_limit("magic_verify", _client_ip(request), MAGIC_VERIFY_MAX,
                       MAGIC_VERIFY_WINDOW, "Trop de tentatives. Réessayez plus tard.")
    token_hash = _hash_magic_token((token or "").strip())
    rec = await db.magic_tokens.find_one({"token_hash": token_hash})
    now = datetime.now(timezone.utc)
    invalid = HTTPException(status_code=400, detail="Lien invalide ou expiré.")
    if not rec or rec.get("used_at"):
        raise invalid
    try:
        if datetime.fromisoformat(rec["expires_at"]) < now:
            raise invalid
    except HTTPException:
        raise
    except Exception:
        raise invalid
    claimed = await db.magic_tokens.update_one(
        {"_id": rec["_id"], "used_at": None},
        {"$set": {"used_at": now.isoformat()}},
    )
    if claimed.modified_count != 1:
        raise invalid
    email = rec["email"]
    is_signup = bool(rec.get("is_signup"))
    user = await db.users.find_one({"email": email})
    if not user:
        user_doc = {
            "id": str(uuid.uuid4()),
            "email": email,
            "name": rec.get("name") or email.split("@")[0],
            "first_name": rec.get("first_name", ""),
            "last_name": rec.get("last_name", ""),
            "password_hash": hash_password(secrets.token_urlsafe(32)),
            "role": "user",
            "token_version": 0,
            "created_at": now.isoformat(),
            "passwordless": True,
            "email_verified": True,
        }
        await db.users.insert_one(user_doc)
        try:
            await db.subscribers.update_one({"email": email}, {"$set": {"converted": True}})
        except Exception as e:  # pragma: no cover
            logging.warning("subscriber conversion flag failed ref=%s error_type=%s", _private_ref(email), type(e).__name__)
        asyncio.create_task(welcome_new_user(email, user_doc["name"], "fr"))
        user = user_doc
    else:
        await db.users.update_one({"email": email}, {"$set": {"email_verified": True}})
        if is_signup:
            asyncio.create_task(welcome_new_user(email, user.get("name", ""), "fr"))
        user = {**user, "email_verified": True}
    await _start_session(response, request, user)
    return {
        "id": user["id"], "email": user["email"], "name": user["name"],
        "role": user["role"], "created_at": user["created_at"],
    }


# ---------------------------------------------------------------------------
# Réinitialisation de mot de passe (mot de passe oublié)
# ---------------------------------------------------------------------------
RESET_TOKEN_TTL_MINUTES = 30
RESET_REQUEST_MAX = 5
RESET_REQUEST_WINDOW = 3600


async def _send_reset_email(email: str, link: str, lang: str) -> None:
    fr = lang.startswith("fr")
    if fr:
        subject = "FIRONOVA — Réinitialisation du mot de passe"
        heading = "Réinitialiser votre mot de passe"
        body = "Vous avez demandé à réinitialiser votre mot de passe. Ce lien expire dans 30 minutes."
        cta = "Choisir un nouveau mot de passe"
        ignore = "Si vous n'êtes pas à l'origine de cette demande, ignorez cet email — votre mot de passe reste inchangé."
    else:
        subject = "FIRONOVA — Password reset"
        heading = "Reset your password"
        body = "You requested a password reset. This link expires in 30 minutes."
        cta = "Choose a new password"
        ignore = "If you didn't request this, ignore this email — your password stays unchanged."
    html = f"""<div style="font-family:Inter,Arial,sans-serif;max-width:480px;margin:0 auto;padding:32px 24px;color:#0B2E4F">
      <h1 style="font-size:20px;margin:0 0 12px">{heading}</h1>
      <p style="font-size:14px;line-height:1.6;color:#3E5C76;margin:0 0 24px">{body}</p>
      <a href="{link}" style="display:inline-block;background:#00B8D4;color:#0B2E4F;font-weight:600;text-decoration:none;padding:12px 28px;border-radius:9999px;font-size:14px">{cta}</a>
      <p style="font-size:12px;line-height:1.6;color:#7A8CA0;margin:24px 0 0">{ignore}</p>
      <p style="font-size:11px;color:#A0AEC0;margin:16px 0 0">For Research Use Only · Réservé à la recherche</p>
    </div>"""
    from_addr = MAGIC_SENDER_EMAIL or SENDER_EMAIL
    await _send_email(email, subject, html, from_email=from_addr)


async def forgot_password(payload: ForgotPasswordIn, request: Request):
    """Émet un lien de réinitialisation. Réponse uniforme (anti-énumération)."""
    email = payload.email.lower().strip()
    await _rate_limit_email("reset_request", email, RESET_REQUEST_MAX,
                             RESET_REQUEST_WINDOW, "Trop de demandes. Réessayez plus tard.")
    user = await db.users.find_one({"email": email})
    if user and not user.get("passwordless", False):
        raw = secrets.token_urlsafe(32)
        await db.magic_tokens.insert_one({
            "id": str(uuid.uuid4()),
            "email": email,
            "token_hash": _hash_magic_token(raw),
            "purpose": "reset",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=RESET_TOKEN_TTL_MINUTES)).isoformat(),
            "used_at": None,
            "ip": _client_ip(request),
        })
        base = _trusted_public_base_url()
        link = f"{base}/reset-password?token={raw}"
        await _send_reset_email(email, link, payload.lang or "fr")
    return {"ok": True}


async def reset_password(payload: ResetPasswordIn, response: Response, request: Request):
    token_hash = _hash_magic_token(payload.token.strip())
    rec = await db.magic_tokens.find_one({"token_hash": token_hash, "purpose": "reset"})
    if not rec:
        raise HTTPException(400, "Invalid or expired reset link")
    if rec.get("used_at"):
        raise HTTPException(400, "This reset link has already been used")
    expires_at = datetime.fromisoformat(rec["expires_at"])
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(400, "This reset link has expired")
    claimed = await db.magic_tokens.update_one(
        {"_id": rec["_id"], "used_at": None},
        {"$set": {"used_at": datetime.now(timezone.utc).isoformat()}},
    )
    if claimed.modified_count != 1:
        raise HTTPException(400, "This reset link has already been used")
    user = await db.users.find_one({"email": rec["email"]})
    if not user:
        raise HTTPException(400, "Invalid or expired reset link")
    new_hash = hash_password(payload.password)
    new_tv = user.get("token_version", 0) + 1
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"password_hash": new_hash, "token_version": new_tv, "passwordless": False}},
    )
    await db.magic_tokens.delete_many({"email": rec["email"], "purpose": "reset", "used_at": None})
    await db.refresh_sessions.update_many(
        {"user_id": user["id"], "revoked_at": None},
        {"$set": {"revoked_at": datetime.now(timezone.utc), "revoke_reason": "password_reset"}},
    )
    await _start_session(response, request, {**user, "token_version": new_tv})
    return {
        "id": user["id"], "email": user["email"], "name": user["name"],
        "role": user["role"], "created_at": user["created_at"],
    }


async def _magic_tokens_cleanup():
    while True:
        try:
            cutoff = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
            res = await db.magic_tokens.delete_many({"expires_at": {"$lt": cutoff}})
            if res.deleted_count:
                logging.info("[magic] purged %d expired tokens", res.deleted_count)
        except Exception as e:
            logging.error("[magic] cleanup error: %s", e)
        try:
            unverified_cutoff = (datetime.now(timezone.utc) - timedelta(hours=72)).isoformat()
            res2 = await db.users.delete_many({
                "email_verified": False,
                "created_at": {"$lte": unverified_cutoff},
            })
            if res2.deleted_count:
                logging.info("[magic] purged %d unactivated accounts", res2.deleted_count)
        except Exception as e:
            logging.error("[magic] unactivated account purge error: %s", e)
        await asyncio.sleep(6 * 3600)


# ---------------------------------------------------------------------------
# Mon Compte — profil, mot de passe, changement d'email (double opt-in),
# adresses sauvegardées, suppression de compte avec anonymisation.
# ---------------------------------------------------------------------------
EMAIL_CHANGE_TTL_HOURS = 24
GUEST_ORDER_ACCESS_TTL_MINUTES = 60 * 24 * 7


async def account_update_profile(payload: ProfileUpdateIn, user: dict = Depends(get_current_user)):
    name = payload.name.strip()
    await db.users.update_one({"id": user["id"]}, {"$set": {"name": name}})
    return {"ok": True, "name": name}


def _assert_current_password(doc: dict, provided: Optional[str]) -> None:
    """Vérifie le mot de passe actuel, sauf pour un compte passwordless où
    l'auth cookie (déjà exigée) fait foi."""
    if doc.get("passwordless", False):
        return
    if not provided or not verify_password(provided, doc["password_hash"]):
        raise HTTPException(status_code=403, detail="Current password is incorrect")


async def account_change_password(payload: PasswordChangeIn, response: Response, request: Request,
                                  user: dict = Depends(get_current_user)):
    await _rate_limit_email("account_password", user["email"], 5, 900,
                             "Too many password attempts. Try again later.")
    doc = await db.users.find_one({"id": user["id"]})
    if not doc:
        raise HTTPException(status_code=404, detail="Account not found")
    _assert_current_password(doc, payload.current_password)
    new_tv = doc.get("token_version", 0) + 1
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"password_hash": hash_password(payload.new_password),
                  "token_version": new_tv, "passwordless": False}},
    )
    await db.refresh_sessions.update_many(
        {"user_id": doc["id"], "revoked_at": None},
        {"$set": {"revoked_at": datetime.now(timezone.utc), "revoke_reason": "password_change"}},
    )
    await _start_session(response, request, {**doc, "token_version": new_tv})
    # Le cookie httpOnly frais suffit — pas de token dans le body.
    return {"ok": True}


def _email_change_html(confirm_url: str, lang: str = "fr") -> str:
    if lang == "fr":
        heading = "Confirmez votre nouvelle adresse email"
        body = ("Vous avez demandé à changer l'adresse email de votre compte FIRONOVA. "
                "Cliquez sur le bouton ci-dessous pour confirmer. "
                f"Ce lien expire dans {EMAIL_CHANGE_TTL_HOURS} heures. "
                "Si vous n'êtes pas à l'origine de cette demande, ignorez ce message.")
        btn = "Confirmer mon email"
    else:
        heading = "Confirm your new email address"
        body = ("You requested to change the email on your FIRONOVA account. "
                "Click the button below to confirm. "
                f"This link expires in {EMAIL_CHANGE_TTL_HOURS} hours. "
                "If you didn't request this, you can safely ignore this message.")
        btn = "Confirm my email"
    return f"""
    <div style="font-family:Georgia,serif;max-width:520px;margin:0 auto;padding:32px;color:#3A0A08;background:#FFFAF6">
      <div style="font-size:22px;font-weight:800;letter-spacing:-0.5px">FIRONOVA<span style="color:#C20114">.</span></div>
      <h2 style="margin-top:24px">{heading}</h2>
      <p style="line-height:1.6">{body}</p>
      <a href="{confirm_url}" style="display:inline-block;margin-top:16px;background:#3A0A08;color:#fff;
         padding:14px 28px;text-decoration:none;font-family:monospace;font-size:12px;
         letter-spacing:0.2em;text-transform:uppercase">{btn} →</a>
      <p style="margin-top:24px;font-size:12px;color:#6B0504">{confirm_url}</p>
    </div>
    """


async def account_request_email_change(payload: EmailChangeRequestIn, request: Request,
                                       user: dict = Depends(get_current_user)):
    await _rate_limit_email("email_change", user["email"], 3, 3600,
                             "Too many email change requests. Try again later.")
    doc = await db.users.find_one({"id": user["id"]})
    if not doc:
        raise HTTPException(status_code=404, detail="Account not found")
    _assert_current_password(doc, payload.current_password)
    new_email = payload.new_email.lower().strip()
    if new_email == doc["email"]:
        raise HTTPException(status_code=400, detail="This is already your email address")
    if await db.users.find_one({"email": new_email}):
        raise HTTPException(status_code=409, detail="Email already in use")

    token = secrets.token_urlsafe(32)
    now = datetime.now(timezone.utc)
    await db.email_change_requests.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "new_email": new_email,
        "token_hash": _hash_token(token),
        "created_at": now.isoformat(),
        "expires_at": (now + timedelta(hours=EMAIL_CHANGE_TTL_HOURS)).isoformat(),
        "used": False,
    })
    base = _trusted_public_base_url()
    confirm_url = f"{base}/api/account/email/confirm?token={token}"
    lang = "fr"  # les emails transactionnels suivent la langue du site ; FR par défaut au Québec
    asyncio.create_task(_send_email(new_email, "FIRONOVA — Confirmez votre nouvelle adresse email",
                                    _email_change_html(confirm_url, lang)))
    return {"ok": True, "sent_to": new_email}


async def account_confirm_email_change(token: str, request: Request):
    """Lien cliqué depuis l'email — pas d'authentification (l'utilisateur peut
    ouvrir le lien sur un autre appareil). Le token à usage unique + TTL fait
    office de preuve."""
    await _rate_limit(
        "email_change_confirm", _client_ip(request), 10, 3600,
        "Too many confirmation attempts. Try again later.",
    )
    now = datetime.now(timezone.utc).isoformat()
    token_hash = _hash_token(token)
    req = await db.email_change_requests.find_one({"token_hash": token_hash, "used": False})
    if not req or req["expires_at"] < now:
        raise HTTPException(status_code=400, detail="Invalid or expired confirmation link")
    # Re-vérifier l'unicité : l'email a pu être pris entre la demande et le clic.
    if await db.users.find_one({"email": req["new_email"]}):
        raise HTTPException(status_code=409, detail="Email already in use")
    user = await db.users.find_one({"id": req["user_id"]})
    if not user:
        raise HTTPException(status_code=404, detail="Account not found")

    await db.users.update_one(
        {"id": req["user_id"]},
        {"$set": {"email": req["new_email"]},
         "$inc": {"token_version": 1}},  # révoque les sessions liées à l'ancien email
    )
    await db.email_change_requests.update_one({"token_hash": token_hash}, {"$set": {"used": True, "used_at": now}})
    # Réponse HTML minimale : le lien est ouvert dans un navigateur, pas via l'app.
    return Response(
        content=f"""<!doctype html><html><body style="font-family:Georgia,serif;background:#FFFAF6;color:#3A0A08;
        display:flex;align-items:center;justify-content:center;height:100vh;margin:0">
        <div style="text-align:center"><div style="font-size:22px;font-weight:800">FIRONOVA<span style="color:#C20114">.</span></div>
        <h2>Email confirmé ✓ / Email confirmed ✓</h2>
        <p>Reconnectez-vous avec votre nouvelle adresse.<br/>Please sign in again with your new address.</p>
        </div></body></html>""",
        media_type="text/html",
    )


# --- Adresses sauvegardées ---------------------------------------------------
async def account_list_addresses(user: dict = Depends(get_current_user)):
    return await db.addresses.find({"user_id": user["id"]}, {"_id": 0}).sort("created_at", -1).to_list(20)


async def account_add_address(payload: SavedAddressIn, user: dict = Depends(get_current_user)):
    count = await db.addresses.count_documents({"user_id": user["id"]})
    if count >= 10:
        raise HTTPException(status_code=400, detail="Address limit reached (10)")
    doc = payload.model_dump()
    doc.update({
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    # Première adresse = défaut automatique ; sinon respecter le flag envoyé.
    if count == 0:
        doc["is_default"] = True
    elif doc["is_default"]:
        await db.addresses.update_many({"user_id": user["id"]}, {"$set": {"is_default": False}})
    await db.addresses.insert_one(doc)
    doc.pop("_id", None)
    return doc


async def account_update_address(address_id: str, payload: SavedAddressIn,
                                 user: dict = Depends(get_current_user)):
    existing = await db.addresses.find_one({"id": address_id, "user_id": user["id"]})
    if not existing:
        raise HTTPException(status_code=404, detail="Address not found")
    doc = payload.model_dump()
    if doc["is_default"]:
        await db.addresses.update_many({"user_id": user["id"]}, {"$set": {"is_default": False}})
    await db.addresses.update_one({"id": address_id, "user_id": user["id"]}, {"$set": doc})
    fresh = await db.addresses.find_one(
        {"id": address_id, "user_id": user["id"]},
        {"_id": 0},
    )
    return fresh


async def account_delete_address(address_id: str, user: dict = Depends(get_current_user)):
    res = await db.addresses.delete_one({"id": address_id, "user_id": user["id"]})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="Address not found")
    # Si on vient de supprimer l'adresse par défaut, promouvoir la plus récente.
    remaining_default = await db.addresses.find_one({"user_id": user["id"], "is_default": True})
    if not remaining_default:
        newest = await db.addresses.find_one({"user_id": user["id"]}, sort=[("created_at", -1)])
        if newest:
            await db.addresses.update_one(
                {"id": newest["id"], "user_id": user["id"]},
                {"$set": {"is_default": True}},
            )
    return {"ok": True}


# --- Suppression de compte (PIPEDA / Loi 25) ---------------------------------
async def account_delete(payload: AccountDeleteIn, response: Response,
                         user: dict = Depends(get_current_user)):
    """Suppression du compte + anonymisation des commandes.
    Les commandes sont CONSERVÉES (obligations comptables/fiscales) mais toutes
    les données personnelles y sont remplacées. Irréversible."""
    await _rate_limit_email("account_delete", user["email"], 3, 900,
                             "Too many account deletion attempts. Try again later.")
    doc = await db.users.find_one({"id": user["id"]})
    if not doc:
        raise HTTPException(status_code=404, detail="Account not found")
    _assert_current_password(doc, payload.current_password)
    if doc.get("role") == "admin":
        raise HTTPException(status_code=400, detail="Admin accounts cannot be self-deleted")

    now = datetime.now(timezone.utc).isoformat()
    anon = f"deleted-user-{doc['id'][:8]}"

    # 1. Anonymiser les commandes (on garde montants, items, dates — pas l'identité)
    await db.orders.update_many(
        {"user_id": doc["id"]},
        {"$set": {
            "email": f"{anon}@anonymized.invalid",
            "shipping_address.full_name": "[deleted]",
            "shipping_address.address1": "[deleted]",
            "shipping_address.address2": "",
            "shipping_address.city": "[deleted]",
            "shipping_address.province": "[deleted]",
            "shipping_address.postal_code": "[deleted]",
            "shipping_address.phone": "",
            "anonymized_at": now,
        }},
    )
    # 2. Purger les données annexes et tokens liés au compte
    await db.addresses.delete_many({"user_id": doc["id"]})
    await db.stock_notifications.delete_many({"email": doc["email"]})
    await db.subscribers.delete_one({"email": doc["email"]})
    await db.email_change_requests.delete_many({"user_id": doc["id"]})
    await db.magic_tokens.delete_many({"email": doc["email"]})
    await db.refresh_sessions.delete_many({"user_id": doc["id"]})
    await db.wishlist.delete_many({"user_id": doc["id"]})
    if hasattr(db, "coupons"):
        await db.coupons.update_many(
            {"used_by.email": doc["email"].lower().strip()},
            {"$pull": {"used_by": {"email": doc["email"].lower().strip()}}},
        )
    order_cursor = db.orders.find({"user_id": doc["id"]}, {"_id": 0, "id": 1})
    order_ids = [o.get("id") async for o in order_cursor if o.get("id")]
    if order_ids:
        await db.order_access_tokens.delete_many({"order_id": {"$in": order_ids}})
    # 3. Supprimer l'utilisateur
    await db.users.delete_one({"id": doc["id"]})
    clear_auth_cookie(response)
    logging.info("Account deleted + orders anonymized for user %s", anon)
    return {"ok": True}


# ---------------------------------------------------------------------------
# Product endpoints
# ---------------------------------------------------------------------------
async def list_products(category: Optional[str] = None, q: Optional[str] = None, featured: Optional[bool] = None):
    filt: dict = {"active": True}
    if category and category != "all":
        filt["category"] = category
    if featured is True:
        filt["featured"] = True
    if q:
        # Neutralise les métacaractères regex : sans ça, un `q` du type
        # "(a+)+$" est injecté tel quel dans MongoDB → ReDoS / scan complet.
        q_safe = re.escape(q.strip())
        filt["$or"] = [
            {"name_en": {"$regex": q_safe, "$options": "i"}},
            {"name_fr": {"$regex": q_safe, "$options": "i"}},
            {"slug": {"$regex": q_safe, "$options": "i"}},
        ]
    products = await db.products.find(filt, {"_id": 0}).sort("name_en", 1).to_list(500)
    return products


async def get_product(slug: str):
    product = await db.products.find_one({"slug": slug, "active": True}, {"_id": 0})
    if not product:
        raise HTTPException(404, "Product not found")
    return product


async def notify_stock_request(payload: StockNotifyIn, request: Request):
    await _rate_limit("notify_stock", _client_ip(request), 10, 3600,
                       "Too many requests. Try again later.")
    """Public endpoint — a visitor asks to be emailed once a sold-out product/variant is back."""
    if (payload.website or "").strip():
        return {"ok": True}
    product = await db.products.find_one({"id": payload.product_id}, {"_id": 0})
    if not product:
        raise HTTPException(404, "Product not found")
    variant_id = payload.variant_id
    if variant_id and not any(v.get("id") == variant_id for v in product.get("variants", [])):
        raise HTTPException(400, "Variant not found for this product")
    email = payload.email.lower().strip()
    await _rate_limit_email("notify_stock", email, 3, 3600,
                             "Too many stock notifications. Try again later.")
    existing = await db.stock_notifications.find_one({
        "email": email, "product_id": payload.product_id, "variant_id": variant_id, "notified": False,
    })
    if existing:
        return {"ok": True, "already_subscribed": True}
    await db.stock_notifications.insert_one({
        "id": str(uuid.uuid4()),
        "email": email,
        "product_id": payload.product_id,
        "variant_id": variant_id,
        "notified": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    return {"ok": True, "already_subscribed": False}


# ---------------------------------------------------------------------------
# BLOC 1 — Catégories : collection réelle + CRUD admin + publication.
# Les produits continuent de référencer une catégorie par slug (string) : aucune
# migration produit, aucune rupture des 12 produits seedés.
# ---------------------------------------------------------------------------
async def list_categories():
    """Vitrine : uniquement les catégories publiées."""
    return await db.categories.find({"published": True}, {"_id": 0}).sort("display_order", 1).to_list(200)


async def admin_list_categories(_admin: dict = Depends(get_admin_user)):
    return await db.categories.find({}, {"_id": 0}).sort("display_order", 1).to_list(200)


async def admin_create_category(payload: CategoryIn, _admin: dict = Depends(get_admin_user)):
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    try:
        await db.categories.insert_one(dict(doc))
    except DuplicateKeyError:
        raise HTTPException(409, f"A category with slug '{doc['slug']}' already exists")
    doc.pop("_id", None)
    return doc


async def admin_update_category(cat_id: str, payload: CategoryIn, _admin: dict = Depends(get_admin_user)):
    existing = await db.categories.find_one({"id": cat_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Category not found")
    new = payload.model_dump()
    old_slug, new_slug = existing["slug"], new["slug"]

    if new_slug != old_slug and await db.categories.find_one({"slug": new_slug, "id": {"$ne": cat_id}}):
        raise HTTPException(409, f"A category with slug '{new_slug}' already exists")

    await db.categories.update_one({"id": cat_id}, {"$set": new})

    # Cascade : renommer un slug ne doit jamais orpheliner les produits.
    migrated = 0
    if new_slug != old_slug:
        res = await db.products.update_many({"category": old_slug}, {"$set": {"category": new_slug}})
        migrated = res.modified_count

    out = {**existing, **new}
    out["products_migrated"] = migrated
    return out


async def admin_delete_category(cat_id: str, _admin: dict = Depends(get_admin_user)):
    cat = await db.categories.find_one({"id": cat_id}, {"_id": 0})
    if not cat:
        raise HTTPException(404, "Category not found")
    in_use = await db.products.count_documents({"category": cat["slug"]})
    if in_use:
        # Même principe que le soft-delete produit lié à des commandes : on masque.
        await db.categories.update_one({"id": cat_id}, {"$set": {"published": False}})
        raise HTTPException(
            409,
            f"Category in use by {in_use} product(s); it has been hidden instead of deleted.",
        )
    await db.categories.delete_one({"id": cat_id})
    return {"ok": True, "deleted": True}


# ---------------------------------------------------------------------------
# BLOC 2 — Menus : navigation header/footer pilotée depuis l'admin.
# ---------------------------------------------------------------------------
def _normalize_menu_items(items: list) -> list:
    out = []
    for it in items:
        d = dict(it)
        if not d.get("id"):
            d["id"] = str(uuid.uuid4())
        out.append(d)
    return out


async def list_menus(location: Optional[str] = None):
    """Vitrine : menus publiés, items non publiés retirés."""
    filt: dict = {"published": True}
    if location:
        filt["location"] = location
    menus = await db.menus.find(filt, {"_id": 0}).sort("display_order", 1).to_list(50)
    for m in menus:
        m["items"] = sorted(
            [i for i in m.get("items", []) if i.get("published", True)],
            key=lambda i: i.get("display_order", 0),
        )
    return menus


async def admin_list_menus(_admin: dict = Depends(get_admin_user)):
    return await db.menus.find({}, {"_id": 0}).sort("display_order", 1).to_list(50)


async def admin_create_menu(payload: MenuIn, _admin: dict = Depends(get_admin_user)):
    doc = payload.model_dump()
    doc["items"] = _normalize_menu_items(doc.get("items") or [])
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    try:
        await db.menus.insert_one(dict(doc))
    except DuplicateKeyError:
        raise HTTPException(409, f"A menu with slug '{doc['slug']}' already exists")
    doc.pop("_id", None)
    return doc


async def admin_update_menu(menu_id: str, payload: MenuIn, _admin: dict = Depends(get_admin_user)):
    existing = await db.menus.find_one({"id": menu_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Menu not found")
    new = payload.model_dump()
    if new["slug"] != existing["slug"] and await db.menus.find_one({"slug": new["slug"], "id": {"$ne": menu_id}}):
        raise HTTPException(409, f"A menu with slug '{new['slug']}' already exists")
    # Les ids d'items existants sont conservés ; seuls les nouveaux en reçoivent un.
    new["items"] = _normalize_menu_items(new.get("items") or [])
    await db.menus.update_one({"id": menu_id}, {"$set": new})
    return {**existing, **new}


async def admin_delete_menu(menu_id: str, _admin: dict = Depends(get_admin_user)):
    res = await db.menus.delete_one({"id": menu_id})
    if not res.deleted_count:
        raise HTTPException(404, "Menu not found")
    return {"ok": True, "deleted": True}


# ---------------------------------------------------------------------------
# BLOC 3 — Prélancement : vérification du jeton d'aperçu.
# Le jeton n'est JAMAIS exposé via /meta — seule cette comparaison serveur existe.
# ---------------------------------------------------------------------------
async def prelaunch_preview(token: str, request: Request):
    await _rate_limit("prelaunch_preview", _client_ip(request), 20, 3600,
                       "Too many attempts. Try again later.")
    if not PRELAUNCH_PREVIEW_TOKEN:
        return {"ok": False}
    return {"ok": hmac.compare_digest(token or "", PRELAUNCH_PREVIEW_TOKEN)}


# ---------------------------------------------------------------------------
# Newsletter / subscribers — conforme LCAP (CASL) : on conserve la preuve de
# consentement (date, IP, source) pour chaque inscription, et un lien de
# désabonnement fonctionne sans authentification (obligatoire par la loi).
# ---------------------------------------------------------------------------
async def newsletter_subscribe(payload: NewsletterSubscribeIn, request: Request):
    await _rate_limit("newsletter_subscribe", _client_ip(request), 10, 3600,
                       "Too many requests. Try again later.")
    email = payload.email.lower().strip()
    if (payload.website or "").strip():
        return {"ok": True, "already_subscribed": False}
    now = datetime.now(timezone.utc)
    now_iso = now.isoformat()
    ip = _client_ip(request)
    lang = (payload.lang or "en").lower()
    if lang not in ("en", "fr"):
        lang = "en"

    existing = await db.subscribers.find_one({"email": email}, {"_id": 0})
    if existing and existing.get("status") == "subscribed":
        return {"ok": True, "already_subscribed": True}

    # Double-opt-in : on n'active pas immédiatement. On génère un token de
    # confirmation valable 7 jours (RGPD/CASL) et on envoie l'email.
    confirmation_token = secrets.token_urlsafe(32)
    confirmation_expires_at = (now + timedelta(days=7)).isoformat()

    if existing:
        await db.subscribers.update_one(
            {"email": email},
            {"$set": {
                "status": "pending_confirmation",
                "lang": lang,
                "source": payload.source,
                "consent_at": now_iso,
                "consent_ip": ip,
                "unsubscribed_at": None,
                "confirmation_token": confirmation_token,
                "confirmation_expires_at": confirmation_expires_at,
                "confirmed_at": None,
            }},
        )
    else:
        await db.subscribers.insert_one({
            "id": str(uuid.uuid4()),
            "email": email,
            "lang": lang,
            "source": payload.source,
            "status": "pending_confirmation",
            "consent_at": now_iso,
            "consent_ip": ip,
            "unsubscribe_token": str(uuid.uuid4()),
            "unsubscribed_at": None,
            "confirmation_token": confirmation_token,
            "confirmation_expires_at": confirmation_expires_at,
            "confirmed_at": None,
            "created_at": now_iso,
            "converted": False,
        })

    # Envoi email de confirmation (double-opt-in RGPD/CASL) — bilingue.
    confirm_url = f"{PUBLIC_BASE_URL.rstrip('/')}/newsletter/confirm/{confirmation_token}?lang={lang}"
    if lang == "fr":
        subject = "Confirmez votre inscription à la newsletter FIRONOVA"
        body_html = (
            f"<p>Merci pour votre intérêt. Un dernier pas : cliquez pour confirmer votre inscription.</p>"
            f"<p><a href=\"{confirm_url}\" style=\"display:inline-block;padding:12px 24px;background:#0f172a;color:#fff;text-decoration:none;border-radius:8px\">Confirmer mon inscription</a></p>"
            f"<p style=\"font-size:12px;color:#64748b\">Le lien est valable 7 jours. "
            f"Si vous n'avez pas demandé cette inscription, ignorez ce message.</p>"
        )
    else:
        subject = "Confirm your FIRONOVA newsletter subscription"
        body_html = (
            f"<p>Thanks for your interest. One final step: click to confirm your subscription.</p>"
            f"<p><a href=\"{confirm_url}\" style=\"display:inline-block;padding:12px 24px;background:#0f172a;color:#fff;text-decoration:none;border-radius:8px\">Confirm my subscription</a></p>"
            f"<p style=\"font-size:12px;color:#64748b\">This link expires in 7 days. "
            f"If you did not request this subscription, please ignore this message.</p>"
        )
    asyncio.create_task(_send_email(email, subject, body_html))

    return {"ok": True, "already_subscribed": False, "confirmation_required": True}


async def newsletter_confirm(token: str, request: Request):
    """Endpoint appelé quand l'utilisateur clique le lien de confirmation
    reçu par email. Marque le subscriber comme confirmé et lui envoie
    l'email de bienvenue pré-launch (contenu réel de la newsletter)."""
    await _rate_limit("newsletter_confirm", _client_ip(request), 30, 3600,
                       "Too many confirmation attempts. Try again later.")
    entry = await db.subscribers.find_one({"confirmation_token": token}, {"_id": 0})
    if not entry:
        raise HTTPException(404, "Lien de confirmation introuvable ou déjà utilisé.")
    if entry.get("status") == "subscribed":
        return {"ok": True, "already_confirmed": True, "email": entry.get("email")}
    # Vérification expiration
    exp = entry.get("confirmation_expires_at")
    if exp:
        try:
            exp_dt = datetime.fromisoformat(exp.replace("Z", "+00:00"))
            if datetime.now(timezone.utc) > exp_dt:
                raise HTTPException(410, "Le lien de confirmation a expiré. Réinscrivez-vous pour recevoir un nouveau lien.")
        except ValueError:
            pass
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.subscribers.update_one(
        {"confirmation_token": token},
        {"$set": {
            "status": "subscribed",
            "confirmed_at": now_iso,
        }, "$unset": {
            "confirmation_token": "",
            "confirmation_expires_at": "",
        }},
    )
    # Envoie l'email de bienvenue réel (l'accueil marketing).
    lang = entry.get("lang") or "en"
    unsubscribe_token = entry.get("unsubscribe_token") or ""
    asyncio.create_task(send_prelaunch_welcome(entry.get("email", ""), lang, unsubscribe_token))
    return {"ok": True, "already_confirmed": False, "email": entry.get("email")}


async def newsletter_unsubscribe(token: str, request: Request):
    """Aucune authentification requise — un lien de désabonnement doit
    fonctionner en un clic, sans login, exigence CASL/LCAP."""
    await _rate_limit(
        "newsletter_unsubscribe", _client_ip(request), 20, 3600,
        "Too many unsubscribe attempts. Try again later.",
    )
    res = await db.subscribers.update_one(
        {"unsubscribe_token": token, "status": "subscribed"},
        {"$set": {"status": "unsubscribed", "unsubscribed_at": datetime.now(timezone.utc).isoformat()}},
    )
    if not res.matched_count:
        # Idempotent : déjà désabonné ou token invalide → même réponse,
        # on ne révèle pas si l'email existe.
        return {"ok": True}
    return {"ok": True}


def _normalize_images(payload_doc: dict) -> dict:
    """Keeps image_url as cover; always first in gallery; caps at 8 entries."""
    image_url = str(payload_doc.get("image_url") or "").strip()
    raw = [str(u) for u in (payload_doc.get("images") or []) if u and isinstance(u, str) and u.strip()]
    images = raw[:8]
    if image_url and image_url not in images:
        images = [image_url] + images[:7]
    if not images and image_url:
        images = [image_url]
    if images:
        payload_doc["images"] = images[:8]
        if not image_url or image_url not in images:
            payload_doc["image_url"] = images[0]
    else:
        payload_doc["images"] = []
        payload_doc["image_url"] = ""
    return payload_doc


def _ensure_variant_ids(payload_doc: dict) -> dict:
    """Ensure every variant has an id. Sync legacy price_cad/stock from first variant if variants present."""
    variants = payload_doc.get("variants") or []
    out_variants = []
    for v in variants:
        v = dict(v)
        if not v.get("id"):
            v["id"] = str(uuid.uuid4())
        out_variants.append(v)
    payload_doc["variants"] = out_variants
    # Sync legacy fields for backward compat / cart fallback
    if out_variants:
        payload_doc["price_cad"] = float(out_variants[0].get("price", 0.0))
        payload_doc["stock"] = sum(int(v.get("stock", 0)) for v in out_variants)
    return payload_doc


async def admin_create_product(payload: ProductIn, _admin: dict = Depends(require_area("products", "manage"))):
    existing = await db.products.find_one({"slug": payload.slug})
    if existing:
        raise HTTPException(409, "Slug already exists")
    doc = payload.model_dump()
    doc = _ensure_variant_ids(doc)
    doc = _normalize_images(doc)
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.products.insert_one(doc)
    doc.pop("_id", None)
    return doc


async def admin_update_product(product_id: str, payload: ProductIn, _admin: dict = Depends(require_area("products", "manage"))):
    before = await db.products.find_one({"id": product_id}, {"_id": 0})
    update = payload.model_dump()
    update = _ensure_variant_ids(update)
    update = _normalize_images(update)
    res = await db.products.update_one({"id": product_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(404, "Product not found")
    after = await db.products.find_one({"id": product_id}, {"_id": 0})

    # Back-in-stock: fire for any variant (or legacy product-level stock) that went from <=0 to >0.
    if before:
        before_variant_stock = {v.get("id"): v.get("stock", 0) for v in before.get("variants", [])}
        for v in after.get("variants", []):
            vid = v.get("id")
            if before_variant_stock.get(vid, 0) <= 0 < v.get("stock", 0):
                asyncio.create_task(_maybe_notify_restock(product_id, vid))
        if before.get("stock", 0) <= 0 < after.get("stock", 0):
            asyncio.create_task(_maybe_notify_restock(product_id, None))

    # Re-evaluate preorders whenever stock or COA/coming-soon badges change.
    asyncio.create_task(_release_ready_preorder_orders())

    return after


async def admin_delete_product(product_id: str, admin: dict = Depends(require_area("products", "manage"))):
    return await _soft_delete("products", product_id, admin)


async def admin_upload_coa(file: UploadFile = File(...), _admin: dict = Depends(require_area("products", "manage"))):
    """Uploads a Certificate of Analysis PDF and returns a URL to paste into a
    product's or variant's coa_url field. Storage is local disk under UPLOAD_DIR,
    served statically at /uploads/coa/<file>."""
    filename = file.filename or ""
    is_pdf = (file.content_type == "application/pdf") or filename.lower().endswith(".pdf")
    if not is_pdf:
        raise HTTPException(400, "Only PDF files are allowed")

    contents = await file.read()
    size_mb = len(contents) / (1024 * 1024)
    if size_mb > MAX_COA_UPLOAD_MB:
        raise HTTPException(400, f"File too large — max {MAX_COA_UPLOAD_MB:.0f} MB")
    if not contents.startswith(b"%PDF"):
        raise HTTPException(400, "File is not a valid PDF")

    safe_name = f"{uuid.uuid4().hex}.pdf"
    dest = COA_UPLOAD_DIR / safe_name
    with open(dest, "wb") as f:
        f.write(contents)

    rel_path = f"/api/uploads/coa/{safe_name}"
    # On stocke le chemin relatif — le client construit l'URL absolue avec son
    # propre `REACT_APP_BACKEND_URL`. Éviter `PUBLIC_BASE_URL` gèle une URL
    # d'ingress spécifique (ex. `*.emergentcf.cloud`) qui peut varier entre
    # environnements ou expirer, laissant des pointeurs morts en BDD.
    return {"url": rel_path, "original_filename": filename, "size_bytes": len(contents)}


async def admin_upload_image(file: UploadFile = File(...), _admin: dict = Depends(require_area("products", "manage"))):
    """Uploads a product image (PNG, JPEG, WebP, GIF) and returns a URL to use in
    the image_url field. Storage is local disk under UPLOAD_DIR/images, served
    statically at /uploads/images/<file>."""
    filename = file.filename or ""

    contents = await file.read()
    size_mb = len(contents) / (1024 * 1024)
    if size_mb > MAX_IMAGE_UPLOAD_MB:
        raise HTTPException(400, f"File too large — max {MAX_IMAGE_UPLOAD_MB:.0f} MB")

    # Validation PAR LE CONTENU (magic bytes), pas par le Content-Type déclaré
    # par le client : un Content-Type est triviable, un header de fichier ne l'est pas.
    _IMAGE_FORMAT_EXT = {
        "PNG": "png",
        "JPEG": "jpg",
        "WEBP": "webp",
        "GIF": "gif",
    }
    try:
        prev_max = PILImage.MAX_IMAGE_PIXELS
        PILImage.MAX_IMAGE_PIXELS = 50_000_000  # borne anti-decompression-bomb
        try:
            with PILImage.open(io.BytesIO(contents)) as im:
                fmt = (im.format or "").upper()
                im.verify()
        finally:
            PILImage.MAX_IMAGE_PIXELS = prev_max
    except Exception:
        raise HTTPException(400, "File is not a valid image")

    ext = _IMAGE_FORMAT_EXT.get(fmt)
    if not ext:
        raise HTTPException(400, "Only PNG, JPEG, WebP, and GIF images are allowed")

    safe_name = f"{uuid.uuid4().hex}.{ext}"
    dest = IMAGE_UPLOAD_DIR / safe_name
    with open(dest, "wb") as f:
        f.write(contents)

    rel_path = f"/api/uploads/images/{safe_name}"
    return {"url": rel_path, "original_filename": filename, "size_bytes": len(contents)}


# ---------------------------------------------------------------------------
# Order / Checkout
# ---------------------------------------------------------------------------
def _resolve_variant(p: dict, variant_id: Optional[str]) -> dict:
    """Return the selected variant subdoc. Falls back to first variant or builds a synthetic one from legacy fields."""
    variants = p.get("variants") or []
    if variant_id:
        for v in variants:
            if v.get("id") == variant_id:
                return v
        raise HTTPException(400, f"Variant {variant_id} not found for product {p['slug']}")
    if variants:
        return variants[0]
    # Legacy synthetic variant
    return {
        "id": "_default",
        "name": f"{p.get('dosage_mg', 0)}mg" if p.get("dosage_mg") else "Default",
        "price": p.get("price_cad", 0.0),
        "stock": p.get("stock", 0),
        "sku": p["slug"].upper(),
        "preorder_enabled": p.get("preorder_allowed", False),
        "preorder_delay_message": "",
        "preorder_price": None,
        "preorder_note": "",
        "coa_status": "available" if p.get("coa_url") else "none",
        "badge_coa_available": bool(p.get("coa_url")),
        "badge_coa_pending": not bool(p.get("coa_url")),
        "badge_coming_soon": False,
        "coa_url": p.get("coa_url", "") or "",
        "sale_price": None,
    }


def _variant_effective_price(v: dict, is_preorder: bool) -> float:
    if is_preorder and v.get("preorder_price"):
        return float(v["preorder_price"])
    sale = v.get("sale_price")
    if sale and float(sale) < float(v.get("price", 0.0)):
        return float(sale)
    return float(v.get("price", 0.0))


# Implementation lives in services/stock.py; re-exported so existing call
# sites (routers/, other server helpers) keep resolving these names here.
try:
    from services.stock import (  # noqa: F401
        _reserve_stock_atomic, _release_stock_atomic, _send_restock_email, _maybe_notify_restock,
        _check_low_stock_alerts, _send_low_stock_admin_email, low_stock_alerts_enriched,
        _restock_order_items,
    )
except ImportError:  # package-relative import (uvicorn backend.server:app)
    from backend.services.stock import (  # noqa: F401
        _reserve_stock_atomic, _release_stock_atomic, _send_restock_email, _maybe_notify_restock,
        _check_low_stock_alerts, _send_low_stock_admin_email, low_stock_alerts_enriched,
        _restock_order_items,
    )


def _coupon_usage_for_email(coupon: dict, email: str) -> int:
    """Nombre d'utilisations d'un coupon par un email donne (tableau used_by)."""
    for entry in coupon.get("used_by") or []:
        if entry and entry.get("email") == email:
            try:
                return int(entry.get("count", 0))
            except (TypeError, ValueError):
                return 0
    return 0


def _is_affiliate_coupon(coupon: dict) -> bool:
    return bool(coupon.get("affiliate_id")) or coupon.get("source") == "affiliate"


def _enforce_standard_coupon_percent_limit(discount_type: str, value: float, is_affiliate: bool) -> None:
    if is_affiliate or STANDARD_COUPON_MAX_PERCENT is None:
        return
    if discount_type == "percent" and float(value) > STANDARD_COUPON_MAX_PERCENT:
        raise HTTPException(400, f"Standard coupon percent cannot exceed {STANDARD_COUPON_MAX_PERCENT:.2f}")


async def _coupon_discount(coupon: dict, subtotal: float,
                           line_items: Optional[list] = None,
                           email: Optional[str] = None) -> tuple[float, dict]:
    """Valide un coupon et retourne (discount, applied_coupon). Leve 400 si invalide.

    Les restrictions contextuelles (emails autorises, premier achat, limite par
    client, restriction produit/categorie) ne sont verifiees QUE si le contexte
    est fourni : le checkout le fournit toujours ; le point public /coupons/validate
    sans email ni items reste strictement retrocompatible."""
    if not coupon or not coupon.get("active"):
        raise HTTPException(400, "Invalid coupon code")
    if coupon.get("expires_at"):
        try:
            if datetime.fromisoformat(coupon["expires_at"].replace("Z", "+00:00")) < datetime.now(timezone.utc):
                raise HTTPException(400, "Coupon expired")
        except ValueError:
            pass
    if coupon.get("start_at"):
        try:
            if datetime.fromisoformat(coupon["start_at"].replace("Z", "+00:00")) > datetime.now(timezone.utc):
                raise HTTPException(400, "Coupon not active yet")
        except ValueError:
            pass
    if coupon.get("usage_limit") and coupon.get("used_count", 0) >= coupon["usage_limit"]:
        raise HTTPException(400, "Coupon usage limit reached")
    if subtotal < coupon.get("min_subtotal", 0):
        raise HTTPException(400, f"Minimum subtotal of ${coupon['min_subtotal']:.2f} required for this coupon")

    email_norm = (email or "").strip().lower()
    allowed = [e.lower().strip() for e in (coupon.get("allowed_emails") or []) if e and e.strip()]
    requires_identity = bool(allowed or coupon.get("first_order_only") or coupon.get("per_customer_limit"))
    if requires_identity and not email_norm:
        raise HTTPException(400, "An email address is required for this coupon")
    if allowed and email_norm not in allowed:
        raise HTTPException(400, "This coupon is not valid for this account")
    if coupon.get("first_order_only"):
        prior = await db.orders.count_documents({"email": email_norm, "payment_status": "paid"})
        if prior > 0:
            raise HTTPException(400, "This coupon is valid for first orders only")
    per_limit = coupon.get("per_customer_limit")
    if per_limit:
        if _coupon_usage_for_email(coupon, email_norm) >= int(per_limit):
            raise HTTPException(400, "Coupon usage limit reached for this account")

    restrict_products = [p for p in (coupon.get("restrict_products") or []) if p]
    restrict_categories = [c for c in (coupon.get("restrict_categories") or []) if c]
    if line_items and (restrict_products or restrict_categories):
        for it in line_items:
            if restrict_products and str(it.get("product_id", "")) not in restrict_products:
                raise HTTPException(400, "This coupon does not apply to all items in your cart")
            if restrict_categories and (it.get("category") or "") not in restrict_categories:
                    raise HTTPException(400, "This coupon does not apply to all items in your cart")

    _enforce_standard_coupon_percent_limit(
        coupon.get("discount_type", ""),
        float(coupon.get("value", 0)),
        _is_affiliate_coupon(coupon),
    )

    if coupon["discount_type"] == "percent":
        discount = round(subtotal * (coupon["value"] / 100.0), 2)
    else:
        discount = round(min(coupon["value"], subtotal), 2)
    max_disc = coupon.get("max_discount_cad")
    if max_disc is not None:
        discount = round(min(discount, float(max_disc)), 2)

    applied = {
        "code": coupon["code"],
        "discount_type": coupon["discount_type"],
        "value": coupon["value"],
        "discount_amount": discount,
    }
    return discount, applied


async def _build_order_totals(items: List[CartItem], coupon_code: Optional[str] = None, customer_email: Optional[str] = None):
    line_items = []
    subtotal = 0.0
    has_preorder = False

    # Un seul aller-retour Mongo au lieu de N (le panier moyen faisait N find_one
    # séquentiels, chacun un RTT réseau complet).
    _ids = list({it.product_id for it in items})
    _docs = await db.products.find({"id": {"$in": _ids}}, {"_id": 0}).to_list(len(_ids))
    _by_id = {d["id"]: d for d in _docs}

    for it in items:
        p = _by_id.get(it.product_id)
        if not p:
            raise HTTPException(400, f"Product {it.product_id} not found")
        if not p.get("active"):
            raise HTTPException(400, f"Product {p['name_en']} unavailable")

        v = _resolve_variant(p, it.variant_id)
        is_preorder = False
        # badge_coa_pending (COA not yet available) also forces preorder path,
        # matching the frontend rule (badge_coa_pending || badge_coming_soon).
        coa_coming = bool(v.get("badge_coming_soon") or v.get("badge_coa_pending"))
        if v.get("preorder_enabled") and (coa_coming or v.get("stock", 0) < it.qty):
            is_preorder = True
            has_preorder = True
        elif v.get("stock", 0) < it.qty:
            if p.get("preorder_allowed"):
                is_preorder = True
                has_preorder = True
            else:
                raise HTTPException(400, f"Insufficient stock for {p['name_en']} ({v.get('name','')})")

        unit_price = _variant_effective_price(v, is_preorder)
        line_total = round(unit_price * it.qty, 2)
        line_items.append({
            "product_id": p["id"],
            "variant_id": v.get("id"),
            "variant_name": v.get("name", ""),
            "slug": p["slug"],
            "sku": v.get("sku", p["slug"].upper()),
            "name_en": p["name_en"],
            "name_fr": p["name_fr"],
            "price_cad": unit_price,
            "qty": it.qty,
            "line_total": line_total,
            "image_url": p.get("image_url", ""),
            "preorder": is_preorder,
            # Poids figé à l'achat : l'étiquette doit refléter ce qui a été vendu,
            # même si la fiche produit change ensuite. Les commandes antérieures
            # sans ce champ retombent sur 50 g/unité dans _order_weight_kg().
            "weight_grams": float(v.get("weight_grams") or 50.0),
        })
        subtotal += line_total
    subtotal = round(subtotal, 2)

    discount = 0.0
    applied_coupon = None
    if coupon_code:
        coupon = await db.coupons.find_one({"code": coupon_code.upper().strip()}, {"_id": 0})
        discount, applied_coupon = await _coupon_discount(
            coupon, subtotal, line_items=line_items, email=customer_email
        )

    tax_rate = 0.0
    shipping = 0.0 if (subtotal - discount) >= FREE_SHIPPING_THRESHOLD_CAD else SHIPPING_FLAT_CAD
    tax = 0.0
    total = round(max(0, subtotal - discount) + shipping, 2)
    return line_items, subtotal, tax_rate, tax, shipping, total, discount, applied_coupon, has_preorder


# Implementation lives in services/canada_post.py; re-exported so existing call
# sites (routers/, other server helpers) keep resolving these names here.
try:
    from services.canada_post import (  # noqa: F401
        _cp_use_openapi, _cp_path_customers, _cp_get_oauth_token, _cp_openapi_call,
        _estimate_parcel_weight_kg, _canada_post_get_rates, _canada_post_track,
        _cp_tracking_indicates_delivered, _sandbox_fallback_ready, is_canada_post_configured,
        _canada_post_create_shipment_openapi, _canada_post_get_artifact_openapi,
        _canada_post_get_manifest_artifact_openapi, _canada_post_shipment_price,
        _canada_post_estimate_openapi, _canada_post_manifest_details,
        _canada_post_transmit_openapi, _canada_post_void_openapi, _canada_post_create_shipment,
        _canada_post_get_artifact, _canada_post_transmit, _canada_post_void,
        _auto_create_dispatch_label, _auto_label_paid_orders_watchdog,
        _auto_sync_delivered_orders_once, _auto_sync_delivered_orders_watchdog,
        UNTRANSMITTED_MATCH, pending_manifest_state,
    )
except ImportError:  # package-relative import (uvicorn backend.server:app)
    from backend.services.canada_post import (  # noqa: F401
        _cp_use_openapi, _cp_path_customers, _cp_get_oauth_token, _cp_openapi_call,
        _estimate_parcel_weight_kg, _canada_post_get_rates, _canada_post_track,
        _cp_tracking_indicates_delivered, _sandbox_fallback_ready, is_canada_post_configured,
        _canada_post_create_shipment_openapi, _canada_post_get_artifact_openapi,
        _canada_post_get_manifest_artifact_openapi, _canada_post_shipment_price,
        _canada_post_estimate_openapi, _canada_post_manifest_details,
        _canada_post_transmit_openapi, _canada_post_void_openapi, _canada_post_create_shipment,
        _canada_post_get_artifact, _canada_post_transmit, _canada_post_void,
        _auto_create_dispatch_label, _auto_label_paid_orders_watchdog,
        _auto_sync_delivered_orders_once, _auto_sync_delivered_orders_watchdog,
        UNTRANSMITTED_MATCH, pending_manifest_state,
    )


# Implementation lives in services/mail.py; re-exported so existing call
# sites (routers/, other server helpers) keep resolving these names here.
try:
    from services.mail import (  # noqa: F401
        _order_email_html, _send_email, _process_email_outbox_job, _email_outbox_worker,
        EMAIL_JANITOR_INTERVAL_S, EMAIL_FAILED_RETRY_AFTER_S, EMAIL_JANITOR_MAX_PER_TICK,
        _email_outbox_janitor, send_order_confirmation, send_payment_received,
        _simple_order_email_html, _prelaunch_email_html, send_prelaunch_welcome,
        send_shipping_notification, send_customer_note_email, send_refund_email, ABANDON_MIN_HOURS,
        ABANDON_MAX_HOURS, ABANDON_COUPON_CODE, ABANDON_SWEEP_MINUTES,
        send_abandoned_cart_reminder, welcome_new_user, _abandoned_cart_watchdog, EMAIL_BLOCKS,
        EMAIL_TEMPLATE_CATALOG, _email_template_get, _render_block, _email_render,
        send_template_email, EmailTemplateIn, EmailTemplateCreateIn, _order_ctx,
    )
except ImportError:  # package-relative import (uvicorn backend.server:app)
    from backend.services.mail import (  # noqa: F401
        _order_email_html, _send_email, _process_email_outbox_job, _email_outbox_worker,
        EMAIL_JANITOR_INTERVAL_S, EMAIL_FAILED_RETRY_AFTER_S, EMAIL_JANITOR_MAX_PER_TICK,
        _email_outbox_janitor, send_order_confirmation, send_payment_received,
        _simple_order_email_html, _prelaunch_email_html, send_prelaunch_welcome,
        send_shipping_notification, send_customer_note_email, send_refund_email, ABANDON_MIN_HOURS,
        ABANDON_MAX_HOURS, ABANDON_COUPON_CODE, ABANDON_SWEEP_MINUTES,
        send_abandoned_cart_reminder, welcome_new_user, _abandoned_cart_watchdog, EMAIL_BLOCKS,
        EMAIL_TEMPLATE_CATALOG, _email_template_get, _render_block, _email_render,
        send_template_email, EmailTemplateIn, EmailTemplateCreateIn, _order_ctx,
    )


# ---------------------------------------------------------------------------
# Admin visibilité + contrôle manuel de la file email
# ---------------------------------------------------------------------------
async def admin_email_outbox_stats(_admin: dict = Depends(require_area("orders", "view"))):  # noqa: F821
    """Retourne l'état de santé de la file email :
      - counts by status (pending, retry, sending, sent, failed)
      - âge du plus ancien job actif (retard file)
      - moyenne des tentatives sur les failed (indique un incident tiers)
    """
    pipeline = [{"$group": {"_id": "$status", "n": {"$sum": 1}}}]
    counts = {"pending": 0, "retry": 0, "sending": 0, "sent": 0, "failed": 0}
    async for row in db.email_outbox.aggregate(pipeline):
        counts[row["_id"]] = row["n"]

    now = datetime.now(timezone.utc)
    oldest_active = await db.email_outbox.find_one(
        {"status": {"$in": ["pending", "retry", "sending"]}},
        {"_id": 0, "created_at": 1, "status": 1, "attempts": 1},
        sort=[("created_at", 1)],
    )
    oldest_age_s = None
    if oldest_active and oldest_active.get("created_at"):
        try:
            ts = datetime.fromisoformat(oldest_active["created_at"])
            oldest_age_s = int((now - ts).total_seconds())
        except Exception:
            oldest_age_s = None

    # Avg attempts on failed (dernier 7 jours)
    since = (now - timedelta(days=7)).isoformat()
    avg_pipeline = [
        {"$match": {"status": "failed", "created_at": {"$gte": since}}},
        {"$group": {"_id": None, "avg": {"$avg": "$attempts"}, "n": {"$sum": 1}}},
    ]
    avg = None
    n_failed_7d = 0
    async for row in db.email_outbox.aggregate(avg_pipeline):
        avg = round(row["avg"], 2) if row.get("avg") is not None else None
        n_failed_7d = row.get("n") or 0

    return {
        "counts": counts,
        "oldest_active_age_seconds": oldest_age_s,
        "oldest_active_status": (oldest_active or {}).get("status"),
        "failed_last_7d": n_failed_7d,
        "avg_attempts_on_failed": avg,
        "janitor_interval_s": EMAIL_JANITOR_INTERVAL_S,
        "failed_retry_after_s": EMAIL_FAILED_RETRY_AFTER_S,
    }


class EmailRequeueIn(BaseModel):
    scope: str = "failed"  # "failed" ou "stuck" ou "all_active"
    max: int = Field(default=100, ge=1, le=1000)


async def admin_email_requeue(payload: EmailRequeueIn,
                               _admin: dict = Depends(require_area("orders", "manage"))):  # noqa: F821
    """Force le rejeu de jobs email. Utile après une panne Resend ou pour
    dépiler manuellement une file bloquée. Retourne le nombre requeué."""
    now_iso = datetime.now(timezone.utc).isoformat()
    query: dict
    if payload.scope == "stuck":
        query = {"status": "sending", "lease_expires_at": {"$lte": now_iso}}
    elif payload.scope == "all_active":
        query = {"status": {"$in": ["failed", "sending"]}}
    else:
        query = {"status": "failed"}

    cursor = db.email_outbox.find(query, {"_id": 0, "id": 1}).limit(payload.max)
    ids = [doc["id"] async for doc in cursor]
    if not ids:
        return {"requeued": 0, "scope": payload.scope}
    await db.email_outbox.update_many(
        {"id": {"$in": ids}},
        {"$set": {"status": "retry", "attempts": 0, "available_at": now_iso,
                  "requeued_at": now_iso, "requeued_by": "admin"},
         "$unset": {"lease_expires_at": ""}},
    )
    return {"requeued": len(ids), "scope": payload.scope}


async def get_shipping_rates(payload: ShippingRateRequest, request: Request):
    """Live Canada Post rates when CANADA_POST_API_KEY is configured; otherwise falls back
    to the existing flat-rate shipping_zones/shipping_methods (same zones used at checkout)."""
    await _rate_limit("shipping_rates", _client_ip(request), 20, 60,
                       "Too many shipping rate requests. Try again later.")
    weight_kg = await _estimate_parcel_weight_kg(payload.items)
    live_rates = await _canada_post_get_rates(payload.postal_code, payload.country, weight_kg)
    if live_rates:
        return {"source": "canada_post_live", "weight_kg": weight_kg, "rates": live_rates}

    zones = await db.shipping_zones.find({"deleted_at": None}, {"_id": 0}).to_list(50)
    zone = next((z for z in zones if payload.country in z.get("countries", [])), None)
    if not zone:
        zone = next((z for z in zones if "INTL" in z.get("countries", [])), None)
    methods = []
    if zone:
        methods = await db.shipping_methods.find(
            {"zone_id": zone["id"], "active": True, "deleted_at": None}, {"_id": 0}
        ).to_list(50)
    rates = [
        {"carrier": m["name"], "service_code": None, "service_name": m["name"],
         "cost_cad": m["cost_cad"], "eta_days": m["eta_days"]}
        for m in methods
    ] or [{"carrier": "FIRONOVA", "service_code": None, "service_name": "Standard",
           "cost_cad": SHIPPING_FLAT_CAD, "eta_days": "3-7 business days"}]
    return {"source": "flat_rate_fallback", "weight_kg": weight_kg, "rates": rates}


# ---------------------------------------------------------------------------
# Confirmation de paiement — point d'entrée UNIQUE et idempotent.
# Le filtre bloque les statuts terminaux (cancelled/failed/refunded) et
# empêche une confirmation tardive de réactiver une commande déjà clôturée.
# Le filtre {"payment_status": {"$nin": [...]}} dans le update_one garantit
# qu'un seul appelant gagne, même avec des callbacks/pollings concurrents
# arrivent à la milliseconde près. Le coupon est décompté ici (et pas à la
# création) : un panier abandonné ne doit jamais consommer un usage_limit.
# ---------------------------------------------------------------------------
async def _claim_coupon_usage(order: dict) -> bool:
    coupon = order.get("coupon")
    order_id = order.get("id")
    if not (coupon and coupon.get("code") and order_id) or order.get("coupon_counted"):
        return False

    email_norm = (order.get("email") or "").strip().lower()
    set_stage: dict = {
        "used_count": {"$add": [{"$ifNull": ["$used_count", 0]}, 1]},
        "counted_order_ids": {
            "$concatArrays": [{"$ifNull": ["$counted_order_ids", []]}, [order_id]],
        },
    }
    if email_norm:
        set_stage["used_by"] = {
            "$let": {
                "vars": {"entries": {"$ifNull": ["$used_by", []]}},
                "in": {
                    "$cond": [
                        {"$in": [email_norm, "$$entries.email"]},
                        {"$map": {
                            "input": "$$entries",
                            "as": "entry",
                            "in": {"$cond": [
                                {"$eq": ["$$entry.email", email_norm]},
                                {"$mergeObjects": [
                                    "$$entry",
                                    {"count": {"$add": [{"$ifNull": ["$$entry.count", 0]}, 1]}},
                                ]},
                                "$$entry",
                            ]},
                        }},
                        {"$concatArrays": ["$$entries", [{"email": email_norm, "count": 1}]]},
                    ]
                },
            }
        }
    coupon_claim = await db.coupons.update_one(
        {
            "code": coupon["code"],
            "counted_order_ids": {"$ne": order_id},
            "$or": [
                {"usage_limit": None},
                {"usage_limit": {"$exists": False}},
                {"$expr": {"$lt": [{"$ifNull": ["$used_count", 0]}, "$usage_limit"]}},
            ],
        },
        [{"$set": set_stage}],
    )
    if not coupon_claim.modified_count:
        return False
    await db.orders.update_one({"id": order_id}, {"$set": {"coupon_counted": True}})
    return True


async def _mark_order_paid(order_id: str, note_text: Optional[str] = None) -> Optional[dict]:
    """Retourne l'order fraîche si CET appel a fait la transition, sinon None."""
    paid_at = datetime.now(timezone.utc).isoformat()
    terminal_statuses = {"paid", "cancelled", "failed", "refunded"}
    # Preorder orders retain 'preorder' fulfillment_status until items ship.
    existing = await db.orders.find_one({"id": order_id}, {"_id": 0, "has_preorder": 1, "payment_status": 1})
    if not existing or existing.get("payment_status") in terminal_statuses:
        return None
    new_fulfillment = "preorder" if existing.get("has_preorder") else "processing"
    update: dict = {
        "$set": {
            "payment_status": "paid",
            "fulfillment_status": new_fulfillment,
            "paid_at": paid_at,
            "dispatch_batch": compute_dispatch_batch(paid_at),
        }
    }
    if note_text:
        update["$push"] = {"notes": {
            "id": str(uuid.uuid4()),
            "text": note_text,
            "author": "system",
            "created_at": paid_at,
        }}

    res = await db.orders.update_one(
        {"id": order_id, "payment_status": {"$nin": list(terminal_statuses)}},
        update,
    )
    if not res.modified_count:
        return None  # déjà payée : un autre chemin a gagné la course

    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        return None

    await _claim_coupon_usage(order)

    if order.get("email"):
        asyncio.create_task(send_payment_received(order))
    # Dispatch manuel : l'étiquette n'est PLUS créée automatiquement au paiement.
    # La commande attend dans « À étiqueter » et l'admin génère l'étiquette
    # depuis l'écran Dispatch.
    # asyncio.create_task(_auto_create_dispatch_label(order_id))
    # --- AFFILIATE: commission pending au paiement confirme ---
    await affiliate_on_order_paid(order)
    return order


async def _flag_late_cancelled_payment(order: dict, provider: str, reference: str = "") -> bool:
    """Mark cancelled orders that receive a payment after timeout.

    This keeps payment_status terminal (cancelled) but raises a clear flag so
    staff can review/reopen through the dedicated workflow.
    """
    if not order:
        return False
    if order.get("payment_status") != "cancelled":
        return False
    if order.get("late_payment_flagged"):
        return False

    now_iso = datetime.now(timezone.utc).isoformat()
    ref = (reference or "").strip()
    provider_label = (provider or "payment").strip().upper()
    ref_txt = f" (ref {ref})" if ref else ""
    note = {
        "id": str(uuid.uuid4()),
        "text": (
                f"⚠️ PAIEMENT TARDIF / LATE PAYMENT detected on cancelled order — {provider_label} confirmation received"
            f"{ref_txt}. Review manually before reopening."
        ),
        "author": "system",
        "created_at": now_iso,
    }

    updates = {
        "late_payment_flagged": True,
        "late_payment_flagged_at": now_iso,
    }
    if ref:
        updates["late_payment_reference"] = ref

    res = await db.orders.update_one(
        {"id": order["id"], "payment_status": "cancelled", "late_payment_flagged": {"$ne": True}},
        {"$set": updates, "$push": {"notes": note}},
    )
    if res.modified_count:
        # Alert admin immediately when late payment is flagged.
        asyncio.create_task(_send_late_payment_admin_alert(order, provider_label, ref))
        logging.warning("Late payment flagged on cancelled order %s via %s", order.get("id"), provider_label)
        return True
    return False


async def _send_late_payment_admin_alert(order: dict, provider_label: str, reference: str = "") -> None:
    """Notify admin when a cancelled order receives a late payment signal."""
    try:
        order_id = order.get("id")
        if not order_id:
            return
        fresh = await db.orders.find_one({"id": order_id}, {"_id": 0}) or order
        order_number = fresh.get("order_number") or order_id
        method = str(fresh.get("payment_method") or "").upper() or "UNKNOWN"
        total = float(fresh.get("total") or 0)
        customer_email = fresh.get("email") or ""
        ref = (reference or fresh.get("late_payment_reference") or "").strip()
        detail_ref = f"<li><strong>Reference:</strong> {ref}</li>" if ref else ""
        customer_html = f"<li><strong>Customer:</strong> {customer_email}</li>" if customer_email else ""
        order_url = f"{(PUBLIC_BASE_URL or '').rstrip('/')}/order/{order_id}" if PUBLIC_BASE_URL else ""
        order_link_html = (
            f"<p style='margin-top:14px'><a href='{order_url}' style='display:inline-block;background:#050505;color:#fff;"
            f"font-family:monospace;font-size:12px;letter-spacing:1px;padding:10px 14px;text-decoration:none'>"
            f"OPEN ORDER PAGE →</a></p>"
            if order_url else ""
        )
        body = (
            f"<p><strong>Late payment flagged</strong> on a cancelled order.</p>"
            f"<ul style='margin:10px 0 0 18px;padding:0;line-height:1.7'>"
            f"<li><strong>Order:</strong> {order_number}</li>"
            f"<li><strong>Payment method:</strong> {method}</li>"
            f"<li><strong>Detected via:</strong> {provider_label}</li>"
            f"<li><strong>Total:</strong> ${total:.2f} CAD</li>"
            f"{customer_html}"
            f"{detail_ref}"
            f"</ul>"
            f"<p style='margin-top:12px'>Please review and reopen manually only after validation.</p>"
            f"{order_link_html}"
        )
        html = _simple_order_email_html(fresh, "Admin alert — late payment flagged", body)
        await _send_email(
            ADMIN_NOTIFICATION_EMAIL,
            f"[FIRONOVA ADMIN] Late payment flagged — {order_number}",
            html,
        )
    except Exception as e:
        logging.error("[email] late-payment admin alert failed order_ref=%s error_type=%s", _private_ref(order.get("id")), type(e).__name__)


async def _send_reconciliation_required_admin_alert(item: dict) -> None:
    """Alert admin when an inbound payment signal needs manual reconciliation."""
    try:
        provider = str(item.get("provider") or "payment").lower()
        provider_label = "Interac" if provider == "interac" else "NOWPayments"
        queue_id = item.get("id") or ""
        amount = item.get("amount_cad")
        amount_txt = f"${float(amount):.2f} CAD" if amount is not None else "Unknown"
        sender = item.get("from_email") or "Unknown"
        received = item.get("received_at") or item.get("detected_at") or ""
        subject = item.get("subject") or ""
        preview = (item.get("preview") or "").strip()
        if len(preview) > 600:
            preview = preview[:600] + "..."
        body = (
            f"<p><strong>{provider_label} payment signal requires manual reconciliation.</strong></p>"
            "<p>Please reconcile this payment manually in Admin > Reconciliation.</p>"
            "<ul style='margin:10px 0 0 18px;padding:0;line-height:1.7'>"
            f"<li><strong>Provider:</strong> {provider_label}</li>"
            f"<li><strong>Reason:</strong> {item.get('reason') or 'unknown'}</li>"
            f"<li><strong>Queue ID:</strong> {queue_id}</li>"
            f"<li><strong>Amount:</strong> {amount_txt}</li>"
            f"<li><strong>From:</strong> {sender}</li>"
            f"<li><strong>Received:</strong> {received}</li>"
            f"<li><strong>Subject:</strong> {subject}</li>"
            "</ul>"
            f"<p style='margin-top:12px'><strong>Email preview</strong><br/>{preview or '(empty)'}</p>"
        )
        html = _prelaunch_email_html("Admin alert — payment reconciliation required", body)
        await _send_email(
            ADMIN_NOTIFICATION_EMAIL,
            f"[FIRONOVA ADMIN] {provider_label} payment needs reconciliation",
            html,
        )
    except Exception as e:
        logging.error("[email] payment reconciliation alert failed: %s", type(e).__name__)


# Implementation lives in services/interac.py; re-exported so existing call
# sites (routers/, other server helpers) keep resolving these names here.
try:
    from services.interac import (  # noqa: F401
        _graph_access_token, _graph_unread_messages, _graph_mark_read, _strip_html,
        _extract_interac_refs, _parse_amounts, _process_interac_deposit_emails,
        _interac_deposit_watchdog,
    )
except ImportError:  # package-relative import (uvicorn backend.server:app)
    from backend.services.interac import (  # noqa: F401
        _graph_access_token, _graph_unread_messages, _graph_mark_read, _strip_html,
        _extract_interac_refs, _parse_amounts, _process_interac_deposit_emails,
        _interac_deposit_watchdog,
    )


def _address_cache_key(addr: dict) -> str:
    """SHA-256 canonique (upper+trim+dedupe whitespace) — chaque adresse
    identique en essence n'appelle Google qu'une fois par 24h."""
    canonical = {
        k: " ".join(str(addr.get(k, "")).strip().upper().split())
        for k in ("address1", "address2", "city", "province", "postal_code", "country")
    }
    return hashlib.sha256(json.dumps(canonical, sort_keys=True).encode()).hexdigest()


async def _validate_shipping_address_google(addr: dict) -> dict:
    """Appelle Google Maps Address Validation API. Retourne un dict :
      {
        "valid": bool,
        "suggestions": [postalAddress dict],  # candidat corrigé si dispo
        "verdict": {...},
        "normalized": {...} | None,
        "provider": "google_maps" | "disabled" | "unavailable",
        "response_id": str | None,
      }

    Politique : accepte SEULEMENT si `addressComplete=true`, aucun composant
    non-confirmé, et `possibleNextAction=ACCEPT`. Sur erreur Google (timeout,
    quota, config), retourne `unavailable` → l'appelant décide (défaut :
    laisser passer sinon Google en panne bloque toutes les ventes).
    """
    if not GOOGLE_MAPS_API_KEY:
        return {"valid": True, "suggestions": [], "verdict": {},
                "normalized": None, "provider": "disabled", "response_id": None}

    country = (addr.get("country") or "CA").upper()
    if country != "CA":
        # On ne valide que le Canada pour le lancement. Les autres pays passent
        # (à revoir si vous ouvrez à l'international).
        return {"valid": True, "suggestions": [], "verdict": {},
                "normalized": None, "provider": "skipped_non_ca", "response_id": None}

    cache_key = _address_cache_key(addr)
    now = time.time()
    cached = _ADDRESS_CACHE.get(cache_key)
    if cached and now - cached[0] < _ADDRESS_CACHE_TTL_SEC:
        return cached[1]

    lines = [addr.get("address1", "")]
    if addr.get("address2"):
        lines.append(addr["address2"])
    payload = {
        "address": {
            "regionCode": "CA",
            "locality": addr.get("city", ""),
            "administrativeArea": addr.get("province", ""),
            "postalCode": addr.get("postal_code", ""),
            "addressLines": [l for l in lines if l],
        }
    }
    url = "https://addressvalidation.googleapis.com/v1:validateAddress"
    try:
        import httpx  # local import — évite la charge au boot si feature off
        async with httpx.AsyncClient(timeout=httpx.Timeout(5.0, connect=2.0)) as client:
            r = await client.post(url, params={"key": GOOGLE_MAPS_API_KEY}, json=payload)
    except Exception as e:
        logging.warning("Google Maps AVS request failed: %s", e)
        return {"valid": True, "suggestions": [], "verdict": {},
                "normalized": None, "provider": "unavailable", "response_id": None}

    if r.status_code >= 400:
        try:
            err = r.json().get("error", {})
        except Exception:
            err = {}
        status = err.get("status", "")
        logging.warning("Google Maps AVS HTTP %s status=%s", r.status_code, status)
        # Quota / auth / config → on ne bloque pas la vente. Alerte visible en log.
        return {"valid": True, "suggestions": [], "verdict": {"error": status},
                "normalized": None, "provider": "unavailable", "response_id": None}

    data = r.json()
    result = data.get("result", {})
    verdict = result.get("verdict", {})
    normalized = result.get("address")
    is_ok = (
        verdict.get("addressComplete") is True
        and not verdict.get("hasUnconfirmedComponents", False)
        and verdict.get("possibleNextAction") == "ACCEPT"
    )
    out = {
        "valid": is_ok,
        "suggestions": [] if is_ok or not normalized else [normalized],
        "verdict": verdict,
        "normalized": normalized,
        "provider": "google_maps",
        "response_id": data.get("responseId"),
    }
    _ADDRESS_CACHE[cache_key] = (now, out)
    # Nettoyage passif : si le cache dépasse 20K entrées, drop 25% les plus vieilles
    if len(_ADDRESS_CACHE) > 20000:
        oldest = sorted(_ADDRESS_CACHE.items(), key=lambda x: x[1][0])[:5000]
        for k, _ in oldest:
            _ADDRESS_CACHE.pop(k, None)
    return out


# --- Checkout compensation registry (Item 1.2 B4 SMART) -------------------
# Trois pièces d'infra qui garantissent que si un checkout partiel plante :
#   1. La CompensationContext exécute des rollbacks en ordre inverse
#   2. Chaque échec de compensation est journalisé dans checkout_compensation_failures
#      → admin peut retry manuellement via l'UI Reconciliation
#   3. Un circuit breaker bloque temporairement /checkout si >5 échecs/heure

_checkout_breaker: dict = {"failures": [], "opened_at": None}
_CHECKOUT_BREAKER_WINDOW_SEC = 3600  # 1h
_CHECKOUT_BREAKER_THRESHOLD = 5
_CHECKOUT_BREAKER_OPEN_SEC = 1800  # 30min


def _checkout_breaker_check() -> None:
    """Raise 503 si le breaker est ouvert. Purge la fenêtre glissante."""
    now = time.time()
    if _checkout_breaker["opened_at"]:
        if now - _checkout_breaker["opened_at"] < _CHECKOUT_BREAKER_OPEN_SEC:
            raise HTTPException(
                503,
                "Le service commande est temporairement indisponible. "
                "Nos équipes ont été alertées, réessayez dans quelques minutes.",
            )
        # Fenêtre de refroidissement passée : on referme le breaker
        _checkout_breaker["opened_at"] = None
        _checkout_breaker["failures"].clear()


def _checkout_breaker_record_failure() -> None:
    """Enregistre un échec de compensation. Ouvre le breaker si seuil atteint."""
    now = time.time()
    _checkout_breaker["failures"] = [
        t for t in _checkout_breaker["failures"] if now - t < _CHECKOUT_BREAKER_WINDOW_SEC
    ]
    _checkout_breaker["failures"].append(now)
    if len(_checkout_breaker["failures"]) >= _CHECKOUT_BREAKER_THRESHOLD:
        _checkout_breaker["opened_at"] = now
        logging.error(
            "[checkout-breaker] OPENED — %d compensation failures in the last hour. "
            "Blocking /checkout for %ds.",
            len(_checkout_breaker["failures"]), _CHECKOUT_BREAKER_OPEN_SEC,
        )


class CompensationContext:
    """Context-manager qui exécute une série d'écritures avec rollback.

    Chaque étape enregistre son compensator :
        async with CompensationContext(label="checkout") as ctx:
            oid = await db.orders.insert_one(order)
            ctx.register("delete_order", db.orders.delete_one, {"id": order["id"]})
            await db.products.update_one(...)
            ctx.register("restore_stock", db.products.update_one, ...)

    Si le body raise, les compensations sont rejouées en ordre inverse.
    Si une compensation elle-même échoue, on écrit dans
    checkout_compensation_failures (audit + queue admin retry).

    Utilise les transactions natives Mongo si dispo (replica set) — dans
    ce cas les compensations ne sont jamais exécutées (rollback DB fait
    par Mongo). Sinon fallback saga applicatif.
    """

    def __init__(self, label: str, order_id: Optional[str] = None):
        self.label = label
        self.order_id = order_id
        self._compensations: list = []
        self._writes: list = []
        self._session = None
        self._use_transaction = False

    async def __aenter__(self):
        # Détection replica set — tentative de transaction native
        try:
            self._session = await client.start_session()
            await self._session.__aenter__()
            self._session.start_transaction()
            self._use_transaction = True
            logging.debug("[compensation] native Mongo transaction for %s", self.label)
        except Exception as e:
            # Standalone Mongo : pas de transactions. Fallback saga.
            if self._session is not None:
                try:
                    await self._session.__aexit__(None, None, None)
                except Exception:
                    pass
                self._session = None
            self._use_transaction = False
            logging.debug(
                "[compensation] saga fallback for %s (no transactions: %s)",
                self.label, type(e).__name__,
            )
        return self

    def register(self, name: str, coro_func, *args, **kwargs):
        """Enregistre une compensation. `coro_func` doit être un callable async."""
        self._compensations.append({"name": name, "func": coro_func, "args": args, "kwargs": kwargs})
        self._writes.append(name)

    async def __aexit__(self, exc_type, exc, tb):
        if self._use_transaction and self._session:
            try:
                if exc_type is None:
                    await self._session.commit_transaction()
                else:
                    await self._session.abort_transaction()
            finally:
                try:
                    await self._session.__aexit__(exc_type, exc, tb)
                except Exception:
                    pass
            return False  # ne pas avaler l'exception

        # Saga path
        if exc_type is None:
            return False
        # Body a raise → rejouer les compensations en ordre inverse
        failed_compensations = []
        for step in reversed(self._compensations):
            try:
                await step["func"](*step["args"], **step["kwargs"])
            except Exception as compensation_err:
                failed_compensations.append({
                    "stage": step["name"],
                    "error": f"{type(compensation_err).__name__}: {compensation_err}",
                })
                logging.error(
                    "[compensation] step %s FAILED during %s rollback: %s",
                    step["name"], self.label, compensation_err,
                )
        if failed_compensations:
            try:
                await db.checkout_compensation_failures.insert_one({
                    "id": str(uuid.uuid4()),
                    "order_id": self.order_id,
                    "label": self.label,
                    "original_error": f"{type(exc).__name__}: {exc}",
                    "collections_written": self._writes,
                    "failed_compensations": failed_compensations,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "status": "pending_manual_review",
                    "assigned_admin": None,
                    "resolved_at": None,
                    "resolution_note": None,
                })
                _checkout_breaker_record_failure()
                # Fire-and-forget admin alert
                try:
                    admin_to = os.environ.get("ADMIN_NOTIFICATION_EMAIL", ADMIN_EMAIL)
                    subject = f"[FIRONOVA] Checkout compensation failed — {self.order_id or 'unknown'}"
                    html = (
                        f"<p>Une compensation checkout a échoué et doit être réconciliée manuellement.</p>"
                        f"<ul><li>Label : {self.label}</li>"
                        f"<li>Order id : {self.order_id or '-'}</li>"
                        f"<li>Étapes en échec : {len(failed_compensations)}</li></ul>"
                        f"<p>Voir /ops-portal-fn7k2q → Reconciliation → Checkout failures.</p>"
                    )
                    await _send_email(admin_to, subject, html)
                except Exception:
                    pass
            except Exception as ledger_err:
                logging.critical(
                    "[compensation] FAILED TO PERSIST LEDGER for %s: %s",
                    self.label, ledger_err,
                )
        return False  # ne pas avaler l'exception d'origine


# --- Admin reconciliation for checkout compensation failures (B4 SMART) --

class CheckoutFailureResolveIn(BaseModel):
    note: str = ""
    action: str = "manual_reconciled"  # future: 'refund_issued', 'no_op'


async def admin_checkout_failures_list(status: Optional[str] = None, limit: int = 50):
    q: dict = {}
    if status:
        q["status"] = status
    limit = max(1, min(200, int(limit)))
    cursor = db.checkout_compensation_failures.find(q, {"_id": 0}).sort("created_at", -1).limit(limit)
    items = [d async for d in cursor]
    total = await db.checkout_compensation_failures.count_documents(q)
    return {"items": items, "total": total}


async def admin_checkout_failures_retry(failure_id: str, admin: dict):
    entry = await db.checkout_compensation_failures.find_one({"id": failure_id}, {"_id": 0})
    if not entry:
        raise HTTPException(404, "Failure entry not found")
    if entry["status"] not in ("pending_manual_review", "retry_failed"):
        raise HTTPException(400, f"Cannot retry entry with status={entry['status']}")
    # Le retry auto reconstruirait chaque compensation ; puisque les args
    # sont sérialisés au ledger mais pas les callables, on marque juste
    # 'retry_attempted' — l'admin voit l'état et résout manuellement.
    now = datetime.now(timezone.utc).isoformat()
    await db.checkout_compensation_failures.update_one(
        {"id": failure_id},
        {"$set": {
            "status": "retry_attempted",
            "retry_attempted_at": now,
            "retry_attempted_by": admin.get("email"),
        }},
    )
    return {"ok": True, "next_step": "review_and_resolve_manually"}


async def admin_checkout_failures_resolve(failure_id: str, payload: CheckoutFailureResolveIn, admin: dict):
    entry = await db.checkout_compensation_failures.find_one({"id": failure_id}, {"_id": 0})
    if not entry:
        raise HTTPException(404, "Failure entry not found")
    if entry["status"] == "resolved":
        raise HTTPException(400, "Already resolved")
    now = datetime.now(timezone.utc).isoformat()
    await db.checkout_compensation_failures.update_one(
        {"id": failure_id},
        {"$set": {
            "status": "resolved",
            "resolved_at": now,
            "assigned_admin": admin.get("email"),
            "resolution_note": payload.note or "",
            "resolution_action": payload.action or "manual_reconciled",
        }},
    )
    return {"ok": True}


async def admin_checkout_breaker_state():
    now = time.time()
    failures_in_window = [t for t in _checkout_breaker["failures"] if now - t < _CHECKOUT_BREAKER_WINDOW_SEC]
    open_until = None
    if _checkout_breaker["opened_at"]:
        open_until = _checkout_breaker["opened_at"] + _CHECKOUT_BREAKER_OPEN_SEC
    return {
        "is_open": _checkout_breaker["opened_at"] is not None
                   and (now - _checkout_breaker["opened_at"] < _CHECKOUT_BREAKER_OPEN_SEC),
        "opened_at": _checkout_breaker["opened_at"],
        "open_until_epoch": open_until,
        "failures_in_window": len(failures_in_window),
        "threshold": _CHECKOUT_BREAKER_THRESHOLD,
        "window_seconds": _CHECKOUT_BREAKER_WINDOW_SEC,
        "cooldown_seconds": _CHECKOUT_BREAKER_OPEN_SEC,
    }


async def admin_checkout_breaker_reset(admin: dict):
    _checkout_breaker["opened_at"] = None
    _checkout_breaker["failures"].clear()
    logging.warning("[checkout-breaker] MANUALLY RESET by %s", admin.get("email"))
    return {"ok": True}


async def checkout(payload: CheckoutIn, request: Request):
    _checkout_breaker_check()
    await _rate_limit("checkout", _client_ip(request), CHECKOUT_MAX_PER_MINUTE, 60, "Too many checkout attempts. Try again shortly.")
    if not (payload.accept_terms and payload.confirm_age and payload.confirm_research_use):
        raise HTTPException(400, "All compliance confirmations are required")
    if not payload.items:
        raise HTTPException(400, "Cart is empty")

    # Address validation (Google Maps AVS) — bloquant si adresse invalide.
    # Le résultat est persisté sur l'ordre pour audit. Cache 24h TTL en interne.
    address_check = await _validate_shipping_address_google(payload.shipping.model_dump())
    if not address_check["valid"]:
        raise HTTPException(
            status_code=422,
            detail={
                "code": "invalid_shipping_address",
                "message": (
                    "L'adresse de livraison n'a pas pu être vérifiée. "
                    "Veuillez la corriger ou accepter la suggestion proposée."
                ),
                "suggestions": address_check["suggestions"],
                "verdict": address_check["verdict"],
            },
        )

    # Priorité au body (évite un préflight CORS déclenché par un en-tête
    # personnalisé, qui échoue derrière certains ingress) — l'en-tête reste
    # accepté pour rétro-compatibilité.
    raw_idem_key = ((getattr(payload, "idempotency_key", None) or request.headers.get("Idempotency-Key", ""))).strip()
    if len(raw_idem_key) > 200:
        raise HTTPException(400, "Idempotency key is too long")
    user = await _resolve_user(request)
    customer_email = ((user["email"] if user else None) or (payload.email.lower().strip() if payload.email else None))
    actor_key = (user or {}).get("id") or customer_email or _client_ip(request)
    idem_key = (
        hmac.new(JWT_SECRET.encode(), f"{actor_key}:{raw_idem_key}".encode(), hashlib.sha256).hexdigest()
        if raw_idem_key else ""
    )
    lock_doc = None
    if idem_key:
        try:
            now = datetime.now(timezone.utc)
            lock_doc = {
                "_id": idem_key,
                "status": "processing",
                "created_at": now.isoformat(),
                "expires_at": now + timedelta(hours=24),
            }
            await db.idempotency.insert_one(lock_doc)
        except DuplicateKeyError:
            cached = await db.idempotency.find_one({"_id": idem_key}, {"_id": 0, "response": 1, "status": 1})
            if cached and cached.get("response"):
                return _checkout_response(cached["response"])
            for _ in range(50):
                await asyncio.sleep(0.05)
                cached = await db.idempotency.find_one({"_id": idem_key}, {"_id": 0, "response": 1, "status": 1})
                if cached and cached.get("response"):
                    return _checkout_response(cached["response"])
            raise HTTPException(409, "Checkout already in progress")

    try:
        line_items, subtotal, tax_rate, tax, shipping, total, discount, applied_coupon, has_preorder = await _build_order_totals(
            payload.items, payload.coupon_code, customer_email=customer_email
        )
    except TypeError as e:
        # Compat legacy pour les doubles/mocks de tests qui n'acceptent pas le
        # nouvel argument optionnel customer_email.
        if "customer_email" not in str(e):
            raise
        line_items, subtotal, tax_rate, tax, shipping, total, discount, applied_coupon, has_preorder = await _build_order_totals(
            payload.items, payload.coupon_code
        )

    order_id = str(uuid.uuid4())
    # Numérotation Fironova : FN-AAMMJJ-XXXXXX (l'ancien préfixe NP- venait
    # de NORDPEP ; les commandes existantes gardent leur numéro).
    order_number = f"FN-{datetime.now(timezone.utc).strftime('%y%m%d')}-{order_id[:6].upper()}"

    # Réservation atomique AVANT tout appel réseau au PSP. Ferme la fenêtre
    # check-then-act qui laissait deux clients acheter le même dernier flacon.
    reserved = await _reserve_stock_atomic(line_items)

    payment_info: dict = {}
    if payload.payment_method == "interac":
        payment_info = {
            "type": "interac",
            "instructions": {
                "send_to": INTERAC_EMAIL,
                "amount_cad": total,
                "reference": order_number,
                "security_question": "What is the brand name? (lowercase)",
                "security_answer_hint": INTERAC_PASSWORD_HINT.lower(),
            },
        }
        payment_status = "awaiting_etransfer"
    elif payload.payment_method == "nowpayments":
        try:
            np = await _nowpayments_create(order_id, total, payload.pay_currency or "btc")
        except Exception:
            await _release_stock_atomic(reserved)
            raise
        payment_info = {"type": "nowpayments", "provider_response": np}
        payment_status = "awaiting_crypto"
    else:  # defensive guard if a client sends an unsupported value
        await _release_stock_atomic(reserved)
        raise HTTPException(400, "Unsupported payment method")

    order_doc = {
        "id": order_id,
        "order_number": order_number,
        "user_id": user["id"] if user else None,
        "email": (user["email"] if user else None) or (payload.email.lower() if payload.email else None),
        "items": line_items,
        "subtotal": subtotal,
        "discount": discount,
        "coupon": applied_coupon,
        "tax_rate": tax_rate,
        "tax": tax,
        "shipping": shipping,
        "total": total,
        "currency": "CAD",
        "shipping_address": payload.shipping.model_dump(),
        # C2 — audit trail Google Maps Address Validation (voir _validate_shipping_address_google)
        "address_verified": bool(address_check.get("valid")),
        "address_suggestions": address_check.get("suggestions") or [],
        "address_verification_provider": address_check.get("provider", "disabled"),
        "shipping_info": {"carrier": "", "tracking_number": "", "shipped_at": None},
        "payment_method": payload.payment_method,
        "payment_status": payment_status,
        "payment_info": payment_info,
        "fulfillment_status": "preorder" if has_preorder else "pending",
        "has_preorder": has_preorder,
        "notes": [],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "payment_ttl_hours": UNPAID_ORDER_TTL_HOURS,
        "payment_deadline": (datetime.now(timezone.utc) + timedelta(hours=UNPAID_ORDER_TTL_HOURS)).isoformat(),
        "compliance": {
            "accept_terms": True,
            "confirm_age": True,
            "confirm_research_use": True,
            "ip": request.client.host if request.client else None,
        },
    }
    # --- AFFILIATE: attribution depuis le cookie fn_ref (champ additif) ---
    created_order = False
    _compensation_failures: list = []
    try:
        await affiliate_attach_to_order(order_doc, request)
        await db.orders.insert_one(order_doc)
        created_order = True
        await db.payment_transactions.insert_one({
            "id": str(uuid.uuid4()),
            "order_id": order_id,
            "session_id": idem_key or order_id,
            "status": payment_status,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    except Exception as _outer_err:
        # Compensation en ordre inverse. Chaque étape est capturée
        # individuellement pour ne pas masquer un échec de rollback.
        if created_order:
            try:
                await db.orders.delete_one({"id": order_id})
            except Exception as e:
                _compensation_failures.append({"stage": "delete_order", "error": f"{type(e).__name__}: {e}"})
        try:
            await _release_stock_atomic(reserved)
        except Exception as e:
            _compensation_failures.append({"stage": "release_stock", "error": f"{type(e).__name__}: {e}"})
        if idem_key:
            try:
                await db.idempotency.delete_one({"_id": idem_key})
            except Exception as e:
                _compensation_failures.append({"stage": "delete_idempotency", "error": f"{type(e).__name__}: {e}"})
        # Ledger B4 SMART : si l'une des compensations a échoué, on trace pour
        # le dashboard admin Reconciliation → Checkout failures.
        if _compensation_failures:
            try:
                await db.checkout_compensation_failures.insert_one({
                    "id": str(uuid.uuid4()),
                    "order_id": order_id,
                    "label": "checkout_writes",
                    "original_error": f"{type(_outer_err).__name__}: {_outer_err}",
                    "collections_written": ["orders", "payment_transactions"] if created_order else [],
                    "reserved_stock_items": reserved,  # noqa: safe pour audit
                    "idem_key": idem_key,
                    "failed_compensations": _compensation_failures,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "status": "pending_manual_review",
                    "assigned_admin": None,
                    "resolved_at": None,
                    "resolution_note": None,
                })
                _checkout_breaker_record_failure()
                try:
                    admin_to = os.environ.get("ADMIN_NOTIFICATION_EMAIL", ADMIN_EMAIL)
                    subject = f"[FIRONOVA] Checkout compensation failed — order {order_id}"
                    body = (
                        f"<p>Une compensation checkout a échoué et doit être réconciliée manuellement.</p>"
                        f"<ul><li>Order id : {order_id}</li>"
                        f"<li>Erreur d'origine : {type(_outer_err).__name__}</li>"
                        f"<li>Étapes en échec : {len(_compensation_failures)}</li></ul>"
                        f"<p>Voir <code>/ops-portal-fn7k2q → Reconciliation → Checkout failures</code>.</p>"
                    )
                    await _send_email(admin_to, subject, body)
                except Exception:
                    pass
            except Exception as ledger_err:
                logging.critical("[compensation] FAILED TO PERSIST LEDGER: %s", ledger_err)
        raise
    order_doc.pop("_id", None)

    # Persist idempotency key so replay returns same order without re-processing.
    if idem_key:
        try:
            await db.idempotency.update_one(
                {"_id": idem_key},
                {"$set": {"response": order_doc, "status": "completed"}},
            )
        except Exception:
            pass

    # Le stock est déjà réservé atomiquement plus haut (_reserve_stock_atomic),
    # avant l'appel au PSP. Rien à faire ici.
    # Le coupon n'est PAS décompté ici : il l'est à la confirmation de paiement
    # (_mark_order_paid), sinon les paniers abandonnés épuisent l'usage_limit.

    # Fire-and-forget order confirmation + admin notification
    response_doc = _checkout_response(order_doc)
    await send_order_confirmation(order_doc)

    return response_doc


def _customer_order_payload(order: dict) -> dict:
    public_order = dict(order)
    compliance = dict(public_order.get("compliance") or {})
    compliance.pop("ip", None)
    public_order["compliance"] = compliance
    public_order["notes"] = [
        dict(note) for note in (public_order.get("notes") or [])
        if note.get("visible_to_customer")
    ]
    payment_info = dict(public_order.get("payment_info") or {})
    provider = payment_info.get("provider_response")
    if isinstance(provider, dict):
        allowed_provider_fields = {
            "invoice_id", "invoice_url", "payment_id", "pay_address",
            "pay_amount", "pay_currency", "payment_status",
        }
        payment_info["provider_response"] = {
            key: value for key, value in provider.items() if key in allowed_provider_fields
        }
    public_order["payment_info"] = payment_info
    return public_order


async def my_orders(user: dict = Depends(get_current_user)):
    items = await db.orders.find(
        {"user_id": user["id"], "deleted_at": None}, {"_id": 0}
    ).sort("created_at", -1).to_list(200)
    return [_customer_order_payload(order) for order in items]


def _guest_order_access_token(order: dict) -> str:
    message = "\0".join([
        str(order.get("id") or ""),
        str(order.get("created_at") or ""),
        str(order.get("email") or "").strip().lower(),
    ])
    return hmac.new(JWT_SECRET.encode(), message.encode(), hashlib.sha256).hexdigest()


def _checkout_response(order: dict) -> dict:
    response = dict(order)
    if not response.get("user_id") and response.get("email"):
        response["guest_access_token"] = _guest_order_access_token(response)
    return response


def _guest_order_accessible(order: dict, request: Request) -> bool:
    provided = (request.headers.get("x-order-access-token") or "").strip()
    if not provided or not order.get("id") or not order.get("email") or not order.get("created_at"):
        return False
    try:
        created_at = datetime.fromisoformat(str(order["created_at"]).replace("Z", "+00:00"))
        if created_at.tzinfo is None:
            created_at = created_at.replace(tzinfo=timezone.utc)
        expires_at = created_at + timedelta(minutes=GUEST_ORDER_ACCESS_TTL_MINUTES)
        if datetime.now(timezone.utc) > expires_at:
            return False
    except (TypeError, ValueError):
        return False
    return secrets.compare_digest(_guest_order_access_token(order), provided)


async def request_guest_order_access(order_id: str, payload: GuestOrderAccessIn, request: Request):
    await _rate_limit(
        "guest_order_access_ip",
        _client_ip(request),
        5,
        3600,
        "Too many access requests. Try again later.",
    )
    await _rate_limit(
        "guest_order_access_order",
        _private_ref(order_id),
        5,
        3600,
        "Too many access requests. Try again later.",
    )
    response = {
        "ok": True,
        "message": "If the order and email match, an access link will be sent.",
    }
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order or order.get("user_id"):
        return response
    expected_email = str(order.get("email") or "").strip().lower()
    provided_email = str(payload.email).strip().lower()
    if not expected_email or not secrets.compare_digest(expected_email, provided_email):
        return response

    token = _guest_order_access_token(order)
    order_url = (
        f"{_trusted_public_base_url()}/order/{quote(str(order['id']), safe='')}"
        f"#access_token={token}"
    )
    safe_url = html.escape(order_url, quote=True)
    body = (
        "<p>Use this private link to view your order and tracking details.</p>"
        f'<p><a href="{safe_url}">View order</a></p>'
        "<p>This link expires seven days after checkout.</p>"
    )
    await _send_email(expected_email, "FIRONOVA - Private order access", body)
    return response


async def get_order(order_id: str, request: Request):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")

    user = await _resolve_user(request)
    if order.get("user_id"):
        if not user or (user["id"] != order["user_id"] and user.get("role") != "admin"):
            raise HTTPException(403, "Forbidden")
        if not (user and user.get("role") == "admin"):
            return _customer_order_payload(order)
        return order

    if not _guest_order_accessible(order, request):
        raise HTTPException(403, "Forbidden")
    order = _customer_order_payload(order)
    order["guest_access_used"] = True
    return order


async def order_tracking(order_id: str, request: Request):
    """Live Canada Post tracking for the order's tracking_number, when configured."""
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    user = await _resolve_user(request)
    if order.get("user_id"):
        if not user or (user["id"] != order["user_id"] and user.get("role") != "admin"):
            raise HTTPException(403, "Forbidden")
    elif not _guest_order_accessible(order, request):
        raise HTTPException(403, "Forbidden")
    pin = (order.get("shipping_info") or {}).get("tracking_number")
    if not pin:
        return {"tracked": False, "reason": "no_tracking_number"}
    live = await _canada_post_track(pin)
    if live:
        return {"tracked": True, "source": "canada_post_live", **live}
    return {"tracked": False, "reason": "unavailable_or_not_configured", "pin": pin}


# --- Refund workflow (Item 5 : A2+B2+C3+D2+E1) ---------------------------
class RefundRequestIn(BaseModel):
    reason: str = Field(min_length=10, max_length=1000)
    amount_requested: Optional[float] = None
    refund_type: str = Field(default="full", pattern="^(full|partial|store_credit)$")


class RefundDecisionIn(BaseModel):
    action: str = Field(pattern="^(approve|deny)$")
    approved_amount: Optional[float] = None
    approved_type: Optional[str] = Field(default=None, pattern="^(full|partial|store_credit)$")
    admin_note: str = ""


class RefundProcessedIn(BaseModel):
    tx_reference: str = Field(min_length=3)
    admin_note: str = ""


def _refund_eligibility_reason(order: dict) -> Optional[str]:
    if not order:
        return "Commande introuvable"
    if order.get("payment_status") != "paid":
        return "La commande n'est pas payée"
    if (order.get("fulfillment_status") or "") not in ("shipped", "delivered"):
        return "Le remboursement est possible après expédition uniquement"
    if order.get("refund_status") in ("requested", "approved", "processed"):
        return f"Une demande est déjà en cours (statut : {order.get('refund_status')})"
    now = datetime.now(timezone.utc)
    delivered_at = (order.get("shipping_info") or {}).get("delivered_at")
    if delivered_at:
        try:
            d = datetime.fromisoformat(delivered_at.replace("Z", "+00:00"))
            if (now - d).days > 14:
                return "Fenêtre de 14 jours après livraison dépassée"
        except ValueError:
            pass
    else:
        paid_at = order.get("paid_at")
        if paid_at:
            try:
                p = datetime.fromisoformat(paid_at.replace("Z", "+00:00"))
                if (now - p).days > 30:
                    return "Fenêtre de 30 jours après paiement dépassée"
            except ValueError:
                pass
    return None


async def order_request_refund(order_id: str, payload: RefundRequestIn, request: Request):
    await _rate_limit("refund_request", _client_ip(request), 5, 3600,
                       "Trop de demandes. Réessayez plus tard.")
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    user = await _resolve_user(request)
    is_owner = user and (order.get("user_id") == user.get("id") or user.get("role") == "admin")
    if not is_owner:
        if order.get("user_id") or not _guest_order_accessible(order, request):
            raise HTTPException(403, "Not authorized")
    ineligible = _refund_eligibility_reason(order)
    if ineligible:
        raise HTTPException(400, ineligible)
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.orders.update_one({"id": order_id}, {"$set": {
        "refund_status": "requested", "refund_requested_at": now_iso,
        "refund_reason": payload.reason.strip(),
        "refund_amount_requested": payload.amount_requested,
        "refund_type_requested": payload.refund_type,
    }})
    try:
        await _send_email(
            os.environ.get("ADMIN_NOTIFICATION_EMAIL", ADMIN_EMAIL),
            f"[FIRONOVA] Demande de remboursement — {order.get('order_number')}",
            f"<p>Commande <b>{order.get('order_number')}</b>. Raison : {payload.reason[:500]}</p>"
            f"<p>Voir /ops-portal-fn7k2q/refunds.</p>",
        )
    except Exception:
        pass
    return {"ok": True, "refund_status": "requested"}


async def admin_refund_decision(order_id: str, payload: RefundDecisionIn, admin: dict):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    if order.get("refund_status") != "requested":
        raise HTTPException(400, f"Cannot decide — status is {order.get('refund_status')}")
    now_iso = datetime.now(timezone.utc).isoformat()
    if payload.action == "deny":
        await db.orders.update_one({"id": order_id}, {"$set": {
            "refund_status": "denied", "refund_decided_at": now_iso,
            "refund_decided_by": admin.get("email"),
            "refund_admin_note": payload.admin_note or "",
        }})
        try:
            await _send_email(order.get("email", ""),
                "Votre demande de remboursement FIRONOVA",
                f"<p>Après examen, la demande sur <b>{order.get('order_number')}</b> n'a pas été approuvée.</p>"
                f"<p>Motif : {payload.admin_note or 'Non spécifié'}</p>")
        except Exception:
            pass
        return {"ok": True, "refund_status": "denied"}
    total = float(order.get("total") or 0)
    amount = payload.approved_amount if payload.approved_amount is not None else total
    if amount <= 0 or amount > total:
        raise HTTPException(400, f"Montant invalide (max {total} CAD)")
    approved_type = payload.approved_type or ("full" if amount == total else "partial")
    await db.orders.update_one({"id": order_id}, {"$set": {
        "refund_status": "approved", "refund_decided_at": now_iso,
        "refund_decided_by": admin.get("email"),
        "refund_approved_amount": amount, "refund_approved_type": approved_type,
        "refund_admin_note": payload.admin_note or "",
    }})
    try:
        await _send_email(order.get("email", ""),
            "Votre remboursement FIRONOVA est approuvé",
            f"<p>Approuvé : <b>{amount:.2f} CAD</b> ({approved_type}).</p>"
            f"<p>Vous recevrez un message dès l'envoi.</p>")
    except Exception:
        pass
    return {"ok": True, "refund_status": "approved", "approved_amount": amount, "approved_type": approved_type}


async def admin_refund_processed(order_id: str, payload: RefundProcessedIn, admin: dict):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    if order.get("refund_status") != "approved":
        raise HTTPException(400, f"Cannot mark processed — status is {order.get('refund_status')}")
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.orders.update_one({"id": order_id}, {"$set": {
        "refund_status": "processed", "refund_processed_at": now_iso,
        "refund_tx_reference": payload.tx_reference.strip(),
        "refund_processed_by": admin.get("email"),
        "refund_admin_note_processed": payload.admin_note or "",
    }})
    try:
        await _send_email(order.get("email", ""),
            "Remboursement FIRONOVA effectué",
            f"<p>Remboursement de <b>{order.get('order_number')}</b> envoyé.</p>"
            f"<p>Référence : <code>{payload.tx_reference}</code> — {order.get('refund_approved_amount','?')} CAD</p>")
    except Exception:
        pass
    return {"ok": True, "refund_status": "processed"}


async def admin_refunds_list(status: Optional[str] = None, limit: int = 50):
    q: dict = {"refund_status": {"$exists": True, "$ne": None}}
    if status:
        q["refund_status"] = status
    limit = max(1, min(200, int(limit)))
    cursor = db.orders.find(q, {
        "_id": 0, "id": 1, "order_number": 1, "email": 1, "total": 1,
        "fulfillment_status": 1, "paid_at": 1,
        "refund_status": 1, "refund_requested_at": 1, "refund_reason": 1,
        "refund_amount_requested": 1, "refund_type_requested": 1,
        "refund_approved_amount": 1, "refund_approved_type": 1,
        "refund_admin_note": 1, "refund_tx_reference": 1, "refund_processed_at": 1,
    }).sort("refund_requested_at", -1).limit(limit)
    return {"items": [d async for d in cursor], "total": await db.orders.count_documents(q)}


async def admin_sync_delivery_status(order_id: str,
                                     _admin: dict = Depends(require_area("orders", "manage"))):
    """Vérifie le repérage CP et passe la commande à 'delivered' si la livraison
    est confirmée par le transporteur."""
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")

    info = order.get("shipping_info") or {}
    pin = str(info.get("tracking_number") or "").strip()
    if not pin:
        raise HTTPException(400, "No tracking number on this order")

    live = await _canada_post_track(pin)
    if not live:
        # Sandbox: fallback temporel pour permettre les tests E2E malgré
        # l'auth tracking CP indisponible.
        shipped_at = str(info.get("shipped_at") or "")
        if _sandbox_fallback_ready(shipped_at):
            now = datetime.now(timezone.utc).isoformat()
            await db.orders.update_one(
                {"id": order_id, "fulfillment_status": "shipped"},
                {
                    "$set": {
                        "fulfillment_status": "delivered",
                        "shipping_info.delivered_at": now,
                        "shipping_info.delivery_source": "sandbox_time_fallback_manual",
                    },
                    "$push": {
                        "notes": {
                            "id": str(uuid.uuid4()),
                            "text": f"Statut livré (fallback sandbox après délai) — {pin}.",
                            "author": "system",
                            "created_at": now,
                        }
                    },
                },
            )
            updated = await db.orders.find_one({"id": order_id}, {"_id": 0, "fulfillment_status": 1, "shipping_info": 1})
            return {
                "ok": True,
                "tracked": False,
                "delivered": True,
                "updated": True,
                "reason": "sandbox_fallback",
                "fulfillment_status": (updated or {}).get("fulfillment_status"),
                "shipping_info": (updated or {}).get("shipping_info") or {},
            }
        return {
            "ok": True,
            "tracked": False,
            "updated": False,
            "reason": "tracking_unavailable",
            "fulfillment_status": order.get("fulfillment_status"),
        }

    delivered, evidence = _cp_tracking_indicates_delivered(live)
    if not delivered:
        return {
            "ok": True,
            "tracked": True,
            "delivered": False,
            "updated": False,
            "reason": "not_delivered_yet",
            "fulfillment_status": order.get("fulfillment_status"),
            "summary": live.get("summary"),
        }

    # Déjà livré -> rien à changer.
    if str(order.get("fulfillment_status") or "").lower() == "delivered":
        return {
            "ok": True,
            "tracked": True,
            "delivered": True,
            "updated": False,
            "reason": "already_delivered",
            "fulfillment_status": "delivered",
            "summary": live.get("summary"),
        }

    now = datetime.now(timezone.utc).isoformat()
    await db.orders.update_one(
        {"id": order_id},
        {
            "$set": {
                "fulfillment_status": "delivered",
                "shipping_info.delivered_at": now,
                "shipping_info.delivery_source": "canada_post_tracking",
            },
            "$push": {
                "notes": {
                    "id": str(uuid.uuid4()),
                    "text": f"Statut livré confirmé par repérage Canada Post ({pin}) — {evidence or 'delivered'}.",
                    "author": "system",
                    "created_at": now,
                }
            },
        },
    )
    updated = await db.orders.find_one({"id": order_id}, {"_id": 0, "fulfillment_status": 1, "shipping_info": 1})
    return {
        "ok": True,
        "tracked": True,
        "delivered": True,
        "updated": True,
        "fulfillment_status": (updated or {}).get("fulfillment_status"),
        "shipping_info": (updated or {}).get("shipping_info") or {},
        "summary": live.get("summary"),
    }


_ORDER_STATUS_GROUPS = {
    "active": {"fulfillment_status": {"$nin": ["delivered", "cancelled", "failed"]}},
    "completed": {"fulfillment_status": "delivered"},
    "cancelled": {"fulfillment_status": {"$in": ["cancelled", "failed"]}},
}


def _status_group_filter(status_group: Optional[str]) -> dict:
    return dict(_ORDER_STATUS_GROUPS.get(status_group or "", {}))


async def admin_orders(status_group: Optional[str] = None, _admin: dict = Depends(require_area("orders", "view"))):
    filt = _status_group_filter(status_group)
    filt["deleted_at"] = None
    return await db.orders.find(filt, {"_id": 0}).sort("created_at", -1).to_list(500)


async def admin_orders_page(
    page: int = 1,
    limit: int = 50,
    status_group: Optional[str] = None,
    query: Optional[str] = None,
    payment_status: Optional[str] = None,
    fulfillment_status: Optional[str] = None,
    late_only: bool = False,
    _admin: dict = Depends(require_area("orders", "view")),
):
    page = max(1, int(page))
    limit = max(1, min(int(limit), 100))
    filt = _status_group_filter(status_group)
    filt["deleted_at"] = None
    if payment_status:
        filt["payment_status"] = payment_status
    if fulfillment_status:
        filt["fulfillment_status"] = fulfillment_status
    if late_only:
        filt["late_payment_flagged"] = True
    search = (query or "").strip()
    if search:
        pattern = re.escape(search)
        filt["$or"] = [
            {"order_number": {"$regex": pattern, "$options": "i"}},
            {"email": {"$regex": pattern, "$options": "i"}},
            {"shipping_address.full_name": {"$regex": pattern, "$options": "i"}},
        ]
    total = await db.orders.count_documents(filt)
    items = await (
        db.orders.find(filt, {"_id": 0})
        .sort([("created_at", -1), ("id", -1)])
        .skip((page - 1) * limit)
        .to_list(limit)
    )
    return {"items": items, "total": total, "page": page, "limit": limit}


async def admin_delete_order(order_id: str, admin: dict = Depends(require_area("orders", "manage"))):
    """Suppression douce — la commande va en corbeille, restaurable, et n'est
    PAS purgée automatiquement (contrairement aux autres types de données),
    car une commande payée est une pièce comptable."""
    return await _soft_delete("orders", order_id, admin)


async def admin_order_counts(_admin: dict = Depends(require_area("orders", "view"))):
    out = {}
    for group, filt in _ORDER_STATUS_GROUPS.items():
        out[group] = await db.orders.count_documents(filt)
    out["all"] = await db.orders.count_documents({})
    return out


async def admin_dispatch_batch(
    date: Optional[str] = None, _admin: dict = Depends(require_area("orders", "view"))
):
    """
    Commandes du lot d'expédition d'un jour donné (défaut : prochain jour
    ouvrable courant). Payées, non annulées/expédiées/livrées, triées par paiement.
    """
    if not date:
        date = compute_dispatch_batch(datetime.now(timezone.utc))
    filt = {
        "dispatch_batch": date,
        "payment_status": "paid",
        "fulfillment_status": {"$nin": ["cancelled", "failed", "shipped", "delivered"]},
    }
    orders = (
        await db.orders.find(filt, {"_id": 0}).sort("paid_at", 1).to_list(500)
    )
    return {"batch_date": date, "count": len(orders), "orders": orders}


async def admin_update_order(
    order_id: str,
    payment_status: Optional[str] = None,
    fulfillment_status: Optional[str] = None,
    _admin: dict = Depends(require_area("orders", "manage")),
):
    valid_payment_statuses = {"awaiting_etransfer", "awaiting_crypto", "paid", "cancelled", "failed", "refunded"}
    if payment_status is not None and payment_status not in valid_payment_statuses:
        raise HTTPException(400, f"Invalid payment status: {payment_status}")

    existing = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Order not found")

    if payment_status == "paid":
        updated = await _mark_order_paid(order_id)
        if fulfillment_status:
            await db.orders.update_one({"id": order_id}, {"$set": {"fulfillment_status": fulfillment_status}})
            updated = await db.orders.find_one({"id": order_id}, {"_id": 0})
        return updated or existing

    update: dict = {}
    if payment_status:
        update["payment_status"] = payment_status
    if fulfillment_status:
        update["fulfillment_status"] = fulfillment_status
    if not update:
        raise HTTPException(400, "No fields to update")

    # GAP 1 & 2 — Effets de bord unifiés lorsque la commande passe à un
    # statut d'annulation. Restock + coupon + reverse de la commission
    # affiliée sont désormais centralisés dans `_cancel_order_side_effects()`
    # pour garantir la cohérence avec l'auto-cancel.
    now_iso = datetime.now(timezone.utc).isoformat()
    cancel_terminal = {"cancelled", "failed", "refunded"}
    payment_going_terminal = (
        payment_status in cancel_terminal
        and existing.get("payment_status") not in cancel_terminal
    )
    fulfillment_going_cancelled = (
        update.get("fulfillment_status") in ("cancelled", "failed")
        and existing.get("fulfillment_status") not in ("cancelled", "failed")
    )
    if payment_going_terminal or fulfillment_going_cancelled:
        # Reverse la commission uniquement si la commande était réellement
        # payée avant cette transition (cancel d'un awaiting → pas de commission).
        was_paid = existing.get("payment_status") == "paid"
        await _cancel_order_side_effects(existing, reverse_affiliate=was_paid)
        # Métadonnées pour audit / réouverture éventuelle
        if payment_going_terminal:
            update["cancelled_at"] = now_iso
            update["cancelled_reason"] = "admin_manual"
            update["prev_payment_status"] = existing.get("payment_status")

    await db.orders.update_one({"id": order_id}, {"$set": update})
    updated = await db.orders.find_one({"id": order_id}, {"_id": 0})
    return updated


async def admin_confirm_payment(order_id: str, _admin: dict = Depends(require_area("orders", "manage"))):
    """One-click 'Mark as Paid' — atomically marks order as paid + processing + sends email."""
    existing = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Order not found")
    if existing.get("payment_status") == "paid":
        return existing  # idempotent
    updated = await _mark_order_paid(order_id, "Payment manually confirmed by admin")
    return updated or existing


class ReopenOrderIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    mark_paid: bool = False
    note: Optional[str] = ""


class InteracReconcileMatchIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    order_number: str = Field(min_length=3, max_length=64)
    mark_paid: bool = True
    force_mismatch: bool = False
    note: Optional[str] = ""


class InteracReconcileDismissIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    note: Optional[str] = ""


async def _reconciliation_match_item(
    item_id: str,
    payload: InteracReconcileMatchIn,
    admin: dict,
):
    item = await db.interac_reconciliation_queue.find_one({"id": item_id}, {"_id": 0})
    if not item:
        raise HTTPException(404, "Reconciliation item not found")
    if item.get("status") != "pending":
        raise HTTPException(400, "Reconciliation item already processed")

    provider = str(item.get("provider") or "interac").lower()
    expected_method = "interac" if provider == "interac" else "nowpayments"

    order_number = payload.order_number.strip().upper()
    order = await db.orders.find_one({"order_number": order_number}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    if order.get("payment_method") != expected_method:
        raise HTTPException(400, f"Target order is not a {expected_method} order")

    amount_cad = item.get("amount_cad")
    amount_matches = True
    if amount_cad is not None:
        amount_matches = abs(float(amount_cad) - float(order.get("total", 0))) <= 0.01
        if not amount_matches and not payload.force_mismatch:
            raise HTTPException(409, "Amount mismatch: use force_mismatch to override")

    note_time = datetime.now(timezone.utc).isoformat()
    source_ref = (item.get("graph_message_id") or item.get("id") or "").strip()
    note = {
        "id": str(uuid.uuid4()),
        "text": (
            f"{provider.upper()} reconciliation matched manually by {admin['email']} "
            f"(queue {item.get('id')}, ref {source_ref})"
            + (f" — {payload.note.strip()}" if payload.note and payload.note.strip() else "")
        ),
        "author": admin["email"],
        "created_at": note_time,
    }
    await db.orders.update_one({"id": order["id"]}, {"$push": {"notes": note}})

    paid_marked = False
    late_flagged = False
    if payload.mark_paid:
        target_awaiting = "awaiting_etransfer" if expected_method == "interac" else "awaiting_crypto"
        if order.get("payment_status") == target_awaiting:
            fresh = await _mark_order_paid(
                order["id"],
                f"{provider.upper()} payment manually reconciled by {admin['email']} (queue {item.get('id')})",
            )
            paid_marked = bool(fresh)
        elif order.get("payment_status") == "cancelled":
            late_flagged = await _flag_late_cancelled_payment(order, f"{provider} manual reconcile", source_ref)

    now_iso = datetime.now(timezone.utc).isoformat()
    await db.interac_reconciliation_queue.update_one(
        {"id": item_id, "status": "pending"},
        {"$set": {
            "status": "matched",
            "matched_at": now_iso,
            "matched_by": admin["email"],
            "matched_order_id": order.get("id"),
            "matched_order_number": order_number,
            "marked_paid": paid_marked,
            "late_flagged": late_flagged,
            "amount_matches": amount_matches,
            "force_mismatch": bool(payload.force_mismatch),
            "match_note": (payload.note or "").strip(),
        }},
    )

    updated_item = await db.interac_reconciliation_queue.find_one({"id": item_id}, {"_id": 0})
    updated_order = await db.orders.find_one({"id": order["id"]}, {"_id": 0})
    return {"item": updated_item, "order": updated_order}


async def _reconciliation_dismiss_item(
    item_id: str,
    payload: InteracReconcileDismissIn,
    admin: dict = Depends(require_area("orders", "manage")),
):
    item = await db.interac_reconciliation_queue.find_one({"id": item_id}, {"_id": 0})
    if not item:
        raise HTTPException(404, "Reconciliation item not found")
    if item.get("status") != "pending":
        raise HTTPException(400, "Reconciliation item already processed")
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.interac_reconciliation_queue.update_one(
        {"id": item_id, "status": "pending"},
        {"$set": {
            "status": "dismissed",
            "dismissed_at": now_iso,
            "dismissed_by": admin["email"],
            "dismiss_note": (payload.note or "").strip(),
        }},
    )
    return await db.interac_reconciliation_queue.find_one({"id": item_id}, {"_id": 0})


async def admin_reopen_order(
    order_id: str,
    payload: ReopenOrderIn,
    admin: dict = Depends(require_area("orders", "manage")),
):
    """GAP 3 — Réouvre une commande annulée après réception tardive du paiement.
    - Ré-décrémente le stock (échoue si un item n'est plus disponible → conseille
      au staff de contacter le client).
    - Rétablit le compteur du coupon (best-effort, respecte usage_limit).
    - Restaure `payment_status` à sa valeur pré-annulation (`prev_payment_status`)
      ou passe directement à "paid" si `mark_paid=True`.
    - Enregistre une note d'audit avec l'auteur et le motif fourni."""
    existing = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Order not found")

    # Garde-fou permissions: cette action est plus sensible qu'une édition
    # commande classique; staff doit avoir le droit dédié.
    if not _has_area_permission(admin, "orders_reopen", "manage"):
        raise HTTPException(403, "Insufficient permissions")

    if existing.get("payment_status") != "cancelled":
        raise HTTPException(400, "Only cancelled orders can be reopened")

    # 1) Ré-décrément atomique du stock
    reserved: list[dict] = []
    for it in (existing.get("items") or []):
        if it.get("preorder"):
            continue
        qty = int(it.get("qty", 1))
        vid = it.get("variant_id")
        if vid in (None, "", "_default"):
            res = await db.products.update_one(
                {"id": it["product_id"], "stock": {"$gte": qty}},
                {"$inc": {"stock": -qty}},
            )
        else:
            res = await db.products.update_one(
                {"id": it["product_id"], "variants": {"$elemMatch": {"id": vid, "stock": {"$gte": qty}}}},
                {"$inc": {"variants.$[v].stock": -qty}},
                array_filters=[{"v.id": vid}],
            )
        if res.modified_count != 1:
            # Rollback des réservations déjà faites
            for rb in reserved:
                rb_qty = int(rb.get("qty", 1))
                rb_vid = rb.get("variant_id")
                if rb_vid in (None, "", "_default"):
                    await db.products.update_one({"id": rb["product_id"]}, {"$inc": {"stock": rb_qty}})
                else:
                    await db.products.update_one(
                        {"id": rb["product_id"], "variants.id": rb_vid},
                        {"$inc": {"variants.$.stock": rb_qty}},
                    )
            raise HTTPException(
                409,
                f"Cannot reopen: insufficient stock for '{it.get('name_en') or it.get('slug') or 'item'}'",
            )
        reserved.append(it)

    # 2) Rétablir le compteur du coupon si applicable
    await _claim_coupon_usage(existing)

    # 3) Rétablir le statut
    prev = existing.get("prev_payment_status") or "awaiting_etransfer"
    if prev not in {"awaiting_etransfer", "awaiting_crypto"}:
        prev = "awaiting_etransfer"
    now_iso = datetime.now(timezone.utc).isoformat()
    reopened_note_text = (
        f"Order reopened by {admin['email']}"
        + (f" — {payload.note.strip()}" if payload.note and payload.note.strip() else "")
    )
    await db.orders.update_one(
        {"id": order_id},
        {"$set": {
            "payment_status": prev,
            "fulfillment_status": "processing" if existing.get("has_preorder") is False else existing.get("fulfillment_status") or "processing",
            "reopened_at": now_iso,
            "late_payment_flagged": False,
         },
         "$unset": {"cancelled_at": "", "cancelled_reason": "", "prev_payment_status": ""},
         "$push": {"notes": {
            "id": str(uuid.uuid4()),
            "text": reopened_note_text,
            "author": admin["email"],
            "created_at": now_iso,
         }}},
    )

    # 4) Si mark_paid, on confirme immédiatement le paiement
    if payload.mark_paid:
        paid = await _mark_order_paid(order_id, f"Paid confirmed at reopen by {admin['email']}")
        return paid or await db.orders.find_one({"id": order_id}, {"_id": 0})

    return await db.orders.find_one({"id": order_id}, {"_id": 0})


# ---------------------------------------------------------------------------
# Gestion des membres staff — réservé au rôle "admin" (owner). Un staff ne
# peut jamais modifier ses propres permissions ni celles d'un autre membre.
# ---------------------------------------------------------------------------
STAFF_INVITE_TTL_HOURS = 72


# ---------------------------------------------------------------------------
# Journal d'audit — qui a fait quoi, quand. Deux sources d'entrées :
#  1. Automatique : toute action "manage" (mutation) passée par require_area()
#     est journalisée génériquement (méthode + route + zone + auteur).
#  2. Explicite : les actions sensibles sur les membres (invite, promotion,
#     permissions, révocation) sont journalisées avec un détail lisible,
#     puisqu'elles ne passent pas par require_area() (réservées owner-only).
# ---------------------------------------------------------------------------
async def _log_action(user: dict, action: str, detail: str = "", area: str = "") -> None:
    try:
        await db.admin_audit_log.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": user.get("id"),
            "user_email": user.get("email"),
            "user_name": user.get("name"),
            "role": user.get("role"),
            "area": area,
            "action": action,
            "detail": detail,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    except Exception as e:
        logging.error("[audit] failed to log action=%s: %s", action, e)


def _staff_invite_html(accept_url: str, inviter_name: str, lang: str = "fr") -> str:
    if lang == "fr":
        heading = "Vous êtes invité·e à rejoindre l'équipe FIRONOVA"
        body = (f"{inviter_name} vous invite à accéder au panneau d'administration FIRONOVA. "
                f"Cliquez ci-dessous pour créer votre mot de passe et activer votre accès. "
                f"Ce lien expire dans {STAFF_INVITE_TTL_HOURS} heures.")
        btn = "Activer mon accès"
    else:
        heading = "You've been invited to the FIRONOVA team"
        body = (f"{inviter_name} invited you to access the FIRONOVA admin panel. "
                f"Click below to set your password and activate your access. "
                f"This link expires in {STAFF_INVITE_TTL_HOURS} hours.")
        btn = "Activate my access"
    return f"""
    <div style="font-family:Georgia,serif;max-width:520px;margin:0 auto;padding:32px;color:#3A0A08;background:#FFFAF6">
      <div style="font-size:22px;font-weight:800;letter-spacing:-0.5px">FIRONOVA<span style="color:#C20114">.</span></div>
      <h2 style="margin-top:24px">{heading}</h2>
      <p style="line-height:1.6">{body}</p>
      <a href="{accept_url}" style="display:inline-block;margin-top:16px;background:#3A0A08;color:#fff;
         padding:14px 28px;text-decoration:none;font-family:monospace;font-size:12px;
         letter-spacing:0.2em;text-transform:uppercase">{btn} →</a>
      <p style="margin-top:24px;font-size:12px;color:#6B0504">{accept_url}</p>
    </div>
    """


async def admin_list_staff(_admin: dict = Depends(get_admin_user)):
    """Membres actifs (admin + staff) — les clients normaux (role=user) sont exclus."""
    staff = await db.users.find(
        {"role": {"$in": ["admin", "staff"]}},
        {"_id": 0, "password_hash": 0, "token_version": 0},
    ).sort("created_at", -1).to_list(200)
    return staff


async def admin_list_staff_invites(_admin: dict = Depends(get_admin_user)):
    """Invitations en attente — équivalent de l'onglet 'Demandes en attente'."""
    now = datetime.now(timezone.utc).isoformat()
    invites = await db.staff_invites.find(
        {"used": False, "expires_at": {"$gt": now}}, {"_id": 0, "token": 0}
    ).sort("created_at", -1).to_list(100)
    return invites


async def admin_invite_staff(payload: StaffInviteIn, request: Request, admin: dict = Depends(get_admin_user)):
    email = payload.email.lower().strip()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=409, detail="A user with this email already exists")
    await _rate_limit("staff_invite", admin["id"], 20, 3600, "Too many invitations sent. Try again later.")

    token = secrets.token_urlsafe(32)
    token_hash = _hash_refresh_token(token)
    now = datetime.now(timezone.utc)
    await db.staff_invites.insert_one({
        "id": str(uuid.uuid4()),
        "email": email,
        "name": payload.name.strip(),
        "permissions": payload.permissions.model_dump(),
        "as_owner": payload.as_owner,
        "invited_by": admin["id"],
        "token_hash": token_hash,
        "created_at": now.isoformat(),
        "expires_at": (now + timedelta(hours=STAFF_INVITE_TTL_HOURS)).isoformat(),
        "used": False,
    })
    base = _trusted_public_base_url()
    accept_url = f"{base}/staff-accept?token={token}"
    asyncio.create_task(_send_email(
        email, "FIRONOVA — Invitation à rejoindre l'équipe",
        _staff_invite_html(accept_url, admin.get("name", "L'équipe FIRONOVA")),
    ))
    await _log_action(admin, "staff.invite", detail=f"Invited {email} as {'owner' if payload.as_owner else 'staff'}")
    return {"ok": True, "sent_to": email}


async def admin_cancel_staff_invite(invite_id: str, admin: dict = Depends(get_admin_user)):
    res = await db.staff_invites.delete_one({"id": invite_id, "used": False})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="Invite not found")
    await _log_action(admin, "staff.invite_cancel", detail=f"Cancelled invite {invite_id}")
    return {"ok": True}


async def staff_accept_invite(payload: StaffAcceptIn, response: Response, request: Request):
    """Lien cliqué depuis l'email — pas d'authentification préalable, le
    token à usage unique + TTL sert de preuve. Crée le compte staff (ou
    owner, si l'invitation le précisait)."""
    now = datetime.now(timezone.utc).isoformat()
    token_hash = _hash_refresh_token(payload.token)
    invite = await db.staff_invites.find_one({"token_hash": token_hash, "used": False})
    if not invite or invite["expires_at"] < now:
        raise HTTPException(status_code=400, detail="Invalid or expired invitation")
    if await db.users.find_one({"email": invite["email"]}):
        raise HTTPException(status_code=409, detail="An account with this email already exists")

    as_owner = invite.get("as_owner", False)
    user_doc = {
        "id": str(uuid.uuid4()),
        "email": invite["email"],
        "name": invite["name"],
        "password_hash": hash_password(payload.password),
        "role": "admin" if as_owner else "staff",
        "token_version": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    if not as_owner:
        user_doc["permissions"] = invite["permissions"]
    consumed = await db.staff_invites.find_one_and_update(
        {"token_hash": token_hash, "used": False},
        {"$set": {"used": True, "used_at": now}},
        return_document=ReturnDocument.AFTER,
    )
    if not consumed:
        raise HTTPException(status_code=400, detail="Invalid or expired invitation")
    await db.users.insert_one(user_doc)

    role = user_doc["role"]
    await _start_session(response, request, user_doc)
    return {
        "id": user_doc["id"], "email": user_doc["email"], "name": user_doc["name"],
        "role": role, "permissions": user_doc.get("permissions"),
        # NOTE: auth cookie-only — pas de token dans le body.
    }


async def admin_update_staff_permissions(user_id: str, payload: StaffPermissionsIn,
                                         admin: dict = Depends(get_admin_user)):
    target = await db.users.find_one({"id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if target.get("role") == "admin":
        raise HTTPException(status_code=400, detail="Cannot restrict another admin's access this way")
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"role": "staff", "permissions": payload.model_dump()},
         "$inc": {"token_version": 1}},  # les changements de droits prennent effet immédiatement
    )
    await _log_action(admin, "staff.permissions_update",
                      detail=f"Updated permissions for {target.get('email')}: {payload.model_dump()}")
    return {"ok": True}


async def admin_promote_to_owner(user_id: str, admin: dict = Depends(get_admin_user)):
    """Fait passer un membre staff au rang de owner (accès total, y compris
    la gestion des autres membres). Réversible uniquement en révoquant puis
    ré-invitant — il n'y a volontairement pas de 'rétrograder un owner' pour
    éviter qu'un owner en soit accidentellement privé de son propre accès."""
    target = await db.users.find_one({"id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if target.get("role") == "admin":
        raise HTTPException(status_code=400, detail="This user is already an owner")
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"role": "admin"}, "$unset": {"permissions": ""}, "$inc": {"token_version": 1}},
    )
    await _log_action(admin, "staff.promote_to_owner", detail=f"Promoted {target.get('email')} to owner")
    return {"ok": True}


async def admin_revoke_staff(user_id: str, admin: dict = Depends(get_admin_user)):
    """Révoque l'accès admin — le compte redevient un compte client normal,
    n'est PAS supprimé (historique de commandes éventuel préservé)."""
    if user_id == admin["id"]:
        raise HTTPException(status_code=400, detail="You cannot revoke your own access")
    target = await db.users.find_one({"id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if target.get("role") == "admin":
        raise HTTPException(status_code=400, detail="Cannot revoke another admin this way")
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"role": "user"}, "$unset": {"permissions": ""}, "$inc": {"token_version": 1}},
    )
    await _log_action(admin, "staff.revoke", detail=f"Revoked access for {target.get('email')}")
    return {"ok": True}


async def admin_audit_log(limit: int = 200, admin: dict = Depends(get_admin_user)):
    """Journal d'audit — owner only. Couvre automatiquement toute action
    'manage' (mutation) sur les 35 endpoints admin, plus les actions
    explicites de gestion d'équipe (invite/promotion/permissions/révocation)."""
    limit = max(1, min(limit, 1000))
    entries = await db.admin_audit_log.find({}, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return entries


async def admin_customers(_admin: dict = Depends(require_area("customers", "view"))):
    """Clients enrichis : dépenses cumulées, nb de commandes payées, dernière commande,
    et segment de rétention déterministe (nouveau / actif / fidèle / à risque / inactif)."""
    users = await _cursor_all(db.users.find(
        {"role": {"$ne": "admin"}},
        {"_id": 0, "password_hash": 0, "token_version": 0},
    ).sort("created_at", -1))

    # Agrégation des commandes payées par email (une seule passe)
    agg = db.orders.aggregate([
        {"$match": {"payment_status": "paid", "email": {"$ne": None}}},
        {"$group": {
            "_id": "$email",
            "orders": {"$sum": 1},
            "spent": {"$sum": "$total"},
            "last_order": {"$max": "$created_at"},
        }},
    ])
    stats = {}
    async for row in agg:
        if row["_id"]:
            stats[row["_id"].lower()] = row

    now = datetime.now(timezone.utc)

    def _segment(orders, last_iso):
        if orders == 0:
            return "prospect"
        days = 999
        if last_iso:
            try:
                days = (now - datetime.fromisoformat(last_iso.replace("Z", "+00:00"))).days
            except Exception:
                days = 999
        if orders >= 3:
            return "loyal" if days <= 120 else "at_risk"
        if days <= 45:
            return "active"
        if days <= 120:
            return "cooling"
        return "dormant"

    out = []
    for u in users:
        st = stats.get((u.get("email") or "").lower())
        orders = st["orders"] if st else 0
        spent = round(st["spent"], 2) if st else 0.0
        last_order = st["last_order"] if st else None
        u["orders_count"] = orders
        u["total_spent"] = spent
        u["last_order_at"] = last_order
        u["segment"] = _segment(orders, last_order)
        out.append(u)

    # Compte par segment (pour les filtres UI)
    seg_counts = {}
    for u in out:
        seg_counts[u["segment"]] = seg_counts.get(u["segment"], 0) + 1

    return {"customers": out, "segment_counts": seg_counts, "total": len(out)}


async def admin_list_subscribers(status: Optional[str] = None, _admin: dict = Depends(require_area("subscribers", "view"))):
    query = {"status": status} if status else {}
    subs = await _cursor_all(db.subscribers.find(
        query, {"_id": 0, "unsubscribe_token": 0}
    ).sort("created_at", -1))
    return subs


async def admin_subscribers_csv(status: Optional[str] = None, _admin: dict = Depends(require_area("subscribers", "view"))):
    query = {"status": status} if status else {}
    cursor = db.subscribers.find(query, {"_id": 0, "unsubscribe_token": 0}).sort("created_at", -1)
    return _csv_cursor_response(
        cursor,
        lambda s: {
            "email": s.get("email", ""),
            "lang": s.get("lang", ""),
            "source": s.get("source", ""),
            "status": s.get("status", ""),
            "consent_at": s.get("consent_at", ""),
            "consent_ip": s.get("consent_ip", ""),
            "unsubscribed_at": s.get("unsubscribed_at") or "",
        },
        ["email", "lang", "source", "status", "consent_at", "consent_ip", "unsubscribed_at"],
        f"fironova-subscribers-{datetime.now().strftime('%Y%m%d')}.csv",
    )


async def _low_stock_variants(limit: int = 200) -> list:
    """Variantes actuellement sous leur seuil, lues en DIRECT depuis products.

    Source unique pour la tuile « stock faible » et pour le panneau du
    dashboard. L'ancien panneau lisait la collection `low_stock_alerts`, qui
    n'est alimentée que par ÉVÉNEMENT — une variante saisie à 2 unités
    directement dans l'admin ne déclenche rien et restait donc invisible.
    La collection reste en place : elle sert aux courriels d'alerte.
    """
    rows: list = []
    cursor = db.products.find(
        {"active": True, "deleted_at": None},
        {"_id": 0, "id": 1, "slug": 1, "name_en": 1, "name_fr": 1,
         "low_stock_threshold": 1, "variants": 1},
    )
    async for p in cursor:
        threshold = int(p.get("low_stock_threshold") or 10)
        for v in (p.get("variants") or []):
            stock = int(v.get("stock") or 0)
            if stock > threshold:
                continue
            rows.append({
                "product_id": p.get("id"),
                "product_slug": p.get("slug"),
                # product_name : conservé pour le widget existant, qui l'affiche
                # tel quel. Les deux langues suivent pour qu'il puisse localiser.
                "product_name": p.get("name_en") or p.get("name_fr") or p.get("slug") or "?",
                "product_name_en": p.get("name_en") or "",
                "product_name_fr": p.get("name_fr") or "",
                "variant_id": v.get("id"),
                "variant_name": v.get("name") or "",
                "variant_sku": v.get("sku") or "",
                "stock": stock,
                "threshold": threshold,
                "out_of_stock": stock <= 0,
            })
    # Le plus critique en premier : rupture d'abord, puis stock croissant.
    rows.sort(key=lambda r: (r["stock"], r["product_name"] or ""))
    return rows[:limit]


async def admin_stats(_admin: dict = Depends(require_area("dashboard", "view"))):
    total_orders = await db.orders.count_documents({})
    pending = await db.orders.count_documents({"fulfillment_status": "pending"})
    paid = await db.orders.count_documents({"payment_status": "paid"})
    users = await db.users.count_documents({"role": "user"})
    products = await db.products.count_documents({})
    # Le compte portait sur products.stock — un champ « legacy/fallback » qui
    # vaut souvent 0 puisque le stock réel vit sur les variantes. La tuile
    # annonçait donc un stock bas que le panneau du dashboard, lui, ne voyait
    # pas. Les deux lisent maintenant le même calcul, au niveau variante.
    low_stock = len(await _low_stock_variants())
    revenue_cursor = db.orders.aggregate([
        {"$match": {"payment_status": "paid"}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}}},
    ])
    revenue_doc = await revenue_cursor.to_list(1)
    revenue = revenue_doc[0]["total"] if revenue_doc else 0
    return {
        "total_orders": total_orders,
        "pending_orders": pending,
        "paid_orders": paid,
        "customers": users,
        "products": products,
        "low_stock": low_stock,
        "revenue_cad": round(revenue, 2),
    }


# ---------------------------------------------------------------------------
# Admin — order notes & shipping & stock
# ---------------------------------------------------------------------------
async def admin_add_order_note(order_id: str, payload: OrderNoteIn, admin: dict = Depends(require_area("orders", "manage"))):
    note = {
        "text": payload.text,
        "admin_email": admin["email"],
        "visible_to_customer": payload.visible_to_customer,
        "ts": datetime.now(timezone.utc).isoformat(),
    }
    res = await db.orders.update_one({"id": order_id}, {"$push": {"notes": note}})
    if res.matched_count == 0:
        raise HTTPException(404, "Order not found")
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if payload.visible_to_customer and order.get("email"):
        asyncio.create_task(send_customer_note_email(order, payload.text))
    return order


async def admin_refund_order(order_id: str, payload: RefundIn, admin: dict = Depends(require_area("orders", "manage"))):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    if order.get("payment_status") != "paid":
        raise HTTPException(409, "Only paid orders can be refunded")
    already = float(order.get("refunded_amount", 0) or 0)
    total = float(order.get("total", 0))
    amount = round(payload.amount, 2)
    if already + amount > total + 0.001:
        raise HTTPException(400, f"Refund exceeds order total (already refunded ${already:.2f} of ${total:.2f})")
    new_refunded = round(already + amount, 2)
    update = {"refunded_amount": new_refunded}
    if new_refunded >= total:
        update["payment_status"] = "refunded"
        update["fulfillment_status"] = "refunded"
    note = {
        "text": f"Refund issued: ${amount:.2f} CAD (total refunded ${new_refunded:.2f} / ${total:.2f})",
        "admin_email": admin["email"],
        "visible_to_customer": False,
        "ts": datetime.now(timezone.utc).isoformat(),
    }
    await db.orders.update_one({"id": order_id}, {"$set": update, "$push": {"notes": note}})
    if new_refunded >= total:
        if order.get("fulfillment_status") not in {"shipped", "delivered"}:
            await _restock_order_items(order)
            await db.orders.update_one({"id": order_id}, {"$set": {"refund_restocked": True}})
        await _decrement_coupon_usage(order)
        await affiliate_on_order_reversed(order_id, full=True)
    elif total > 0:
        remaining_ratio = max(0.0, (total - new_refunded) / total)
        now_iso = datetime.now(timezone.utc).isoformat()
        await db.affiliate_referrals.update_many(
            {"order_id": order_id, "status": {"$in": ["pending", "approved"]}},
            [
                {"$set": {
                    "original_commission_amount": {
                        "$ifNull": ["$original_commission_amount", "$commission_amount"]
                    },
                }},
                {"$set": {
                    "commission_amount": {
                        "$round": [{"$multiply": ["$original_commission_amount", remaining_ratio]}, 2]
                    },
                    "refund_adjusted_at": now_iso,
                }},
            ],
        )
    updated = await db.orders.find_one({"id": order_id}, {"_id": 0})
    provider_refund_id = (
        updated.get("provider_refund_id")
        or (updated.get("payment_info") or {}).get("provider_refund_id")
        or ((updated.get("payment_info") or {}).get("provider_response") or {}).get("refund_id")
    )
    if updated.get("email") and provider_refund_id:
        asyncio.create_task(send_refund_email(updated, amount, new_refunded))
    return updated


async def admin_resend_order_email(order_id: str, _admin: dict = Depends(require_area("orders", "manage"))):
    """Re-sends the order details / payment-instructions email to the customer."""
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    if not order.get("email"):
        raise HTTPException(400, "Order has no customer email")
    heading = "Payment received" if order.get("payment_status") == "paid" else "Order received"
    html = _order_email_html(order, heading)
    await _send_email(order["email"], f"FIRONOVA — Order {order['order_number']} details", html)
    return {"ok": True, "sent_to": order["email"]}


async def admin_set_shipping_info(order_id: str, payload: ShippingInfoIn, _admin: dict = Depends(require_area("orders", "manage"))):
    existing_order = await db.orders.find_one(
        {"id": order_id}, {"_id": 0, "shipping_info": 1, "payment_status": 1},
    )
    if not existing_order:
        raise HTTPException(404, "Order not found")
    if payload.tracking_number and existing_order.get("payment_status") != "paid":
        raise HTTPException(409, "Cannot ship an unpaid order")
    prev = existing_order.get("shipping_info") or {}

    shipped_at = payload.shipped_at or (datetime.now(timezone.utc).isoformat() if payload.tracking_number else None)
    # PIÈGE : ce PUT remplaçait shipping_info EN BLOC, ce qui effaçait label_url,
    # cp_group_id et surtout cp_transmitted. La commande sortait alors de la
    # requête du manifeste → surcharge de 2 $/article encourue en silence.
    # On repart donc de l'état existant et on ne surcharge que les champs manuels.
    shipping_info = {
        **prev,
        "carrier": payload.carrier or "",
        "tracking_number": payload.tracking_number or "",
        "shipped_at": shipped_at,
    }
    update = {"shipping_info": shipping_info}
    if payload.tracking_number:
        update["fulfillment_status"] = "shipped"
    res = await db.orders.update_one({"id": order_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(404, "Order not found")
    return await db.orders.find_one({"id": order_id}, {"_id": 0})


# ---------------------------------------------------------------------------
# BLOC 4 — Postes Canada : étiquettes, manifeste, annulation.
# Le tarif live et le suivi existaient déjà. Ce qui manquait : générer
# l'étiquette, et surtout TRANSMETTRE LE MANIFESTE. Sans manifeste, Postes
# Canada facture chaque envoi non payé avec 2 $ de surcharge par article et
# retire le rabais d'automatisation.
# ---------------------------------------------------------------------------
def _order_weight_kg(order: dict) -> float:
    """Poids d'après les line_items figés sur la commande (pas de relecture produit :
    le poids doit refléter ce qui a été vendu, même si la fiche a changé depuis)."""
    total_g = 0.0
    for it in _order_items(order):
        total_g += float(it.get("weight_grams") or 50.0) * int(it.get("qty") or 1)
    return max(0.1, round(total_g / 1000.0, 3)) if total_g else 0.5


async def admin_order_shipping_rates(order_id: str, _admin: dict = Depends(require_area("orders", "view"))):
    """Services disponibles pour CETTE commande, afin de peupler le sélecteur admin."""
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    if not is_canada_post_configured():
        return {"configured": False, "rates": []}
    ship = order.get("shipping_address") or {}
    rates = await _canada_post_get_rates(
        ship.get("postal_code", ""), ship.get("country") or "CA", _order_weight_kg(order)
    )
    return {"configured": True, "rates": rates}


async def admin_create_label(order_id: str, payload: CreateLabelIn,
                             _admin: dict = Depends(require_area("orders", "manage"))):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    if order.get("payment_status") != "paid":
        raise HTTPException(409, "Cannot create a label for an unpaid order")

    info = order.get("shipping_info") or {}
    # Idempotence exigée : ne jamais recréer une étiquette déjà émise —
    # sinon on paie deux fois le même colis.
    if info.get("label_url") and info.get("tracking_number"):
        return {"already_existed": True, "shipping_info": info}

    if not is_canada_post_configured():
        raise HTTPException(503, "Canada Post is not configured")

    res = await _canada_post_create_shipment(order, payload.service_code, _order_weight_kg(order))
    label_url = await _canada_post_get_artifact(res["label_href"], order_id)

    now = datetime.now(timezone.utc).isoformat()
    shipping_info = {
        **info,
        "carrier": "Canada Post",
        "tracking_number": res["pin"],
        "label_url": label_url,
        "cp_shipment_id": res["shipment_id"],
        # Coût réel de l'étiquette côté transporteur (pas le montant client).
        "cost": await _canada_post_shipment_price(res["shipment_id"], preferred_service_code=payload.service_code),
        "cp_group_id": res["group_id"],
        "cp_transmitted": False,   # tant que False → surcharge de 2 $ encourue
        "service_code": payload.service_code,
        "shipped_at": now,
    }
    await db.orders.update_one(
        {"id": order_id},
        {"$set": {"shipping_info": shipping_info, "fulfillment_status": "shipped"},
         "$push": {"notes": {
             "id": str(uuid.uuid4()),
             "text": f"Étiquette Postes Canada créée — suivi {res['pin']} (lot {res['group_id']}).",
             "author": "system",
             "created_at": now,
         }}},
    )
    fresh = await db.orders.find_one({"id": order_id}, {"_id": 0})
    asyncio.create_task(send_shipping_notification(fresh))
    return {"already_existed": False, "shipping_info": shipping_info}


async def admin_shipping_label(filename: str, _admin: dict = Depends(require_area("orders", "view"))):
    if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._-]*\.pdf", filename or ""):
        raise HTTPException(404, "Label not found")
    label_path = LABEL_UPLOAD_DIR / filename
    if not label_path.is_file():
        raise HTTPException(404, "Label not found")
    return FileResponse(
        label_path,
        media_type="application/pdf",
        filename=filename,
        headers={"Cache-Control": "private, no-store"},
    )


async def admin_void_label(order_id: str, _admin: dict = Depends(require_area("orders", "manage"))):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    info = order.get("shipping_info") or {}
    if info.get("cp_transmitted"):
        raise HTTPException(400, "Label already transmitted to Canada Post — it can no longer be voided.")
    ok = await _canada_post_void(info.get("cp_shipment_id", ""))
    if not ok:
        raise HTTPException(502, "Canada Post refused to void this label")
    await db.orders.update_one(
        {"id": order_id},
        {"$set": {"shipping_info": {"carrier": "", "tracking_number": "", "shipped_at": None},
                  "fulfillment_status": "processing"},
         "$push": {"notes": {
             "id": str(uuid.uuid4()),
             "text": "Étiquette Postes Canada annulée (void).",
             "author": "system",
             "created_at": datetime.now(timezone.utc).isoformat(),
         }}},
    )
    return {"ok": True, "voided": True}


async def admin_pending_manifest(_admin: dict = Depends(require_area("orders", "view"))):
    """Alimente la bannière rouge de l'admin : combien d'étiquettes non transmises."""
    return await pending_manifest_state()


async def admin_shipping_config_status(_admin: dict = Depends(require_area("shipping", "view"))):
    """Expose un état lisible de la config Postes Canada (sans secrets)
    pour faciliter le diagnostic côté interface admin."""
    using_openapi = _cp_use_openapi()
    mailed_by, mobo = _cp_path_customers()
    missing_required = []
    if using_openapi:
        if not mailed_by:
            missing_required.append("CANADA_POST_MAILED_BY/CANADA_POST_CUSTOMER_NUMBER")
        if not mobo:
            missing_required.append("CANADA_POST_MOBO/CANADA_POST_CUSTOMER_NUMBER")
        if not CANADA_POST_ORIGIN_POSTAL_CODE:
            missing_required.append("CANADA_POST_ORIGIN_POSTAL_CODE")
        if not CANADA_POST_OAUTH_CLIENT_ID:
            missing_required.append("CANADA_POST_OAUTH_CLIENT_ID")
        if not CANADA_POST_OAUTH_CLIENT_SECRET:
            missing_required.append("CANADA_POST_OAUTH_CLIENT_SECRET")
    else:
        if not CANADA_POST_API_KEY:
            missing_required.append("CANADA_POST_API_KEY")
        if not CANADA_POST_CUSTOMER_NUMBER:
            missing_required.append("CANADA_POST_CUSTOMER_NUMBER")
        if not CANADA_POST_ORIGIN_POSTAL_CODE:
            missing_required.append("CANADA_POST_ORIGIN_POSTAL_CODE")

    return {
        "configured": is_canada_post_configured(),
        "api_mode": CANADA_POST_API_MODE,
        "using_openapi": using_openapi,
        "environment": CANADA_POST_ENVIRONMENT,
        "base_url": CANADA_POST_OPENAPI_BASE_URL if using_openapi else CANADA_POST_BASE_URL,
        "has_legacy_api_key": bool(CANADA_POST_API_KEY),
        "has_customer_number": bool(CANADA_POST_CUSTOMER_NUMBER),
        "has_origin_postal_code": bool(CANADA_POST_ORIGIN_POSTAL_CODE),
        "has_oauth_client_id": bool(CANADA_POST_OAUTH_CLIENT_ID),
        "has_oauth_client_secret": bool(CANADA_POST_OAUTH_CLIENT_SECRET),
        "has_platform_id": bool(CANADA_POST_PLATFORM_ID),
        "has_mailed_by": bool(mailed_by),
        "has_mobo": bool(mobo),
        "missing_required": missing_required,
    }


async def admin_backfill_label_costs(limit: int = 300,
                                     force: bool = False,
                                     _admin: dict = Depends(require_area("shipping", "manage"))):
    """Récupère le coût réel des étiquettes existantes via shipment_id.

    - force=False: ne traite que les commandes sans shipping_info.cost
    - force=True: recalcule même si un coût est déjà présent
    """
    if not _cp_use_openapi():
        raise HTTPException(400, "Label cost backfill requires Canada Post OpenAPI mode")

    safe_limit = max(1, min(int(limit or 1), 2000))
    base_filter = {
        "shipping_info.cp_shipment_id": {"$nin": [None, ""]},
    }
    if not force:
        base_filter["$or"] = [
            {"shipping_info.cost": {"$exists": False}},
            {"shipping_info.cost": None},
            {"shipping_info.cost": {}},
        ]

    rows = await db.orders.find(
        base_filter,
        {"_id": 0, "id": 1, "order_number": 1, "shipping_info": 1},
    ).sort("created_at", -1).to_list(safe_limit)

    updated = 0
    unchanged = 0
    failed = []
    for o in rows:
        info = o.get("shipping_info") or {}
        shipment_id = info.get("cp_shipment_id")
        if not shipment_id:
            unchanged += 1
            continue

        try:
            price = await _canada_post_shipment_price(str(shipment_id))
            if not price:
                unchanged += 1
                continue

            await db.orders.update_one(
                {"id": o["id"]},
                {"$set": {"shipping_info.cost": price,
                          "shipping_info.cost_refreshed_at": datetime.now(timezone.utc).isoformat()}},
            )
            updated += 1
        except Exception as ex:
            failed.append({
                "order_id": o.get("id"),
                "order_number": o.get("order_number"),
                "error": str(ex),
            })

    return {
        "ok": True,
        "processed": len(rows),
        "updated": updated,
        "unchanged": unchanged,
        "failed": failed,
        "limit": safe_limit,
        "force": force,
    }


async def admin_transmit_manifest(_admin: dict = Depends(require_area("orders", "manage"))):
    """À faire CHAQUE JOUR. Oublier = 2 $/article de surcharge + perte du rabais."""
    if not is_canada_post_configured():
        raise HTTPException(503, "Canada Post is not configured")
    group_cursor = db.orders.aggregate([
        # $ne: True et non == False : une étiquette dont le champ est absent
        # est tout aussi non transmise, et restait invisible ici.
        {"$match": {
            "shipping_info.cp_group_id": {"$nin": [None, ""]},
            "shipping_info.cp_transmitted": {"$ne": True},
        }},
        {"$group": {"_id": "$shipping_info.cp_group_id"}},
        {"$sort": {"_id": 1}},
    ])
    group_ids = [group["_id"] async for group in group_cursor if group.get("_id")]
    if not group_ids:
        return {"ok": True, "transmitted_groups": [], "manifests": [], "orders_marked": 0}

    manifests: list = []
    done_groups: list = []
    for gid in group_ids:
        try:
            hrefs = await _canada_post_transmit(gid)
        except HTTPException:
            logging.error("Transmission du manifeste échouée pour le lot %s", gid)
            continue
        manifests.extend(hrefs)
        done_groups.append(gid)

    marked = 0
    if done_groups:
        res = await db.orders.update_many(
            {"shipping_info.cp_group_id": {"$in": done_groups},
             "shipping_info.cp_transmitted": {"$ne": True}},
            {"$set": {"shipping_info.cp_transmitted": True,
                      "shipping_info.cp_transmitted_at": datetime.now(timezone.utc).isoformat()}},
        )
        marked = res.modified_count
        # Détails de coût du manifeste : total réellement facturé par Postes
        # Canada pour la journée (rapprochement comptable).
        cost_details = None
        for href in manifests:
            cost_details = await _canada_post_manifest_details(href)
            if cost_details:
                break
        await db.manifests.insert_one({
            "id": str(uuid.uuid4()),
            "group_ids": done_groups,
            "manifest_links": manifests,
            "cost": cost_details,
            # Date du lot (TZ cutoff) : permet de retrouver le manifeste du jour.
            "dispatch_date": _local_today_iso(),
            "created_at": datetime.now(timezone.utc).isoformat(),
        })

    if not done_groups:
        raise HTTPException(502, "No manifest could be transmitted — check the Canada Post logs.")
    return {"ok": True, "transmitted_groups": done_groups, "manifests": manifests, "orders_marked": marked}


async def admin_dispatch_manifest_pdf(date: str,
                                      _admin: dict = Depends(require_area("orders", "view"))):
    """Manifeste(s) transmis pour la date donnée.
    Priorité: 1) PDF local sauvegardé, 2) hrefs CP en direct, 3) récapitulatif interne."""
    docs = await db.manifests.find({"dispatch_date": date}, {"_id": 0}).sort("created_at", 1).to_list(50)
    if not docs:
        raise HTTPException(404, "Aucun manifeste transmis pour cette date.")

    # 1) PDF déjà téléchargé localement (par retry-manifest)
    for d in docs:
        local_url = d.get("local_pdf_url") or ""
        if local_url and local_url.startswith("/uploads/labels/"):
            fpath = LABEL_UPLOAD_DIR / local_url.split("/")[-1]
            if fpath.exists():
                return Response(
                    content=fpath.read_bytes(),
                    media_type="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="manifeste-{date}.pdf"'},
                )

    # 2) Hrefs CP disponibles — tentative de téléchargement direct
    hrefs = []
    for d in docs:
        for h in (d.get("manifest_links") or []):
            if h and h not in hrefs:
                hrefs.append(h)

    pdfs = []
    for href in hrefs:
        try:
            r = await _cp_openapi_call("GET", href, accept="application/pdf")
            if r.status_code < 400 and r.content:
                pdfs.append(r.content)
        except Exception as ex:  # pragma: no cover
            logging.error("manifest artifact failed (%s): %s", href, ex)

    if pdfs:
        merged = _merge_pdfs(pdfs) if len(pdfs) > 1 else pdfs[0]
        return Response(
            content=merged,
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="manifeste-{date}.pdf"'},
        )

    # 3) Fallback : récapitulatif interne des commandes du lot
    orders = await db.orders.find(
        {"payment_status": "paid", "dispatch_batch": date,
         "shipping_info.label_url": {"$nin": [None, ""]}},
        {"_id": 0, "order_number": 1, "shipping_info": 1, "shipping_address": 1, "total": 1},
    ).sort("created_at", 1).to_list(500)
    if not orders:
        raise HTTPException(404, "Aucune commande étiquetée pour cette date.")

    lines = [
        f"<h1>Manifeste — lot du {date}</h1>",
        f"<p style='font-size:11px;color:#888'>Récapitulatif interne — utilisez « Récupérer manifeste CP » pour le document officiel</p>",
        "<table style='width:100%;border-collapse:collapse;font-family:monospace;font-size:11px'>",
        "<tr style='background:#f0f0f0'><th style='padding:4px;border:1px solid #ccc'>Commande</th>"
        "<th style='padding:4px;border:1px solid #ccc'>Destinataire</th>"
        "<th style='padding:4px;border:1px solid #ccc'>Suivi</th>"
        "<th style='padding:4px;border:1px solid #ccc'>Total</th></tr>",
    ]
    total_sum = 0.0
    for o in orders:
        info = o.get("shipping_info") or {}
        addr = o.get("shipping_address") or {}
        dest = f"{addr.get('full_name', '')} — {addr.get('city', '')}, {addr.get('province', '')}"
        pin = info.get("tracking_number", "")
        total = o.get("total") or 0
        total_sum += float(total)
        lines.append(
            f"<tr><td style='padding:4px;border:1px solid #ccc'>{o.get('order_number','')}</td>"
            f"<td style='padding:4px;border:1px solid #ccc'>{dest}</td>"
            f"<td style='padding:4px;border:1px solid #ccc'>{pin}</td>"
            f"<td style='padding:4px;border:1px solid #ccc;text-align:right'>${float(total):.2f}</td></tr>"
        )
    lines += [
        f"<tr style='font-weight:bold'><td colspan='3' style='padding:4px;border:1px solid #ccc;text-align:right'>Total</td>"
        f"<td style='padding:4px;border:1px solid #ccc;text-align:right'>${total_sum:.2f}</td></tr>",
        "</table>",
        f"<p style='font-size:10px;margin-top:12px'>{len(orders)} envoi(s) · Postes Canada — transmis le {date}</p>",
    ]
    html = (
        "<!DOCTYPE html><html><head><meta charset='utf-8'>"
        "<style>body{font-family:sans-serif;margin:24px;color:#111}</style></head>"
        f"<body>{''.join(lines)}</body></html>"
    )
    from weasyprint import HTML as WP
    try:
        pdf_bytes = WP(string=html).write_pdf()
    except Exception:
        return Response(
            content=html.encode("utf-8"),
            media_type="text/html",
            headers={"Content-Disposition": f'inline; filename="manifeste-{date}.html"'},
        )
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="manifeste-{date}.pdf"'},
    )


async def admin_retry_manifest(date: str,
                               _admin: dict = Depends(require_area("orders", "manage"))):
    """Récupère le manifeste officiel Postes Canada pour la date donnée.
    Appelle GET /manifests pour lister les vrais hrefs existants, puis
    télécharge le PDF via le flux en 2 étapes (JSON → artifact → PDF).
    N'essaie pas de re-transmettre les lots déjà transmis."""
    if not is_canada_post_configured():
        raise HTTPException(503, "Canada Post is not configured")
    if not _cp_use_openapi():
        raise HTTPException(400, "retry-manifest nécessite le mode OpenAPI")

    mailed_by, mobo = _cp_path_customers()

    # Étape 1: lister les manifests existants via GET /manifests
    try:
        r = await _cp_openapi_call("GET", f"/{mailed_by}/{mobo}/manifests")
    except Exception as ex:
        logging.error("Canada Post manifest lookup failed: %s", type(ex).__name__)
        raise HTTPException(502, "Impossible de contacter Canada Post") from ex
    if r.status_code >= 400:
        logging.error("Canada Post GET /manifests status=%s response_ref=%s", r.status_code, _private_ref(r.text))
        raise HTTPException(502, f"Canada Post GET /manifests failed ({r.status_code})")

    raw = r.json() if r.headers.get("content-type","").startswith("application/json") else []
    if not isinstance(raw, list):
        raw = []
    manifest_hrefs = [l.get("href") for l in raw
                      if isinstance(l, dict) and l.get("rel") == "manifest" and l.get("href")]
    if not manifest_hrefs:
        # Fallback: essayer aussi les hrefs stockés en base pour cette date
        manifest_doc = await db.manifests.find_one({"dispatch_date": date}, {"_id": 0})
        stored_hrefs = [h for h in (manifest_doc or {}).get("manifest_links", [])
                        if h and "workgroup1" not in h and "0000000001" not in h]
        manifest_hrefs = stored_hrefs

    if not manifest_hrefs:
        raise HTTPException(404, "Aucun manifeste trouvé auprès de Canada Post pour cette date.")

    # Étape 2: télécharger le PDF de chaque manifest et sauvegarder le premier valide
    local_pdf_url: Optional[str] = None
    for href in manifest_hrefs:
        pdf_url = await _canada_post_get_manifest_artifact_openapi(href, date)
        if pdf_url:
            local_pdf_url = pdf_url
            logging.info("Manifest PDF saved for %s: %s", date, pdf_url)
            break

    # Mettre à jour le document manifeste en base
    now = datetime.now(timezone.utc).isoformat()
    group_ids = await db.orders.distinct(
        "shipping_info.cp_group_id",
        {"payment_status": "paid", "dispatch_batch": date,
         "shipping_info.cp_group_id": {"$nin": [None, ""]}}
    )
    await db.manifests.update_one(
        {"dispatch_date": date},
        {"$set": {
            "manifest_links": manifest_hrefs,
            "local_pdf_url": local_pdf_url,
            "refreshed_at": now,
            **({"group_ids": group_ids} if group_ids else {}),
        }},
        upsert=True,
    )

    return {
        "ok": True,
        "date": date,
        "manifest_hrefs": manifest_hrefs,
        "local_pdf_url": local_pdf_url,
        "pdf_source": "cp" if local_pdf_url else "fallback",
    }


async def admin_dispatch_manifest_status(date: str,
                                         _admin: dict = Depends(require_area("orders", "view"))):
    """Indique si un manifeste a été transmis pour cette date (pilote le bouton
    de téléchargement dans l'écran Dispatch)."""
    count = await db.manifests.count_documents({"dispatch_date": date})
    # Vérifie aussi si le PDF manifeste est déjà sauvegardé localement
    doc = await db.manifests.find_one({"dispatch_date": date}, {"_id": 0})
    local_pdf = (doc or {}).get("local_pdf_url")
    return {"date": date, "transmitted": count > 0, "manifests": count, "local_pdf_url": local_pdf}


# ---------------------------------------------------------------------------
# Dispatch batch — traitement des commandes payées par fenêtre journalière
# (cutoff 13h ET). Modèle ShipStation : on regroupe par dispatch_batch, on
# génère toutes les étiquettes d'un lot en une passe, puis on transmet le
# manifeste une fois par jour (voir /admin/shipping/transmit).
# ---------------------------------------------------------------------------
def _iso_day_shift(day: str, delta: int) -> str:
    """Renvoie la date ISO décalée de <delta> jours (bornes de recherche UTC)."""
    try:
        d = datetime.strptime(day, "%Y-%m-%d").date() + timedelta(days=delta)
        return d.isoformat()
    except Exception:
        return day


def _local_today_iso() -> str:
    """Date locale (ORDER_CUTOFF_TZ) au format YYYY-MM-DD."""
    return datetime.now(ZoneInfo(ORDER_CUTOFF_TZ)).date().isoformat()


class DispatchLabelsIn(BaseModel):
    service_code: str = Field(default="DOM.EP", min_length=1, max_length=40)


async def admin_dispatch_today(date: Optional[str] = None,
                               service_code: Optional[str] = None,
                               _admin: dict = Depends(require_area("orders", "view"))):
    """File d'expédition du jour : commandes payées dont le dispatch_batch
    tombe le <date> (défaut = aujourd'hui, TZ cutoff)."""
    day = date or _local_today_iso()
    cursor = db.orders.find(
        {
            "payment_status": "paid",
            "$or": [
                {
                    # File de travail : ce qui reste à traiter pour ce lot.
                    "dispatch_batch": day,
                    "fulfillment_status": {"$in": ["processing", "pending", "packed", "shipped"]},
                },
                {
                    # HISTORIQUE : étiquettes réellement émises ce jour-là.
                    # Indispensable car le report de minuit modifie
                    # dispatch_batch, alors que shipped_at ne bouge jamais.
                    # Fenêtre élargie (J-1 → J+1) car shipped_at est en UTC :
                    # le tri fin par heure locale est fait plus bas.
                    "shipping_info.shipped_at": {"$gte": _iso_day_shift(day, -1),
                                                 "$lte": _iso_day_shift(day, 1) + "T23:59:59"},
                },
            ],
        },
        {"_id": 0},
    ).sort("created_at", 1)
    orders = await cursor.to_list(2000)

    to_label, labeled = [], []
    selected_service = (service_code or CANADA_POST_DEFAULT_SERVICE_CODE or "DOM.EP").strip().upper()
    rate_cache: dict[tuple, Any] = {}

    # Charger les emballages actifs une seule fois pour tout le lot.
    available_boxes = await db.shipping_boxes.find(
        {"deleted_at": None, "active": True}, {"_id": 0}
    ).sort("max_units", 1).to_list(200)

    for o in orders:
        info = o.get("shipping_info") or {}
        cost_due = (info.get("cost") or {}).get("due_amount")
        row = {
            "id": o["id"],
            "order_number": o.get("order_number"),
            "email": o.get("email"),
            "items": len(o.get("items", [])),
            "total": o.get("total"),
            "shipping_charged": o.get("shipping"),
            "city": (o.get("shipping_address") or {}).get("city"),
            "province": (o.get("shipping_address") or {}).get("province"),
            "tracking_number": info.get("tracking_number") or "",
            "label_url": info.get("label_url") or "",
            "cp_transmitted": bool(info.get("cp_transmitted")),
            # Coût réel facturé par Postes Canada pour cette étiquette.
            "cost_due": cost_due,
            "rated_weight_kg": (info.get("cost") or {}).get("rated_weight_kg"),
            # Champ unifié pour affichage ligne par ligne dans Dispatch.
            "line_label_cost": cost_due,
            "line_label_cost_source": "actual_cp" if cost_due is not None else None,
        }
        # « Étiquetées » = étiquettes émises CE JOUR (shipped_at), pas les
        # Critère principal : dispatch_batch == day (lot commandé ce jour).
        # Critère secondaire : shipped_at tombe ce jour en heure locale.
        # Les deux sont valides — évite la perte silencieuse des commandes dont
        # shipped_at est en UTC (ex. 00h32 UTC = veille en heure locale).
        shipped_at = info.get("shipped_at") or ""
        labeled_today = o.get("dispatch_batch") == day  # critère primaire
        if not labeled_today and shipped_at:
            try:
                shipped_dt = datetime.fromisoformat(shipped_at.replace("Z", "+00:00"))
                if shipped_dt.tzinfo is None:
                    shipped_dt = shipped_dt.replace(tzinfo=timezone.utc)
                labeled_today = shipped_dt.astimezone(ZoneInfo(ORDER_CUTOFF_TZ)).date().isoformat() == day
            except Exception:
                labeled_today = shipped_at.startswith(day)
        if row["label_url"] and row["tracking_number"]:
            if labeled_today:
                labeled.append(row)
        else:
            # Ne montrer en "À étiqueter" que les commandes qui ont terminé
            # le flux Journée (état packed).
            if o.get("fulfillment_status") != "packed":
                continue
            # Estimation du prix Canada Post (avant création réelle d'étiquette)
            # selon le service sélectionné dans l'écran Dispatch.
            row["estimated_cost_due"] = None
            row["estimated_eta_days"] = None
            if is_canada_post_configured():
                ship = o.get("shipping_address") or {}
                dest_pc = str(ship.get("postal_code") or "").replace(" ", "").upper()
                dest_country = str(ship.get("country") or "CA").upper()
                weight_kg = _order_weight_kg(o)
                if _cp_use_openapi():
                    cache_key = ("openapi", o.get("id"), selected_service)
                    if cache_key not in rate_cache:
                        rate_cache[cache_key] = await _canada_post_estimate_openapi(o, selected_service, weight_kg)
                    chosen = rate_cache.get(cache_key)
                    if chosen is not None:
                        row["estimated_cost_due"] = chosen.get("cost_cad")
                        row["estimated_eta_days"] = chosen.get("eta_days")
                        row["line_label_cost"] = chosen.get("cost_cad")
                        chosen_code = str(chosen.get("service_code") or "").upper()
                        row["line_label_cost_source"] = (
                            "estimated_cp"
                            if (not chosen_code) or (chosen_code == selected_service)
                            else "estimated_cp_alt"
                        )
                else:
                    cache_key = (dest_pc, dest_country, weight_kg)
                    if cache_key not in rate_cache:
                        rate_cache[cache_key] = await _canada_post_get_rates(dest_pc, dest_country, weight_kg)
                    rates = rate_cache.get(cache_key) or []
                    chosen = next((r for r in rates if str(r.get("service_code") or "").upper() == selected_service), None)
                    # Si le service exact n'est pas renvoyé, on prend un devis CP
                    # alternatif plutôt qu'un montant checkout interne.
                    if chosen is None and rates:
                        chosen = rates[0]
                    if chosen is not None:
                        row["estimated_cost_due"] = chosen.get("cost_cad")
                        row["estimated_eta_days"] = chosen.get("eta_days")
                        row["line_label_cost"] = chosen.get("cost_cad")
                        chosen_code = str(chosen.get("service_code") or "").upper()
                        row["line_label_cost_source"] = (
                            "estimated_cp"
                            if chosen_code == selected_service
                            else "estimated_cp_alt"
                        )
            # Emballage sélectionné (auto ou override) pour cette commande.
            chosen_box = await _select_box_for_order(o, all_boxes=available_boxes)
            row["box_id"] = chosen_box.get("id") if chosen_box else None
            row["box_name"] = chosen_box.get("name") if chosen_box else None
            to_label.append(row)

    overdue = await db.orders.count_documents({
        "payment_status": "paid",
        "dispatch_batch": {"$lt": day},
        "fulfillment_status": {"$in": ["processing", "pending"]},
    })
    # Bilan financier du jour (règle métier Dispatch) :
    # - estimated_labels_total: somme des coûts par ligne de commande affichée
    #   dans Dispatch ("à étiqueter" + "étiquetées").
    # - customer_charged_total: somme des frais facturés aux clients (champ shipping de la facture)
    # - gap_total: écart = manifeste total dû - facturé aux clients
    dispatch_lines = [*to_label, *labeled]
    estimated_labels_total = round(
        sum(float(r.get("line_label_cost") or 0) for r in dispatch_lines if r.get("line_label_cost") is not None),
        2,
    )
    customer_charged_total = round(sum(float(r.get("shipping_charged") or 0) for r in labeled), 2)
    manifest_doc = await db.manifests.find_one({"dispatch_date": day}, {"_id": 0}, sort=[("created_at", -1)])
    manifest_total_due = None
    try:
        manifest_total_due = float(((manifest_doc or {}).get("cost") or {}).get("total_due"))
    except Exception:
        manifest_total_due = None
    gap_total = (round(manifest_total_due - customer_charged_total, 2)
                 if manifest_total_due is not None else None)
    return {
        "date": day,
        "configured": is_canada_post_configured(),
        "totals": {
            "labels_cost": estimated_labels_total,
            "shipping_charged": round(sum(float(r.get("cost_due") or 0) for r in labeled), 2),
            "customer_shipping_charged": customer_charged_total,
            "margin": gap_total,
            "manifest": (manifest_doc or {}).get("cost"),
        },
        "counts": {"to_label": len(to_label), "labeled": len(labeled), "overdue": overdue},
        "to_label": to_label,
        "labeled": labeled,
        "available_boxes": [
            {"id": b.get("id"), "name": b.get("name"), "max_units": b.get("max_units"),
             "tare_grams": b.get("tare_grams"), "length_cm": b.get("length_cm"),
             "width_cm": b.get("width_cm"), "height_cm": b.get("height_cm")}
            for b in available_boxes
        ],
    }


class OrderShippingBoxIn(BaseModel):
    box_id: Optional[str] = None  # None = réinitialiser à la sélection automatique


async def admin_order_set_shipping_box(
    order_id: str,
    payload: OrderShippingBoxIn,
    _admin: dict = Depends(require_area("orders", "manage")),
):
    """Définit (ou réinitialise) l'emballage d'une commande en vue de l'étiquetage."""
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")

    if payload.box_id:
        box = await db.shipping_boxes.find_one({"id": payload.box_id, "active": True, "deleted_at": None}, {"_id": 0})
        if not box:
            raise HTTPException(404, "Packaging not found or inactive")
        await db.orders.update_one({"id": order_id}, {"$set": {"shipping_box_override_id": payload.box_id}})
        return {"order_id": order_id, "box_id": payload.box_id, "box_name": box.get("name")}
    else:
        await db.orders.update_one({"id": order_id}, {"$unset": {"shipping_box_override_id": ""}})
        return {"order_id": order_id, "box_id": None, "box_name": None}


async def admin_order_refresh_dispatch_estimate(
    order_id: str,
    service_code: Optional[str] = None,
    _admin: dict = Depends(require_area("orders", "view")),
):
    """Recalcule l'estimation CP d'une ligne Dispatch avec le poids produits + tare emballage."""
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")

    selected_service = (service_code or CANADA_POST_DEFAULT_SERVICE_CODE or "DOM.EP").strip().upper()
    chosen_box = await _select_box_for_order(order)
    products_weight_kg = _order_weight_kg(order)
    box_tare_kg = round(float((chosen_box or {}).get("tare_grams") or 0.0) / 1000.0, 3)
    total_weight_kg = max(0.1, round(products_weight_kg + box_tare_kg, 3))

    out = {
        "order_id": order_id,
        "service_requested": selected_service,
        "line_label_cost": None,
        "line_label_cost_source": None,
        "estimated_cost_due": None,
        "estimated_eta_days": None,
        "box_id": (chosen_box or {}).get("id"),
        "box_name": (chosen_box or {}).get("name"),
        "products_weight_kg": products_weight_kg,
        "box_tare_kg": box_tare_kg,
        "packaged_weight_kg": total_weight_kg,
    }

    if not is_canada_post_configured():
        return out

    ship = order.get("shipping_address") or {}
    dest_pc = str(ship.get("postal_code") or "").replace(" ", "").upper()
    dest_country = str(ship.get("country") or "CA").upper()

    if _cp_use_openapi():
        chosen = await _canada_post_estimate_openapi(order, selected_service, products_weight_kg)
    else:
        rates = await _canada_post_get_rates(dest_pc, dest_country, total_weight_kg)
        chosen = next((r for r in rates if str(r.get("service_code") or "").upper() == selected_service), None)
        if chosen is None and rates:
            chosen = rates[0]

    if chosen is None:
        return out

    out["estimated_cost_due"] = chosen.get("cost_cad")
    out["estimated_eta_days"] = chosen.get("eta_days")
    out["line_label_cost"] = chosen.get("cost_cad")
    chosen_code = str(chosen.get("service_code") or "").upper()
    out["line_label_cost_source"] = "estimated_cp" if (not chosen_code) or (chosen_code == selected_service) else "estimated_cp_alt"
    return out


async def admin_dispatch_labels(date: str, payload: DispatchLabelsIn,
                                _admin: dict = Depends(require_area("orders", "manage"))):
    """Génère en une passe les étiquettes des commandes payées du lot <date>
    qui n'en ont pas encore. Idempotent : commande déjà étiquetée = sautée."""
    if not is_canada_post_configured():
        raise HTTPException(503, "Canada Post is not configured")

    orders = db.orders.find(
        {
            "payment_status": "paid",
            "dispatch_batch": date,
            # Garde-fou métier: seule une commande entièrement préparée
            # (état "packed" dans l'écran Journée) peut passer en Dispatch.
            "fulfillment_status": "packed",
        },
        {"_id": 0},
    ).sort("created_at", 1)

    created, skipped, failed = [], [], []
    async for order in orders:
        if order.get("fulfillment_status") != "packed":
            skipped.append({
                "order_number": order.get("order_number"),
                "reason": "not_fully_processed_in_today",
            })
            continue
        info = order.get("shipping_info") or {}
        if info.get("label_url") and info.get("tracking_number"):
            skipped.append({"order_number": order.get("order_number"),
                            "tracking_number": info["tracking_number"],
                            "label_url": info["label_url"]})
            continue
        try:
            res = await _canada_post_create_shipment(order, payload.service_code, _order_weight_kg(order))
            label_url = await _canada_post_get_artifact(res["label_href"], order["id"])
            now = datetime.now(timezone.utc).isoformat()
            shipping_info = {
                **info,
                "carrier": "Canada Post",
                "tracking_number": res["pin"],
                "label_url": label_url,
                "cp_shipment_id": res["shipment_id"],
                "cost": await _canada_post_shipment_price(res["shipment_id"], preferred_service_code=payload.service_code),
                "cp_group_id": res["group_id"],
                "cp_transmitted": False,
                "service_code": payload.service_code,
                "shipped_at": now,
            }
            await db.orders.update_one(
                {"id": order["id"]},
                {"$set": {"shipping_info": shipping_info, "fulfillment_status": "shipped"},
                 "$push": {"notes": {
                     "id": str(uuid.uuid4()),
                     "text": f"Étiquette Postes Canada créée (lot {date}) — suivi {res['pin']}.",
                     "author": "system",
                     "created_at": now,
                 }}},
            )
            fresh = await db.orders.find_one({"id": order["id"]}, {"_id": 0})
            asyncio.create_task(send_shipping_notification(fresh))
            created.append({"order_number": order.get("order_number"),
                            "tracking_number": res["pin"],
                            "label_url": label_url})
        except HTTPException as ex:
            failed.append({"order_number": order.get("order_number"), "error": ex.detail})
        except Exception as ex:
            logging.error("[dispatch] label failed for %s: %s", order.get("order_number"), ex)
            failed.append({"order_number": order.get("order_number"), "error": str(ex)})

    all_labels = [c["label_url"] for c in created] + [s["label_url"] for s in skipped]
    return {
        "date": date,
        "created": created,
        "skipped": skipped,
        "failed": failed,
        "label_urls": all_labels,
        "counts": {"created": len(created), "skipped": len(skipped), "failed": len(failed)},
    }


def _merge_pdfs(pdf_bytes_list: list) -> bytes:
    """Fusionne plusieurs PDF en un seul. Requiert pypdf."""
    from pypdf import PdfWriter, PdfReader
    import io
    writer = PdfWriter()
    for b in pdf_bytes_list:
        if not b:
            continue
        reader = PdfReader(io.BytesIO(b))
        for page in reader.pages:
            writer.add_page(page)
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


async def _dispatch_labeled_orders(date: str) -> list:
    """Commandes du lot <date> déjà étiquetées (label_url présent), triées."""
    return await _cursor_all(db.orders.find(
        {
            "payment_status": "paid",
            "dispatch_batch": date,
            "shipping_info.label_url": {"$nin": [None, ""]},
        },
        {"_id": 0},
    ).sort("created_at", 1))


async def admin_dispatch_labels_pdf(date: str, _admin: dict = Depends(require_area("orders", "view"))):
    """Toutes les étiquettes du lot fusionnées en UN pdf 4x6 (imprimante thermique)."""
    orders = await _dispatch_labeled_orders(date)
    pdfs = []
    for o in orders:
        url = (o.get("shipping_info") or {}).get("label_url") or ""
        if url.startswith("/uploads/labels/"):
            fpath = LABEL_UPLOAD_DIR / url.split("/")[-1]
            if fpath.exists():
                pdfs.append(fpath.read_bytes())
    if not pdfs:
        raise HTTPException(404, "Aucune étiquette pour ce lot.")
    merged = _merge_pdfs(pdfs)
    return Response(
        content=merged,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="etiquettes-{date}.pdf"'},
    )


async def admin_dispatch_slips_pdf(date: str, _admin: dict = Depends(require_area("orders", "view"))):
    """Tous les bons de commande du lot fusionnés en UN pdf (imprimante papier)."""
    orders = await _dispatch_labeled_orders(date)
    pdfs = [_generate_invoice_pdf(o) for o in orders]
    pdfs = [p for p in pdfs if p]
    if not pdfs:
        raise HTTPException(404, "Aucun bon de commande pour ce lot.")
    merged = _merge_pdfs(pdfs)
    return Response(
        content=merged,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="bons-{date}.pdf"'},
    )


# ===========================================================================
# POSTE D'EXPÉDITION « FIRONOVA FULFILLMENT » — Phase 1 (v2, amélioré)
#   payé → processing (à préparer) → packing → packed → shipped (étiqueté)
# Améliorations v2 :
#  - capte aussi les commandes payées SANS dispatch_batch (rétro-remplissage
#    à la volée) : plus de commande « fantôme » invisible dans Journée ;
#  - avance groupée d'une colonne entière ;
#  - retards remontés en tête de colonne.
# ===========================================================================
FULFILLMENT_FLOW = ["processing", "packing", "packed", "shipped", "delivered"]
FULFILLMENT_LABELS = {
    "processing": {"fr": "À préparer", "en": "To prepare"},
    "packing":    {"fr": "En préparation", "en": "Preparing"},
    "packed":     {"fr": "Empaquetée", "en": "Packed"},
    "shipped":    {"fr": "Étiquetée", "en": "Labeled"},
    "delivered":  {"fr": "Livrée", "en": "Delivered"},
}


class FulfillmentTransitionIn(BaseModel):
    to: str = Field(min_length=1)


class FulfillmentBulkIn(BaseModel):
    date: str = Field(min_length=1)
    from_status: str = Field(min_length=1)
    to: str = Field(min_length=1)


def _order_items(order: dict) -> list:
    """Articles d'une commande. Les commandes sont stockées avec la clé
    "items" ; on accepte aussi "line_items" par compatibilité."""
    return order.get("items") or order.get("line_items") or []


def _picking_lines(order: dict) -> list:
    out = []
    for it in _order_items(order):
        out.append({
            "sku": it.get("sku") or "",
            "slug": it.get("slug") or "",
            "name_en": it.get("name_en") or "",
            "name_fr": it.get("name_fr") or "",
            "variant_name": it.get("variant_name") or "",
            "qty": int(it.get("qty") or 1),
        })
    return out


async def _fulfillment_backfill_batches(day: str) -> None:
    """Répare les commandes payées sans dispatch_batch (créées avant la logique
    de lot, ou par un flux qui ne l'a pas posé). Sans ça elles restent
    invisibles dans Journée."""
    cursor = db.orders.find(
        {
            "payment_status": "paid",
            "fulfillment_status": {"$in": ["processing", "packing", "packed"]},
            "$or": [{"dispatch_batch": {"$in": [None, ""]}}, {"dispatch_batch": {"$exists": False}}],
        },
        {"_id": 0, "id": 1, "paid_at": 1, "created_at": 1},
    )
    async for o in cursor:
        ref = o.get("paid_at") or o.get("created_at") or datetime.now(timezone.utc).isoformat()
        try:
            batch = compute_dispatch_batch(ref)
        except Exception:
            batch = day
        await db.orders.update_one({"id": o["id"]}, {"$set": {"dispatch_batch": batch}})


async def admin_fulfillment_day(date: Optional[str] = None,
                                _admin: dict = Depends(require_area("orders", "view"))):
    """Tableau de bord « Journée » : lot du jour ventilé par étape physique."""
    def _is_local_day(ts: str, target_day: str) -> bool:
        if not ts:
            return False
        try:
            dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(ZoneInfo(ORDER_CUTOFF_TZ)).date().isoformat() == target_day
        except Exception:
            return ts.startswith(target_day)

    day = date or _local_today_iso()
    await _fulfillment_backfill_batches(day)

    # Étapes en cours : on inclut les retards (<= jour) pour ne rien perdre.
    # Historique journalier : bornes calculées dans le fuseau local métier,
    # puis converties en UTC pour interroger les timestamps stockés.
    local_tz = ZoneInfo(ORDER_CUTOFF_TZ)
    day_local_start = datetime.fromisoformat(day).replace(tzinfo=local_tz)
    day_local_end = day_local_start + timedelta(days=1)
    day_start_utc = day_local_start.astimezone(timezone.utc).isoformat()
    day_end_utc = day_local_end.astimezone(timezone.utc).isoformat()
    orders = db.orders.find(
        {
            "payment_status": "paid",
            "$or": [
                {
                    # À préparer / en préparation : retards inclus pour qu'une
                    # commande oubliée ne disparaisse jamais de l'écran.
                    "dispatch_batch": {"$lte": day},
                    "fulfillment_status": {"$in": ["processing", "packing"]},
                },
                {
                    # Empaquetée : historique du jour. On se base sur packed_at
                    # (et non sur le statut) pour que la commande RESTE visible
                    # même après l'émission de son étiquette dans Dispatch.
                    "packed_at": {"$gte": day_start_utc, "$lt": day_end_utc},
                },
                {
                    # Commandes étiquetées ce jour dans Dispatch (même logique
                    # métier que l'écran Dispatch pour éviter les écarts de
                    # comptage entre les deux vues).
                    "dispatch_batch": day,
                    "fulfillment_status": "shipped",
                    "shipping_info.label_url": {"$nin": [None, ""]},
                    "shipping_info.tracking_number": {"$nin": [None, ""]},
                    "shipping_info.shipped_at": {"$gte": day_start_utc, "$lt": day_end_utc},
                },
            ],
        },
        {"_id": 0},
    ).sort("created_at", 1).to_list(3000)

    buckets = {"processing": [], "packing": [], "packed": []}
    for o in orders:
        st = o.get("fulfillment_status")
        info = o.get("shipping_info") or {}
        has_label = bool(info.get("label_url") and info.get("tracking_number"))
        packed_today = _is_local_day(o.get("packed_at") or "", day)
        labeled_today = has_label and _is_local_day(info.get("shipped_at") or "", day)
        # Une commande empaquetée aujourd'hui reste dans « Empaquetée » même
        # une fois étiquetée (shipped) : c'est l'historique de la journée.
        if packed_today or labeled_today:
            st = "packed"
        if st not in buckets:
            continue
        addr = o.get("shipping_address") or {}
        buckets[st].append({
            "id": o["id"],
            "order_number": o.get("order_number"),
            "email": o.get("email"),
            "created_at": o.get("created_at"),
            "dispatch_batch": o.get("dispatch_batch"),
            "units": sum(int(i.get("qty") or 1) for i in _order_items(o)),
            "items": len(_order_items(o)),
            "total": o.get("total"),
            "city": addr.get("city"),
            "province": addr.get("province"),
            "tracking_number": info.get("tracking_number") or "",
            "label_url": info.get("label_url") or "",
            "picking": _picking_lines(o),
            "is_overdue": bool(o.get("dispatch_batch") and o["dispatch_batch"] < day),
        })

    # Retards en tête de chaque colonne.
    for k in buckets:
        buckets[k].sort(key=lambda r: (not r["is_overdue"], r["created_at"] or ""))

    counts = {k: len(v) for k, v in buckets.items()}
    counts["total"] = sum(counts.values())
    counts["overdue"] = sum(1 for k in buckets for r in buckets[k] if r["is_overdue"])
    return {
        "date": day,
        "configured": is_canada_post_configured(),
        "labels": {k: FULFILLMENT_LABELS[k] for k in buckets},
        "counts": counts,
        "buckets": buckets,
    }


async def _advance_one(order: dict, target: str, admin_email: str) -> None:
    now = datetime.now(timezone.utc).isoformat()
    ts_field = {"packing": "packing_at", "packed": "packed_at"}.get(target)
    update = {"fulfillment_status": target}
    if ts_field:
        update[ts_field] = now
    await db.orders.update_one(
        {"id": order["id"]},
        {"$set": update,
         "$push": {"notes": {
             "id": str(uuid.uuid4()),
             "text": f"Préparation : {FULFILLMENT_LABELS.get(target, {}).get('fr', target)}.",
             "author": admin_email,
             "created_at": now,
         }}},
    )


async def admin_fulfillment_advance(order_id: str, payload: FulfillmentTransitionIn,
                                    _admin: dict = Depends(require_area("orders", "manage"))):
    """Fait avancer une commande dans le flux physique (jusqu'à empaquetée)."""
    target = payload.to
    if target not in {"processing", "packing", "packed"}:
        raise HTTPException(400, "Étape gérée par l'étiquetage (Dispatch), pas ici.")
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Commande introuvable")
    if order.get("payment_status") != "paid":
        raise HTTPException(400, "La commande doit être payée avant préparation.")
    await _advance_one(order, target, _admin.get("email", "system"))
    return await db.orders.find_one({"id": order_id}, {"_id": 0})


async def admin_fulfillment_bulk(payload: FulfillmentBulkIn,
                                 _admin: dict = Depends(require_area("orders", "manage"))):
    """Avance TOUTES les commandes d'une colonne (ex. tout passer « en
    préparation » d'un coup)."""
    if payload.to not in {"packing", "packed"}:
        raise HTTPException(400, "Cible invalide.")
    orders = await db.orders.find(
        {
            "payment_status": "paid",
            "dispatch_batch": {"$lte": payload.date},
            "fulfillment_status": payload.from_status,
        },
        {"_id": 0},
    )
    advanced = 0
    async for o in orders:
        await _advance_one(o, payload.to, _admin.get("email", "system"))
        advanced += 1
    return {"advanced": advanced, "to": payload.to}


async def admin_fulfillment_picking_pdf(date: str,
                                        _admin: dict = Depends(require_area("orders", "view"))):
    """Liste de prélèvement consolidée : total par article à sortir du stock."""
    orders = await _cursor_all(db.orders.find(
        {
            "payment_status": "paid",
            "dispatch_batch": {"$lte": date},
            "fulfillment_status": {"$in": ["processing", "packing"]},
        },
        {"_id": 0},
    ).sort("created_at", 1))

    totals: dict = {}
    for o in orders:
        for it in _order_items(o):
            key = (it.get("sku") or it.get("slug") or "?", it.get("variant_name") or "")
            if key not in totals:
                totals[key] = {"sku": it.get("sku") or "",
                               "name": it.get("name_fr") or it.get("name_en") or it.get("slug") or "",
                               "variant_name": it.get("variant_name") or "", "qty": 0}
            totals[key]["qty"] += int(it.get("qty") or 1)
    if not totals:
        raise HTTPException(404, "Aucune commande à préparer pour ce lot.")

    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas as _canvas
    import io as _io
    buf = _io.BytesIO()
    c = _canvas.Canvas(buf, pagesize=letter)
    w, h = letter
    y = h - 60
    c.setFont("Helvetica-Bold", 16); c.drawString(50, y, f"Liste de prélèvement — lot {date}")
    c.setFont("Helvetica", 9); y -= 18
    c.drawString(50, y, "FIRONOVA · FOR LABORATORY RESEARCH USE ONLY · 19+"); y -= 30
    c.setFont("Helvetica-Bold", 10)
    c.drawString(50, y, "Qté"); c.drawString(90, y, "SKU"); c.drawString(220, y, "Produit"); y -= 6
    c.line(50, y, w - 50, y); y -= 18
    c.setFont("Helvetica", 10)
    for row in sorted(totals.values(), key=lambda r: r["name"]):
        if y < 60:
            c.showPage(); y = h - 60; c.setFont("Helvetica", 10)
        label = row["name"] + (f" — {row['variant_name']}" if row["variant_name"] else "")
        c.drawString(50, y, f"{row['qty']}×"); c.drawString(90, y, (row["sku"] or "")[:22])
        c.drawString(220, y, label[:60]); y -= 18
    c.showPage(); c.save(); buf.seek(0)
    return Response(content=buf.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="prelevement-{date}.pdf"'})


# ---------------------------------------------------------------------------
# Report des étiquettes non imprimées + signaux de navigation
# ---------------------------------------------------------------------------
async def admin_dispatch_mark_printed(date: str,
                                      _admin: dict = Depends(require_area("orders", "manage"))):
    """Marque les étiquettes du lot comme imprimées. Une étiquette imprimée
    n'est plus reportée au lendemain par le report de minuit."""
    now = datetime.now(timezone.utc).isoformat()
    res = await db.orders.update_many(
        {
            "payment_status": "paid",
            "dispatch_batch": date,
            "shipping_info.label_url": {"$nin": [None, ""]},
        },
        {"$set": {"shipping_info.label_printed_at": now}},
    )
    return {"ok": True, "marked": res.modified_count, "date": date}


async def _rollover_unprinted_labels() -> int:
    """Reporte au prochain jour ouvrable les commandes d'un lot passé dont
    l'étiquette n'a pas été imprimée. Évite qu'un lot oublié disparaisse."""
    today = _local_today_iso()
    cursor = db.orders.find(
        {
            "payment_status": "paid",
            "dispatch_batch": {"$lt": today},
            "fulfillment_status": {"$in": ["processing", "pending", "packing", "packed", "shipped"]},
            "$or": [
                {"shipping_info.label_printed_at": {"$in": [None, ""]}},
                {"shipping_info.label_printed_at": {"$exists": False}},
            ],
        },
        {"_id": 0, "id": 1, "order_number": 1},
    )
    moved = 0
    async for o in cursor:
        await db.orders.update_one(
            {"id": o["id"]},
            {"$set": {"dispatch_batch": today},
             "$push": {"notes": {
                 "id": str(uuid.uuid4()),
                 "text": f"Reportée au lot {today} — étiquette non imprimée.",
                 "author": "system",
                 "created_at": datetime.now(timezone.utc).isoformat(),
             }}},
        )
        moved += 1
    if moved:
        logging.info("[rollover] %d commande(s) reportée(s) au lot %s", moved, today)
    return moved


async def _rollover_watchdog() -> None:
    """Passe une fois par heure : à la première exécution après minuit, les
    lots non imprimés de la veille basculent sur le jour courant."""
    last_day = None
    while True:
        try:
            today = _local_today_iso()
            if last_day != today:
                await _rollover_unprinted_labels()
                last_day = today
        except Exception as ex:  # pragma: no cover
            logging.error("rollover watchdog failed: %s", ex)
        await asyncio.sleep(3600)


async def admin_ops_signals(_admin: dict = Depends(require_area("orders", "view"))):
    """Compteurs pour les pastilles de navigation (Journée / Dispatch) et
    l'alerte manifeste du tableau de bord."""
    day = _local_today_iso()
    to_prepare = await db.orders.count_documents({
        "payment_status": "paid",
        "dispatch_batch": {"$lte": day},
        "fulfillment_status": {"$in": ["processing", "packing"]},
    })
    to_label = await db.orders.count_documents({
        "payment_status": "paid",
        "dispatch_batch": day,
        # Notification Dispatch: seulement les commandes prêtes (flux Journée complété).
        "fulfillment_status": "packed",
        "$or": [
            {"shipping_info.label_url": {"$in": [None, ""]}},
            {"shipping_info.label_url": {"$exists": False}},
        ],
    })
    to_print = await db.orders.count_documents({
        "payment_status": "paid",
        "dispatch_batch": day,
        "shipping_info.label_url": {"$nin": [None, ""]},
        "$or": [
            {"shipping_info.label_printed_at": {"$in": [None, ""]}},
            {"shipping_info.label_printed_at": {"$exists": False}},
        ],
    })
    # Même source que la bannière de la page Commandes, sinon les deux
    # affichent des nombres différents pour la même chose.
    pending_manifest = await db.orders.count_documents(UNTRANSMITTED_MATCH)
    # Retards : lots antérieurs encore non expédiés.
    overdue = await db.orders.count_documents({
        "payment_status": "paid",
        "dispatch_batch": {"$lt": day},
        # Retards de Dispatch = commandes prêtes mais non étiquetées sur lot antérieur.
        "fulfillment_status": "packed",
    })
    return {
        "date": day,
        "fulfillment": to_prepare,
        # Le badge Dispatch signale ce qui reste À ÉTIQUETER (+ retards),
        # pas les étiquettes déjà créées en attente d'impression.
        "dispatch": to_label + overdue,
        "overdue": overdue,
        "to_label": to_label,
        "to_print": to_print,
        "pending_manifest": pending_manifest,
    }


async def admin_adjust_stock(product_id: str, payload: StockAdjustIn, _admin: dict = Depends(require_area("products", "manage"))):
    res = await db.products.update_one({"id": product_id}, {"$inc": {"stock": payload.delta}})
    if res.matched_count == 0:
        raise HTTPException(404, "Product not found")
    if payload.delta > 0:
        asyncio.create_task(_maybe_notify_restock(product_id, None))
    return await db.products.find_one({"id": product_id}, {"_id": 0})


async def admin_bulk_restock(product_id: str, payload: StockRestockIn,
                              admin: dict = Depends(require_area("products", "manage"))):
    """Restock atomique multi-variantes avec audit trail.

    - `quantity > 0` : ajoute du stock (`$inc`).
    - `quantity < 0` : retire du stock (perte/casse/retour) avec garde
      atomique `stock >= abs(qty)` pour empêcher un stock négatif.
    - Chaque ligne écrit un doc `stock_movements` (audit qui/quand/combien
      + `movement_type` in ('restock', 'adjustment')).
    - Déclenche `_maybe_notify_restock` si stock passe de 0 → >0.
    - Déclenche `_maybe_send_low_stock_alert` si stock franchit le seuil bas.
    - Réponse : produit à jour + `{applied[], skipped[]}` (raisons de skip
      typées : `variant_not_found`, `insufficient_stock`, `update_failed`).
    """
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(404, "Product not found")

    now_iso = datetime.now(timezone.utc).isoformat()
    admin_email = admin.get("email") or ""
    variant_index = {v.get("id"): v for v in (product.get("variants") or []) if v.get("id")}

    applied: list = []
    skipped: list = []
    affected_variant_ids: set = set()

    for entry in payload.deltas:
        vid = entry.variant_id
        qty = int(entry.quantity)
        movement_type = "restock" if qty > 0 else "adjustment"
        if vid:
            variant = variant_index.get(vid)
            if not variant:
                skipped.append({"variant_id": vid, "reason": "variant_not_found"})
                continue
            before = int(variant.get("stock") or 0)
            if qty < 0:
                # Garde atomique : refuse si stock < |qty|
                res = await db.products.update_one(
                    {"id": product_id,
                     "variants": {"$elemMatch": {"id": vid, "stock": {"$gte": -qty}}}},
                    {"$inc": {"variants.$.stock": qty}},
                )
                if not res.modified_count:
                    skipped.append({"variant_id": vid, "reason": "insufficient_stock",
                                    "current": before, "requested": qty})
                    continue
            else:
                res = await db.products.update_one(
                    {"id": product_id, "variants.id": vid},
                    {"$inc": {"variants.$.stock": qty}},
                )
                if not res.modified_count:
                    skipped.append({"variant_id": vid, "reason": "update_failed"})
                    continue
            after = before + qty
            variant_name = variant.get("name") or vid
        else:
            before = int(product.get("stock") or 0)
            if qty < 0 and before < -qty:
                skipped.append({"variant_id": None, "reason": "insufficient_stock",
                                "current": before, "requested": qty})
                continue
            res = await db.products.update_one(
                {"id": product_id}, {"$inc": {"stock": qty}}
            )
            if not res.modified_count:
                skipped.append({"variant_id": None, "reason": "update_failed"})
                continue
            after = before + qty
            variant_name = None

        movement = {
            "id": str(uuid.uuid4()),
            "product_id": product_id,
            "product_name": product.get("name_en") or product.get("name_fr") or product.get("slug"),
            "variant_id": vid,
            "variant_name": variant_name,
            "delta": qty,
            "movement_type": movement_type,
            "stock_before": before,
            "stock_after": after,
            "reason": (entry.note or payload.reason or "").strip() or None,
            "admin_email": admin_email,
            "created_at": now_iso,
        }
        await db.stock_movements.insert_one(movement)
        movement.pop("_id", None)
        applied.append(movement)
        affected_variant_ids.add(vid)

        # Restock notifications : si le stock passait de 0 à >0, avertir les abonnés
        if before == 0 and after > 0:
            asyncio.create_task(_maybe_notify_restock(product_id, vid))

    # Post-check low stock alerts (une fois toutes les modifs appliquées)
    if applied:
        asyncio.create_task(_check_low_stock_alerts(product_id, affected_variant_ids))

    fresh = await db.products.find_one({"id": product_id}, {"_id": 0})
    return {"ok": True, "product": fresh, "applied": applied, "skipped": skipped}


async def admin_bulk_restock_csv(payload: StockBulkRestockIn,
                                  admin: dict = Depends(require_area("products", "manage"))):
    """Restock multi-produits depuis un CSV. Chaque ligne référence une
    variante via `sku` (préféré) ou via `(product_slug, variant_name)`.

    Retour : `{ok, total, applied[], failed[]}` — chaque `applied` contient
    le mouvement inséré, chaque `failed` explique la raison de rejet."""
    now_iso = datetime.now(timezone.utc).isoformat()
    admin_email = admin.get("email") or ""

    applied: list = []
    failed: list = []
    # Groupement par (product_id, variant_id) pour post-check alerte
    affected: dict = {}

    for idx, row in enumerate(payload.rows):
        line_no = idx + 1
        # Résolution variant → (product, variant, product_id, variant_id)
        product = None
        variant = None
        vid = None
        if row.sku:
            product = await db.products.find_one(
                {"variants.sku": row.sku}, {"_id": 0}
            )
            if not product:
                # fallback : SKU sur produit sans variantes
                product = await db.products.find_one({"sku": row.sku}, {"_id": 0})
            if product:
                for v in product.get("variants") or []:
                    if v.get("sku") == row.sku:
                        variant = v
                        vid = v.get("id")
                        break
        elif row.product_slug:
            product = await db.products.find_one({"slug": row.product_slug}, {"_id": 0})
            if product and row.variant_name:
                for v in product.get("variants") or []:
                    if (v.get("name") or "").strip().lower() == row.variant_name.strip().lower():
                        variant = v
                        vid = v.get("id")
                        break

        if not product:
            failed.append({"line": line_no, "reason": "product_not_found",
                           "sku": row.sku, "product_slug": row.product_slug})
            continue
        if product.get("variants") and not variant:
            failed.append({"line": line_no, "reason": "variant_not_found",
                           "sku": row.sku, "product_slug": row.product_slug,
                           "variant_name": row.variant_name})
            continue

        qty = int(row.quantity)
        movement_type = "restock" if qty > 0 else "adjustment"
        before = int((variant.get("stock") if variant else product.get("stock")) or 0)

        if vid:
            if qty < 0:
                res = await db.products.update_one(
                    {"id": product["id"],
                     "variants": {"$elemMatch": {"id": vid, "stock": {"$gte": -qty}}}},
                    {"$inc": {"variants.$.stock": qty}},
                )
            else:
                res = await db.products.update_one(
                    {"id": product["id"], "variants.id": vid},
                    {"$inc": {"variants.$.stock": qty}},
                )
        else:
            if qty < 0 and before < -qty:
                failed.append({"line": line_no, "reason": "insufficient_stock",
                               "current": before, "requested": qty})
                continue
            res = await db.products.update_one(
                {"id": product["id"]}, {"$inc": {"stock": qty}}
            )

        if not res.modified_count:
            failed.append({"line": line_no, "reason": "insufficient_stock" if qty < 0 else "update_failed",
                           "current": before, "requested": qty})
            continue

        after = before + qty
        movement = {
            "id": str(uuid.uuid4()),
            "product_id": product["id"],
            "product_name": product.get("name_en") or product.get("name_fr") or product.get("slug"),
            "variant_id": vid,
            "variant_name": (variant.get("name") if variant else None),
            "delta": qty,
            "movement_type": movement_type,
            "stock_before": before,
            "stock_after": after,
            "reason": (row.note or payload.reason or "").strip() or None,
            "admin_email": admin_email,
            "source": "csv_bulk",
            "created_at": now_iso,
        }
        await db.stock_movements.insert_one(movement)
        movement.pop("_id", None)
        applied.append(movement)
        affected.setdefault(product["id"], set()).add(vid)

        if before == 0 and after > 0:
            asyncio.create_task(_maybe_notify_restock(product["id"], vid))

    for pid, vids in affected.items():
        asyncio.create_task(_check_low_stock_alerts(pid, vids))

    return {"ok": True, "total": len(payload.rows),
            "applied": applied, "failed": failed,
            "counts": {"applied": len(applied), "failed": len(failed)}}


async def admin_list_low_stock_alerts(_admin: dict = Depends(require_area("products", "view"))):
    """Variantes actuellement sous leur seuil, pour le widget du dashboard.

    Lecture EN DIRECT de products (via _low_stock_variants) plutôt que de la
    collection low_stock_alerts : celle-ci n'est alimentée que lorsqu'une
    commande ou un réapprovisionnement fait franchir le seuil, si bien qu'un
    stock saisi bas à la main n'y apparaissait jamais. La tuile du dashboard
    lit exactement le même calcul, ce qui met fin à la contradiction entre
    « 1 stock faible » en haut et « aucune alerte » juste en dessous.
    """
    items = await _low_stock_variants()
    return {"items": items, "count": len(items)}


async def admin_product_stock_history(product_id: str, limit: int = 50,
                                       _admin: dict = Depends(require_area("products", "view"))):
    """Retourne l'historique des mouvements de stock d'un produit
    (restocks admin), du plus récent au plus ancien."""
    limit = max(1, min(int(limit or 50), 500))
    cursor = db.stock_movements.find(
        {"product_id": product_id}, {"_id": 0}
    ).sort("created_at", -1).limit(limit)
    items = await cursor.to_list(limit)
    return {"product_id": product_id, "items": items, "count": len(items)}


async def admin_list_stock_notifications(_admin: dict = Depends(require_area("products", "view"))):
    """Overview of pending back-in-stock subscriptions, grouped implicitly by product/variant."""
    pending = await db.stock_notifications.find({"notified": False}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return pending


# ---------------------------------------------------------------------------
# Admin — coupons
# ---------------------------------------------------------------------------
async def admin_list_coupons(_admin: dict = Depends(require_area("coupons", "view"))):
    return await db.coupons.find({"deleted_at": None}, {"_id": 0}).sort("created_at", -1).to_list(500)


async def admin_create_coupon(payload: CouponIn, _admin: dict = Depends(require_area("coupons", "manage"))):
    code = payload.code.upper().strip()
    if await db.coupons.find_one({"code": code}):
        raise HTTPException(409, "Coupon code already exists")
    _enforce_standard_coupon_percent_limit(payload.discount_type, payload.value, is_affiliate=False)
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["code"] = code
    doc["allowed_emails"] = [e.lower().strip() for e in (doc.get("allowed_emails") or []) if e and e.strip()]
    doc["restrict_products"] = [p for p in (doc.get("restrict_products") or []) if p]
    doc["restrict_categories"] = [c for c in (doc.get("restrict_categories") or []) if c]
    doc["used_count"] = 0
    doc["used_by"] = []
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.coupons.insert_one(doc)
    doc.pop("_id", None)
    return doc


async def admin_update_coupon(coupon_id: str, payload: CouponIn, _admin: dict = Depends(require_area("coupons", "manage"))):
    # exclude_unset : seuls les champs envoyes par le formulaire sont appliques,
    # les compteurs (used_count/used_by) et champs avances absents sont preserves.
    update = payload.model_dump(exclude_unset=True)
    existing = await db.coupons.find_one({"id": coupon_id}, {"_id": 0, "affiliate_id": 1, "source": 1})
    if not existing:
        raise HTTPException(404, "Coupon not found")
    _enforce_standard_coupon_percent_limit(
        update.get("discount_type", payload.discount_type),
        float(update.get("value", payload.value)),
        _is_affiliate_coupon(existing),
    )
    if "code" in update:
        update["code"] = update["code"].upper().strip()
    if "allowed_emails" in update:
        update["allowed_emails"] = [e.lower().strip() for e in (update["allowed_emails"] or []) if e and e.strip()]
    if "restrict_products" in update:
        update["restrict_products"] = [p for p in (update["restrict_products"] or []) if p]
    if "restrict_categories" in update:
        update["restrict_categories"] = [c for c in (update["restrict_categories"] or []) if c]
    res = await db.coupons.update_one({"id": coupon_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(404, "Coupon not found")
    return await db.coupons.find_one({"id": coupon_id}, {"_id": 0})


async def admin_delete_coupon(coupon_id: str, admin: dict = Depends(require_area("coupons", "manage"))):
    return await _soft_delete("coupons", coupon_id, admin)


async def validate_coupon(code: str, subtotal: float,
                          email: Optional[str] = None,
                          items: Optional[str] = None):
    """Validation cote client. email et items (JSON) sont OPTIONNELS : sans eux,
    seules les regles de base sont verifiees (retrocompatible avec l'appel actuel).
    La validation finale complete est toujours faite au checkout."""
    coupon = await db.coupons.find_one({"code": code.upper().strip()}, {"_id": 0})
    line_items = None
    if items:
        try:
            raw = json.loads(items)
            if isinstance(raw, list):
                ids = [str(x.get("product_id")) for x in raw if isinstance(x, dict) and x.get("product_id")]
                docs = await db.products.find({"id": {"$in": ids}}, {"_id": 0}).to_list(len(ids))
                by_id = {d["id"]: d for d in docs}
                line_items = [
                    {"product_id": str(x.get("product_id")),
                     "category": (by_id.get(str(x.get("product_id"))) or {}).get("category", "")}
                    for x in raw if isinstance(x, dict) and x.get("product_id")
                ]
        except Exception:
            line_items = None
    discount, applied = await _coupon_discount(coupon, subtotal, line_items=line_items, email=email)
    return {
        "code": coupon["code"], "discount_type": coupon["discount_type"],
        "value": coupon["value"], "discount_amount": discount,
        "min_subtotal": coupon.get("min_subtotal", 0),
        "restricted": bool(coupon.get("allowed_emails") or coupon.get("first_order_only")
                           or coupon.get("per_customer_limit")
                           or coupon.get("restrict_products") or coupon.get("restrict_categories")),
    }


# ---------------------------------------------------------------------------
# Admin — shipping zones & methods
# ---------------------------------------------------------------------------
async def admin_list_zones(_admin: dict = Depends(require_area("shipping", "view"))):
    zones = await db.shipping_zones.find({"deleted_at": None}, {"_id": 0}).sort("name", 1).to_list(200)
    # attach methods
    out = []
    for z in zones:
        methods = await db.shipping_methods.find({"zone_id": z["id"], "deleted_at": None}, {"_id": 0}).to_list(200)
        z["methods"] = methods
        out.append(z)
    return out


async def admin_create_zone(payload: ShippingZoneIn, _admin: dict = Depends(require_area("shipping", "manage"))):
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.shipping_zones.insert_one(doc)
    doc.pop("_id", None)
    return doc


async def admin_update_zone(zone_id: str, payload: ShippingZoneIn, _admin: dict = Depends(require_area("shipping", "manage"))):
    res = await db.shipping_zones.update_one({"id": zone_id}, {"$set": payload.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(404, "Zone not found")
    return await db.shipping_zones.find_one({"id": zone_id}, {"_id": 0})


async def admin_delete_zone(zone_id: str, admin: dict = Depends(require_area("shipping", "manage"))):
    # Cascade : les méthodes de cette zone vont aussi en corbeille, pour
    # pouvoir tout restaurer ensemble si la suppression était une erreur.
    methods = await db.shipping_methods.find({"zone_id": zone_id, "deleted_at": None}, {"_id": 0}).to_list(200)
    for m in methods:
        await _soft_delete("shipping_methods", m["id"], admin)
    return await _soft_delete("shipping_zones", zone_id, admin)


async def admin_create_method(payload: ShippingMethodIn, _admin: dict = Depends(require_area("shipping", "manage"))):
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.shipping_methods.insert_one(doc)
    doc.pop("_id", None)
    return doc


async def admin_update_method(method_id: str, payload: ShippingMethodIn, _admin: dict = Depends(require_area("shipping", "manage"))):
    res = await db.shipping_methods.update_one({"id": method_id}, {"$set": payload.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(404, "Method not found")
    return await db.shipping_methods.find_one({"id": method_id}, {"_id": 0})


async def admin_delete_method(method_id: str, admin: dict = Depends(require_area("shipping", "manage"))):
    return await _soft_delete("shipping_methods", method_id, admin)


# ---------------------------------------------------------------------------
# Contenants d'expédition (enveloppes / boîtes) configurables depuis l'admin.
# Chaque contenant a des dimensions (cm), un poids à vide (tare, g) et une
# capacité max en nombre d'unités. Le calcul de colis choisit le plus petit
# contenant dont la capacité >= nombre total d'unités de la commande, ajoute
# sa tare au poids des produits, et transmet ses dimensions à Canada Post
# (nécessaire pour le poids volumétrique).
# ---------------------------------------------------------------------------
class ShippingBoxIn(BaseModel):
    name: str = Field(min_length=1, max_length=80)      # ex. "Enveloppe à bulles"
    length_cm: float = Field(gt=0)
    width_cm: float = Field(gt=0)
    height_cm: float = Field(default=2.0, gt=0)          # épaisseur (enveloppe ~2 cm)
    tare_grams: float = Field(default=20.0, ge=0)        # poids à vide du contenant
    max_units: int = Field(default=10, ge=1)             # capacité en nombre d'articles
    active: bool = True


async def admin_list_boxes(_admin: dict = Depends(require_area("shipping", "view"))):
    return await db.shipping_boxes.find({"deleted_at": None}, {"_id": 0}).sort("max_units", 1).to_list(200)


async def admin_create_box(payload: ShippingBoxIn, _admin: dict = Depends(require_area("shipping", "manage"))):
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["deleted_at"] = None
    await db.shipping_boxes.insert_one(doc)
    doc.pop("_id", None)
    return doc


async def admin_update_box(box_id: str, payload: ShippingBoxIn, _admin: dict = Depends(require_area("shipping", "manage"))):
    res = await db.shipping_boxes.update_one({"id": box_id}, {"$set": payload.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(404, "Box not found")
    return await db.shipping_boxes.find_one({"id": box_id}, {"_id": 0})


async def admin_delete_box(box_id: str, admin: dict = Depends(require_area("shipping", "manage"))):
    return await _soft_delete("shipping_boxes", box_id, admin)


async def _select_box_for_order(order: dict, all_boxes: Optional[list] = None) -> Optional[dict]:
    """Plus petit contenant actif dont la capacité >= nb total d'unités.
    Si shipping_box_override_id est défini sur la commande, on l'utilise en priorité."""
    boxes = all_boxes
    if boxes is None:
        boxes = await db.shipping_boxes.find(
            {"deleted_at": None, "active": True}, {"_id": 0}
        ).sort("max_units", 1).to_list(200)

    override_id = order.get("shipping_box_override_id")
    if override_id:
        for b in boxes:
            if b.get("id") == override_id:
                return b

    units = 0
    for it in _order_items(order):
        units += int(it.get("qty") or 1)
    units = max(1, units)
    for b in sorted(boxes, key=lambda b: int(b.get("max_units") or 0)):
        if int(b.get("max_units") or 0) >= units:
            return b
    return boxes[-1] if boxes else None


# ---------------------------------------------------------------------------
# Admin — Analytics
# ---------------------------------------------------------------------------
DASHBOARD_PERIODS = (7, 30, 90, 180, 365)


def _order_day(value) -> str:
    """Jour AAAA-MM-JJ d'une commande, que created_at soit une chaîne ISO ou
    un objet datetime. Les deux formes coexistent en base selon l'ancienneté
    de la commande et le chemin qui l'a créée."""
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).strftime("%Y-%m-%d")
    return str(value or "")[:10]


def _series_granularity(period: int) -> str:
    """Pas de temps du graphique selon la période demandée.

    Un an en barres quotidiennes fait 365 traits de deux pixels : on voit une
    texture, pas une tendance. On agrège donc au-delà de 3 mois.
    """
    if period <= 90:
        return "day"
    if period <= 180:
        return "week"
    return "month"


def _bucket_key(day: str, granularity: str) -> str:
    """Étiquette du seau auquel appartient un jour AAAA-MM-JJ."""
    if granularity == "month":
        return day[:7]                      # AAAA-MM
    if granularity == "week":
        d = datetime.strptime(day, "%Y-%m-%d")
        monday = d - timedelta(days=d.weekday())
        return monday.strftime("%Y-%m-%d")  # lundi de la semaine
    return day


async def admin_dashboard_pulse(_admin: dict = Depends(require_area("dashboard", "view"))):
    """Ce qui demande une action maintenant, et où en est l'argent.

    Une seule requête pour toute la partie haute du tableau de bord : le
    widget appelait sinon cinq endpoints différents pour afficher six
    compteurs. Tout est calculé en direct — aucun de ces chiffres ne doit
    dépendre d'une collection d'événements qui pourrait ne pas s'être
    déclenchée.
    """
    now = datetime.now(timezone.utc)
    now_s = now.isoformat()
    soon_s = (now + timedelta(hours=3)).isoformat()

    # --- Commandes vendues mais pas encore encaissées, TOUS circuits -------
    awaiting = await db.orders.find(
        {"payment_status": {"$in": ["awaiting_etransfer", "awaiting_crypto"]},
         "deleted_at": None},
        {"_id": 0, "total": 1, "payment_method": 1, "payment_deadline": 1},
    ).to_list(2000)
    pending_total = round(sum(float(o.get("total") or 0) for o in awaiting), 2)
    pending_expiring = sum(
        1 for o in awaiting
        if o.get("payment_deadline") and now_s < str(o["payment_deadline"]) <= soon_s
    )

    # --- Réconciliation : une seule file, deux fournisseurs ---------------
    # La collection porte un nom historique (interac_) mais reçoit aussi les
    # signaux crypto, distingués par le champ provider.
    recon = await db.interac_reconciliation_queue.find(
        {"status": "pending"}, {"_id": 0, "provider": 1},
    ).to_list(1000)
    recon_by_provider: dict = {}
    for r in recon:
        prov = (r.get("provider") or "interac").lower()
        recon_by_provider[prov] = recon_by_provider.get(prov, 0) + 1

    # --- Circuits de paiement : encaissé / en attente / à réconcilier -----
    rails: dict = {}
    for method in ("interac", "nowpayments"):
        paid_rows = await db.orders.aggregate([
            {"$match": {"payment_status": "paid", "payment_method": method}},
            {"$group": {"_id": None, "amount": {"$sum": "$total"}, "count": {"$sum": 1}}},
        ]).to_list(1)
        pend = [o for o in awaiting if (o.get("payment_method") or "") == method]
        key = "crypto" if method == "nowpayments" else "interac"
        rails[key] = {
            "paid_amount": round(float(paid_rows[0]["amount"]), 2) if paid_rows else 0.0,
            "paid_count": int(paid_rows[0]["count"]) if paid_rows else 0,
            "pending_amount": round(sum(float(o.get("total") or 0) for o in pend), 2),
            "pending_count": len(pend),
            "reconcile_count": recon_by_provider.get(
                "crypto" if key == "crypto" else "interac", 0),
        }

    # --- Opérations du jour ----------------------------------------------
    to_ship = await db.orders.count_documents({
        "payment_status": "paid",
        "fulfillment_status": {"$in": ["processing", "pending"]},
        "deleted_at": None,
    })
    low_stock_rows = await _low_stock_variants(limit=50)
    late_payments = await db.orders.count_documents({"late_payment_flagged": True})
    emails_failed = await db.email_outbox.count_documents({"status": "failed"})

    return {
        "money": {
            "pending_payment": {
                "amount": pending_total,
                "count": len(awaiting),
                "expiring_soon": pending_expiring,
                "by_method": {
                    "interac": sum(1 for o in awaiting if o.get("payment_method") == "interac"),
                    "crypto": sum(1 for o in awaiting if o.get("payment_method") == "nowpayments"),
                },
            },
            "reconcile": {"count": len(recon), "by_provider": recon_by_provider},
        },
        "rails": rails,
        "ops": {
            "to_ship": to_ship,
            "low_stock": len(low_stock_rows),
            "low_stock_top": low_stock_rows[:1],
            "late_payments": late_payments,
            "emails_failed": emails_failed,
            # Billets d'affiliés en attente de réponse. Exposé ICI, sur le
            # pouls quotidien, et pas seulement dans l'écran des billets :
            # un billet non relevé est pire qu'un courriel oublié, parce que
            # l'affilié le voit « ouvert » et attend. Le seul moyen que ce
            # système tienne sa promesse est qu'on ne puisse pas l'ignorer.
            "tickets_open": await db.affiliate_tickets.count_documents(
                {"status": "open"}
            ),
        },
    }


async def admin_analytics(period: int = 30,
                          _admin: dict = Depends(require_area("dashboard", "view"))):
    # --- Série de revenu sur la période demandée ----------------------------
    # Le regroupement se fait en Python, volontairement : l'agrégation Mongo
    # comparait created_at (parfois stocké en Date) à une chaîne ISO, ce qui
    # est fragile selon l'ordre des types BSON. Le volume (quelques centaines
    # de commandes) rend le tri côté application sans coût mesurable.
    #
    # La série suivait auparavant 30 jours EN DUR : le sélecteur de période du
    # dashboard changeait les tuiles mais jamais le graphique.
    if period not in DASHBOARD_PERIODS:
        period = 30
    granularity = _series_granularity(period)
    since_day = (datetime.now(timezone.utc) - timedelta(days=period)).strftime("%Y-%m-%d")
    paid = await db.orders.find(
        {"payment_status": "paid"},
        {"_id": 0, "created_at": 1, "total": 1},
    ).to_list(20000)

    buckets: dict = {}
    for o in paid:
        day = _order_day(o.get("created_at"))
        if not day or day < since_day:
            continue
        key = _bucket_key(day, granularity)
        b = buckets.setdefault(key, {"revenue": 0.0, "orders": 0})
        b["revenue"] += float(o.get("total") or 0)
        b["orders"] += 1
    daily = [
        {"date": d, "revenue": round(v["revenue"], 2), "orders": v["orders"]}
        for d, v in sorted(buckets.items())
    ]

    # --- Meilleures ventes, PAR VARIANTE ------------------------------------
    # Le regroupement se faisait sur le seul slug produit, si bien que deux
    # dosages du même composé apparaissaient comme deux lignes identiques aux
    # montants différents — illisible. On groupe désormais sur le couple
    # (slug, variante) et on renvoie le nom de la variante pour l'afficher.
    top_cursor = db.orders.aggregate([
        {"$match": {"payment_status": "paid"}},
        {"$unwind": "$items"},
        {"$group": {
            "_id": {"slug": "$items.slug", "variant": "$items.variant_name"},
            "name_en": {"$first": "$items.name_en"},
            "name_fr": {"$first": "$items.name_fr"},
            "units_sold": {"$sum": "$items.qty"},
            "revenue": {"$sum": "$items.line_total"},
        }},
        # Trie par REVENU, pas par unités : 17 unités à 1 $ passaient devant
        # 15 unités à 70 $, ce qui donnait un classement trompeur.
        {"$sort": {"revenue": -1}},
        {"$limit": 10},
    ])
    top = await top_cursor.to_list(10)
    top = [{
        "slug": t["_id"].get("slug"),
        "variant_name": t["_id"].get("variant") or "",
        "name_en": t.get("name_en"),
        "name_fr": t.get("name_fr"),
        "units_sold": t["units_sold"],
        "revenue": round(t["revenue"], 2),
    } for t in top]

    # Recent orders
    recent = await db.orders.find({}, {"_id": 0}).sort("created_at", -1).limit(10).to_list(10)

    return {
        "daily_revenue": daily,
        "granularity": granularity,   # day | week | month — pilote l'étiquetage côté UI
        "period": period,
        "top_products": top,
        "recent_orders": recent,
    }


# ---------------------------------------------------------------------------
# Admin — CSV exports
# ---------------------------------------------------------------------------
def _csv_safe_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    if value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _csv_safe_row(row: dict) -> dict:
    return {key: _csv_safe_value(value) for key, value in row.items()}


def _csv_response(rows: list, filename: str) -> StreamingResponse:
    buf = io.StringIO()
    if not rows:
        buf.write("\n")
    else:
        writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(_csv_safe_row(row) for row in rows)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _csv_cursor_response(cursor, row_mapper, fieldnames: list[str], filename: str) -> StreamingResponse:
    async def generate():
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=fieldnames)
        writer.writeheader()
        yield buffer.getvalue()
        async for document in cursor:
            buffer.seek(0)
            buffer.truncate(0)
            writer.writerow(_csv_safe_row(row_mapper(document)))
            yield buffer.getvalue()

    return StreamingResponse(
        generate(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def admin_orders_csv(status_group: Optional[str] = None, _admin: dict = Depends(require_area("orders", "view"))):
    cursor = db.orders.find(_status_group_filter(status_group), {"_id": 0}).sort("created_at", -1)
    fieldnames = list(_orders_export_rows([{}])[0].keys())
    return _csv_cursor_response(
        cursor,
        lambda order: _orders_export_rows([order])[0],
        fieldnames,
        f"fironova-orders-{datetime.now().strftime('%Y%m%d')}.csv",
    )


def _orders_export_rows(orders: list) -> list:
    rows = []
    for o in orders:
        addr = o.get("shipping_address", {})
        rows.append({
            "order_number": o.get("order_number", ""),
            "created_at": o.get("created_at", ""),
            "email": o.get("email") or "",
            "customer": addr.get("full_name", ""),
            "city": addr.get("city", ""),
            "province": addr.get("province", ""),
            "postal_code": addr.get("postal_code", ""),
            "country": addr.get("country", ""),
            "payment_method": o.get("payment_method", ""),
            "payment_status": o.get("payment_status", ""),
            "fulfillment_status": o.get("fulfillment_status", ""),
            "items_count": len(o.get("items", [])),
            "subtotal_cad": o.get("subtotal", 0),
            "discount_cad": o.get("discount", 0),
            "shipping_cad": o.get("shipping", 0),
            "total_cad": o.get("total", 0),
            "refunded_cad": o.get("refunded_amount", 0),
            "tracking_number": (o.get("shipping_info") or {}).get("tracking_number", ""),
            "carrier": (o.get("shipping_info") or {}).get("carrier", ""),
        })
    return rows


def _xlsx_response(rows: list, filename: str) -> Response:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    headers = list(rows[0].keys()) if rows else ["no_data"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="050505", end_color="050505", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    for col_idx, h in enumerate(headers, start=1):
        width = max([len(str(h))] + [len(str(r.get(h, ""))) for r in rows[:500]]) + 2
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(width, 50)
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def admin_orders_xlsx(status_group: Optional[str] = None, _admin: dict = Depends(require_area("orders", "view"))):
    orders = await _cursor_all(db.orders.find(
        _status_group_filter(status_group), {"_id": 0}
    ).sort("created_at", -1))
    return _xlsx_response(_orders_export_rows(orders), f"fironova-orders-{datetime.now().strftime('%Y%m%d')}.xlsx")


async def admin_products_csv(_admin: dict = Depends(require_area("products", "view"))):
    products = await _cursor_all(db.products.find({}, {"_id": 0}).sort("name_en", 1))
    rows = []
    for p in products:
        rows.append({
            "slug": p.get("slug", ""),
            "name_en": p.get("name_en", ""),
            "name_fr": p.get("name_fr", ""),
            "category": p.get("category", ""),
            "sequence": p.get("sequence", ""),
            "purity": p.get("purity", ""),
            "dosage_mg": p.get("dosage_mg", 0),
            "price_cad": p.get("price_cad", 0),
            "stock": p.get("stock", 0),
            "low_stock_threshold": p.get("low_stock_threshold", 10),
            "featured": p.get("featured", False),
            "preorder_allowed": p.get("preorder_allowed", False),
            "lab_tested": p.get("lab_tested", False),
            "coa_url": p.get("coa_url", ""),
            "coa_lot": p.get("coa_lot", ""),
            "coa_date": p.get("coa_date", ""),
            "active": p.get("active", True),
            # Scientific data fields (new)
            "molecular_formula": p.get("molecular_formula", ""),
            "molecular_weight": p.get("molecular_weight", ""),
            "cas_number": p.get("cas_number", ""),
            "sequence_length": p.get("sequence_length", ""),
            "storage": p.get("storage", ""),
            "solubility": p.get("solubility", ""),
            "appearance": p.get("appearance", ""),
            "mechanism": p.get("mechanism", ""),
            "research_areas": "; ".join(p.get("research_areas", [])),
            "synonyms": "; ".join(p.get("synonyms", [])),
        })
    return _csv_response(rows, f"fironova-products-{datetime.now().strftime('%Y%m%d')}.csv")


async def admin_products_xlsx(_admin: dict = Depends(require_area("products", "view"))):
    products = await db.products.find({}, {"_id": 0}).sort("name_en", 1).to_list(2000)
    rows = []
    for p in products:
        rows.append({
            "slug": p.get("slug", ""),
            "name_en": p.get("name_en", ""),
            "name_fr": p.get("name_fr", ""),
            "category": p.get("category", ""),
            "purity": p.get("purity", ""),
            "dosage_mg": p.get("dosage_mg", 0),
            "price_cad": p.get("price_cad", 0),
            "stock": p.get("stock", 0),
            "variants": "; ".join(
                f"{v.get('name','')} ${v.get('price',0)} stock={v.get('stock',0)}" for v in (p.get("variants") or [])
            ),
            "low_stock_threshold": p.get("low_stock_threshold", 10),
            "featured": p.get("featured", False),
            "lab_tested": p.get("lab_tested", False),
            "active": p.get("active", True),
            # Scientific data fields (new)
            "molecular_formula": p.get("molecular_formula", ""),
            "molecular_weight": p.get("molecular_weight", ""),
            "cas_number": p.get("cas_number", ""),
            "sequence_length": p.get("sequence_length", ""),
            "storage": p.get("storage", ""),
            "solubility": p.get("solubility", ""),
            "appearance": p.get("appearance", ""),
            "mechanism": p.get("mechanism", ""),
            "research_areas": "; ".join(p.get("research_areas", [])),
            "synonyms": "; ".join(p.get("synonyms", [])),
        })
    return _xlsx_response(rows, f"fironova-products-{datetime.now().strftime('%Y%m%d')}.xlsx")


# ---------------------------------------------------------------------------
# PDF Invoice
# ---------------------------------------------------------------------------
def _generate_invoice_pdf(order: dict) -> bytes:
    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=LETTER)
    w, h = LETTER

    # Header band
    c.setFillColor(rl_colors.black)
    c.rect(0, h - 18 * mm, w, 18 * mm, fill=1, stroke=0)
    c.setFillColor(rl_colors.white)
    c.setFont("Helvetica-Bold", 22)
    c.drawString(20 * mm, h - 13 * mm, "FIRONOVA")
    c.setFillColor(rl_colors.HexColor("#C20114"))
    c.circle(20 * mm + 41 * mm, h - 13 * mm + 1.5 * mm, 1.5 * mm, fill=1, stroke=0)
    c.setFillColor(rl_colors.white)
    c.setFont("Courier", 8)
    c.drawRightString(w - 20 * mm, h - 9 * mm, "// INVOICE")
    c.drawRightString(w - 20 * mm, h - 14 * mm, f"ORDER {order.get('order_number','')}")

    # Meta
    y = h - 30 * mm
    c.setFillColor(rl_colors.black)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(20 * mm, y, "INVOICE")
    c.setFont("Helvetica", 9)
    y -= 6 * mm
    c.drawString(20 * mm, y, f"Order #: {order.get('order_number','')}")
    y -= 4 * mm
    c.drawString(20 * mm, y, f"Date: {(order.get('created_at') or '')[:10]}")
    y -= 4 * mm
    c.drawString(20 * mm, y, f"Status: {order.get('payment_status','')} / {order.get('fulfillment_status','')}")

    # Bill to
    addr = order.get("shipping_address") or {}
    c.setFont("Helvetica-Bold", 10)
    c.drawString(110 * mm, h - 36 * mm, "SHIP TO")
    c.setFont("Helvetica", 9)
    sy = h - 42 * mm
    for line in [
        addr.get("full_name", ""),
        addr.get("address1", ""),
        addr.get("address2", "") or None,
        f"{addr.get('city','')}, {addr.get('province','')} {addr.get('postal_code','')}",
        addr.get("country", ""),
        order.get("email") or "",
    ]:
        if line:
            c.drawString(110 * mm, sy, str(line))
            sy -= 4 * mm

    # Items table
    y -= 20 * mm
    c.setFillColor(rl_colors.black)
    c.rect(20 * mm, y - 1, w - 40 * mm, 7 * mm, fill=1, stroke=0)
    c.setFillColor(rl_colors.white)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(22 * mm, y + 1.5 * mm, "ITEM")
    c.drawString(110 * mm, y + 1.5 * mm, "QTY")
    c.drawRightString(150 * mm, y + 1.5 * mm, "UNIT")
    c.drawRightString(w - 22 * mm, y + 1.5 * mm, "TOTAL")
    y -= 6 * mm
    c.setFillColor(rl_colors.black)
    c.setFont("Helvetica", 9)
    for it in order.get("items", []):
        y -= 6 * mm
        c.drawString(22 * mm, y, f"{it.get('name_en','')} ({it.get('slug','')})")
        c.drawString(110 * mm, y, str(it.get("qty", 0)))
        c.drawRightString(150 * mm, y, f"${it.get('price_cad',0):.2f}")
        c.drawRightString(w - 22 * mm, y, f"${it.get('line_total',0):.2f}")
        c.setStrokeColor(rl_colors.HexColor("#e0e0e0"))
        c.line(20 * mm, y - 2 * mm, w - 20 * mm, y - 2 * mm)

    # Totals
    y -= 12 * mm
    c.setFont("Helvetica", 9)
    c.drawRightString(150 * mm, y, "Subtotal")
    c.drawRightString(w - 22 * mm, y, f"${order.get('subtotal',0):.2f}")
    if order.get("discount", 0) > 0:
        y -= 4 * mm
        c.drawRightString(150 * mm, y, f"Discount ({(order.get('coupon') or {}).get('code','')})")
        c.drawRightString(w - 22 * mm, y, f"-${order.get('discount',0):.2f}")
    y -= 4 * mm
    c.drawRightString(150 * mm, y, "Shipping")
    c.drawRightString(w - 22 * mm, y, f"${order.get('shipping',0):.2f}")
    y -= 6 * mm
    c.setStrokeColor(rl_colors.black)
    c.line(110 * mm, y + 2 * mm, w - 20 * mm, y + 2 * mm)
    c.setFont("Helvetica-Bold", 12)
    c.drawRightString(150 * mm, y - 3 * mm, "TOTAL CAD")
    c.drawRightString(w - 22 * mm, y - 3 * mm, f"${order.get('total',0):.2f}")

    # Footer disclaimer
    c.setFont("Courier", 7)
    c.setFillColor(rl_colors.HexColor("#666666"))
    c.drawCentredString(w / 2, 20 * mm, "FOR LABORATORY RESEARCH USE ONLY · NOT FOR HUMAN OR VETERINARY CONSUMPTION · 19+ ONLY")
    c.drawCentredString(w / 2, 16 * mm, f"FIRONOVA · CANADA · INVOICE GENERATED {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")

    c.showPage()
    c.save()
    buf.seek(0)
    return buf.getvalue()


async def order_invoice_pdf(order_id: str, request: Request):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    user = await _resolve_user(request)
    if order.get("user_id"):
        if not user or (user["id"] != order["user_id"] and user.get("role") != "admin"):
            raise HTTPException(403, "Forbidden")
    elif not _guest_order_accessible(order, request):
        raise HTTPException(403, "Forbidden")
    pdf = _generate_invoice_pdf(order)
    return Response(
        content=pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="invoice-{order.get("order_number","order")}.pdf"'},
    )


# ---------------------------------------------------------------------------
# Interac Autodeposit — confirmation automatique via Microsoft Graph API
# ---------------------------------------------------------------------------

_GRAPH_TOKEN_URL = "https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
_GRAPH_API_URL = "https://graph.microsoft.com/v1.0"


INTERAC_GRAPH_POLL_SECONDS = int(os.environ.get("INTERAC_GRAPH_POLL_SECONDS", "120"))


@api.get("/admin/interac/graph-health")
async def admin_interac_graph_health(_admin: dict = Depends(get_admin_user)):
    """Diagnostic Microsoft Graph — vérifie que l'app peut obtenir un token
    et lire la boîte INTERAC_GRAPH_USER. Retourne latence + code d'erreur.
    Ne modifie aucune donnée. Réservé aux administrateurs."""
    import time as _time
    result: dict = {
        "mode": INTERAC_AUTOCONFIRM_MODE,
        "tenant_id_configured": bool(INTERAC_GRAPH_TENANT_ID),
        "client_id_configured": bool(INTERAC_GRAPH_CLIENT_ID),
        "client_secret_configured": bool(INTERAC_GRAPH_CLIENT_SECRET),
        "graph_user": INTERAC_GRAPH_USER,
        "trusted_sender_configured": bool(INTERAC_TRUSTED_SENDER),
        "poll_interval_seconds": INTERAC_GRAPH_POLL_SECONDS,
    }
    if INTERAC_AUTOCONFIRM_MODE != "strict":
        result.update({"ok": False, "error": "INTERAC_AUTOCONFIRM_MODE is not 'strict'"})
        return result
    if not (INTERAC_GRAPH_TENANT_ID and INTERAC_GRAPH_CLIENT_ID and INTERAC_GRAPH_CLIENT_SECRET):
        result.update({"ok": False, "error": "One or more Azure AD credentials are missing"})
        return result
    t0 = _time.perf_counter()
    try:
        async with httpx.AsyncClient(timeout=20) as cx:
            tok_resp = await cx.post(
                _GRAPH_TOKEN_URL.format(tenant=INTERAC_GRAPH_TENANT_ID),
                data={
                    "grant_type": "client_credentials",
                    "client_id": INTERAC_GRAPH_CLIENT_ID,
                    "client_secret": INTERAC_GRAPH_CLIENT_SECRET,
                    "scope": "https://graph.microsoft.com/.default",
                },
            )
        if tok_resp.status_code != 200:
            result.update({
                "ok": False,
                "stage": "token",
                "http_status": tok_resp.status_code,
                "error": tok_resp.text[:400],
                "latency_ms": int((_time.perf_counter() - t0) * 1000),
            })
            return result
        access_token = tok_resp.json().get("access_token")
        async with httpx.AsyncClient(timeout=20) as cx:
            msg_resp = await cx.get(
                f"{_GRAPH_API_URL}/users/{INTERAC_GRAPH_USER}/messages",
                params={"$top": 1, "$select": "id,subject,receivedDateTime"},
                headers={"Authorization": f"Bearer {access_token}"},
            )
        latency_ms = int((_time.perf_counter() - t0) * 1000)
        if msg_resp.status_code != 200:
            result.update({
                "ok": False,
                "stage": "messages",
                "http_status": msg_resp.status_code,
                "error": msg_resp.text[:400],
                "latency_ms": latency_ms,
            })
            return result
        payload = msg_resp.json()
        result.update({
            "ok": True,
            "stage": "messages",
            "http_status": 200,
            "latency_ms": latency_ms,
            "messages_returned": len(payload.get("value") or []),
        })
        return result
    except Exception as ex:
        result.update({
            "ok": False,
            "error": str(ex)[:400],
            "latency_ms": int((_time.perf_counter() - t0) * 1000),
        })
        return result


# ---------------------------------------------------------------------------
# Public meta
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Seed data
# ---------------------------------------------------------------------------
# =============================================================================
# FIRONOVA — Catalogue produits (19 peptides de recherche)
# Données scientifiques vérifiées (PubChem, ChemicalBook, Sigma-Aldrich).
# active: True  → 5 produits au lancement
# active: False → 14 produits, à activer depuis l'admin
# image_url: "" → à remplir via POST /admin/upload/image
# =============================================================================
SEED_PRODUCTS = [
    {
        "slug": "bpc-157-5mg",
        "name_en": "BPC-157",
        "name_fr": "BPC-157",
        "category": "healing",
        "sequence": "GEPPPGKPADDAGLV",
        "molecular_formula": "C62H98N16O22",
        "molecular_weight": 1419.55,
        "cas_number": "137525-51-0",
        "sequence_length": 15,
        "purity": "≥ 99.0%",
        "dosage_mg": 5.0,
        "description_en": "Body Protection Compound, a 15-amino-acid synthetic peptide derived from gastric protein. Widely studied in research models for tissue repair pathways.",
        "description_fr": "Composé de protection corporelle, peptide synthétique de 15 acides aminés dérivé d'une protéine gastrique. Étudié dans des modèles de recherche sur les voies de réparation tissulaire.",
        "price_cad": 64.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": True,
    },
    {
        "slug": "tb-500-5mg",
        "name_en": "TB-500",
        "name_fr": "TB-500",
        "category": "healing",
        "sequence": "Ac-SDKPDMAEIEKFDKSKLKKTETQEKNPLPSKETIEQEKQAGES",
        "molecular_formula": "C212H350N56O78S",
        "molecular_weight": 4963.44,
        "cas_number": "77591-33-4",
        "sequence_length": 43,
        "purity": "≥ 99.0%",
        "dosage_mg": 5.0,
        "description_en": "Full-length Thymosin Beta-4 (43 amino acids). Investigated in research for actin sequestration and cellular migration studies.",
        "description_fr": "Thymosine bêta-4 complète (43 acides aminés). Étudiée en recherche pour la séquestration de l'actine et les études de migration cellulaire.",
        "price_cad": 79.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": True,
    },
    {
        "slug": "semaglutide-5mg",
        "name_en": "Semaglutide",
        "name_fr": "Sémaglutide",
        "category": "weight-loss",
        "sequence": "GLP-1 analog (31 aa, C18 diacid modified)",
        "molecular_formula": "C187H291N45O59",
        "molecular_weight": 4113.58,
        "cas_number": "910463-68-2",
        "sequence_length": 31,
        "purity": "≥ 98.0%",
        "dosage_mg": 5.0,
        "description_en": "GLP-1 receptor agonist analog under extensive research in metabolic studies.",
        "description_fr": "Analogue agoniste du récepteur GLP-1 largement étudié dans les recherches métaboliques.",
        "price_cad": 189.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": True,
    },
    {
        "slug": "tirzepatide-10mg",
        "name_en": "Tirzepatide",
        "name_fr": "Tirzépatide",
        "category": "weight-loss",
        "sequence": "GIP/GLP-1 dual agonist (39 aa, C20 diacid modified)",
        "molecular_formula": "C225H348N48O68",
        "molecular_weight": 4813.45,
        "cas_number": "2023788-19-2",
        "sequence_length": 39,
        "purity": "≥ 98.0%",
        "dosage_mg": 10.0,
        "description_en": "Dual GIP and GLP-1 receptor agonist studied in metabolic and body-weight research models.",
        "description_fr": "Double agoniste des récepteurs GIP et GLP-1 étudié dans les modèles de recherche métabolique et pondérale.",
        "price_cad": 219.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": True,
    },
    {
        "slug": "ipamorelin-5mg",
        "name_en": "Ipamorelin",
        "name_fr": "Ipamoréline",
        "category": "cognitive",
        "sequence": "Aib-His-D-2-Nal-D-Phe-Lys-NH2",
        "molecular_formula": "C38H49N9O5",
        "molecular_weight": 711.85,
        "cas_number": "170851-70-4",
        "sequence_length": 5,
        "purity": "≥ 98.0%",
        "dosage_mg": 5.0,
        "description_en": "Selective growth hormone secretagogue (ghrelin receptor agonist) studied in endocrine research.",
        "description_fr": "Sécrétagogue sélectif de l'hormone de croissance (agoniste du récepteur de la ghréline) étudié en recherche endocrinienne.",
        "price_cad": 74.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": True,
    },
    {
        "slug": "cjc-1295-no-dac-5mg",
        "name_en": "CJC-1295 No-DAC",
        "name_fr": "CJC-1295 sans DAC",
        "category": "cognitive",
        "sequence": "Tyr-D-Ala-Asp-Ala-Ile-Phe-Thr-Gln-Ser-Tyr-Arg-Lys-Val-Leu-Ala-Gln-Leu-Ser-Ala-Arg-Lys-Leu-Leu-Gln-Asp-Ile-Leu-Ser-Arg-NH2",
        "molecular_formula": "C152H252N44O42",
        "molecular_weight": 3367.9,
        "cas_number": "863288-34-0",
        "sequence_length": 29,
        "purity": "≥ 98.0%",
        "dosage_mg": 5.0,
        "description_en": "Modified GRF (1-29), a GHRH analog studied for pulsatile growth-hormone-release research.",
        "description_fr": "GRF modifié (1-29), analogue de la GHRH étudié pour la recherche sur la libération pulsatile de l'hormone de croissance.",
        "price_cad": 69.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": False,
    },
    {
        "slug": "selank-5mg",
        "name_en": "Selank",
        "name_fr": "Sélank",
        "category": "cognitive",
        "sequence": "Thr-Lys-Pro-Arg-Pro-Gly-Pro",
        "molecular_formula": "C33H57N11O9",
        "molecular_weight": 751.87,
        "cas_number": "129954-34-3",
        "sequence_length": 7,
        "purity": "≥ 99.0%",
        "dosage_mg": 5.0,
        "description_en": "Synthetic heptapeptide derived from tuftsin, studied for anxiolytic and nootropic pathways.",
        "description_fr": "Heptapeptide synthétique dérivé de la tuftsine, étudié pour les voies anxiolytiques et nootropiques.",
        "price_cad": 59.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": False,
    },
    {
        "slug": "semax-10mg",
        "name_en": "Semax",
        "name_fr": "Sémax",
        "category": "cognitive",
        "sequence": "Met-Glu-His-Phe-Pro-Gly-Pro",
        "molecular_formula": "C37H51N9O10S",
        "molecular_weight": 813.92,
        "cas_number": "80714-61-0",
        "sequence_length": 7,
        "purity": "≥ 99.0%",
        "dosage_mg": 10.0,
        "description_en": "Synthetic heptapeptide ACTH(4-7) analog studied for BDNF modulation and neuroprotection.",
        "description_fr": "Heptapeptide synthétique analogue de l'ACTH(4-7) étudié pour la modulation du BDNF et la neuroprotection.",
        "price_cad": 84.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": False,
    },
    {
        "slug": "ghk-cu-50mg",
        "name_en": "GHK-Cu",
        "name_fr": "GHK-Cu",
        "category": "healing",
        "sequence": "Gly-His-Lys · Cu(II)",
        "molecular_formula": "C14H22CuN6O4",
        "molecular_weight": 403.93,
        "cas_number": "89030-95-5",
        "sequence_length": 3,
        "purity": "≥ 99.0%",
        "dosage_mg": 50.0,
        "description_en": "Copper(II) complex of the tripeptide Gly-His-Lys, studied for collagen synthesis and matrix remodeling.",
        "description_fr": "Complexe cuivre(II) du tripeptide Gly-His-Lys, étudié pour la synthèse du collagène et le remodelage matriciel.",
        "price_cad": 89.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": False,
    },
    {
        "slug": "retatrutide-10mg",
        "name_en": "Retatrutide",
        "name_fr": "Rétatrutide",
        "category": "weight-loss",
        "sequence": "GIP/GLP-1/GCGR triple agonist (39 aa, lipid diacid modified)",
        "molecular_formula": "C221H342N46O68",
        "molecular_weight": 4731.33,
        "cas_number": "2381089-83-2",
        "sequence_length": 39,
        "purity": "≥ 98.0%",
        "dosage_mg": 10.0,
        "description_en": "Triple GIP/GLP-1/glucagon receptor agonist studied in metabolic and body-weight research.",
        "description_fr": "Triple agoniste des récepteurs GIP/GLP-1/glucagon étudié en recherche métabolique et pondérale.",
        "price_cad": 249.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": False,
    },
    {
        "slug": "pt-141-10mg",
        "name_en": "PT-141",
        "name_fr": "PT-141",
        "category": "cognitive",
        "sequence": "Ac-Nle-cyclo[Asp-His-D-Phe-Arg-Trp-Lys]-OH",
        "molecular_formula": "C50H68N14O10",
        "molecular_weight": 1025.18,
        "cas_number": "189691-06-3",
        "sequence_length": 7,
        "purity": "≥ 98.0%",
        "dosage_mg": 10.0,
        "description_en": "Cyclic melanocortin receptor agonist (bremelanotide) studied in neuroscience research.",
        "description_fr": "Agoniste cyclique des récepteurs de la mélanocortine (bremélanotide) étudié en recherche en neurosciences.",
        "price_cad": 79.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": False,
    },
    {
        "slug": "melanotan-2-10mg",
        "name_en": "Melanotan-2",
        "name_fr": "Mélanotan-2",
        "category": "cognitive",
        "sequence": "Ac-Nle-cyclo[Asp-His-D-Phe-Arg-Trp-Lys]-NH2",
        "molecular_formula": "C50H69N15O9",
        "molecular_weight": 1024.18,
        "cas_number": "121062-08-6",
        "sequence_length": 7,
        "purity": "≥ 98.0%",
        "dosage_mg": 10.0,
        "description_en": "Cyclic heptapeptide α-MSH analog and non-selective melanocortin receptor agonist studied in research.",
        "description_fr": "Heptapeptide cyclique analogue de l'α-MSH et agoniste non sélectif des récepteurs de la mélanocortine étudié en recherche.",
        "price_cad": 69.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": False,
    },
    {
        "slug": "mots-c-10mg",
        "name_en": "MOTS-c",
        "name_fr": "MOTS-c",
        "category": "weight-loss",
        "sequence": "Met-Arg-Trp-Gln-Glu-Met-Gly-Tyr-Ile-Phe-Tyr-Pro-Arg-Lys-Leu-Arg",
        "molecular_formula": "C101H152N28O22S2",
        "molecular_weight": 2174.6,
        "cas_number": "1627580-64-6",
        "sequence_length": 16,
        "purity": "≥ 98.0%",
        "dosage_mg": 10.0,
        "description_en": "Mitochondrial-derived 16-amino-acid peptide studied for AMPK-linked metabolic regulation.",
        "description_fr": "Peptide mitochondrial de 16 acides aminés étudié pour la régulation métabolique liée à l'AMPK.",
        "price_cad": 99.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": False,
    },
    {
        "slug": "epithalon-10mg",
        "name_en": "Epithalon",
        "name_fr": "Épithalon",
        "category": "healing",
        "sequence": "Ala-Glu-Asp-Gly",
        "molecular_formula": "C14H22N4O9",
        "molecular_weight": 390.35,
        "cas_number": "307297-39-8",
        "sequence_length": 4,
        "purity": "≥ 99.0%",
        "dosage_mg": 10.0,
        "description_en": "Synthetic tetrapeptide (AEDG) studied for telomerase activity and circadian gene expression.",
        "description_fr": "Tétrapeptide synthétique (AEDG) étudié pour l'activité télomérase et l'expression des gènes circadiens.",
        "price_cad": 64.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": False,
    },
    {
        "slug": "tesamorelin-5mg",
        "name_en": "Tesamorelin",
        "name_fr": "Tésamoréline",
        "category": "weight-loss",
        "sequence": "trans-3-hexenoyl-GRF(1-44)-NH2 (44 aa)",
        "molecular_formula": "C221H366N72O67S",
        "molecular_weight": 5135.86,
        "cas_number": "218949-48-5",
        "sequence_length": 44,
        "purity": "≥ 98.0%",
        "dosage_mg": 5.0,
        "description_en": "Stabilized GHRH(1-44) analog studied in metabolic and adipose-tissue research models.",
        "description_fr": "Analogue stabilisé de la GHRH(1-44) étudié dans les modèles de recherche métabolique et adipeuse.",
        "price_cad": 179.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": False,
    },
    {
        "slug": "hexarelin-5mg",
        "name_en": "Hexarelin",
        "name_fr": "Hexaréline",
        "category": "cognitive",
        "sequence": "His-D-2-Me-Trp-Ala-Trp-D-Phe-Lys-NH2",
        "molecular_formula": "C47H58N12O6",
        "molecular_weight": 887.04,
        "cas_number": "140703-51-1",
        "sequence_length": 6,
        "purity": "≥ 98.0%",
        "dosage_mg": 5.0,
        "description_en": "Synthetic hexapeptide growth hormone secretagogue (GHS-R agonist) studied in endocrine research.",
        "description_fr": "Hexapeptide synthétique sécrétagogue de l'hormone de croissance (agoniste GHS-R) étudié en recherche endocrinienne.",
        "price_cad": 74.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": False,
    },
    {
        "slug": "ghrp-2-5mg",
        "name_en": "GHRP-2",
        "name_fr": "GHRP-2",
        "category": "cognitive",
        "sequence": "D-Ala-D-2-Nal-Ala-Trp-D-Phe-Lys-NH2",
        "molecular_formula": "C45H55N9O6",
        "molecular_weight": 817.97,
        "cas_number": "158861-67-7",
        "sequence_length": 6,
        "purity": "≥ 98.0%",
        "dosage_mg": 5.0,
        "description_en": "Synthetic hexapeptide (pralmorelin) growth hormone secretagogue studied in endocrine research.",
        "description_fr": "Hexapeptide synthétique (pralmoréline) sécrétagogue de l'hormone de croissance étudié en recherche endocrinienne.",
        "price_cad": 69.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": False,
    },
    {
        "slug": "ghrp-6-5mg",
        "name_en": "GHRP-6",
        "name_fr": "GHRP-6",
        "category": "cognitive",
        "sequence": "His-D-Trp-Ala-Trp-D-Phe-Lys-NH2",
        "molecular_formula": "C46H56N12O6",
        "molecular_weight": 873.03,
        "cas_number": "87616-84-0",
        "sequence_length": 6,
        "purity": "≥ 98.0%",
        "dosage_mg": 5.0,
        "description_en": "Synthetic hexapeptide growth hormone secretagogue and ghrelin receptor agonist studied in research.",
        "description_fr": "Hexapeptide synthétique sécrétagogue de l'hormone de croissance et agoniste du récepteur de la ghréline étudié en recherche.",
        "price_cad": 64.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": False,
    },
    {
        "slug": "aod-9604-5mg",
        "name_en": "AOD-9604",
        "name_fr": "AOD-9604",
        "category": "weight-loss",
        "sequence": "Tyr-Leu-Arg-Ile-Val-Gln-Cys-Arg-Ser-Val-Glu-Gly-Ser-Cys-Gly-Phe (disulfide Cys7-Cys14)",
        "molecular_formula": "C78H123N23O23S2",
        "molecular_weight": 1815.1,
        "cas_number": "221231-10-3",
        "sequence_length": 16,
        "purity": "≥ 99.0%",
        "dosage_mg": 5.0,
        "description_en": "Modified hGH fragment (176-191) with N-terminal tyrosine, studied in lipid-metabolism research.",
        "description_fr": "Fragment modifié de l'hGH (176-191) avec tyrosine N-terminale, étudié en recherche sur le métabolisme lipidique.",
        "price_cad": 89.99,
        "stock": 0,
        "image_url": "",
        "lab_tested": True,
        "active": False,
    },
]


BPC_157_CANONICAL_ID = "5dc85d91-7089-4e6c-88b5-67ff79c6ef92"
BPC_157_CANONICAL_SLUG = "bpc-157-5mg"
BPC_157_LEGACY_SLUG = "bpc-157"


def _variant_stock_or_default(variants: list[dict], name: str, fallback: int) -> int:
    for variant in variants:
        if variant.get("name") == name:
            try:
                stock = int(variant.get("stock", fallback) or 0)
                if stock > 0:
                    return stock
            except (TypeError, ValueError):
                pass
    return fallback


async def _repair_nonprod_bpc_seed() -> None:
    """Converge legacy preview/demo BPC seed drift to a single canonical doc.

    Older preview data used slug `bpc-157` with the fixed product id, while a
    later seed inserted a second `bpc-157-5mg` row. Tests and the frontend
    expect the fixed id and canonical slug to point to the same product.
    """
    if IS_PRODUCTION:
        return

    canonical = await db.products.find_one({"slug": BPC_157_CANONICAL_SLUG})
    legacy = await db.products.find_one({"id": BPC_157_CANONICAL_ID})
    legacy_slug = await db.products.find_one({"slug": BPC_157_LEGACY_SLUG})

    source = legacy or legacy_slug or canonical
    if not source:
        return

    all_variants = []
    for doc in (legacy, canonical, legacy_slug):
        if doc:
            all_variants.extend(doc.get("variants") or [])
    five_variant = next((v for v in all_variants if v.get("name") == "5.0mg"), None)
    ten_variant = next((v for v in all_variants if v.get("name") == "10.0mg"), None)
    canonical_ok = (
        source.get("id") == BPC_157_CANONICAL_ID
        and source.get("slug") == BPC_157_CANONICAL_SLUG
        and five_variant is not None
        and ten_variant is not None
        and bool(ten_variant.get("preorder_enabled"))
        and bool(ten_variant.get("badge_coa_pending"))
    )
    duplicate_slug_doc = canonical and canonical.get("id") != BPC_157_CANONICAL_ID
    if canonical_ok and not duplicate_slug_doc:
        return

    bpc_seed = next(p for p in SEED_PRODUCTS if p["slug"] == BPC_157_CANONICAL_SLUG)
    five_stock = _variant_stock_or_default(all_variants, "5.0mg", 12)
    ten_stock = _variant_stock_or_default(all_variants, "10.0mg", 6)
    repaired_variants = [
        {
            "id": (five_variant or {}).get("id") or str(uuid.uuid4()),
            "name": "5.0mg",
            "price": 64.99,
            "sale_price": 49.99,
            "stock": five_stock,
            "sku": "BPC-157-5MG",
            "coa_status": "available",
            "badge_coa_available": True,
            "badge_coa_pending": False,
            "badge_coming_soon": False,
            "preorder_enabled": False,
            "preorder_delay_message": "",
            "preorder_price": None,
            "preorder_note": "",
            "coa_url": (five_variant or {}).get("coa_url") or "https://example.com/coa-bpc157-5mg.pdf",
        },
        {
            "id": (ten_variant or {}).get("id") or str(uuid.uuid4()),
            "name": "10.0mg",
            "price": 100.0,
            "sale_price": None,
            "stock": ten_stock,
            "sku": "BPC-157-10MG",
            "coa_status": "pending",
            "badge_coa_available": False,
            "badge_coa_pending": True,
            "badge_coming_soon": False,
            "preorder_enabled": True,
            "preorder_delay_message": "",
            "preorder_price": 85.0,
            "preorder_note": "COA pending",
            "coa_url": "",
        },
    ]

    if canonical and canonical.get("_id") != source.get("_id"):
        await db.products.delete_one({"_id": canonical["_id"]})

    await db.products.update_one(
        {"_id": source["_id"]},
        {"$set": {
            "id": BPC_157_CANONICAL_ID,
            "slug": BPC_157_CANONICAL_SLUG,
            "name_en": bpc_seed["name_en"],
            "name_fr": bpc_seed["name_fr"],
            "category": bpc_seed["category"],
            "sequence": bpc_seed["sequence"],
            "molecular_formula": bpc_seed["molecular_formula"],
            "molecular_weight": bpc_seed["molecular_weight"],
            "cas_number": bpc_seed["cas_number"],
            "sequence_length": bpc_seed["sequence_length"],
            "purity": bpc_seed["purity"],
            "dosage_mg": bpc_seed["dosage_mg"],
            "description_en": bpc_seed["description_en"],
            "description_fr": bpc_seed["description_fr"],
            "price_cad": 64.99,
            "stock": five_stock + ten_stock,
            "active": True,
            "featured": True,
            "preorder_allowed": False,
            "coa_url": "",
            "coa_lot": source.get("coa_lot", ""),
            "coa_date": source.get("coa_date", ""),
            "variants": repaired_variants,
        }},
    )


async def seed_admin_and_products():
    # Indexes
    await db.rate_limit_counters.create_index("expires_at", expireAfterSeconds=0)
    await db.idempotency.create_index("expires_at", expireAfterSeconds=0)
    await db.users.create_index("email", unique=True)
    await db.users.create_index("id", unique=True)
    await db.refresh_sessions.create_index("token_hash", unique=True)
    await db.refresh_sessions.create_index([("user_id", 1), ("revoked_at", 1)])
    await db.refresh_sessions.create_index([("family_id", 1), ("revoked_at", 1)])
    await db.refresh_sessions.create_index("expires_at", expireAfterSeconds=0)
    await db.products.create_index("slug", unique=True)
    await affiliate_ensure_indexes()
    await db.products.create_index("id", unique=True)
    await db.products.create_index([("active", 1), ("category", 1)])
    await db.products.create_index([("active", 1), ("featured", 1)])
    await db.orders.create_index("order_number")
    await db.orders.create_index("user_id")
    await db.orders.create_index("id", unique=True)
    # Écrans admin & watchdog : sans ces index, chaque poll = COLLSCAN complet.
    await db.orders.create_index([("payment_status", 1), ("created_at", -1)])
    await db.orders.create_index([("fulfillment_status", 1), ("created_at", -1)])
    await db.orders.create_index([("dispatch_batch", 1), ("payment_status", 1)])
    await db.orders.create_index([("created_at", -1)])
    # Recherche par email client (admin customers screen) & filtre par affilié
    # (dashboard commissions, calcul de payouts) : sans index chaque appel = COLLSCAN.
    await db.orders.create_index("email")
    await db.orders.create_index([("affiliate_id", 1), ("payment_status", 1)])
    await db.orders.create_index([("paid_at", -1), ("payment_status", 1)])
    # Filtrage admin par rôle (staff, affiliés, clients) — évite un scan complet users.
    await db.users.create_index("role")
    await db.coupons.create_index("code", unique=True)
    # counted_order_ids : lookup index (non-unique).
    # L'ancienne définition `unique=True, sparse=True` plantait au cancel :
    # $setDifference sur le dernier order_id produit `[]`, MongoDB indexait
    # cela comme `null`, et deux coupons avec `[]` violaient l'unique multikey
    # (`E11000 dup key: counted_order_ids: undefined`).
    # L'invariant "un même order_id n'apparaît que dans un seul coupon" est
    # déjà garanti par le filtre `counted_order_ids: {$ne: order_id}` dans
    # `_apply_coupon_usage`. L'index reste utile pour retrouver rapidement
    # le coupon lors du décrément — pas besoin d'unicité au niveau BDD.
    try:
        await db.coupons.drop_index("counted_order_ids_1")
    except Exception:
        pass  # index absent : silencieux
    await db.coupons.create_index("counted_order_ids")
    await db.payment_transactions.create_index("session_id", unique=True)
    await db.payment_transactions.create_index("order_id")
    await db.stock_notifications.create_index([("product_id", 1), ("variant_id", 1), ("notified", 1)])
    await db.subscribers.create_index("email", unique=True)
    await db.subscribers.create_index("unsubscribe_token", unique=True)
    await db.subscribers.create_index("status")
    await db.addresses.create_index([("user_id", 1), ("created_at", -1)])
    await db.categories.create_index("slug", unique=True)
    await db.menus.create_index("slug", unique=True)
    await db.menus.create_index([("location", 1), ("published", 1), ("display_order", 1)])
    # Bannière « manifeste non transmis » : sans index, chaque chargement de
    # l'admin scanne toute la collection orders.
    await db.orders.create_index([("shipping_info.cp_transmitted", 1), ("shipping_info.cp_group_id", 1)])
    await db.email_change_requests.create_index("token_hash", unique=True)
    await db.email_change_requests.create_index("user_id")
    await db.staff_invites.create_index("token_hash", unique=True)
    await db.staff_invites.create_index("email")
    await db.admin_audit_log.create_index([("created_at", -1)])
    await db.admin_audit_log.create_index("user_id")
    await db.interac_reconciliation_queue.create_index("id", unique=True)
    await db.interac_reconciliation_queue.create_index("graph_message_id", unique=True)
    await db.interac_reconciliation_queue.create_index([("status", 1), ("detected_at", -1)])
    await db.webhook_events.create_index("event_key", unique=True)
    await db.webhook_events.create_index("created_at_dt", expireAfterSeconds=172800)
    await db.email_outbox.create_index("id", unique=True)
    await db.email_outbox.create_index([("status", 1), ("available_at", 1), ("created_at", 1)])
    await db.email_outbox.create_index("expires_at", expireAfterSeconds=0)
    # Affiliate payout deferrals — Item 3.2 : audit + idempotence des notifications
    # de report envoyées quand le montant cumulé d'un affilié pour une période
    # est en dessous de AFFILIATE_PAYOUT_MIN_CAD. Unique (affiliate_id, period) →
    # un affilié ne reçoit qu'UN seul email par période même si le run est rejoué.
    await db.affiliate_payout_deferrals.create_index(
        [("affiliate_id", 1), ("period", 1)], unique=True
    )
    await db.affiliate_payout_deferrals.create_index([("created_at", -1)])
    # Stock movements — audit trail restocks admin (qui, quand, combien).
    # Un mouvement par ligne (product_id, variant_id?, delta, admin_email, timestamp).
    await db.stock_movements.create_index([("product_id", 1), ("created_at", -1)])
    await db.stock_movements.create_index([("created_at", -1)])
    # Low stock alerts — un doc par (product_id, variant_id), unique.
    # `active: true` = alerte en cours (email envoyé, on ne re-notifie pas).
    # `active: false` = seuil regagné → prêt à re-déclencher plus tard.
    await db.low_stock_alerts.create_index(
        [("product_id", 1), ("variant_id", 1)], unique=True
    )
    await db.low_stock_alerts.create_index([("active", 1), ("triggered_at", -1)])
    await db.products.create_index("deleted_at")
    await db.coupons.create_index("deleted_at")
    await db.shipping_zones.create_index("deleted_at")
    await db.shipping_methods.create_index("deleted_at")
    await db.orders.create_index("deleted_at")
    # Failure ledger — Item 1.2 B4 SMART : audit trail des compensations qui
    # ont échoué au checkout. Indexé pour la vue admin (list + status filter).
    await db.checkout_compensation_failures.create_index([("status", 1), ("created_at", -1)])
    await db.checkout_compensation_failures.create_index("order_id")
    # Admin
    existing = await db.users.find_one({"email": ADMIN_EMAIL.lower()})
    hashed = hash_password(ADMIN_PASSWORD)
    if not existing:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": ADMIN_EMAIL.lower(),
            "name": "FIRONOVA Admin",
            "password_hash": hashed,
            "role": "admin",
            "token_version": 0,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    elif not verify_password(ADMIN_PASSWORD, existing["password_hash"]):
        # Rotation de mot de passe détectée (ADMIN_PASSWORD changé côté env) :
        # on incrémente token_version pour révoquer immédiatement toute
        # session admin ouverte avec l'ancien mot de passe.
        await db.users.update_one(
            {"email": ADMIN_EMAIL.lower()},
            {"$set": {"password_hash": hashed, "role": "admin"},
             "$inc": {"token_version": 1}},
        )

    # Products
    featured_slugs = {"bpc-157-5mg", "semaglutide-5mg", "tirzepatide-10mg", "ipamorelin-5mg", "ghk-cu-50mg", "epithalon-10mg"}
    for p in SEED_PRODUCTS:
        default_variant = {
            "id": str(uuid.uuid4()),
            "name": f"{p['dosage_mg']}mg",
            "price": p["price_cad"],
            "stock": p["stock"],
            "sku": p["slug"].upper(),
            "coa_status": "available",
            "badge_coa_available": True,
            "badge_coa_pending": False,
            "badge_coming_soon": False,
            "preorder_enabled": False,
            "preorder_delay_message": "",
            "preorder_price": None,
            "preorder_note": "",
        }
        defaults = {
            **p,
            "id": str(uuid.uuid4()),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "featured": p["slug"] in featured_slugs,
            "preorder_allowed": False,
            "low_stock_threshold": 10,
            "coa_url": "",
            "coa_lot": "",
            "coa_date": "",
            "variants": [default_variant],
        }
        await db.products.update_one({"slug": p["slug"]}, {"$setOnInsert": defaults}, upsert=True)
        # Backfill featured flag and scientific fields on existing docs (non-destructive)
        sci_backfill = {k: p[k] for k in (
            "molecular_formula", "molecular_weight", "cas_number", "sequence_length",
            "sequence", "purity", "description_en", "description_fr",
        ) if k in p}
        sci_backfill["featured"] = p["slug"] in featured_slugs
        await db.products.update_one(
            {"slug": p["slug"]},
            {"$set": sci_backfill},
        )
        for field, value in {"preorder_allowed": False, "low_stock_threshold": 10,
                              "coa_url": "", "coa_lot": "", "coa_date": ""}.items():
            await db.products.update_one(
                {"slug": p["slug"], field: {"$exists": False}},
                {"$set": {field: value}},
            )
        # Ensure at least one variant exists on legacy products
        await db.products.update_one(
            {"slug": p["slug"], "$or": [{"variants": {"$exists": False}}, {"variants": []}]},
            {"$set": {"variants": [default_variant]}},
        )

    # Default shipping zone: Canada
    if await db.shipping_zones.count_documents({}) == 0:
        canada_zone_id = str(uuid.uuid4())
        await db.shipping_zones.insert_one({
            "id": canada_zone_id,
            "name": "Canada",
            "countries": ["CA"],
            "provinces": PROVINCES_CA,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        intl_zone_id = str(uuid.uuid4())
        await db.shipping_zones.insert_one({
            "id": intl_zone_id,
            "name": "International",
            "countries": ["INTL"],
            "provinces": [],
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        await db.shipping_methods.insert_many([
            {
                "id": str(uuid.uuid4()),
                "zone_id": canada_zone_id,
                "name": "Canada Post Xpresspost",
                "cost_cad": 20.0,
                "eta_days": "2-3 business days",
                "active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
            },
            {
                "id": str(uuid.uuid4()),
                "zone_id": canada_zone_id,
                "name": "Canada Post Expedited",
                "cost_cad": 12.0,
                "eta_days": "5-7 business days",
                "active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
            },
            {
                "id": str(uuid.uuid4()),
                "zone_id": intl_zone_id,
                "name": "International Tracked",
                "cost_cad": 45.0,
                "eta_days": "10-20 business days",
                "active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
            },
        ])

    # ------------------------------------------------------------------
    # MIGRATION — variant.coa_status (single source of truth)
    # Converts legacy per-variant booleans to the new coa_status field, once,
    # only on variants that don't yet have it. Non-destructive: legacy booleans
    # are left in place for rollback. Mapping:
    #   coa_url present OR badge_coa_available  -> "available"
    #   else badge_coa_pending                  -> "pending"
    #   else                                    -> "none"
    # ------------------------------------------------------------------
    async for prod in db.products.find({"variants": {"$exists": True, "$ne": []}}, {"variants": 1}):
        vs = prod.get("variants") or []
        changed = False
        for v in vs:
            if "coa_status" in v and v["coa_status"] in ("available", "pending", "none"):
                continue
            if v.get("coa_url") or v.get("badge_coa_available"):
                v["coa_status"] = "available"
            elif v.get("badge_coa_pending"):
                v["coa_status"] = "pending"
            else:
                v["coa_status"] = "none"
            changed = True
        if changed:
            await db.products.update_one({"_id": prod["_id"]}, {"$set": {"variants": vs}})

    await _repair_nonprod_bpc_seed()


async def _decrement_coupon_usage(order: dict) -> bool:
    """Rend l'usage global + par-client d'un coupon compté. Idempotent via
    le flag `coupon_counted` sur la commande. Retourne True si un décrément
    a réellement eu lieu."""
    c = order.get("coupon")
    if not (c and c.get("code") and order.get("coupon_counted")):
        return False
    code = c["code"]
    email_norm = (order.get("email") or "").strip().lower()
    set_stage: dict = {
        "used_count": {"$max": [0, {"$subtract": [{"$ifNull": ["$used_count", 0]}, 1]}]},
        "counted_order_ids": {
            "$setDifference": [{"$ifNull": ["$counted_order_ids", []]}, [order["id"]]],
        },
    }
    if email_norm:
        set_stage["used_by"] = {
            "$map": {
                "input": {"$ifNull": ["$used_by", []]},
                "as": "entry",
                "in": {"$cond": [
                    {"$eq": ["$$entry.email", email_norm]},
                    {"$mergeObjects": [
                        "$$entry",
                        {"count": {"$max": [0, {"$subtract": [{"$ifNull": ["$$entry.count", 0]}, 1]}]}},
                    ]},
                    "$$entry",
                ]},
            }
        }
    coupon_release = await db.coupons.update_one(
        {"code": code, "counted_order_ids": order["id"]},
        [{"$set": set_stage}],
    )
    if not coupon_release.modified_count:
        return False
    await db.orders.update_one(
        {"id": order["id"]},
        {"$set": {"coupon_counted": False}},
    )
    return True


async def _cancel_order_side_effects(order: dict, *, reverse_affiliate: bool = True) -> None:
    """Effets de bord standards d'une annulation :
      1. Restock des lignes non-preorder
      2. Décrément du coupon (si compté)
      3. Reverse de la commission affiliée (si commande était payée)

    Idempotent — sûr à appeler plusieurs fois grâce aux gardes internes.
    À utiliser depuis TOUS les chemins d'annulation (auto-cancel, cancel
    manuel admin, refund complet) pour garantir cohérence."""
    await _restock_order_items(order)
    await _decrement_coupon_usage(order)
    if reverse_affiliate and order.get("payment_status") == "paid":
        try:
            await affiliate_on_order_reversed(order["id"], full=True)  # noqa: F821
        except Exception as e:
            logging.warning("[cancel] affiliate reverse failed for %s: %s", order.get("id"), e)


PAYMENT_REMINDER_HOURS = float(os.environ.get("PAYMENT_REMINDER_HOURS", "6"))
OPERATIONAL_BATCH_SIZE = max(1, min(int(os.environ.get("OPERATIONAL_BATCH_SIZE", "500")), 2000))
OPERATIONAL_MAX_BATCHES_PER_RUN = max(1, min(int(os.environ.get("OPERATIONAL_MAX_BATCHES_PER_RUN", "20")), 100))

async def send_payment_reminders(limit: int = OPERATIONAL_BATCH_SIZE):
    """Envoie un rappel unique aux commandes impayées ayant dépassé la moitié
    du délai d'expiration, avant leur annulation automatique."""
    now = datetime.now(timezone.utc)
    cutoff = (now - timedelta(hours=PAYMENT_REMINDER_HOURS)).isoformat()
    floor = (now - timedelta(hours=UNPAID_ORDER_TTL_HOURS)).isoformat()
    pending = await db.orders.find(
        {
            "payment_status": {"$in": ["awaiting_etransfer", "awaiting_crypto"]},
            "created_at": {"$lt": cutoff, "$gte": floor},
            "payment_reminder_sent": {"$ne": True},
        }, {"_id": 0}
    ).to_list(limit)
    for order in pending:
        marked = await db.orders.update_one(
            {"id": order["id"], "payment_reminder_sent": {"$ne": True}},
            {"$set": {"payment_reminder_sent": True}},
        )
        if marked.modified_count and order.get("email"):
            lang, to, ctx = _order_ctx(order)
            await send_template_email("payment_reminder", to, lang, ctx, order)
    return len(pending)

async def cancel_stale_unpaid_orders(limit: int = OPERATIONAL_BATCH_SIZE):
    cutoff = (datetime.now(timezone.utc) - timedelta(hours=UNPAID_ORDER_TTL_HOURS)).isoformat()
    stale = await db.orders.find(
        {
            "payment_status": {"$in": ["awaiting_etransfer", "awaiting_crypto"]},
            "created_at": {"$lt": cutoff},
        },
        {"_id": 0},
    ).to_list(limit)
    for order in stale:
        note = {
            "id": str(uuid.uuid4()),
            "text": f"Auto-cancelled: payment not received within {int(UNPAID_ORDER_TTL_HOURS)}h",
            "author": "system",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        res = await db.orders.update_one(
            {"id": order["id"], "payment_status": order["payment_status"]},
            {"$set": {
                "payment_status": "cancelled",
                "fulfillment_status": "cancelled",
                "cancelled_at": datetime.now(timezone.utc).isoformat(),
                "cancelled_reason": "auto_unpaid_timeout",
                "prev_payment_status": order["payment_status"],
             },
             "$push": {"notes": note}},
        )
        if res.modified_count:
            # Effets de bord centralisés (restock + coupon + affiliate reverse).
            # reverse_affiliate=False car une commande auto-annulée n'était jamais payée.
            await _cancel_order_side_effects(order, reverse_affiliate=False)
            logging.info("Auto-cancelled unpaid order %s", order.get("order_number", order["id"]))
            if order.get("email"):
                lang, to, ctx = _order_ctx(order)
                asyncio.create_task(send_template_email("order_expired", to, lang, ctx, order))
    return len(stale)


async def _drain_unpaid_order_batches() -> dict:
    totals = {"reminders": 0, "cancelled": 0, "backlog": False}
    for operation, key in (
        (send_payment_reminders, "reminders"),
        (cancel_stale_unpaid_orders, "cancelled"),
    ):
        for _ in range(OPERATIONAL_MAX_BATCHES_PER_RUN):
            processed = await operation(OPERATIONAL_BATCH_SIZE)
            totals[key] += processed
            if processed < OPERATIONAL_BATCH_SIZE:
                break
        else:
            totals["backlog"] = True
            logging.warning(
                "Unpaid order %s backlog remains after %d batches of %d",
                key,
                OPERATIONAL_MAX_BATCHES_PER_RUN,
                OPERATIONAL_BATCH_SIZE,
            )
    return totals


async def _unpaid_orders_watchdog():
    while True:
        try:
            await _drain_unpaid_order_batches()
        except Exception as e:
            logging.error("Unpaid order watchdog error: %s", e)
        await asyncio.sleep(3600)


async def _release_ready_preorder_orders() -> int:
    """Flip paid preorder orders to processing once every preorder line has
    stock >= qty and no coming-soon / COA-pending badge. Idempotent."""
    released = 0
    cursor = db.orders.find(
        {"payment_status": "paid", "fulfillment_status": "preorder"},
        {"_id": 0, "id": 1, "order_number": 1, "items": 1},
    )
    async for order in cursor:
        items = order.get("items") or []
        preorder_lines = [it for it in items if it.get("preorder")]
        if not preorder_lines:
            continue
        all_ready = True
        for it in preorder_lines:
            p = await db.products.find_one({"id": it.get("product_id")}, {"_id": 0, "variants": 1, "stock": 1})
            if not p:
                all_ready = False
                break
            variant_id = it.get("variant_id")
            if variant_id:
                v = next((v for v in p.get("variants", []) if v.get("id") == variant_id), None)
            else:
                v = None
            stock = int((v or p).get("stock", 0)) if (v or p) else 0
            if v and (v.get("badge_coming_soon") or v.get("badge_coa_pending")):
                all_ready = False
                break
            if stock < it.get("qty", 1):
                all_ready = False
                break
        if all_ready:
            reserved_lines = []
            for it in preorder_lines:
                qty = int(it.get("qty", 1))
                vid = it.get("variant_id")
                if vid in (None, "", "_default"):
                    res = await db.products.update_one(
                        {"id": it.get("product_id"), "stock": {"$gte": qty}},
                        {"$inc": {"stock": -qty}},
                    )
                else:
                    res = await db.products.update_one(
                        {"id": it.get("product_id"), "variants": {"$elemMatch": {"id": vid, "stock": {"$gte": qty}}}},
                        {"$inc": {"variants.$[v].stock": -qty}},
                        array_filters=[{"v.id": vid}],
                    )
                if res.modified_count != 1:
                    for rollback in reserved_lines:
                        rb_qty = int(rollback.get("qty", 1))
                        rb_vid = rollback.get("variant_id")
                        if rb_vid in (None, "", "_default"):
                            await db.products.update_one({"id": rollback.get("product_id")}, {"$inc": {"stock": rb_qty}})
                        else:
                            await db.products.update_one(
                                {"id": rollback.get("product_id"), "variants.id": rb_vid},
                                {"$inc": {"variants.$.stock": rb_qty}},
                            )
                    all_ready = False
                    break
                reserved_lines.append(it)
        if not all_ready:
            continue
        now = datetime.now(timezone.utc).isoformat()
        res = await db.orders.update_one(
            {"id": order["id"], "fulfillment_status": "preorder"},
            {"$set": {
                "fulfillment_status": "processing",
                "dispatch_batch": compute_dispatch_batch(now),
                "preorder_released_at": now,
            }, "$push": {"notes": {
                "id": str(uuid.uuid4()),
                "text": "Preorder auto-released: all items back in stock.",
                "author": "system",
                "created_at": now,
            }}},
        )
        if res.modified_count:
            released += 1
            logging.info("Preorder order %s released to processing", order.get("order_number", order["id"]))
    return released


async def _release_preorders_watchdog():
    while True:
        try:
            await _release_ready_preorder_orders()
        except Exception as e:
            logging.error("Preorder release watchdog error: %s", e)
        await asyncio.sleep(PREORDER_RELEASE_INTERVAL_SECONDS)


# Identité de CE processus pour le verrou des tâches de fond. Permet de
# renouveler le bail sans risquer de voler celui d'un autre worker.
_WORKER_LOCK_OWNER = str(uuid.uuid4())
_WORKER_LOCK_TTL = 120
_WORKER_LOCK_RENEW_EVERY = 45   # < TTL, pour que le bail ne se périme jamais
_WORKER_LOCK_TASK = None        # référence forte au superviseur


async def _renew_worker_lock(name: str) -> bool:
    """Prolonge le bail, uniquement si ce processus le détient encore."""
    now = datetime.now(timezone.utc)
    try:
        res = await db.locks.update_one(
            {"_id": name, "owner": _WORKER_LOCK_OWNER},
            {"$set": {"expires_at": (now + timedelta(seconds=_WORKER_LOCK_TTL)).isoformat()}},
        )
        return res.matched_count == 1
    except Exception:
        return False


async def _acquire_worker_lock(name: str, ttl_seconds: int = 120) -> bool:
    """Verrou coopératif Mongo : un seul worker exécute les tâches de fond.
    Sans ça, N workers uvicorn lancent N boucles concurrentes sur les mêmes
    commandes (auto-cancel, backfill)."""
    now = datetime.now(timezone.utc)
    try:
        await db.locks.update_one(
            {"_id": name, "expires_at": {"$lt": now.isoformat()}},
            {"$set": {"expires_at": (now + timedelta(seconds=ttl_seconds)).isoformat(),
                      "owner": _WORKER_LOCK_OWNER}},
            upsert=True,
        )
        return True
    except Exception:
        return False  # duplicate key = un autre worker détient déjà le verrou


async def _seed_categories_from_products() -> None:
    """Dérive les catégories réelles à partir des slugs déjà portés par les produits.
    Idempotent ($setOnInsert) : une catégorie éditée ensuite par l'admin n'est
    jamais réécrite au redéploiement. Aucun produit n'est touché."""
    try:
        slugs = [x for x in await db.products.distinct("category") if x and _SLUG_RE.match(str(x))]
        # Libellés par défaut alignés sur les clés i18n existantes.
        labels = {
            "healing": ("Healing & Recovery", "Guérison et récupération"),
            "weight-loss": ("Weight Management", "Gestion du poids"),
            "gh-secretagogues": ("GH Secretagogues", "Sécrétagogues de GH"),
            "cognitive": ("Cognitive", "Cognitif"),
            "longevity": ("Longevity", "Longévité"),
        }
        for i, slug in enumerate(sorted(slugs)):
            en, fr = labels.get(slug, (slug.replace("-", " ").title(), slug.replace("-", " ").title()))
            await db.categories.update_one(
                {"slug": slug},
                {"$setOnInsert": {
                    "id": str(uuid.uuid4()),
                    "slug": slug,
                    "name_en": en,
                    "name_fr": fr,
                    "published": True,
                    "display_order": i,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                }},
                upsert=True,
            )
        if slugs:
            logging.info("categories seed: %d slug(s) vérifié(s)", len(slugs))
    except Exception as e:  # pragma: no cover
        logging.error("categories seed failed: %s", e)


async def _seed_default_menus() -> None:
    """Reproduit à l'identique la navigation actuellement codée en dur, pour que
    rien ne bouge visuellement au premier déploiement. $setOnInsert uniquement."""
    defaults = [
        {
            "slug": "main-navigation", "name_en": "Main navigation", "name_fr": "Navigation principale",
            "location": "header", "display_order": 0,
            "items": [
                {"label_en": "Catalog", "label_fr": "Catalogue", "url": "/catalog", "display_order": 0},
                {"label_en": "Lab", "label_fr": "Labo", "url": "/lab", "display_order": 1},
                {"label_en": "About", "label_fr": "À propos", "url": "/about", "display_order": 2},
            ],
        },
        {
            "slug": "footer-shop", "name_en": "Shop", "name_fr": "Boutique",
            "location": "footer", "display_order": 1,
            # Plus de raccourcis par catégorie : leurs libellés (« Guérison et
            # récupération », « Gestion du poids ») décrivaient un effet
            # physiologique, ce qui rattache un composé à un usage humain.
            "items": [
                {"label_en": "Catalog", "label_fr": "Catalogue", "url": "/catalog", "display_order": 0},
                {"label_en": "Lab & COA", "label_fr": "Labo & COA", "url": "/lab", "display_order": 1},
            ],
        },
        {
            "slug": "footer-legal", "name_en": "Legal", "name_fr": "Légal",
            "location": "footer", "display_order": 2,
            "items": [
                {"label_en": "Terms", "label_fr": "Conditions", "url": "/compliance", "display_order": 0},
                {"label_en": "Privacy", "label_fr": "Confidentialité", "url": "/privacy", "display_order": 1},
                {"label_en": "Shipping", "label_fr": "Expédition", "url": "/compliance#shipping", "display_order": 2},
                {"label_en": "FAQ", "label_fr": "FAQ", "url": "/faq", "display_order": 3},
            ],
        },
    ]
    try:
        for m in defaults:
            items = []
            for it in m["items"]:
                items.append({
                    "id": str(uuid.uuid4()), "published": True, "open_new_tab": False, **it,
                })
            await db.menus.update_one(
                {"slug": m["slug"]},
                {"$setOnInsert": {
                    "id": str(uuid.uuid4()),
                    "slug": m["slug"], "name_en": m["name_en"], "name_fr": m["name_fr"],
                    "location": m["location"], "published": True,
                    "display_order": m["display_order"], "items": items,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                }},
                upsert=True,
            )
    except Exception as e:  # pragma: no cover
        logging.error("menus seed failed: %s", e)


async def _seed_launch_coupon() -> None:
    """Provisionne le coupon de lancement 15 %. $setOnInsert : si tu l'édites
    ensuite dans l'admin, le redéploiement ne l'écrase pas."""
    if not LAUNCH_COUPON_ENABLED:
        return
    try:
        await db.coupons.update_one(
            {"code": LAUNCH_COUPON_CODE},
            {"$setOnInsert": {
                "id": str(uuid.uuid4()),
                "code": LAUNCH_COUPON_CODE,
                "discount_type": "percent",
                "value": 15.0,
                "min_subtotal": 0.0,
                "usage_limit": None,
                "used_count": 0,
                "active": True,
                "expires_at": None,
                "deleted_at": None,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }},
            upsert=True,
        )
    except Exception as e:  # pragma: no cover
        logging.error("launch coupon seed failed: %s", e)


async def _backfill_affiliate_coupons() -> None:
    """Crée le coupon lié pour les affiliés déjà actifs qui n'en ont pas
    (les nouveaux le reçoivent à l'activation). Idempotent, sûr à relancer."""
    if AFFILIATE_COUPON_PERCENT <= 0:
        return
    try:
        actives = db.affiliates.find(
            {"status": "active", "code": {"$ne": None}}, {"_id": 0, "id": 1, "code": 1}
        )
        created = 0
        async for a in actives:
            c = await _affiliate_ensure_coupon(a.get("code"), a["id"])
            if c:
                created += 1
        if created:
            logging.info("[affiliate] backfill coupons: %d affiliés vérifiés", created)
    except Exception as e:
        logging.warning("[affiliate] backfill coupons échoué: %s", e)


@app.on_event("startup")
async def startup_event():
    # Warn if APP_ENV is not set — /docs and /openapi.json remain public, and
    # any other production-only guards keyed on IS_PRODUCTION stay off.
    if not APP_ENV:
        logging.warning(
            "[config] APP_ENV is not set — running in DEV mode. "
            "/docs, /redoc and /openapi.json are publicly exposed. "
            "Set APP_ENV=production before deploying."
        )
    # Warn about unconfigured optional services so issues are visible in logs.
    _svc_checks = {
        "RESEND_API_KEY (emails)": RESEND_API_KEY,
        "PUBLIC_BASE_URL (email links)": PUBLIC_BASE_URL,
        "NOWPAYMENTS_API_KEY (crypto)": NOWPAYMENTS_API_KEY,
    }
    if INTERAC_AUTOCONFIRM_MODE == "strict":
        _svc_checks["INTERAC_GRAPH_TENANT_ID (Interac auto-confirm)"] = INTERAC_GRAPH_TENANT_ID
    else:
        logging.info("[config] INTERAC_AUTOCONFIRM_MODE=off — Interac auto-confirm disabled by configuration")
    for label, val in _svc_checks.items():
        if not val:
            logging.warning("[config] %s is not set — related features disabled", label)
    await seed_admin_and_products()
    await _seed_categories_from_products()
    await _seed_default_menus()
    await _seed_launch_coupon()
    await _backfill_affiliate_coupons()
    await _drain_unpaid_order_batches()
    # Superviseur du verrou : réessaie jusqu'à l'obtenir, puis renouvelle le
    # bail tant que ce processus vit.
    #
    # Avant, l'acquisition était tentée UNE SEULE FOIS au démarrage. Le bail
    # dure 120 s et n'était jamais renouvelé : redémarrer le backend moins de
    # deux minutes après le démarrage précédent laissait le verrou détenu par
    # le processus mort, la tentative échouait, et AUCUNE tâche de fond ne
    # démarrait — ni l'envoi des courriels, ni l'auto-annulation des impayés,
    # ni la confirmation Interac, ni les versements mensuels. En silence, et
    # jusqu'au prochain redémarrage tombant par chance après l'expiration.
    # Référence gardée au niveau module : asyncio ne retient qu'une référence
    # faible, une tâche détachée sans référence peut être ramassée en vol.
    global _WORKER_LOCK_TASK
    _WORKER_LOCK_TASK = asyncio.create_task(_worker_lock_supervisor())


async def _start_background_workers() -> None:
    """Lance les tâches de fond. Appelé une seule fois, verrou en main."""
    asyncio.create_task(_unpaid_orders_watchdog())
    asyncio.create_task(_monthly_payouts_scheduler())
    asyncio.create_task(_backfill_dispatch_batch())
    asyncio.create_task(_release_preorders_watchdog())
    if INTERAC_AUTOCONFIRM_MODE == "strict":
        asyncio.create_task(_interac_deposit_watchdog())
    # Dispatch manuel : watchdog d'auto-étiquetage désactivé.
    # asyncio.create_task(_auto_label_paid_orders_watchdog())
    asyncio.create_task(_auto_sync_delivered_orders_watchdog())
    asyncio.create_task(_trash_auto_purge_watchdog())
    asyncio.create_task(_rollover_watchdog())
    asyncio.create_task(_magic_tokens_cleanup())
    asyncio.create_task(_abandoned_cart_watchdog())
    asyncio.create_task(affiliate_maintenance_watchdog())
    asyncio.create_task(_affiliate_email_worker())
    asyncio.create_task(_email_outbox_worker())
    asyncio.create_task(_email_outbox_janitor())


async def _worker_lock_supervisor() -> None:
    """Obtient le verrou des taches de fond, puis le renouvelle.

    Reessaie toutes les 30 s : si un processus mort detient encore le bail,
    la tentative suivante reussit des son expiration. Sans cette boucle, un
    redemarrage malchanceux laissait le serveur tourner sans AUCUNE tache de
    fond, en silence, jusqu'au prochain redemarrage.
    """
    started = False
    while True:
        if await _acquire_worker_lock("background_tasks", _WORKER_LOCK_TTL):
            if not started:
                logging.info("[worker-lock] verrou obtenu (%s) — demarrage des taches de fond",
                             _WORKER_LOCK_OWNER[:8])
                await _start_background_workers()
                started = True   # les boucles tournent deja : ne jamais relancer
            else:
                logging.info("[worker-lock] verrou repris (%s)", _WORKER_LOCK_OWNER[:8])
            # Renouvellement tant que ce processus vit. Si le bail est perdu
            # (horloge, panne Mongo), on repasse en phase d'acquisition.
            while await _renew_worker_lock("background_tasks"):
                await asyncio.sleep(_WORKER_LOCK_RENEW_EVERY)
            logging.warning("[worker-lock] bail perdu — nouvelle tentative d'acquisition")
        await asyncio.sleep(30)


async def _backfill_dispatch_batch():
    """Renseigne dispatch_batch pour les commandes payées qui n'en ont pas (non destructif)."""
    try:
        cursor = db.orders.find(
            {"payment_status": "paid", "paid_at": {"$ne": None},
             "dispatch_batch": {"$in": [None, ""]}},
            {"_id": 0, "id": 1, "paid_at": 1},
        )
        n = 0
        async for o in cursor:
            batch = compute_dispatch_batch(o["paid_at"])
            await db.orders.update_one({"id": o["id"]}, {"$set": {"dispatch_batch": batch}})
            n += 1
        if n:
            logging.info("dispatch_batch backfill: %d commande(s) mise(s) à jour", n)
    except Exception as e:  # pragma: no cover
        logging.error("dispatch_batch backfill failed: %s", e)


@app.on_event("shutdown")
async def shutdown_event():
    client.close()


# ---------------------------------------------------------------------------
# App wiring
# ---------------------------------------------------------------------------

# ===== FIRONOVA_AFFILIATE_BLOCK_START =====
# ===========================================================================
# CONSTANTES DU PROGRAMME
# ===========================================================================

AFFILIATE_INVITE_TTL_HOURS = 168          # 7 jours
# Version des conditions du programme d'affiliation, au format ISO d'une date.
# Datée plutôt que numérotée : on doit pouvoir dire QUEL texte a été accepté, et
# une date se recoupe avec l'archive du document. Changer cette valeur redemande
# l'acceptation à tout le monde — c'est le seul mécanisme de redemande, donc ne
# la modifier que lorsque le texte change réellement.
AFFILIATE_TERMS_VERSION = os.environ.get("AFFILIATE_TERMS_VERSION", "2026-08-01")

# Fenêtre d'attribution du clic. Portée de 30 à 365 jours : le rattachement
# durable ne couvre que les clients qui REVIENNENT, jamais leur première
# commande. Un visiteur qui cliquait un lien, hésitait, puis achetait au 35e
# jour sans saisir le code ne rapportait donc rien à l'affilié — alors que
# c'est lui qui l'avait amené. Trois à six mois de réflexion sont ordinaires
# sur ce type de produit.
#
# 365 et non davantage : les navigateurs plafonnent la durée de vie d'un témoin
# autour de 400 jours et rogneraient silencieusement toute valeur supérieure.
AFFILIATE_COOKIE_DAYS = int(os.environ.get("AFFILIATE_COOKIE_DAYS", "365"))
# Rétention des enregistrements de clics (purge auto). Volontairement DÉCOUPLÉ
# de AFFILIATE_COOKIE_DAYS malgré l'écart apparent : ces documents servent aux
# statistiques, pas à l'attribution — celle-ci lit la valeur du témoin, jamais
# cette collection. Les allonger à un an ne changerait donc aucune commission,
# et reviendrait à conserver des empreintes d'adresses IP douze mois durant,
# à rebours de ce que promet la politique de confidentialité.
AFFILIATE_CLICK_TTL_DAYS = 45
AFFILIATE_COOKIE_NAME = "fn_ref"
AFFILIATE_INVITE_MAX = 5                  # rate-limit renvois / fenêtre
AFFILIATE_INVITE_WINDOW = 3600            # 1 h

# Paliers cumulés (seuil bas inclus, seuil haut exclu sauf Diamond).
# (rate, floor, ceil)  — ceil None = illimité
AFFILIATE_TIERS = [
    ("standard", 0.10, 0.0, 2000.0),
    ("bronze", 0.12, 2001.0, 5000.0),
    ("silver", 0.14, 5001.0, 10000.0),
    ("gold", 0.16, 10001.0, 20000.0),
    ("platinum", 0.18, 20001.0, 35000.0),
    ("diamond", 0.20, 35001.0, None),
]

AFFILIATE_TIER_LABELS = {
    "standard": {"fr": "Standard", "en": "Standard"},
    "bronze": {"fr": "Bronze", "en": "Bronze"},
    "silver": {"fr": "Argent", "en": "Silver"},
    "gold": {"fr": "Or", "en": "Gold"},
    "platinum": {"fr": "Platine", "en": "Platinum"},
    "diamond": {"fr": "Diamant", "en": "Diamond"},
}


# Implementation lives in services/affiliate.py; re-exported so existing call
# sites (routers/, other server helpers) keep resolving these names here.
try:
    from services.affiliate import (  # noqa: F401
        _affiliate_tier_for_revenue, _affiliate_tier_index, _affiliate_tier_bounds,
        _affiliate_hash_token, _affiliate_hash_ip, _affiliate_referrer_domain,
        _affiliate_normalize_custom_code, _affiliate_gen_code_v2, _fetch_cad_to_usd_rate,
        _normalize_payout, _detect_payout_network, _affiliate_quarter_start, _affiliate_compute_metrics,
        _affiliate_compute_list_metrics, _affiliate_public, _affiliate_send_invite,
        _affiliate_ensure_coupon, affiliate_capture_click, affiliate_attach_to_order,
        affiliate_on_order_paid, affiliate_on_order_reversed, affiliate_maintenance_watchdog,
        _process_affiliate_email_job, _affiliate_email_worker, affiliate_ensure_indexes,
        _defer_affiliate_payout_below_threshold, _monthly_payouts_scheduler,
        _generate_payouts_for_period,
    )
except ImportError:  # package-relative import (uvicorn backend.server:app)
    from backend.services.affiliate import (  # noqa: F401
        _affiliate_tier_for_revenue, _affiliate_tier_index, _affiliate_tier_bounds,
        _affiliate_hash_token, _affiliate_hash_ip, _affiliate_referrer_domain,
        _affiliate_normalize_custom_code, _affiliate_gen_code_v2, _fetch_cad_to_usd_rate,
        _normalize_payout, _detect_payout_network, _affiliate_quarter_start, _affiliate_compute_metrics,
        _affiliate_compute_list_metrics, _affiliate_public, _affiliate_send_invite,
        _affiliate_ensure_coupon, affiliate_capture_click, affiliate_attach_to_order,
        affiliate_on_order_paid, affiliate_on_order_reversed, affiliate_maintenance_watchdog,
        _process_affiliate_email_job, _affiliate_email_worker, affiliate_ensure_indexes,
        _defer_affiliate_payout_below_threshold, _monthly_payouts_scheduler,
        _generate_payouts_for_period,
    )


# ===========================================================================
# PAYOUT — USDT/USDC (ERC-20) + validation Ethereum EIP-55
# ===========================================================================
AFFILIATE_PAYOUT_CURRENCIES = ("usdt", "usdc")


# ===========================================================================
# MODÈLES Pydantic
# ===========================================================================

class AffiliateInviteIn(BaseModel):
    """Fiche d'invitation admin — nouveau schéma :
    Prénom + Nom obligatoires, Entreprise optionnelle. Le code affilié sera
    généré automatiquement à l'activation : `BASE + 10` par défaut
    (BASE = entreprise si fournie sinon prénom). L'admin peut modifier
    le rabais ensuite ; le code sera renommé et l'ancien archivé en alias."""
    email: EmailStr
    first_name: Optional[str] = Field(default=None, min_length=1, max_length=80)
    last_name: Optional[str] = Field(default=None, min_length=1, max_length=80)
    company: Optional[str] = Field(default="", max_length=120)
    commission_note: Optional[str] = ""        # note interne admin
    payout_currency: Optional[str] = "usdt"    # USDT par défaut (ERC-20)
    code: Optional[str] = None
    coupon_percent: Optional[float] = Field(default=None, ge=0, le=100)
    lang: str = "fr"

    # Rétrocompat : accepte encore `name` sur le wire — on le splittera si
    # les nouveaux champs sont absents (pour ne pas casser l'ancien front).
    name: Optional[str] = None


class AffiliateJoinIn(BaseModel):
    """Payload magic-link : SEUL le token est requis. Aucune auth préalable.
    L'email et le nom viennent de l'invitation. L'adresse de payout est
    optionnelle (on la demandera plus tard depuis le dashboard)."""
    token: str = Field(min_length=10)
    payout_address: Optional[str] = ""
    payout_currency: Optional[str] = None


class AffiliateTermsAcceptIn(BaseModel):
    """Acceptation des conditions du programme.

    Les trois cases sont `Literal[True]` : une valeur absente ou fausse fait
    échouer la validation. On ne veut surtout pas d'un booléen optionnel qui
    enregistrerait une acceptation partielle sans que personne s'en aperçoive.
    """
    accept_terms: Literal[True]
    confirm_age: Literal[True]
    accept_research_use: Literal[True]
    # La version acceptée est imposée par le SERVEUR, jamais reçue du client :
    # sinon n'importe qui pourrait déclarer avoir accepté une version obsolète
    # pour échapper à une redemande.


class AffiliateTicketIn(BaseModel):
    """Ouverture d'un billet d'assistance par un affilié."""
    subject: str = Field(min_length=3, max_length=140)
    body: str = Field(min_length=10, max_length=4000)
    # Page d'où part la demande. Recueillie automatiquement : une personne qui
    # écrit « ça ne marche pas » depuis l'onglet des versements pose une
    # question différente de la même phrase écrite depuis les paramètres.
    context_path: Optional[str] = Field(default="", max_length=200)


class AffiliateTicketReplyIn(BaseModel):
    body: str = Field(min_length=1, max_length=4000)


class AffiliateTicketStatusIn(BaseModel):
    status: Literal["open", "pending", "resolved"]


class AffiliatePayoutSettingsIn(BaseModel):
    payout_address: str = Field(min_length=4, max_length=200)
    payout_currency: str = Field(min_length=2, max_length=12)


class AffiliateAdminUpdateIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    # Identité
    first_name: Optional[str] = Field(default=None, min_length=1, max_length=80)
    last_name: Optional[str] = Field(default=None, min_length=1, max_length=80)
    company: Optional[str] = Field(default=None, max_length=120)
    name: Optional[str] = Field(default=None, min_length=1, max_length=160)   # legacy
    # Statut
    status: Optional[Literal["invited", "active", "suspended"]] = None
    compliance_status: Optional[Literal["compliant", "review", "suspended"]] = None
    manual_tier: Optional[str] = None
    clear_manual_tier: bool = False
    # Entente négociée. DISTINCT de manual_tier, et volontairement : figer un
    # palier à la main est un geste administratif, révocable et parfois
    # accidentel ; une entente est un engagement pris envers la personne. Seul
    # ce drapeau autorise l'écran affilié à écrire « accordé par entente » et
    # « ne peut pas redescendre ». Sans lui, un palier forcé par erreur
    # promettait quelque chose qui n'avait jamais été convenu.
    tier_agreement: Optional[bool] = None
    commission_note: Optional[str] = None
    # Payout
    payout_address: Optional[str] = Field(default=None, max_length=200)
    payout_currency: Optional[str] = Field(default=None, max_length=12)
    # Rabais client — DÉCLENCHE un rename de code si modifié
    code: Optional[str] = None
    coupon_percent: Optional[float] = Field(default=None, ge=0, le=100)
    # Notes internes
    admin_notes: Optional[str] = Field(default=None, max_length=2000)
    suspension_reason: Optional[str] = Field(default=None, max_length=500)


class AffiliatePayoutMarkIn(BaseModel):
    reference: str = Field(min_length=1, max_length=200)   # tx hash / réf
    note: Optional[str] = ""


# ===========================================================================
# TÂCHE : approbation automatique après fenêtre de rétractation
# ===========================================================================

AFFILIATE_APPROVAL_HOLD_DAYS = 14  # délai avant qu'une commission 'pending'


# ===========================================================================
# DÉPENDANCE : affilié courant (compte user avec is_affiliate + affiliate lié)
# ===========================================================================

async def get_current_affiliate(request: Request) -> dict:
    user = await get_current_user(request)  # noqa: F821
    aff = await db.affiliates.find_one(
        {"user_id": user["id"], "status": {"$in": ["active", "suspended"]}},
        {"_id": 0},
    )
    if not aff:
        raise HTTPException(403, "Not an affiliate")
    if aff.get("status") == "suspended":
        raise HTTPException(403, "Affiliate account suspended")
    return aff


# ===========================================================================
# ENDPOINTS — AFFILIÉ (dashboard)
# ===========================================================================

async def affiliate_join(payload: AffiliateJoinIn, request: Request,
                          response: Response):
    """Active un affilié via son token d'invitation — MAGIC-LINK 1-CLIC.

    Ne requiert AUCUNE auth préalable. Le clic vaut connexion :
      - crée un compte passwordless si aucun n'existe pour l'email invité
        (email vérifié : l'admin a invité) ;
      - bascule l'affilié `invited` → `active`, génère son code à partir du
        prénom (ou entreprise si fournie) + suffixe 10 (rabais par défaut) ;
      - pose la session (cookie + JWT retourné dans le body).
    Token d'invitation consommé atomiquement."""
    token_hash = _affiliate_hash_token(payload.token.strip())
    now = datetime.now(timezone.utc)
    invite = await db.affiliates.find_one(
        {"invite_token_hash": token_hash, "status": "invited"}, {"_id": 0}
    )
    if not invite:
        raise HTTPException(400, "Invalid or already-used invitation")

    exp = invite.get("invite_expires_at")
    if exp:
        try:
            exp_dt = datetime.fromisoformat(exp.replace("Z", "+00:00"))
            if exp_dt.tzinfo is None:
                exp_dt = exp_dt.replace(tzinfo=timezone.utc)
            if now > exp_dt:
                raise HTTPException(400, "Invitation expired")
        except HTTPException:
            raise
        except Exception:
            pass

    invite_email = (invite.get("email") or "").lower().strip()
    if not invite_email:
        raise HTTPException(400, "Invitation is missing an email")

    # --- Compte user cible : créé passwordless si absent (magic-link) -------
    from pymongo.errors import DuplicateKeyError as _DupKey
    user = await db.users.find_one({"email": invite_email})
    if not user:
        display_name = (invite.get("name")
                        or f"{invite.get('first_name','')} {invite.get('last_name','')}".strip()
                        or invite_email.split("@")[0]).strip()
        user_doc = {
            "id": str(uuid.uuid4()),
            "email": invite_email,
            "name": display_name,
            "password_hash": hash_password(secrets.token_urlsafe(32)),
            "role": "user",
            "token_version": 0,
            "created_at": now.isoformat(),
            "passwordless": True,
            "email_verified": True,     # invitation admin = email vérifié
        }
        try:
            await db.users.insert_one(user_doc)
            user = user_doc
            try:
                await db.subscribers.update_one(
                    {"email": invite_email}, {"$set": {"converted": True}}
                )
            except Exception as _e:  # pragma: no cover
                logging.warning("subscriber conversion flag failed for %s: %s",
                                invite_email, _e)
            try:
                asyncio.create_task(welcome_new_user(invite_email, display_name, "fr"))
            except Exception:
                pass
        except _DupKey:
            user = await db.users.find_one({"email": invite_email})
    else:
        await db.users.update_one(
            {"email": invite_email}, {"$set": {"email_verified": True}}
        )
        user = {**user, "email_verified": True}

    # --- Un même compte ne peut pas être déjà lié à un autre affilié actif --
    other = await db.affiliates.find_one(
        {"user_id": user["id"], "status": {"$in": ["active", "suspended"]}},
        {"_id": 1, "id": 1},
    )
    if other and other.get("id") != invite["id"]:
        raise HTTPException(409, "This account is already linked to an affiliate")

    # --- Payout optionnel à l'activation (0 friction) -----------------------
    _payout_addr = (payload.payout_address or "").strip()
    _payout_network = ""
    if _payout_addr:
        _payout_addr, _payout_cur, _payout_network = _normalize_payout(
            _payout_addr,
            payload.payout_currency or invite.get("payout_currency") or "usdt",
        )
    else:
        _payout_cur = ((payload.payout_currency or invite.get("payout_currency") or "usdt")
                       .strip().lower())
        if _payout_cur not in AFFILIATE_PAYOUT_CURRENCIES:
            _payout_cur = "usdt"

    # --- Génération du code v2 : BASE + rabais ------------------------------
    # Priorité : coupon_percent pré-défini sur la fiche (ex. via bulk CSV),
    # sinon AFFILIATE_DEFAULT_DISCOUNT_PERCENT (10% par défaut).
    base_source = (invite.get("company") or "").strip() or (invite.get("first_name") or "").strip()
    if not base_source:
        # Fallback : dérive du champ name legacy
        base_source = (invite.get("name") or invite_email.split("@")[0]).split()[0]
    if invite.get("coupon_percent") is not None:
        try:
            discount_percent_effective = float(invite["coupon_percent"])
        except Exception:
            discount_percent_effective = float(os.environ.get("AFFILIATE_DEFAULT_DISCOUNT_PERCENT", "10"))
    else:
        discount_percent_effective = float(os.environ.get("AFFILIATE_DEFAULT_DISCOUNT_PERCENT", "10"))
    code = invite.get("code") or await _affiliate_gen_code_v2(
        base_source, discount_percent_effective,
        email=invite_email, exclude_id=invite["id"],
    )

    update = {
        "$set": {
            "status": "active",
            "user_id": user["id"],
            "code": code,
            "coupon_percent": discount_percent_effective,
            "activated_at": now.isoformat(),
            "ip_hash": _affiliate_hash_ip(_client_ip(request)),
            "known_addresses": [],
            "payout_address": _payout_addr,
            "payout_currency": _payout_cur,
            "invite_token_hash": None,
            "invite_expires_at": None,
        }
    }
    res = await db.affiliates.update_one(
        {"id": invite["id"], "status": "invited"}, update
    )
    if not res.modified_count:
        raise HTTPException(409, "Invitation already consumed")

    await _start_session(response, request, user)

    aff = await db.affiliates.find_one({"id": invite["id"]}, {"_id": 0})
    # Coupon lié auto (idempotent) — utilise le rabais courant de l'affilié.
    try:
        await _affiliate_ensure_coupon(aff.get("code"), aff["id"],
                                       percent=discount_percent_effective)
    except Exception as _e:
        logging.warning("[affiliate] échec création coupon pour %s: %s",
                        aff.get("code"), _e)
    metrics = await _affiliate_compute_metrics(aff["id"])
    return _affiliate_public(aff, metrics)


async def affiliate_terms_accept(payload: AffiliateTermsAcceptIn, request: Request):
    """Enregistre l'acceptation des conditions du programme.

    L'adresse IP est conservée EN CLAIR, contrairement au reste du module
    affiliation qui n'en garde qu'une empreinte salée. Les deux finalités sont
    distinctes : le hachage sert à détecter l'auto-parrainage, où seule
    l'égalité entre deux adresses importe ; ici il s'agit de prouver un
    engagement contractuel, et une empreinte ne prouve rien — elle établit
    qu'une acceptation a eu lieu, pas d'où elle venait. C'est le même choix que
    pour le consentement à l'infolettre, qui garde déjà `consent_ip`.
    """
    aff = await get_current_affiliate(request)
    now = datetime.now(timezone.utc).isoformat()
    await db.affiliates.update_one(
        {"id": aff["id"]},
        {"$set": {
            "terms_accepted_at": now,
            "terms_version": AFFILIATE_TERMS_VERSION,
            "age_confirmed": True,
            "terms_accepted_ip": _client_ip(request),
        }},
    )
    logging.info("[affiliate] conditions acceptées code=%s version=%s",
                 aff.get("code"), AFFILIATE_TERMS_VERSION)
    return {"ok": True, "terms_version": AFFILIATE_TERMS_VERSION}


def _ticket_message(auteur: str, body: str) -> dict:
    return {
        "id": str(uuid.uuid4()),
        "from": auteur,                       # "affiliate" | "admin"
        "body": body.strip()[:4000],
        "at": datetime.now(timezone.utc).isoformat(),
    }


async def affiliate_ticket_create(payload: AffiliateTicketIn, request: Request):
    """Ouvre un billet. Le fil complet vit dans le document lui-même : une
    conversation d'assistance se lit d'un bloc, et la séparer en deux
    collections obligerait à deux requêtes pour afficher trois messages."""
    aff = await get_current_affiliate(request)
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "affiliate_id": aff["id"],
        "affiliate_code": aff.get("code", ""),
        "affiliate_email": aff.get("email", ""),
        "affiliate_name": " ".join(
            x for x in [aff.get("first_name"), aff.get("last_name")] if x
        ) or aff.get("name", ""),
        "subject": payload.subject.strip(),
        "status": "open",
        "context_path": (payload.context_path or "").strip(),
        # Contexte figé à l'ouverture : le palier et le solde d'aujourd'hui
        # expliquent la question d'aujourd'hui. Les relire au moment de
        # répondre donnerait un état qui a peut-être changé entre-temps.
        "snapshot": {
            "tier": aff.get("manual_tier") or "",
            "payout_address": aff.get("payout_address", ""),
            "payout_currency": aff.get("payout_currency", ""),
        },
        "messages": [_ticket_message("affiliate", payload.body)],
        "created_at": now,
        "updated_at": now,
        "last_from": "affiliate",
    }
    await db.affiliate_tickets.insert_one(doc)
    logging.info("[ticket] ouvert code=%s sujet=%r", aff.get("code"), doc["subject"][:60])
    doc.pop("_id", None)
    return doc


async def affiliate_tickets_list(request: Request):
    aff = await get_current_affiliate(request)
    rows = await db.affiliate_tickets.find(
        {"affiliate_id": aff["id"]}, {"_id": 0}
    ).sort("updated_at", -1).to_list(100)
    return rows


async def affiliate_ticket_reply(ticket_id: str, payload: AffiliateTicketReplyIn,
                                  request: Request):
    aff = await get_current_affiliate(request)
    # Le filtre porte AUSSI sur affiliate_id : sans lui, connaître un
    # identifiant suffirait à écrire dans le billet de quelqu'un d'autre.
    res = await db.affiliate_tickets.find_one_and_update(
        {"id": ticket_id, "affiliate_id": aff["id"]},
        {"$push": {"messages": _ticket_message("affiliate", payload.body)},
         "$set": {"updated_at": datetime.now(timezone.utc).isoformat(),
                  "last_from": "affiliate",
                  # Répondre à un billet résolu le rouvre : la personne n'a
                  # pas eu satisfaction, et la laisser écrire dans le vide
                  # serait pire que de ne pas lui offrir de réponse.
                  "status": "open"}},
        projection={"_id": 0},
        return_document=ReturnDocument.AFTER,
    )
    if not res:
        raise HTTPException(404, "Billet introuvable")
    return res


async def admin_affiliate_tickets(admin: dict = Depends(get_admin_user),  # noqa: F821
                                   status: Optional[str] = None):
    filt = {"status": status} if status else {}
    rows = await db.affiliate_tickets.find(filt, {"_id": 0}).sort("updated_at", -1).to_list(300)
    return rows


async def admin_affiliate_ticket_reply(ticket_id: str, payload: AffiliateTicketReplyIn,
                                        admin: dict = Depends(get_admin_user)):  # noqa: F821
    res = await db.affiliate_tickets.find_one_and_update(
        {"id": ticket_id},
        {"$push": {"messages": _ticket_message("admin", payload.body)},
         "$set": {"updated_at": datetime.now(timezone.utc).isoformat(),
                  "last_from": "admin",
                  "status": "pending"}},
        projection={"_id": 0},
        return_document=ReturnDocument.AFTER,
    )
    if not res:
        raise HTTPException(404, "Billet introuvable")
    return res


async def admin_affiliate_ticket_status(ticket_id: str, payload: AffiliateTicketStatusIn,
                                         admin: dict = Depends(get_admin_user)):  # noqa: F821
    res = await db.affiliate_tickets.find_one_and_update(
        {"id": ticket_id},
        {"$set": {"status": payload.status,
                  "updated_at": datetime.now(timezone.utc).isoformat()}},
        projection={"_id": 0},
        return_document=ReturnDocument.AFTER,
    )
    if not res:
        raise HTTPException(404, "Billet introuvable")
    return res


async def affiliate_me(request: Request, lang: str = "fr"):
    aff = await get_current_affiliate(request)
    metrics = await _affiliate_compute_metrics(aff["id"])
    out = _affiliate_public(aff, metrics, lang=lang)
    # Taux de change CAD→USD transparent (Banque du Canada) — utilisé dans
    # l'aperçu Payments pour montrer combien 1 CAD = X USDT/USDC.
    try:
        fx_rate, fx_source = await _fetch_cad_to_usd_rate()
        out["fx_rate_cad_to_usd"] = fx_rate
        out["fx_source"] = fx_source
        out["fx_captured_at"] = datetime.now(timezone.utc).isoformat()
    except Exception:
        # ne bloque pas le dashboard si l'API Banque du Canada est down
        pass
    return out


async def affiliate_referrals(request: Request, limit: int = 200):
    aff = await get_current_affiliate(request)
    rows = await db.affiliate_referrals.find(
        {"affiliate_id": aff["id"], "status": {"$ne": "excluded"}},
        {"_id": 0, "order_email": 0, "affiliate_ip_hash": 0},
    ).sort("created_at", -1).to_list(min(limit, 500))
    return rows


async def affiliate_payouts(request: Request):
    aff = await get_current_affiliate(request)
    rows = await db.affiliate_payouts.find(
        {"affiliate_id": aff["id"]}, {"_id": 0}
    ).sort("created_at", -1).to_list(200)
    return rows


async def affiliate_payout_settings(payload: AffiliatePayoutSettingsIn, request: Request):
    aff = await get_current_affiliate(request)
    address, currency, network = _normalize_payout(payload.payout_address, payload.payout_currency)
    await db.affiliates.update_one(
        {"id": aff["id"]},
        {"$set": {
            "payout_address": address,
            "payout_currency": currency,
            "payout_network": network,
        }},
    )
    fresh = await db.affiliates.find_one({"id": aff["id"]}, {"_id": 0})
    return _affiliate_public(fresh)


async def affiliate_performance(request: Request):
    """Séries mensuelles (12 derniers mois) de CA validé pour les graphiques."""
    aff = await get_current_affiliate(request)
    buckets: dict = {}
    cursor = db.affiliate_referrals.find(
        {"affiliate_id": aff["id"], "status": {"$in": ["approved", "paid"]}},
        {"_id": 0, "base_amount": 1, "approved_at": 1, "created_at": 1},
    )
    async for r in cursor:
        ts = r.get("approved_at") or r.get("created_at")
        if not ts:
            continue
        try:
            dt = datetime.fromisoformat(str(ts).replace("Z", "+00:00"))
        except Exception:
            continue
        key = f"{dt.year}-{dt.month:02d}"
        buckets[key] = round(buckets.get(key, 0.0) + float(r.get("base_amount", 0.0)), 2)
    series = [{"month": k, "revenue": v} for k, v in sorted(buckets.items())]
    return {"series": series}


# ===========================================================================
# ENDPOINTS — ATTRIBUTION PUBLIQUE (pose du cookie)
# ===========================================================================

async def affiliate_insights(request: Request):
    """Indicateurs synthétiques pour le tableau de bord affilié :
    mois courant, meilleur mois, clics, taux de conversion, commandes
    validées, panier moyen. Données réelles agrégées depuis les referrals
    et les clics."""
    aff = await get_current_affiliate(request)
    aff_id = aff["id"]

    # Referrals non exclus
    referrals = await _cursor_all(db.affiliate_referrals.find(
        {"affiliate_id": aff_id, "status": {"$ne": "excluded"}},
        {"_id": 0, "order_total": 1, "commission_amount": 1, "status": 1, "created_at": 1},
    ))

    # Un referral "validé" = commande payée (approved/paid/approuvé)
    VALIDATED = {"approved", "paid"}
    by_month = {}
    validated_orders = 0
    validated_revenue = 0.0
    for r in referrals:
        st = str(r.get("status", "")).lower()
        comm = float(r.get("commission_amount", 0.0))
        rev = float(r.get("order_total", 0.0))
        ts = str(r.get("created_at", ""))[:7]  # YYYY-MM
        if st in VALIDATED:
            validated_orders += 1
            validated_revenue += rev
            m = by_month.setdefault(ts, {"commission": 0.0, "revenue": 0.0})
            m["commission"] += comm
            m["revenue"] += rev

    now_month = datetime.now(timezone.utc).isoformat()[:7]
    current_month = by_month.get(now_month, {"commission": 0.0, "revenue": 0.0})

    best_month = None
    if by_month:
        bm_key = max(by_month, key=lambda k: by_month[k]["commission"])
        if by_month[bm_key]["commission"] > 0:
            best_month = {"month": bm_key, "commission": round(by_month[bm_key]["commission"], 2)}

    # Clics (collection affiliate_clicks)
    clicks = await db.affiliate_clicks.count_documents({"affiliate_id": aff_id})
    conversion_rate = (validated_orders / clicks) if clicks > 0 else None
    avg_order_value = (validated_revenue / validated_orders) if validated_orders > 0 else None

    return {
        "current_month": {
            "commission": round(current_month["commission"], 2),
            "revenue": round(current_month["revenue"], 2),
        },
        "best_month": best_month,
        "clicks": clicks,
        "conversion_rate": conversion_rate,
        "validated_orders": validated_orders,
        "avg_order_value": round(avg_order_value, 2) if avg_order_value is not None else None,
    }


async def affiliate_top_products(request: Request, limit: int = 5):
    """Meilleurs produits de l'affilié : agrège les articles vendus dans les
    commandes payées attribuées à cet affilié, triés par revenu décroissant.
    Retourne un lien de partage `?ref=<code>` pour chaque produit."""
    aff = await get_current_affiliate(request)  # noqa: F821
    aff_id = aff["id"]
    limit = max(1, min(int(limit or 5), 20))

    # Commandes payées attribuées à cet affilié (attribution stockée sur la commande).
    cursor = db.orders.find(
        {"affiliate_id": aff_id, "payment_status": "paid"},
        {"_id": 0, "items": 1},
    ).limit(2000)

    stats: dict = {}
    total_orders = 0
    async for order in cursor:
        total_orders += 1
        for it in (order.get("items") or []):
            pid = it.get("product_id") or it.get("slug")
            if not pid:
                continue
            row = stats.setdefault(pid, {
                "product_id": it.get("product_id"),
                "slug": it.get("slug") or "",
                "name_fr": it.get("name_fr") or it.get("name_en") or "",
                "name_en": it.get("name_en") or it.get("name_fr") or "",
                "image_url": it.get("image_url") or "",
                "qty": 0,
                "revenue": 0.0,
                "orders": 0,
            })
            row["qty"] += int(it.get("qty", 0) or 0)
            row["revenue"] += float(it.get("line_total", 0.0) or 0.0)
            row["orders"] += 1

    ranked = sorted(stats.values(), key=lambda r: r["revenue"], reverse=True)[:limit]
    for r in ranked:
        r["revenue"] = round(r["revenue"], 2)

    code = aff.get("code") or ""
    base = _trusted_public_base_url()  # noqa: F821
    for r in ranked:
        if r.get("slug") and code:
            r["share_url"] = f"{base}/product/{r['slug']}?ref={code}"
        else:
            r["share_url"] = ""

    return {
        "items": ranked,
        "orders_analyzed": total_orders,
        "code": code,
    }


async def affiliate_clicks(request: Request, days: int = 30):
    aff = await get_current_affiliate(request)
    days = max(7, min(int(days), 90))
    start_date = datetime.now(timezone.utc).date() - timedelta(days=days - 1)
    start = start_date.isoformat() + "T00:00:00"
    counts = {}
    cursor = db.affiliate_clicks.find(
        {"affiliate_id": aff["id"], "created_at": {"$gte": start}},
        {"_id": 0, "created_at": 1},
    )
    async for click in cursor:
        day = str(click.get("created_at") or "")[:10]
        counts[day] = counts.get(day, 0) + 1

    series = []
    for offset in range(days):
        day = (start_date + timedelta(days=offset)).isoformat()
        series.append({"date": day, "clicks": counts.get(day, 0)})
    total_clicks = sum(counts.values())
    validated_orders = await db.affiliate_referrals.count_documents({
        "affiliate_id": aff["id"],
        "status": {"$in": ["approved", "paid"]},
        "created_at": {"$gte": start},
    })
    return {
        "days": days,
        "series": series,
        "summary": {
            "total_clicks": total_clicks,
            "validated_orders": validated_orders,
            "conversion_rate": (validated_orders / total_clicks) if total_clicks else None,
        },
    }


async def affiliate_clicks_sources(request: Request, days: int = 30):
    """Top sources des clics de l'affilié : pages d'atterrissage, domaines
    référents et types d'appareil (30 derniers jours par défaut)."""
    aff = await get_current_affiliate(request)  # noqa: F821
    days = max(7, min(int(days), 90))
    start = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    pages, refs, devices = {}, {}, {}
    total = 0
    cursor = db.affiliate_clicks.find(
        {"affiliate_id": aff["id"], "created_at": {"$gte": start}},
        {"_id": 0, "page": 1, "referrer": 1, "device": 1},
    )
    async for c in cursor:
        total += 1
        page = str(c.get("page") or "").strip() or "direct"
        pages[page] = pages.get(page, 0) + 1
        ref = _affiliate_referrer_domain(str(c.get("referrer") or "")) or "direct"
        refs[ref] = refs.get(ref, 0) + 1
        dev = str(c.get("device") or "").strip() or "unknown"
        devices[dev] = devices.get(dev, 0) + 1

    def _top(d, n=8):
        return [{"source": k, "clicks": v}
                for k, v in sorted(d.items(), key=lambda kv: kv[1], reverse=True)[:n]]

    return {
        "days": days,
        "total_clicks": total,
        "top_pages": _top(pages),
        "top_referrers": _top(refs),
        "devices": devices,
    }


async def affiliate_activity(request: Request, limit: int = 20):
    """Flux d'activité récent de l'affilié : clics, commandes et paiements
    fusionnés, triés par date décroissante (type/at/label/status/amount)."""
    aff = await get_current_affiliate(request)  # noqa: F821
    limit = max(5, min(int(limit), 50))
    events = []

    clicks = await db.affiliate_clicks.find(
        {"affiliate_id": aff["id"]},
        {"_id": 0, "created_at": 1, "page": 1, "device": 1},
    ).sort("created_at", -1).to_list(limit)
    for c in clicks:
        events.append({
            "type": "click", "at": c.get("created_at"),
            "label": str(c.get("page") or ""), "device": str(c.get("device") or ""),
        })

    referrals = await db.affiliate_referrals.find(
        {"affiliate_id": aff["id"], "status": {"$ne": "excluded"}},
        {"_id": 0, "order_number": 1, "status": 1, "commission_amount": 1,
         "base_amount": 1, "created_at": 1},
    ).sort("created_at", -1).to_list(limit)
    for r in referrals:
        events.append({
            "type": "referral", "at": r.get("created_at"),
            "label": str(r.get("order_number") or ""), "status": r.get("status"),
            "amount": round(float(r.get("commission_amount") or 0), 2),
            "base": round(float(r.get("base_amount") or 0), 2),
        })

    payouts = await db.affiliate_payouts.find(
        {"affiliate_id": aff["id"]},
        {"_id": 0, "period": 1, "status": 1, "amount": 1, "created_at": 1},
    ).sort("created_at", -1).to_list(limit)
    for p in payouts:
        events.append({
            "type": "payout", "at": p.get("created_at"),
            "label": str(p.get("period") or ""), "status": p.get("status"),
            "amount": round(float(p.get("amount") or 0), 2),
        })

    events.sort(key=lambda e: str(e.get("at") or ""), reverse=True)
    return events[:limit]


async def affiliate_ref(code: str, request: Request, response: Response,
                        page: str = "", referrer: str = "", device: str = ""):
    """Endpoint de tracking : pose le cookie et renvoie ok. Le frontend appelle
    ceci au 1er chargement quand ?ref=CODE est présent dans l'URL.
    page/referrer/device (optionnels) alimentent l'analyse des sources."""
    await _rate_limit("affiliate_ref", _client_ip(request), 60, 60,  # noqa: F821
                "Trop de requêtes.")
    normalized = (code or "").strip().upper()
    if not normalized:
        raise HTTPException(404, "Referral code missing")
    # 404 explicite pour code inconnu — évite la pollution analytique et permet
    # au frontend de fallback proprement (masquer le badge / prévenir l'affilié).
    matched = await db.affiliates.find_one(
        {
            "status": "active",
            "$or": [
                {"code": normalized},
                {"aliases": {"$elemMatch": {"code": normalized, "active": True}}},
            ],
        },
        {"_id": 1},
    )
    if not matched:
        raise HTTPException(404, "Referral code not found or inactive")
    await affiliate_capture_click(request, response, normalized,
                                  page=page, referrer=referrer, device=device)
    return {"ok": True}


# ===========================================================================
# ENDPOINTS — ADMIN
# ===========================================================================

async def admin_affiliate_invite(payload: AffiliateInviteIn,
                                 admin: dict = Depends(get_admin_user)):  # noqa: F821
    """Crée (ou ré-invite) un affilié verrouillé à un email.
    Nouveau schéma : Prénom + Nom obligatoires, Entreprise optionnelle.
    Rétrocompat : accepte encore `name` seul (ancien front) — auto-split."""
    email = payload.email.lower().strip()
    # Réconciliation prénom / nom : nouveau schéma prioritaire, fallback sur legacy `name`.
    first_name = (payload.first_name or "").strip()
    last_name = (payload.last_name or "").strip()
    if not first_name and payload.name:
        parts = payload.name.strip().split(None, 1)
        first_name = parts[0] if parts else ""
        last_name = parts[1] if len(parts) > 1 else ""
    if not first_name or not last_name:
        raise HTTPException(422, "Prénom et nom sont obligatoires")
    company = (payload.company or "").strip()
    display_name = f"{first_name} {last_name}".strip()
    custom_code = _affiliate_normalize_custom_code(payload.code) if payload.code else None
    if custom_code:
        duplicate = await db.affiliates.find_one(
            {"$or": [{"code": custom_code}, {"aliases.code": custom_code}],
             "email": {"$ne": email}},
            {"_id": 1},
        )
        if duplicate:
            raise HTTPException(409, "Affiliate code already in use")

    payout_currency = (payload.payout_currency or "usdt").strip().lower()
    if payout_currency not in AFFILIATE_PAYOUT_CURRENCIES:
        payout_currency = "usdt"

    raw = secrets.token_urlsafe(32)
    now = datetime.now(timezone.utc)
    expires = (now + timedelta(hours=AFFILIATE_INVITE_TTL_HOURS)).isoformat()

    existing = await db.affiliates.find_one({"email": email}, {"_id": 0})
    if existing and existing.get("status") == "active":
        raise HTTPException(409, "This email is already an active affiliate")

    if existing:
        await db.affiliates.update_one(
            {"id": existing["id"]},
            {"$set": {
                "invite_token_hash": _affiliate_hash_token(raw),
                "invite_expires_at": expires,
                "name": display_name,
                "first_name": first_name,
                "last_name": last_name,
                "company": company,
                "code": custom_code,
                "coupon_percent": payload.coupon_percent,
                "status": "invited",
                "payout_currency": payout_currency,
                "invite_last_sent_at": now.isoformat(),
            },
             "$inc": {"invite_sent_count": 1}},
        )
        aff_id = existing["id"]
    else:
        aff_id = str(uuid.uuid4())
        await db.affiliates.insert_one({
            "id": aff_id,
            "email": email,
            "name": display_name,
            "first_name": first_name,
            "last_name": last_name,
            "company": company,
            "code": custom_code,
            "coupon_percent": payload.coupon_percent,
            "user_id": None,
            "status": "invited",
            "compliance_status": "compliant",
            "manual_tier": None,
            "commission_note": payload.commission_note or "",
            "payout_currency": payout_currency,
            "payout_address": "",
            "ip_hash": None,
            "known_addresses": [],
            "aliases": [],
            "invite_token_hash": _affiliate_hash_token(raw),
            "invite_expires_at": expires,
            "invite_sent_count": 1,
            "invite_last_sent_at": now.isoformat(),
            "created_at": now.isoformat(),
            "activated_at": None,
        })

    base = _trusted_public_base_url()
    link = f"{base}/affiliate/join?token={raw}"
    await _affiliate_send_invite(email, display_name, link, payload.lang or "fr")
    return {"ok": True, "affiliate_id": aff_id, "invite_link": link}


async def admin_affiliate_resend(affiliate_id: str,
                                 admin: dict = Depends(get_admin_user)):  # noqa: F821
    aff = await db.affiliates.find_one({"id": affiliate_id}, {"_id": 0})
    if not aff:
        raise HTTPException(404, "Affiliate not found")
    if aff.get("status") == "active":
        raise HTTPException(400, "Affiliate already active — no invite to resend")
    await _rate_limit("affiliate_invite", affiliate_id, AFFILIATE_INVITE_MAX,  # noqa: F821
                AFFILIATE_INVITE_WINDOW, "Trop de renvois. Réessayez plus tard.")
    raw = secrets.token_urlsafe(32)
    now = datetime.now(timezone.utc)
    expires = (now + timedelta(hours=AFFILIATE_INVITE_TTL_HOURS)).isoformat()
    await db.affiliates.update_one(
        {"id": affiliate_id},
        {"$set": {"invite_token_hash": _affiliate_hash_token(raw),
                  "invite_expires_at": expires,
                  "status": "invited",
                  "invite_last_sent_at": now.isoformat()},
         "$inc": {"invite_sent_count": 1}},
    )
    base = _trusted_public_base_url()
    link = f"{base}/affiliate/join?token={raw}"
    await _affiliate_send_invite(aff["email"], aff.get("name", ""), link,
                                 "fr")
    return {"ok": True, "invite_link": link, "sent_to": aff["email"]}


class AffiliateBulkInviteRow(BaseModel):
    model_config = ConfigDict(extra="ignore")
    email: str = Field(min_length=1, max_length=320)
    first_name: Optional[str] = ""
    last_name: Optional[str] = ""
    company: Optional[str] = ""
    discount_percent: Optional[float] = None
    # Rétrocompat ancien CSV : accepte encore `name` unique (auto-split)
    name: Optional[str] = ""


class AffiliateBulkInviteIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    rows: List[AffiliateBulkInviteRow] = Field(min_length=1, max_length=500)
    commission_note: Optional[str] = ""
    payout_currency: Optional[str] = "usdt"
    lang: str = "fr"
    # Rabais par défaut appliqué aux lignes sans `discount_percent` explicite.
    default_discount_percent: float = 10.0


async def admin_affiliate_bulk_invite(payload: AffiliateBulkInviteIn,
                                      admin: dict = Depends(get_admin_user)):  # noqa: F821
    """Invite en masse depuis un CSV parsé côté client.
    Nouveau schéma 5 colonnes : first_name, last_name, company, email, discount_percent.
    - Rétrocompat : si l'ancien champ `name` est fourni, il est splitté.
    - Le rabais renseigné (ou default_discount_percent) sera appliqué à
      l'activation via le code v2 `BASE + %`.
    """
    lang = (payload.lang or "fr").lower()
    if lang not in ("fr", "en"):
        lang = "fr"
    payout_currency = (payload.payout_currency or "usdt").lower()
    if payout_currency not in AFFILIATE_PAYOUT_CURRENCIES:
        payout_currency = "usdt"

    base = _trusted_public_base_url()
    now = datetime.now(timezone.utc)
    expires = (now + timedelta(hours=AFFILIATE_INVITE_TTL_HOURS)).isoformat()

    results: list = []
    skipped: list = []
    failed: list = []
    seen_emails: set = set()

    for row in payload.rows:
        email = (row.email or "").lower().strip()
        # Résolution prénom/nom : nouveau schéma prioritaire, fallback sur legacy `name`.
        fn = (row.first_name or "").strip()
        ln = (row.last_name or "").strip()
        if (not fn or not ln) and row.name:
            parts = (row.name or "").strip().split(None, 1)
            fn = fn or (parts[0] if parts else "")
            ln = ln or (parts[1] if len(parts) > 1 else "")
        company = (row.company or "").strip()
        display_name = f"{fn} {ln}".strip() or email.split("@")[0]
        # Rabais utilisé pour ce futur affilié (persisté sur la fiche pour l'activation).
        raw_pct = row.discount_percent
        if raw_pct is None:
            discount_percent = float(payload.default_discount_percent or 10.0)
        else:
            try:
                discount_percent = float(raw_pct)
            except Exception:
                discount_percent = float(payload.default_discount_percent or 10.0)
        discount_percent = max(0.0, min(100.0, discount_percent))

        if not email:
            failed.append({"email": row.email, "error": "Missing email"})
            continue
        if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
            failed.append({"email": email, "error": "Invalid email format"})
            continue
        if not fn or not ln:
            failed.append({"email": email, "error": "first_name and last_name are required"})
            continue
        if email in seen_emails:
            skipped.append({"email": email, "reason": "Duplicate in CSV"})
            continue
        seen_emails.add(email)

        try:
            existing = await db.affiliates.find_one({"email": email}, {"_id": 0})
            if existing and existing.get("status") == "active":
                skipped.append({"email": email, "reason": "Already active"})
                continue

            raw = secrets.token_urlsafe(32)
            token_hash = _affiliate_hash_token(raw)

            common_fields = {
                "invite_token_hash": token_hash,
                "invite_expires_at": expires,
                "name": display_name,
                "first_name": fn,
                "last_name": ln,
                "company": company,
                "coupon_percent": discount_percent,
                "status": "invited",
                "payout_currency": payout_currency,
                "invite_last_sent_at": now.isoformat(),
            }
            if existing:
                await db.affiliates.update_one(
                    {"id": existing["id"]},
                    {"$set": common_fields, "$inc": {"invite_sent_count": 1}},
                )
                aff_id = existing["id"]
            else:
                aff_id = str(uuid.uuid4())
                await db.affiliates.insert_one({
                    "id": aff_id,
                    "email": email,
                    **common_fields,
                    "code": None,      # sera généré à l'activation par _affiliate_gen_code_v2
                    "user_id": None,
                    "compliance_status": "compliant",
                    "manual_tier": None,
                    "commission_note": payload.commission_note or "",
                    "payout_address": "",
                    "ip_hash": None,
                    "known_addresses": [],
                    "aliases": [],
                    "invite_sent_count": 1,
                    "created_at": now.isoformat(),
                    "activated_at": None,
                    "source": "bulk_csv",
                })

            link = f"{base}/affiliate/join?token={raw}"
            try:
                await db.affiliate_email_jobs.insert_one({
                    "id": str(uuid.uuid4()),
                    "kind": "bulk_invite",
                    "status": "pending",
                    "email": email,
                    "name": display_name,
                    "link": link,
                    "lang": lang,
                    "attempts": 0,
                    "available_at": now.isoformat(),
                    "created_at": now.isoformat(),
                    "expires_at": now + timedelta(days=7),
                })
                email_status = "queued"
            except Exception as e:
                logging.warning(
                    "[affiliate] bulk-invite email queue failed ref=%s error_type=%s",
                    _private_ref(email),
                    type(e).__name__,
                )
                email_status = "queue_error"

            results.append({
                "email": email,
                "name": display_name,
                "first_name": fn,
                "last_name": ln,
                "company": company,
                "discount_percent": discount_percent,
                "affiliate_id": aff_id,
                "invite_link": link,
                "email_status": email_status,
            })
        except Exception as e:
            logging.error(
                "[affiliate] bulk-invite failed ref=%s error_type=%s",
                _private_ref(email),
                type(e).__name__,
            )
            failed.append({"email": email, "error": "invite_processing_failed"})

    return {
        "ok": True,
        "total": len(payload.rows),
        "sent": len(results),
        "queued": sum(1 for row in results if row["email_status"] == "queued"),
        "skipped": skipped,
        "failed": failed,
        "results": results,
    }


async def admin_affiliates_list(admin: dict = Depends(get_admin_user)):  # noqa: F821
    rows = await _cursor_all(db.affiliates.find(
        {}, {"_id": 0, "invite_token_hash": 0}
    ).sort("created_at", -1))
    metrics_by_affiliate = await _affiliate_compute_list_metrics(rows)
    out = []
    for aff in rows:
        metrics = metrics_by_affiliate.get(aff["id"])
        item = dict(aff)
        if metrics:
            item.update({
                "cumulative_revenue": metrics["cumulative_revenue"],
                "quarter_revenue": metrics["quarter_revenue"],
                "rolling12_revenue": metrics["rolling12_revenue"],
                "tier": metrics["tier"],
                # Champs de provenance du palier. Cette fusion est une liste
                # BLANCHE : un champ ajouté aux métriques et oublié ici
                # n'atteint jamais l'interface. Le marqueur « forcé / entente »
                # de la liste dependait de tier_is_manual, absent d'ici : sa
                # condition etait donc toujours fausse et le marqueur ne
                # s'affichait jamais, sans qu'aucune erreur ne le signale.
                "tier_is_manual": metrics["tier_is_manual"],
                "tier_agreement": metrics["tier_agreement"],
                "tier_theoretical": metrics["tier_theoretical"],
                "commission_rate": metrics["commission_rate"],
                "pending_commission": metrics["pending_commission"],
                "approved_commission": metrics["approved_commission"],
                "paid_commission": metrics["paid_commission"],
            })
        out.append(item)

    # Clics + dernier clic par affilié (colonne liste + tri)
    click_stats = {}
    async for grp in db.affiliate_clicks.aggregate([
        {"$group": {"_id": "$affiliate_id",
                    "clicks": {"$sum": 1},
                    "last": {"$max": "$created_at"}}}
    ]):
        click_stats[grp["_id"]] = {
            "clicks": int(grp.get("clicks", 0)),
            "last": grp.get("last"),
        }
    for item in out:
        cs = click_stats.get(item["id"], {})
        item["clicks"] = cs.get("clicks", 0)
        item["last_click_at"] = cs.get("last")
    return out


# ===========================================================================
# ===== FIRONOVA_AFFILIATE_ADMIN_OVERVIEW_START =====
# Endpoints manquants requis par AdminAffiliates.jsx (sinon la page casse).

async def admin_affiliates_overview(admin: dict = Depends(get_admin_user)):  # noqa: F821
    """Vue d'ensemble du programme d'affiliation : finances, effectifs,
    alertes, attribution, série mensuelle, top affiliés, distribution tiers."""
    now = datetime.now(timezone.utc)
    quarter_start = _affiliate_quarter_start()
    tier_window_start = now - timedelta(days=365)

    # --- Effectifs ---
    active = await db.affiliates.count_documents({"status": "active"})
    invited = await db.affiliates.count_documents({"status": "invited"})
    suspended = await db.affiliates.count_documents({"status": "suspended"})
    active_affiliates = await _cursor_all(db.affiliates.find(
        {"status": "active"}, {"_id": 0, "id": 1, "manual_tier": 1}
    ))

    # --- Finances (agrégées dans MongoDB) ---
    validated_statuses = ["approved", "paid"]
    referral_facets = await _cursor_all(db.affiliate_referrals.aggregate([
        {"$match": {"status": {"$ne": "excluded"}}},
        {"$facet": {
            "financial": [{"$group": {
                "_id": None,
                "commission_pending": {"$sum": {"$cond": [
                    {"$eq": ["$status", "pending"]}, {"$ifNull": ["$commission_amount", 0]}, 0
                ]}},
                "commission_due": {"$sum": {"$cond": [
                    {"$eq": ["$status", "approved"]}, {"$ifNull": ["$commission_amount", 0]}, 0
                ]}},
                "commission_paid": {"$sum": {"$cond": [
                    {"$eq": ["$status", "paid"]}, {"$ifNull": ["$commission_amount", 0]}, 0
                ]}},
                "commission_reversed": {"$sum": {"$cond": [
                    {"$in": ["$status", ["reversed", "refunded"]]},
                    {"$ifNull": ["$commission_amount", 0]}, 0
                ]}},
                "validated_orders": {"$sum": {"$cond": [
                    {"$in": ["$status", validated_statuses]}, 1, 0
                ]}},
                "validated_revenue": {"$sum": {"$cond": [
                    {"$in": ["$status", validated_statuses]},
                    {"$ifNull": ["$order_total", 0]}, 0
                ]}},
            }}],
            "monthly": [
                {"$match": {"status": {"$in": validated_statuses}}},
                {"$group": {
                    "_id": {"$substrBytes": [{"$ifNull": ["$created_at", ""]}, 0, 7]},
                    "revenue": {"$sum": {"$ifNull": ["$order_total", 0]}},
                    "commission": {"$sum": {"$ifNull": ["$commission_amount", 0]}},
                }},
                {"$sort": {"_id": -1}},
                {"$limit": 12},
                {"$sort": {"_id": 1}},
            ],
            "per_affiliate": [
                {"$match": {"status": {"$in": validated_statuses}}},
                {"$group": {
                    "_id": "$affiliate_id",
                    "revenue": {"$sum": {"$ifNull": ["$order_total", 0]}},
                    "commission": {"$sum": {"$ifNull": ["$commission_amount", 0]}},
                    "cumulative": {"$sum": {"$ifNull": ["$base_amount", 0]}},
                    # 365 derniers jours — c'est CETTE somme qui fixe le palier,
                    # comme dans _affiliate_compute_metrics(). Sans elle, l'admin
                    # classait sur le cumul a vie pendant que l'affilie voyait un
                    # palier calcule sur douze mois glissants : deux reponses
                    # differentes a la meme question.
                    "rolling12": {"$sum": {"$cond": [
                        {"$gte": [
                            {"$ifNull": ["$approved_at", "$created_at"]},
                            tier_window_start.isoformat(),
                        ]},
                        {"$ifNull": ["$base_amount", 0]}, 0,
                    ]}},
                    "quarter": {"$sum": {"$cond": [
                        {"$gte": [
                            {"$ifNull": ["$approved_at", "$created_at"]},
                            quarter_start.isoformat(),
                        ]},
                        {"$ifNull": ["$base_amount", 0]}, 0,
                    ]}},
                }},
            ],
        }},
    ]))
    facets = referral_facets[0] if referral_facets else {}
    financial = (facets.get("financial") or [{}])[0]
    commission_pending = float(financial.get("commission_pending", 0.0))
    commission_due = float(financial.get("commission_due", 0.0))
    commission_paid = float(financial.get("commission_paid", 0.0))
    commission_reversed = float(financial.get("commission_reversed", 0.0))
    validated_orders = int(financial.get("validated_orders", 0))
    validated_revenue = float(financial.get("validated_revenue", 0.0))
    avg_order_value = (validated_revenue / validated_orders) if validated_orders else 0.0
    monthly_series = [
        {"month": row.get("_id", ""), "revenue": row.get("revenue", 0.0),
         "commission": row.get("commission", 0.0)}
        for row in facets.get("monthly", [])
    ]
    for m in monthly_series:
        m["revenue"] = round(m["revenue"], 2)
        m["commission"] = round(m["commission"], 2)

    per_aff = {
        row.get("_id"): {
            "revenue": float(row.get("revenue", 0.0)),
            "commission": float(row.get("commission", 0.0)),
        }
        for row in facets.get("per_affiliate", []) if row.get("_id")
    }
    tier_revenue = {
        row.get("_id"): {
            "cumulative": float(row.get("cumulative", 0.0)),
            "rolling12": float(row.get("rolling12", 0.0)),
            "quarter": float(row.get("quarter", 0.0)),
        }
        for row in facets.get("per_affiliate", []) if row.get("_id")
    }

    # --- Top affiliés (par revenu validé) ---
    top_ids = sorted(per_aff.keys(), key=lambda k: per_aff[k]["revenue"], reverse=True)[:5]
    top_documents = await _cursor_all(db.affiliates.find(
        {"id": {"$in": top_ids}}, {"_id": 0, "id": 1, "code": 1, "name": 1}
    )) if top_ids else []
    affiliates_by_id = {affiliate.get("id"): affiliate for affiliate in top_documents}
    top_affiliates = []
    for aid in top_ids:
        a = affiliates_by_id.get(aid)
        if a:
            top_affiliates.append({
                "code": a.get("code"), "name": a.get("name"),
                "revenue": round(per_aff[aid]["revenue"], 2),
                "commission": round(per_aff[aid]["commission"], 2),
            })

    # --- Distribution des tiers ---
    tier_distribution = {}
    valid_tiers = {tier[0] for tier in AFFILIATE_TIERS}
    for affiliate in active_affiliates:
        values = tier_revenue.get(affiliate.get("id"), {})
        rolling12 = float(values.get("rolling12", 0.0))
        manual_tier = str(affiliate.get("manual_tier") or "").strip().lower() or None
        if manual_tier not in valid_tiers:
            manual_tier = None
        # Meme regle que _affiliate_compute_metrics() : palier sur les douze mois
        # glissants, surcharge manuelle prioritaire, aucune retrogradation
        # trimestrielle. L'ancienne version classait sur le cumul a vie puis
        # retrogradait si le trimestre passait sous le plancher CUMULE du palier
        # — un seuil qu'un affilie regulier ne pouvait pas tenir chaque
        # trimestre. L'admin voyait donc un palier que l'affilie n'avait pas.
        tier = manual_tier or _affiliate_tier_for_revenue(rolling12)
        tier_distribution[tier] = tier_distribution.get(tier, 0) + 1

    # --- Attribution (clics / conversion) ---
    total_clicks = await db.affiliate_clicks.count_documents({})
    conversion_rate = (validated_orders / total_clicks) if total_clicks else None

    # --- Alertes ---
    ready_summary = await db.affiliate_payouts.aggregate([
        {"$match": {"status": "ready"}},
        {"$group": {"_id": None, "count": {"$sum": 1}, "amount": {"$sum": "$amount"}}},
    ]).to_list(1)
    payouts_ready = int(ready_summary[0].get("count", 0)) if ready_summary else 0
    payouts_ready_amount = round(float(ready_summary[0].get("amount", 0)), 2) if ready_summary else 0.0
    invites_expired = await db.affiliates.count_documents(
        {"status": "invited", "invite_expires_at": {"$lt": now.isoformat()}}
    )
    compliance_review = await db.affiliates.count_documents({"compliance_status": "review"})
    commissions_maturing = await db.affiliate_referrals.count_documents({"status": "pending"})

    return {
        "financial": {
            "commission_pending": round(commission_pending, 2),
            "commission_due": round(commission_due, 2),
            "commission_paid": round(commission_paid, 2),
            "commission_reversed": round(commission_reversed, 2),
            "validated_orders": validated_orders,
            "validated_revenue": round(validated_revenue, 2),
            "avg_order_value": round(avg_order_value, 2),
        },
        "affiliates": {"active": active, "invited": invited, "suspended": suspended},
        "alerts": {
            "payouts_ready": payouts_ready,
            "payouts_ready_amount": payouts_ready_amount,
            "invites_expired": invites_expired,
            "compliance_review": compliance_review,
            "commissions_maturing": commissions_maturing,
        },
        "attribution": {
            "total_clicks": total_clicks,
            "conversion_rate": conversion_rate,
        },
        "monthly_series": monthly_series,
        "top_affiliates": top_affiliates,
        "tier_distribution": tier_distribution,
    }


async def admin_affiliates_clicks(admin: dict = Depends(get_admin_user),  # noqa: F821
                                  days: int = 30):
    """Analyse d'attribution à l'échelle du programme : volume de clics,
    tendance quotidienne, top pages/référents/appareils et top affiliés par
    volume de clics."""
    days = max(7, min(int(days), 90))
    start_date = datetime.now(timezone.utc).date() - timedelta(days=days - 1)
    q = {"created_at": {"$gte": start_date.isoformat() + "T00:00:00"}}
    trend, per_aff, pages, refs, devices = {}, {}, {}, {}, {}
    total = 0
    cursor = db.affiliate_clicks.find(
        q, {"_id": 0, "created_at": 1, "affiliate_id": 1, "page": 1,
            "referrer": 1, "device": 1},
    )
    async for c in cursor:
        total += 1
        day = str(c.get("created_at", ""))[:10]
        trend[day] = trend.get(day, 0) + 1
        aid = c.get("affiliate_id")
        if aid:
            per_aff[aid] = per_aff.get(aid, 0) + 1
        page = str(c.get("page") or "").strip() or "direct"
        pages[page] = pages.get(page, 0) + 1
        ref = _affiliate_referrer_domain(str(c.get("referrer") or "")) or "direct"
        refs[ref] = refs.get(ref, 0) + 1
        dev = str(c.get("device") or "").strip() or "unknown"
        devices[dev] = devices.get(dev, 0) + 1

    top_affiliates = []
    for aid, n in sorted(per_aff.items(), key=lambda kv: kv[1], reverse=True)[:10]:
        a = await db.affiliates.find_one({"id": aid}, {"_id": 0, "name": 1, "code": 1})
        top_affiliates.append({
            "id": aid, "name": (a or {}).get("name") or "—",
            "code": (a or {}).get("code") or "", "clicks": n,
        })

    series = []
    for i in range(days):
        d = (start_date + timedelta(days=i)).isoformat()
        series.append({"date": d, "clicks": trend.get(d, 0)})

    def _top(d, n=8):
        return [{"source": k, "clicks": v}
                for k, v in sorted(d.items(), key=lambda kv: kv[1], reverse=True)[:n]]

    # Conversions validées sur la même fenêtre (approved|paid)
    conversions_30d = 0
    rcursor = db.affiliate_referrals.find(
        {"status": {"$in": ["approved", "paid"]},
         "$or": [{"approved_at": {"$gte": start_date.isoformat() + "T00:00:00"}},
                 {"created_at": {"$gte": start_date.isoformat() + "T00:00:00"}}]},
        {"_id": 0, "approved_at": 1, "created_at": 1},
    )
    async for r in rcursor:
        ts = r.get("approved_at") or r.get("created_at")
        if str(ts or "")[:10] >= start_date.isoformat():
            conversions_30d += 1

    return {
        "days": days,
        "total_clicks": total,
        "conversions_30d": conversions_30d,
        "active_affiliates": len(per_aff),
        "trend": series,
        "top_pages": _top(pages),
        "top_referrers": _top(refs),
        "devices": devices,
        "top_affiliates": top_affiliates,
    }


async def admin_affiliates_risk(admin: dict = Depends(get_admin_user)):  # noqa: F821
    """Détection de risque : referrals exclus (auto-parrainage, IP partagée)
    et affiliés à surveiller. Retourne la liste + le compte signalé."""
    flagged = await db.affiliate_referrals.find(
        {"status": "excluded"},
        {"_id": 0, "id": 1, "order_number": 1, "commission_amount": 1,
         "base_amount": 1, "excluded_reason": 1, "status": 1, "affiliate_id": 1},
    ).sort("created_at", -1).to_list(200)
    # Enrichir avec le code affilié
    for f in flagged:
        a = await db.affiliates.find_one({"id": f.get("affiliate_id")}, {"_id": 0, "code": 1})
        f["affiliate_code"] = a.get("code") if a else None
    return {"affiliates": flagged, "flagged_count": len(flagged)}
# ===== FIRONOVA_AFFILIATE_ADMIN_OVERVIEW_END =====


async def admin_affiliate_detail(affiliate_id: str,
                                 admin: dict = Depends(get_admin_user)):  # noqa: F821
    aff = await db.affiliates.find_one(
        {"id": affiliate_id}, {"_id": 0, "invite_token_hash": 0}
    )
    if not aff:
        raise HTTPException(404, "Affiliate not found")
    # Calculé quel que soit le statut. La condition « active » faisait qu'un
    # palier forcé sur un affilié encore invité était bien enregistré en base
    # mais jamais renvoyé : l'admin voyait « — » et concluait que la sauvegarde
    # avait échoué. Le coût est nul — c'est le détail d'UN affilié.
    metrics = await _affiliate_compute_metrics(affiliate_id)
    referrals = await db.affiliate_referrals.find(
        {"affiliate_id": affiliate_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    payouts = await db.affiliate_payouts.find(
        {"affiliate_id": affiliate_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(200)
    return {"affiliate": aff, "metrics": metrics,
            "referrals": referrals, "payouts": payouts}


async def admin_affiliate_update(affiliate_id: str, payload: AffiliateAdminUpdateIn,
                                 admin: dict = Depends(get_admin_user)):  # noqa: F821
    aff = await db.affiliates.find_one({"id": affiliate_id}, {"_id": 0})
    if not aff:
        raise HTTPException(404, "Affiliate not found")
    update = {k: v for k, v in payload.model_dump(exclude_none=True).items()}
    if "coupon_percent" in payload.model_fields_set and payload.coupon_percent is None:
        update["coupon_percent"] = None
    clear_manual_tier = bool(update.pop("clear_manual_tier", False))
    if clear_manual_tier:
        update["manual_tier"] = None
        # Une entente sans palier fige n'a pas d'objet : revenir au calcul
        # automatique retire donc aussi l'engagement. Sans cette regle, un
        # affilie garderait le drapeau « entente » tout en voyant son palier
        # bouger avec son chiffre d'affaires — exactement ce que l'engagement
        # dit qui n'arrivera pas.
        update["tier_agreement"] = False
    elif "manual_tier" in update:
        manual_tier = str(update["manual_tier"] or "").strip().lower()
        if manual_tier not in {tier[0] for tier in AFFILIATE_TIERS}:
            raise HTTPException(400, "Unknown affiliate tier")
        update["manual_tier"] = manual_tier

    explicit_code = None
    if "code" in update:
        explicit_code = _affiliate_normalize_custom_code(update["code"])
        duplicate = await db.affiliates.find_one(
            {"id": {"$ne": affiliate_id},
             "$or": [{"code": explicit_code}, {"aliases.code": explicit_code}]},
            {"_id": 1},
        )
        if duplicate:
            raise HTTPException(409, "Affiliate code already in use")
        update["code"] = explicit_code

    # Payout : validation stricte (USDT/USDC + Ethereum EIP-55 OU Tron TRC-20)
    if "payout_address" in update or "payout_currency" in update:
        addr, cur, net = _normalize_payout(
            update.get("payout_address", aff.get("payout_address", "")),
            update.get("payout_currency", aff.get("payout_currency", "usdt")),
        )
        update["payout_address"] = addr
        update["payout_currency"] = cur
        update["payout_network"] = net

    # Rename atomique du code affilié si le rabais (coupon_percent) change.
    # L'ancien code passe dans `aliases[]` (actif par défaut → attribution
    # continue de fonctionner sur les vieux liens partagés).
    new_pct = update.get("coupon_percent")
    old_pct = aff.get("coupon_percent")
    if (new_pct is not None
            and explicit_code is None
            and aff.get("status") == "active"
            and aff.get("code")
            and float(new_pct) != float(old_pct or 0)):
        base_source = (update.get("company", aff.get("company") or "").strip()
                       or update.get("first_name", aff.get("first_name") or "").strip()
                       or (aff.get("name") or aff.get("email", "")).split()[0])
        new_code = await _affiliate_gen_code_v2(
            base_source, float(new_pct),
            email=aff.get("email", ""), exclude_id=affiliate_id,
        )
        if new_code != aff.get("code"):
            old_code = aff["code"]
            old_alias = {
                "code": old_code,
                "active": True,
                "discount_percent_at_creation": float(old_pct or 0),
                "archived_at": datetime.now(timezone.utc).isoformat(),
                "archived_by": admin.get("email", ""),
            }
            update["code"] = new_code
            # Désactive l'ancien coupon (single source of truth = coupon actuel).
            try:
                await db.coupons.update_one(
                    {"code": old_code, "source": "affiliate"},
                    {"$set": {"active": False,
                              "deactivated_at": datetime.now(timezone.utc).isoformat(),
                              "deactivated_reason": "code_renamed"}},
                )
            except Exception as e:
                logging.warning("[affiliate] désactivation coupon %s: %s", old_code, e)
            # Crée / réactive le coupon au nouveau code + nouveau %
            try:
                await _affiliate_ensure_coupon(new_code, affiliate_id, percent=float(new_pct))
            except Exception as e:
                logging.warning("[affiliate] création coupon %s: %s", new_code, e)
            # Push alias (dédup si déjà présent)
            await db.affiliates.update_one(
                {"id": affiliate_id, "aliases.code": {"$ne": old_code}},
                {"$push": {"aliases": old_alias}},
            )

    if explicit_code and explicit_code != aff.get("code"):
        old_code = aff.get("code")
        if old_code:
            await db.affiliates.update_one(
                {"id": affiliate_id, "aliases.code": {"$ne": old_code}},
                {"$push": {"aliases": {
                    "code": old_code,
                    "active": True,
                    "discount_percent_at_creation": float(old_pct or 0),
                    "archived_at": datetime.now(timezone.utc).isoformat(),
                    "archived_by": admin.get("email", ""),
                }}},
            )
        await _affiliate_ensure_coupon(
            explicit_code, affiliate_id,
            percent=float(new_pct if new_pct is not None else old_pct or 0),
        )

    # Reconstruit `name` legacy si first/last changent
    if "first_name" in update or "last_name" in update:
        fn = update.get("first_name", aff.get("first_name", ""))
        ln = update.get("last_name", aff.get("last_name", ""))
        update["name"] = f"{fn} {ln}".strip() or aff.get("name") or ""

    if update:
        await db.affiliates.update_one({"id": affiliate_id}, {"$set": update})
    fresh = await db.affiliates.find_one(
        {"id": affiliate_id}, {"_id": 0, "invite_token_hash": 0}
    )
    return fresh


class AffiliateAliasToggleIn(BaseModel):
    active: bool


async def admin_affiliate_alias_toggle(affiliate_id: str, alias_code: str,
                                        payload: AffiliateAliasToggleIn,
                                        admin: dict = Depends(get_admin_user)):  # noqa: F821
    """Active/désactive un alias historique. Un alias inactif ne permet plus
    l'attribution ni l'application du rabais."""
    aff = await db.affiliates.find_one({"id": affiliate_id}, {"_id": 0})
    if not aff:
        raise HTTPException(404, "Affiliate not found")
    res = await db.affiliates.update_one(
        {"id": affiliate_id, "aliases.code": alias_code.upper()},
        {"$set": {"aliases.$.active": bool(payload.active),
                  "aliases.$.toggled_at": datetime.now(timezone.utc).isoformat(),
                  "aliases.$.toggled_by": admin.get("email", "")}},
    )
    if not res.matched_count:
        raise HTTPException(404, "Alias not found")
    return await db.affiliates.find_one(
        {"id": affiliate_id}, {"_id": 0, "invite_token_hash": 0}
    )


async def admin_affiliate_run_payouts(admin: dict = Depends(get_admin_user),  # noqa: F821
                                      period: Optional[str] = None,
                                      dry_run: bool = False):
    """Génère les payouts mensuels : agrège les commissions 'approved' par
    affilié en un relevé de payout (status 'ready'), marque les référrals
    comme rattachés. Le paiement crypto réel se fait ensuite hors-système,
    puis l'admin confirme via /mark-paid avec la référence de transaction.

    dry_run=true : retourne l'aperçu (montants, FX, groupes) SANS écrire
    en base — utilisé pour valider la sortie avant un vrai run mensuel."""
    now = datetime.now(timezone.utc)
    period = period or now.strftime("%Y-%m")
    created = []

    # Regroupe par affilié les commissions approuvées non encore payées
    pipeline = [
        {"$match": {"status": "approved", "payout_id": None}},
        {"$group": {"_id": "$affiliate_id",
                    "total": {"$sum": "$commission_amount"},
                    "ids": {"$push": "$id"}}},
    ]
    # Taux de change unique pour tout le batch (transparence + cohérence intra-batch)
    fx_rate, fx_source = await _fetch_cad_to_usd_rate()
    fx_captured_at = now.isoformat()

    payout_groups = await _cursor_all(db.affiliate_referrals.aggregate(pipeline))
    affiliate_ids = [group.get("_id") for group in payout_groups if group.get("_id")]
    affiliate_documents = await _cursor_all(db.affiliates.find(
        {"id": {"$in": affiliate_ids}}, {"_id": 0}
    )) if affiliate_ids else []
    affiliates_by_id = {
        affiliate.get("id"): affiliate for affiliate in affiliate_documents
    }

    for grp in payout_groups:
        affiliate_id = grp["_id"]
        total = round(float(grp["total"]), 2)
        if total <= 0:
            continue
        aff = affiliates_by_id.get(affiliate_id)
        if not aff or aff.get("status") != "active":
            continue
        # ---- Item 3.2 : seuil minimum de payout (skip + notification) --------
        if total < AFFILIATE_PAYOUT_MIN_CAD:
            deferral_entry = {
                "affiliate_id": affiliate_id,
                "affiliate_code": aff.get("code"),
                "amount_cad": total,
                "threshold_cad": AFFILIATE_PAYOUT_MIN_CAD,
                "referral_count": len(grp["ids"]),
                "deferred": True,
            }
            if not dry_run:
                notified = await _defer_affiliate_payout_below_threshold(
                    aff, period, total, len(grp["ids"]), AFFILIATE_PAYOUT_MIN_CAD,
                )
                deferral_entry["notified"] = notified
            created.append(deferral_entry)
            continue
        # Devise cible : USDT/USDC → 1:1 avec USD (peggé). Sinon on garde la valeur legacy.
        payout_currency = (aff.get("payout_currency") or "usdt").lower()
        amount_cad = total
        if payout_currency in AFFILIATE_PAYOUT_CURRENCIES:
            amount_usd = round(amount_cad * fx_rate, 2)
            amount_target = amount_usd    # USDT/USDC ≈ USD
        else:
            amount_usd = None
            amount_target = amount_cad
        payout_id = str(uuid.uuid4())
        if dry_run:
            # aperçu uniquement — pas d'écriture BDD, pas de payout_id assigné
            created.append({
                "affiliate_id": affiliate_id,
                "affiliate_code": aff.get("code"),
                "amount": amount_target,
                "amount_cad": amount_cad,
                "amount_usd": amount_usd,
                "currency": payout_currency,
                "referral_count": len(grp["ids"]),
                "payout_id": None,
                "dry_run": True,
            })
            continue
        try:
            await db.affiliate_payouts.insert_one({
                "id": payout_id,
                "affiliate_id": affiliate_id,
                "affiliate_code": aff.get("code"),
                "period": period,
                "amount": amount_target,             # montant à envoyer dans la devise cible
                "amount_cad": amount_cad,            # référence brute CAD (transparence)
                "amount_usd": amount_usd,            # équivalent USD (pour USDT/USDC)
                "currency": payout_currency,
                "fx_rate_cad_to_usd": fx_rate if payout_currency in AFFILIATE_PAYOUT_CURRENCIES else None,
                "fx_source": fx_source if payout_currency in AFFILIATE_PAYOUT_CURRENCIES else None,
                "fx_captured_at": fx_captured_at,
                "payout_address": aff.get("payout_address", ""),
                "referral_ids": grp["ids"],
                "referral_count": len(grp["ids"]),
                "status": "ready",           # ready → paid
                "reference": None,
                "note": "",
                "created_at": now.isoformat(),
                "paid_at": None,
            })
        except DuplicateKeyError:
            existing_payout = await db.affiliate_payouts.find_one(
                {"affiliate_id": affiliate_id, "period": period},
                {"_id": 0, "id": 1, "referral_ids": 1},
            )
            recover_ids = list(set(grp["ids"]) & set((existing_payout or {}).get("referral_ids", [])))
            if existing_payout and recover_ids:
                await db.affiliate_referrals.update_many(
                    {"id": {"$in": recover_ids}, "payout_id": None},
                    {"$set": {"payout_id": existing_payout["id"]}},
                )
            continue
        await db.affiliate_referrals.update_many(
            {"id": {"$in": grp["ids"]}},
            {"$set": {"payout_id": payout_id}},
        )
        created.append({"affiliate_id": affiliate_id, "amount": amount_target,
                        "amount_cad": amount_cad, "currency": payout_currency,
                        "payout_id": payout_id})
    payouts_only = [c for c in created if c.get("payout_id")]
    deferred = [c for c in created if c.get("deferred")]
    return {"ok": True, "period": period, "payouts_created": len(payouts_only),
            "payouts_deferred": len(deferred),
            "threshold_cad": AFFILIATE_PAYOUT_MIN_CAD,
            "fx_rate_cad_to_usd": fx_rate, "fx_source": fx_source,
            "fx_captured_at": fx_captured_at,
            "dry_run": bool(dry_run),
            "detail": created}


# =============================================================================
# NOWPayments Mass Payouts — envoi crypto batch (1 seul 2FA)
# =============================================================================
NOWPAYMENTS_PAYOUT_ENABLED = os.environ.get("NOWPAYMENTS_PAYOUT_ENABLED", "false").lower() == "true"
NOWPAYMENTS_JWT = os.environ.get("NOWPAYMENTS_JWT", "")   # session token JWT payout API
NOWPAYMENTS_EMAIL = os.environ.get("NOWPAYMENTS_EMAIL", "")
NOWPAYMENTS_PASSWORD = os.environ.get("NOWPAYMENTS_PASSWORD", "")


class AffiliatePayoutBatchIn(BaseModel):
    payout_ids: List[str] = Field(min_length=1, max_length=200)
    # 2FA/OTP obligatoire pour Mass Payouts NOWPayments — passé au vol par l'admin.
    otp: Optional[str] = None


# Code de devise NOWPayments par (jeton, réseau). Le réseau est DÉDUIT DE
# L'ADRESSE, jamais supposé : c'est l'adresse qui détermine où les fonds
# atterrissent. Le code envoyait « usdterc20 » en dur, donc un affilié ayant
# saisi une adresse Tron — que l'interface accepte et confirme explicitement —
# se voyait dispatché sur Ethereum vers une adresse T…, qui n'y existe pas.
#
# Toute combinaison absente de cette table est IGNORÉE et signalée, jamais
# envoyée sur un réseau approchant : un versement crypto est irréversible, et
# une adresse invalide sur le mauvais réseau ne revient pas.
NOWPAYMENTS_PAYOUT_CURRENCY = {
    ("usdt", "erc20"): "usdterc20",
    ("usdt", "trc20"): "usdttrc20",
    ("usdc", "erc20"): "usdcerc20",
    # ("usdc", "trc20"): "usdctrc20",  # à activer après vérification du code
    #   exact dans GET /v1/currencies de NOWPayments — un code inexistant fait
    #   échouer le lot entier, pas seulement la ligne concernée.
}


async def admin_affiliate_batch_payout(payload: AffiliatePayoutBatchIn,
                                        admin: dict = Depends(get_admin_user)):  # noqa: F821
    """Envoie plusieurs payouts USDT/USDC en lot via NOWPayments Mass Payouts.
    - Filtre les payouts sans adresse configurée → renvoyés dans `skipped`.
    - Marque les envoyés comme `status: 'processing'` avec `np_batch_id`.
    - Nécessite `NOWPAYMENTS_PAYOUT_ENABLED=true` + `NOWPAYMENTS_JWT` en env.
    - Le webhook `/api/webhook/nowpayments-payout` (déjà en place) confirmera
      le passage à `paid`."""
    # Charge les payouts demandés
    payouts = await db.affiliate_payouts.find(
        {"id": {"$in": payload.payout_ids}, "status": "ready"},
        {"_id": 0},
    ).to_list(200)
    if not payouts:
        raise HTTPException(400, "No eligible payouts (status must be 'ready')")

    withdrawals = []
    skipped = []
    for p in payouts:
        addr = (p.get("payout_address") or "").strip()
        cur = (p.get("currency") or "").lower()
        amt = p.get("amount")
        if not addr or cur not in AFFILIATE_PAYOUT_CURRENCIES or not amt:
            skipped.append({"payout_id": p["id"], "reason": "missing address or unsupported currency"})
            continue
        # Réseau déduit de l'adresse elle-même, pas du champ stocké : c'est
        # l'adresse qui décide de la destination réelle des fonds.
        net = _detect_payout_network(addr)
        np_currency = NOWPAYMENTS_PAYOUT_CURRENCY.get((cur, net or ""))
        if not np_currency:
            skipped.append({
                "payout_id": p["id"],
                "reason": (f"{cur.upper()} non supporté sur le réseau "
                           f"{(net or 'inconnu').upper()} — versement non envoyé"),
            })
            continue
        withdrawals.append({
            "address": addr,
            "currency": np_currency,
            "amount": float(amt),
            "ipn_callback_url": f"{PUBLIC_BASE_URL}/api/webhook/nowpayments-payout",
            "unique_external_id": p["id"],
        })

    if not withdrawals:
        return {"ok": False, "batch_id": None, "sent": 0, "skipped": skipped,
                "error": "No withdrawable payouts (missing addresses or invalid currency)"}

    claimed_withdrawals = []
    claim_time = datetime.now(timezone.utc).isoformat()
    for withdrawal in withdrawals:
        payout_id = withdrawal["unique_external_id"]
        claim = await db.affiliate_payouts.update_one(
            {"id": payout_id, "status": "ready"},
            {"$set": {
                "status": "dispatching",
                "np_dispatched_by": admin.get("email"),
                "np_dispatch_claimed_at": claim_time,
            }},
        )
        if claim.modified_count:
            claimed_withdrawals.append(withdrawal)
        else:
            skipped.append({"payout_id": payout_id, "reason": "already claimed"})
    withdrawals = claimed_withdrawals
    if not withdrawals:
        raise HTTPException(409, "All eligible payouts are already being processed")

    if not NOWPAYMENTS_PAYOUT_ENABLED:
        # Mode démo/manuel : on marque les payouts comme "queued_manual" pour que
        # l'admin sache qu'ils sont prêts à être envoyés via l'export CSV.
        now_iso = datetime.now(timezone.utc).isoformat()
        await db.affiliate_payouts.update_many(
            {"id": {"$in": [w["unique_external_id"] for w in withdrawals]}, "status": "dispatching"},
            {"$set": {"status": "queued_manual", "queued_at": now_iso}},
        )
        return {
            "ok": False,
            "batch_id": None,
            "sent": 0,
            "queued_manual": len(withdrawals),
            "skipped": skipped,
            "error": ("NOWPayments Mass Payouts non activé (NOWPAYMENTS_PAYOUT_ENABLED=false) — "
                      "les payouts ont été mis en file d'attente manuelle. Utilisez l'export CSV."),
        }

    async def _do_call(token: str):
        import httpx
        headers = {
            "x-api-key": NOWPAYMENTS_API_KEY,
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        }
        if payload.otp:
            headers["ncw-verify"] = payload.otp
        async with httpx.AsyncClient(timeout=30.0) as client:
            return await client.post(
                f"{NOWPAYMENTS_BASE_URL}/payout",
                headers=headers,
                json={"ipn_callback_url": f"{PUBLIC_BASE_URL}/api/webhook/nowpayments-payout",
                      "withdrawals": withdrawals},
            )

    ids_sent = [w["unique_external_id"] for w in withdrawals]
    try:
        # Obtient un JWT valide (rotation auto si EMAIL+PASSWORD config, sinon lit env)
        jwt_token = await _refresh_np_jwt()
        r = await _do_call(jwt_token)
        # Si 401/403 → JWT expiré, refresh forcé et 2nd try
        if r.status_code in (401, 403):
            logging.warning("[nowpayments] JWT rejected (%d), forcing refresh", r.status_code)
            jwt_token = await _refresh_np_jwt(force=True)
            r = await _do_call(jwt_token)
        r.raise_for_status()
        resp = r.json()
    except HTTPException:
        await db.affiliate_payouts.update_many(
            {"id": {"$in": ids_sent}, "status": "dispatching"},
            {"$set": {"status": "ready", "np_dispatch_error": "NOWPayments authentication failed"}},
        )
        raise
    except Exception as e:
        await db.affiliate_payouts.update_many(
            {"id": {"$in": ids_sent}, "status": "dispatching"},
            {"$set": {"status": "ready", "np_dispatch_error": "NOWPayments request failed"}},
        )
        logging.error("[nowpayments payout] batch failed: %s", type(e).__name__)
        raise HTTPException(502, "NOWPayments Mass Payouts request failed") from e

    batch_id = str(resp.get("id") or resp.get("batch_id") or f"np_{uuid.uuid4().hex[:12]}")
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.affiliate_payouts.update_many(
        {"id": {"$in": ids_sent}, "status": "dispatching"},
        {"$set": {"status": "processing", "np_batch_id": batch_id,
                  "np_dispatched_at": now_iso, "np_dispatched_by": admin.get("email")}},
    )
    return {"ok": True, "batch_id": batch_id, "sent": len(ids_sent),
            "skipped": skipped, "response": resp}


async def admin_affiliate_payouts_csv(admin: dict = Depends(get_admin_user)) -> Response:  # noqa: F821
    """Export CSV format NOWPayments (Mass Payouts import) pour envoi manuel.
    Colonnes : Address, Currency, Amount, ExternalId
    Ne prend que les payouts status 'ready' ou 'queued_manual' avec adresse valide."""
    cursor = db.affiliate_payouts.find(
        {"status": {"$in": ["ready", "queued_manual"]}},
        {"_id": 0, "id": 1, "payout_address": 1, "currency": 1, "amount": 1,
         "affiliate_code": 1, "period": 1},
    )
    rows = [["Address", "Currency", "Amount", "ExternalId", "AffiliateCode", "Period"]]
    async for p in cursor:
        addr = (p.get("payout_address") or "").strip()
        cur = (p.get("currency") or "").lower()
        amt = p.get("amount")
        if not addr or cur not in AFFILIATE_PAYOUT_CURRENCIES or not amt:
            continue
        np_currency = "usdterc20" if cur == "usdt" else "usdcerc20"
        rows.append([addr, np_currency, f"{float(amt):.2f}", p["id"],
                     p.get("affiliate_code", ""), p.get("period", "")])
    import io, csv
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows(rows)
    return Response(
        content="\ufeff" + buf.getvalue(),   # BOM UTF-8 pour Excel
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=fironova-payouts-nowpayments.csv"},
    )


class AffiliatePayoutRunForceIn(BaseModel):
    period: str = Field(pattern=r"^\d{4}-\d{2}$", description="YYYY-MM")


async def admin_affiliate_force_monthly_run(payload: AffiliatePayoutRunForceIn,
                                             admin: dict = Depends(get_admin_user)):  # noqa: F821
    """Force un run manuel du scheduler pour une période donnée."""
    fx_rate, fx_source = await _fetch_cad_to_usd_rate()
    fx_captured_at = datetime.now(timezone.utc).isoformat()
    run_id = str(uuid.uuid4())
    # `auto: False` autorise plusieurs re-runs manuels de la même période
    # (utile pour tester la génération itérativement). Le suffixe run_id
    # garantit l'unicité malgré l'index (period, auto).
    from pymongo.errors import DuplicateKeyError as _DupKey
    try:
        await db.payout_runs.insert_one({
            "id": run_id, "period": payload.period, "auto": False,
            "started_at": fx_captured_at, "status": "running",
            "triggered_by": admin.get("email", ""),
        })
    except _DupKey:
        # Rejoue en injectant un discriminant unique dans la clé `auto`
        await db.payout_runs.insert_one({
            "id": run_id, "period": payload.period,
            "auto": f"manual_{run_id[:8]}",
            "started_at": fx_captured_at, "status": "running",
            "triggered_by": admin.get("email", ""),
        })
    try:
        count = await _generate_payouts_for_period(payload.period, fx_rate,
                                                    fx_source, fx_captured_at)
        await db.payout_runs.update_one(
            {"id": run_id},
            {"$set": {"status": "done", "payouts_created": count,
                      "ended_at": datetime.now(timezone.utc).isoformat()}},
        )
        return {"ok": True, "period": payload.period, "payouts_created": count,
                "fx_rate_cad_to_usd": fx_rate, "fx_source": fx_source, "run_id": run_id}
    except Exception as e:
        await db.payout_runs.update_one(
            {"id": run_id}, {"$set": {"status": "failed", "error": "Payout generation failed"}},
        )
        logging.error("Manual payout run failed run=%s error_type=%s", run_id, type(e).__name__)
        raise HTTPException(500, "Payout generation failed") from e


async def admin_affiliate_payout_runs(admin: dict = Depends(get_admin_user),  # noqa: F821
                                       limit: int = 50):
    """Historique des runs du scheduler mensuel (auto + manuels) pour audit."""
    limit = max(1, min(int(limit or 50), 200))
    cursor = db.payout_runs.find(
        {}, {"_id": 0},
    ).sort("started_at", -1).limit(limit)
    runs = []
    async for r in cursor:
        # Normalise `auto` en booléen affichable + label
        auto_val = r.get("auto")
        r["is_auto"] = auto_val is True
        r["kind"] = "auto" if auto_val is True else "manual"
        runs.append(r)
    return {"runs": runs, "count": len(runs)}


async def admin_affiliate_mark_paid(payout_id: str, payload: AffiliatePayoutMarkIn,
                                    admin: dict = Depends(get_admin_user)):  # noqa: F821
    """Confirme le paiement crypto manuellement : enregistre la référence de
    transaction (tx hash) — statut `paid_manual` pour distinguer d'un paiement
    confirmé automatiquement via NOWPayments (statut `paid`)."""
    payout = await db.affiliate_payouts.find_one({"id": payout_id}, {"_id": 0})
    if not payout:
        raise HTTPException(404, "Payout not found")
    if payout.get("status") in ("paid", "paid_manual"):
        raise HTTPException(400, "Payout already marked paid")
    now = datetime.now(timezone.utc).isoformat()
    await db.affiliate_payouts.update_one(
        {"id": payout_id},
        {"$set": {"status": "paid_manual", "reference": payload.reference.strip(),
                  "note": payload.note or "", "paid_at": now,
                  "paid_by": admin.get("email")}},
    )
    await db.affiliate_referrals.update_many(
        {"payout_id": payout_id, "status": {"$in": ["pending", "approved"]}},
        {"$set": {"status": "paid", "paid_at": now}},
    )
    fresh = await db.affiliate_payouts.find_one({"id": payout_id}, {"_id": 0})
    return fresh


async def admin_affiliate_payouts_all(admin: dict = Depends(get_admin_user),  # noqa: F821
                                      status: Optional[str] = None):
    filt = {}
    if status:
        filt["status"] = status
    rows = await _cursor_all(db.affiliate_payouts.find(filt, {"_id": 0}).sort("created_at", -1))
    return rows

# ===== FIRONOVA_AFFILIATE_BLOCK_END =====


# ===== FIRONOVA_SEO_BACKEND_START =====
async def seo_health():
    return {"ok": True, "source": "fironova"}

async def seo_sitemap():
    return {"sitemap": ["/", "/catalog", "/about", "/faq", "/privacy", "/compliance"]}

# ===== FIRONOVA_SEO_BACKEND_END =====


# ===== FIRONOVA_RELATED_PRODUCTS_START =====
async def get_related_products(slug: str, limit: int = 4):
    """Produits reliés : même catégorie d'abord, complétés par des vedettes.
    Déterministe, sans ML. Sert le cross-sell « souvent recherché avec »."""
    base = await db.products.find_one({"slug": slug}, {"_id": 0, "category": 1, "id": 1})
    if not base:
        raise HTTPException(404, "Product not found")
    limit = max(1, min(limit, 8))
    seen = {base.get("id")}
    out = []
    # 1) même catégorie, en stock en priorité
    same = await db.products.find(
        {"active": True, "category": base.get("category"), "slug": {"$ne": slug}},
        {"_id": 0},
    ).sort("featured", -1).to_list(50)
    for p in same:
        if p.get("id") in seen:
            continue
        seen.add(p.get("id")); out.append(p)
        if len(out) >= limit:
            return out
    # 2) compléter avec des vedettes d'autres catégories
    feat = await db.products.find(
        {"active": True, "featured": True, "slug": {"$ne": slug}},
        {"_id": 0},
    ).to_list(50)
    for p in feat:
        if p.get("id") in seen:
            continue
        seen.add(p.get("id")); out.append(p)
        if len(out) >= limit:
            break
    return out
# ===== FIRONOVA_RELATED_PRODUCTS_END =====


# ===== FIRONOVA_CUSTOMERS_ENRICHED_START =====
async def admin_customer_detail(user_id: str, _admin: dict = Depends(require_area("customers", "view"))):
    """Fiche client complète avec historique de commandes."""
    u = await db.users.find_one(
        {"id": user_id}, {"_id": 0, "password_hash": 0, "token_version": 0}
    )
    if not u:
        raise HTTPException(404, "Customer not found")
    orders = await db.orders.find(
        {"email": (u.get("email") or "").lower()}, {"_id": 0}
    ).sort("created_at", -1).to_list(200)
    paid = [o for o in orders if o.get("payment_status") == "paid"]
    total_spent = round(sum(o.get("total", 0) for o in paid), 2)
    aov = round(total_spent / len(paid), 2) if paid else 0.0
    return {
        "customer": u,
        "orders": orders,
        "summary": {
            "total_spent": total_spent,
            "paid_orders": len(paid),
            "all_orders": len(orders),
            "aov": aov,
        },
    }
# ===== FIRONOVA_CUSTOMERS_ENRICHED_END =====


# ===== FIRONOVA_ANALYTICS_ENHANCED_START =====
TAX_THRESHOLD_CAD = 30000.0          # seuil d'inscription TPS/TVQ (petit fournisseur)
TAX_ALERT_RATIO = 0.80               # alerte "approche" à 80 % du seuil
_UNPAID = ["pending", "awaiting_etransfer", "awaiting_crypto"]


def _pct_change(cur: float, prev: float):
    """Variation en % ; None si base précédente nulle (évite division par zéro)."""
    if prev == 0:
        return None
    return round((cur - prev) / prev * 100, 1)


async def admin_analytics_enhanced(period: int = 30,
                                   _admin: dict = Depends(require_area("dashboard", "view"))):  # noqa: F821
    """Métriques de pilotage avec comparaison période courante vs précédente."""
    if period not in DASHBOARD_PERIODS:
        period = 30
    now = datetime.now(timezone.utc)
    cur_start = now - timedelta(days=period)
    prev_start = now - timedelta(days=period * 2)
    cur_start_s = cur_start.isoformat()
    prev_start_s = prev_start.isoformat()

    # ---- Période courante : commandes payées ----
    async def _period_stats(start_s: str, end_s: str) -> dict:
        cur = db.orders.aggregate([  # noqa: F821
            {"$match": {"payment_status": "paid",
                        "created_at": {"$gte": start_s, "$lt": end_s}}},
            {"$group": {"_id": None,
                        "revenue": {"$sum": "$total"},
                        "orders": {"$sum": 1}}},
        ])
        doc = await cur.to_list(1)
        rev = round(doc[0]["revenue"], 2) if doc else 0.0
        n = doc[0]["orders"] if doc else 0
        return {"revenue": rev, "orders": n, "aov": round(rev / n, 2) if n else 0.0}

    current = await _period_stats(cur_start_s, now.isoformat())
    previous = await _period_stats(prev_start_s, cur_start_s)

    # ---- Conversion : payées / créées (toutes, sur la période courante) ----
    created = await db.orders.count_documents(  # noqa: F821
        {"created_at": {"$gte": cur_start_s}})
    paid = await db.orders.count_documents(  # noqa: F821
        {"created_at": {"$gte": cur_start_s}, "payment_status": "paid"})
    abandoned = await db.orders.count_documents(  # noqa: F821
        {"created_at": {"$gte": cur_start_s}, "payment_status": {"$in": _UNPAID}})
    conversion = round(paid / created * 100, 1) if created else None

    # ---- Nouveaux vs récurrents (clients ayant payé sur la période) ----
    cur = db.orders.aggregate([  # noqa: F821
        {"$match": {"payment_status": "paid", "created_at": {"$gte": cur_start_s},
                    "email": {"$nin": [None, ""]}}},
        {"$group": {"_id": "$email"}},
    ])
    period_emails = [d["_id"] async for d in cur]
    new_customers = 0
    returning_customers = 0
    for em in period_emails:
        prior = await db.orders.count_documents(  # noqa: F821
            {"email": em, "payment_status": "paid", "created_at": {"$lt": cur_start_s}})
        if prior > 0:
            returning_customers += 1
        else:
            new_customers += 1

    # ---- Alerte seuil de taxe : CA payé sur 12 mois glissants ----
    twelve_start = (now - timedelta(days=365)).isoformat()
    ytd_cur = db.orders.aggregate([  # noqa: F821
        {"$match": {"payment_status": "paid", "created_at": {"$gte": twelve_start}}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}}},
    ])
    ytd_doc = await ytd_cur.to_list(1)
    rolling_12mo = round(ytd_doc[0]["total"], 2) if ytd_doc else 0.0
    ratio = rolling_12mo / TAX_THRESHOLD_CAD if TAX_THRESHOLD_CAD else 0
    if rolling_12mo >= TAX_THRESHOLD_CAD:
        tax_level = "exceeded"
    elif ratio >= TAX_ALERT_RATIO:
        tax_level = "approaching"
    else:
        tax_level = "ok"

    return {
        "period_days": period,
        "current": current,
        "previous": previous,
        "changes": {
            "revenue": _pct_change(current["revenue"], previous["revenue"]),
            "orders": _pct_change(current["orders"], previous["orders"]),
            "aov": _pct_change(current["aov"], previous["aov"]),
        },
        "conversion": {
            "orders_created": created,
            "orders_paid": paid,
            "orders_abandoned": abandoned,
            "conversion_rate": conversion,
        },
        "customers": {
            "new": new_customers,
            "returning": returning_customers,
            "total_active": len(period_emails),
        },
        "tax_threshold": {
            "rolling_12mo_revenue": rolling_12mo,
            "threshold": TAX_THRESHOLD_CAD,
            "ratio": round(ratio, 3),
            "level": tax_level,            # ok | approaching | exceeded
            "remaining": round(max(0.0, TAX_THRESHOLD_CAD - rolling_12mo), 2),
        },
    }

# ===== FIRONOVA_ANALYTICS_ENHANCED_END =====


# ===== FIRONOVA_EMAIL_AUTOMATION_END =====


# ===== FIRONOVA_SEO_BLOCK_START =====
SEO_ORIGIN = os.environ.get("PUBLIC_BASE_URL", "https://fironova.com").rstrip("/")  # noqa: F821

# Pages statiques indexables (doit rester cohérent avec robots.txt)
SEO_STATIC_PAGES = [
    ("/", "weekly", "1.0"),
    ("/catalog", "weekly", "0.9"),
    ("/about", "monthly", "0.6"),
    ("/compliance", "monthly", "0.7"),
    ("/faq", "monthly", "0.7"),
    ("/privacy", "yearly", "0.4"),
]


def _seo_xml_escape(s: str) -> str:
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


async def dynamic_sitemap():
    """Sitemap généré à la volée : pages statiques + toutes les fiches produits
    actives. Un produit ajouté apparaît automatiquement — plus de sitemap figé."""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    parts = ['<?xml version="1.0" encoding="UTF-8"?>',
             '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']

    for path, freq, prio in SEO_STATIC_PAGES:
        parts.append(
            f"  <url><loc>{SEO_ORIGIN}{path}</loc>"
            f"<changefreq>{freq}</changefreq><priority>{prio}</priority></url>"
        )

    # Produits actifs
    cursor = db.products.find(  # noqa: F821
        {"active": True},
        {"_id": 0, "slug": 1, "updated_at": 1, "created_at": 1},
    )
    async for p in cursor:
        slug = p.get("slug")
        if not slug:
            continue
        lastmod = (p.get("updated_at") or p.get("created_at") or "")[:10]
        lastmod_tag = f"<lastmod>{lastmod}</lastmod>" if lastmod else ""
        parts.append(
            f"  <url><loc>{SEO_ORIGIN}/product/{_seo_xml_escape(slug)}</loc>"
            f"{lastmod_tag}<changefreq>weekly</changefreq><priority>0.8</priority></url>"
        )

    parts.append("</urlset>")
    return Response(content="\n".join(parts), media_type="application/xml")


async def product_seo(slug: str):
    """Métadonnées SEO d'un produit : title/description/og + JSON-LD Product
    prêt à injecter. Le front appelle ça et remplit le <head> de la fiche."""
    p = await db.products.find_one({"slug": slug, "active": True}, {"_id": 0})  # noqa: F821
    if not p:
        raise HTTPException(404, "Product not found")  # noqa: F821

    variants = p.get("variants", []) or []
    prices = [float(v.get("price", 0)) for v in variants if v.get("price")]
    low_price = min(prices) if prices else float(p.get("price_cad", 0) or 0)
    in_stock = any(int(v.get("stock", 0) or 0) > 0 for v in variants) or int(p.get("stock", 0) or 0) > 0
    img = p.get("og_image_url") or p.get("image_url") or ""
    if img and img.startswith("/"):
        img = SEO_ORIGIN + img

    def _meta(lang: str) -> dict:
        name = p.get(f"name_{lang}") or p.get("name_en") or slug
        title = p.get(f"meta_title_{lang}") or f"{name} — Fironova"
        desc = p.get(f"meta_description_{lang}") or ""
        if not desc:
            raw = p.get(f"description_{lang}") or p.get("description_en") or ""
            desc = (raw[:157] + "…") if len(raw) > 158 else raw
        return {"title": title, "description": desc, "name": name}

    # JSON-LD Product (schema.org) — anglais comme langue canonique du balisage
    en = _meta("en")
    jsonld = {
        "@context": "https://schema.org",
        "@type": "Product",
        "name": en["name"],
        "description": en["description"],
        "sku": p.get("slug"),
        "brand": {"@type": "Brand", "name": "Fironova"},
        "category": p.get("category", ""),
    }
    if img:
        jsonld["image"] = img
    if p.get("molecular_formula"):
        jsonld["additionalProperty"] = [{
            "@type": "PropertyValue", "name": "Molecular formula",
            "value": p.get("molecular_formula"),
        }]
    if low_price > 0:
        jsonld["offers"] = {
            "@type": "Offer",
            "priceCurrency": "CAD",
            "price": round(low_price, 2),
            "availability": ("https://schema.org/InStock" if in_stock
                             else "https://schema.org/OutOfStock"),
            "url": f"{SEO_ORIGIN}/product/{slug}",
        }

    return {
        "slug": slug,
        "canonical": f"{SEO_ORIGIN}/product/{slug}",
        "image": img,
        "en": en,
        "fr": _meta("fr"),
        "jsonld": jsonld,
    }

# ===== FIRONOVA_SEO_BLOCK_END =====


# ===== FIRONOVA_SEO_ADMIN_BLOCK_START =====
# Valeurs par défaut si aucun réglage n'a encore été enregistré.
SEO_DEFAULTS = {
    "site_title_en": "Fironova — Research Peptides (For Research Use Only)",
    "site_title_fr": "Fironova — Peptides de recherche (RUO)",
    "site_description_en": "Fironova supplies research-grade peptides for laboratory use, with certificate-of-analysis documentation. For Research Use Only. Not for human consumption.",
    "site_description_fr": "Fironova fournit des peptides de qualité recherche pour usage en laboratoire, avec documentation de certificat d'analyse. Réservé à la recherche (RUO). Ne pas consommer.",
    "default_og_image": "",
    "keywords_en": "research peptides, laboratory, certificate of analysis, RUO",
    "keywords_fr": "peptides de recherche, laboratoire, certificat d'analyse, RUO",
}


class SeoSettingsIn(BaseModel):
    site_title_en: str = ""
    site_title_fr: str = ""
    site_description_en: str = ""
    site_description_fr: str = ""
    default_og_image: str = ""
    keywords_en: str = ""
    keywords_fr: str = ""


class ProductSeoIn(BaseModel):
    meta_title_en: str = ""
    meta_title_fr: str = ""
    meta_description_en: str = ""
    meta_description_fr: str = ""
    og_image_url: str = ""


async def _seo_get_settings() -> dict:
    doc = await db.seo_settings.find_one({"id": "global"}, {"_id": 0})  # noqa: F821
    if not doc:
        return dict(SEO_DEFAULTS)
    merged = dict(SEO_DEFAULTS)
    for k in SEO_DEFAULTS:
        if doc.get(k):
            merged[k] = doc[k]
    return merged


async def admin_seo_get_settings(admin: dict = Depends(get_admin_user)):  # noqa: F821
    return await _seo_get_settings()


async def admin_seo_set_settings(payload: SeoSettingsIn,
                                 admin: dict = Depends(get_admin_user)):  # noqa: F821
    data = {k: v for k, v in payload.model_dump().items()}
    data.update({
        "id": "global",
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": admin.get("email"),
    })
    await db.seo_settings.update_one({"id": "global"}, {"$set": data}, upsert=True)  # noqa: F821
    return await _seo_get_settings()


async def admin_seo_health(admin: dict = Depends(get_admin_user)):  # noqa: F821
    """État de santé SEO : combien de produits actifs, combien sans meta."""
    total = await db.products.count_documents({"active": True})  # noqa: F821
    missing_title = 0
    missing_desc = 0
    missing_image = 0
    async for p in db.products.find(  # noqa: F821
        {"active": True},
        {"_id": 0, "meta_title_en": 1, "meta_title_fr": 1,
         "meta_description_en": 1, "meta_description_fr": 1,
         "og_image_url": 1, "image_url": 1},
    ):
        if not (p.get("meta_title_en") or p.get("meta_title_fr")):
            missing_title += 1
        if not (p.get("meta_description_en") or p.get("meta_description_fr")):
            missing_desc += 1
        if not (p.get("og_image_url") or p.get("image_url")):
            missing_image += 1

    settings = await _seo_get_settings()
    global_ok = bool(settings.get("site_title_en") and settings.get("site_description_en"))

    return {
        "products_active": total,
        "products_missing_meta_title": missing_title,
        "products_missing_meta_description": missing_desc,
        "products_missing_image": missing_image,
        "products_fully_optimized": max(0, total - max(missing_title, missing_desc)),
        "global_seo_configured": global_ok,
        "default_og_image_set": bool(settings.get("default_og_image")),
        "sitemap_url": "/api/sitemap.xml",
    }


async def admin_seo_products(admin: dict = Depends(get_admin_user)):  # noqa: F821
    """SEO de tous les produits, une ligne par produit (pour édition centralisée)."""
    rows = []
    async for p in db.products.find(  # noqa: F821
        {}, {"_id": 0, "slug": 1, "name_en": 1, "name_fr": 1, "active": 1,
             "meta_title_en": 1, "meta_title_fr": 1,
             "meta_description_en": 1, "meta_description_fr": 1,
             "og_image_url": 1, "image_url": 1},
    ):
        has_title = bool(p.get("meta_title_en") or p.get("meta_title_fr"))
        has_desc = bool(p.get("meta_description_en") or p.get("meta_description_fr"))
        rows.append({
            "slug": p.get("slug"),
            "name_en": p.get("name_en"),
            "name_fr": p.get("name_fr"),
            "active": p.get("active", False),
            "meta_title_en": p.get("meta_title_en", ""),
            "meta_title_fr": p.get("meta_title_fr", ""),
            "meta_description_en": p.get("meta_description_en", ""),
            "meta_description_fr": p.get("meta_description_fr", ""),
            "og_image_url": p.get("og_image_url", ""),
            "image_url": p.get("image_url", ""),
            "optimized": has_title and has_desc,
        })
    rows.sort(key=lambda r: (r["optimized"], not r["active"]))  # non-optimisés d'abord
    return {"products": rows}


async def admin_seo_update_product(slug: str, payload: ProductSeoIn,
                                   admin: dict = Depends(get_admin_user)):  # noqa: F821
    """Édite uniquement les champs SEO d'un produit (sans toucher au reste)."""
    p = await db.products.find_one({"slug": slug}, {"_id": 0, "id": 1})  # noqa: F821
    if not p:
        raise HTTPException(404, "Product not found")  # noqa: F821
    await db.products.update_one(  # noqa: F821
        {"slug": slug},
        {"$set": {
            "meta_title_en": payload.meta_title_en,
            "meta_title_fr": payload.meta_title_fr,
            "meta_description_en": payload.meta_description_en,
            "meta_description_fr": payload.meta_description_fr,
            "og_image_url": payload.og_image_url,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }},
    )
    return {"ok": True, "slug": slug}


async def admin_email_templates_list(_admin: dict = Depends(require_area("settings", "view"))):
    keys = list(EMAIL_TEMPLATE_CATALOG.keys())
    custom = await db.email_templates.find({"key": {"$nin": keys}}, {"_id": 0, "key": 1}).to_list(100)
    keys += [c["key"] for c in custom]
    templates = [t for t in [await _email_template_get(k) for k in keys] if t]
    return {"templates": templates}


async def admin_email_template_create(payload: EmailTemplateCreateIn,
                                      _admin: dict = Depends(require_area("settings", "manage"))):
    key = (payload.key or "").strip().lower().replace(" ", "_")
    if not key or not key.replace("_", "").isalnum():
        raise HTTPException(400, "Clé invalide (lettres, chiffres, underscores).")
    if key in EMAIL_TEMPLATE_CATALOG or await db.email_templates.find_one({"key": key}):
        raise HTTPException(409, "Un courriel avec cette clé existe déjà.")
    doc = payload.model_dump(exclude_none=True)
    doc["key"] = key
    doc.setdefault("label", key)
    doc.setdefault("block", "none")
    doc.setdefault("variables", [])
    doc["custom"] = True
    doc["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.email_templates.insert_one(dict(doc))
    asyncio.create_task(_log_action(_admin, "email_template_create", f"key={key}", "settings"))
    return await _email_template_get(key)


async def admin_email_template_update(key: str, payload: EmailTemplateIn,
                                      _admin: dict = Depends(require_area("settings", "manage"))):
    exists = key in EMAIL_TEMPLATE_CATALOG or await db.email_templates.find_one({"key": key})
    if not exists:
        raise HTTPException(404, "Unknown template key")
    updates = {k: v for k, v in payload.model_dump(exclude_none=True).items()}
    updates["key"] = key
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.email_templates.update_one({"key": key}, {"$set": updates}, upsert=True)
    asyncio.create_task(_log_action(_admin, "email_template_update",
                                    f"key={key} fields={list(payload.model_dump(exclude_none=True).keys())}", "settings"))
    return await _email_template_get(key)


async def admin_email_template_delete(key: str, _admin: dict = Depends(require_area("settings", "manage"))):
    if key in EMAIL_TEMPLATE_CATALOG:
        raise HTTPException(400, "Ce courriel est intégré et ne peut pas être supprimé. Utilisez Réinitialiser.")
    res = await db.email_templates.delete_one({"key": key})
    if res.deleted_count == 0:
        raise HTTPException(404, "Template not found")
    asyncio.create_task(_log_action(_admin, "email_template_delete", f"key={key}", "settings"))
    return {"deleted": True, "key": key}


async def admin_email_template_reset(key: str, _admin: dict = Depends(require_area("settings", "manage"))):
    if key not in EMAIL_TEMPLATE_CATALOG:
        raise HTTPException(400, "Seuls les courriels intégrés peuvent être réinitialisés.")
    await db.email_templates.delete_one({"key": key})
    asyncio.create_task(_log_action(_admin, "email_template_reset", f"key={key}", "settings"))
    return await _email_template_get(key)


async def admin_email_template_preview(key: str, lang: str = "fr",
                                       _admin: dict = Depends(require_area("settings", "view"))):
    tpl = await _email_template_get(key)
    if not tpl:
        raise HTTPException(404, "Unknown template key")
    sample_order = {
        "payment_info": {"instructions": {"send_to": "orders@fironova.com", "amount_cad": 149.00,
                          "reference": "FN-26042", "security_question": "What is the brand name? (lowercase)",
                          "security_answer_hint": "fironova"},
                          "pay_address": "0xA1b2C3d4E5f6...", "pay_amount": "0.0021", "pay_currency": "eth",
                          "payment_url": "https://nowpayments.io/payment/xyz"},
        "items": [{"name_fr": "BPC-157 5mg", "name_en": "BPC-157 5mg", "qty": 2},
                  {"name_fr": "TB-500 5mg", "name_en": "TB-500 5mg", "qty": 1}],
        "total": 149.00, "tracking_number": "1Z999AA10123456784", "_refund_amount": 149.00,
    }
    ctx = {"order_number": "FN-26042", "customer_name": "Alex", "total": "149,00 $",
           "tracking_number": "1Z999AA10123456784", "tracking_url": "https://example.com/track",
           "amount": "149,00 $", "cart_url": "https://fironova.com/cart",
           "catalog_url": "https://fironova.com/catalog",
           "product_name": "BPC-157 5mg", "product_url": "https://fironova.com/p/bpc-157"}
    subject, html = _email_render(tpl, lang, ctx, sample_order)
    return {"subject": subject, "html": html}
# ===== FIRONOVA_EMAIL_WIRING_END =====


# ===========================================================================
# ===== FIRONOVA_NOWPAYMENTS_PAYOUT_START =====
# Client Payouts NOWPayments + flux semi-automatique sécurisé.
#
# Flux (option "semi-auto sécurisé", cf. décision produit) :
#   1) L'admin exécute un payout "ready" -> le système obtient un JWT
#      (email/password NOWPayments), crée le payout via POST /v1/payout,
#      NOWPayments envoie un code 2FA à l'email du compte marchand.
#      -> statut interne: "creating" (en attente du 2FA).
#   2) L'admin saisit le code 2FA -> POST /v1/payout/{id}/verify.
#      -> statut interne: "processing" puis "paid" (via statut/IPN).
#
# Sécurité : aucun secret 2FA stocké. Le JWT est éphémère (5 min), obtenu à
# la demande. L'IP du serveur doit être whitelistée chez NOWPayments.
# Interrupteur: NOWPAYMENTS_PAYOUT_ENABLED (off => flux manuel conservé).
# ===========================================================================

NOWPAYMENTS_PAYOUT_ENABLED = os.environ.get("NOWPAYMENTS_PAYOUT_ENABLED", "false").lower() == "true"
NOWPAYMENTS_EMAIL = os.environ.get("NOWPAYMENTS_EMAIL", "")
NOWPAYMENTS_PASSWORD = os.environ.get("NOWPAYMENTS_PASSWORD", "")


# ===== FIRONOVA_NOWPAYMENTS_PAYOUT_END =====


# ===========================================================================
# ===== FIRONOVA_PAYOUT_ENDPOINTS_START =====

class PayoutVerifyIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    verification_code: str = Field(min_length=3, max_length=32)


async def admin_payout_execute(payout_id: str, admin: dict = Depends(get_admin_user)):  # noqa: F821
    """Étape 1/2 — crée le payout crypto via NOWPayments. Un code 2FA est
    envoyé par NOWPayments à l'email marchand. Le payout passe en 'creating'."""
    if not NOWPAYMENTS_PAYOUT_ENABLED:
        raise HTTPException(400, "Les payouts automatiques NOWPayments sont désactivés (NOWPAYMENTS_PAYOUT_ENABLED=false). Utilisez le paiement manuel + mark-paid.")
    payout = await db.affiliate_payouts.find_one({"id": payout_id}, {"_id": 0})
    if not payout:
        raise HTTPException(404, "Payout introuvable")
    if payout.get("status") not in ("ready", "failed"):
        raise HTTPException(400, f"Payout non exécutable (statut: {payout.get('status')})")
    address = (payout.get("payout_address") or "").strip()
    if not address:
        raise HTTPException(400, "Adresse de versement manquante pour cet affilié.")
    amount = float(payout.get("amount", 0))
    if amount <= 0:
        raise HTTPException(400, "Montant de payout invalide.")
    currency = (payout.get("currency") or "btc").lower()

    # Transition atomique de l'état "ready" vers "creating" pour éviter les
    # exécutions concurrentes et les doubles batches NOWPayments.
    now = datetime.now(timezone.utc).isoformat()
    claimed = await db.affiliate_payouts.update_one(
        {"id": payout_id, "status": {"$in": ["ready", "failed"]}},
        {"$set": {
            "status": "creating",
            "np_batch_id": None,
            "np_error": None,
            "executed_by": admin.get("email"),
            "executed_at": now,
            "updated_at": now,
        }},
    )
    if not claimed.modified_count:
        raise HTTPException(400, "Payout déjà en cours d'exécution ou déjà traité.")

    withdrawals = [{"address": address, "currency": currency, "amount": amount}]
    desc = f"Fironova affiliate {payout.get('affiliate_code')} — {payout.get('period')}"
    try:
        resp = await _np_create_payout(withdrawals, description=desc)
    except NowPaymentsPayoutError as e:
        await db.affiliate_payouts.update_one({"id": payout_id},
            {"$set": {"status": "failed", "np_error": "NOWPayments payout request failed", "updated_at": datetime.now(timezone.utc).isoformat()}})
        raise HTTPException(502, "NOWPayments payout request failed") from e

    batch_id = str(resp.get("id") or resp.get("batch_withdrawal_id") or "")
    if not batch_id:
        await db.affiliate_payouts.update_one({"id": payout_id},
            {"$set": {"status": "failed", "np_error": "NOWPayments n'a pas renvoyé d'identifiant de payout.", "updated_at": datetime.now(timezone.utc).isoformat()}})
        raise HTTPException(502, "NOWPayments n'a pas renvoyé d'identifiant de payout.")
    await db.affiliate_payouts.update_one({"id": payout_id}, {"$set": {
        "status": "creating", "np_batch_id": batch_id, "np_error": None,
        "executed_by": admin.get("email"), "executed_at": now,
    }})
    asyncio.create_task(_log_action(admin, "affiliate_payout_execute", f"payout={payout_id} batch={batch_id} amount={amount} {currency}", "settings"))
    return {"ok": True, "status": "creating", "np_batch_id": batch_id,
            "message": "Payout créé. Un code de vérification 2FA a été envoyé à l'adresse courriel du compte NOWPayments. Saisissez-le pour finaliser."}


async def admin_payout_verify(payout_id: str, payload: PayoutVerifyIn,
                              admin: dict = Depends(get_admin_user)):  # noqa: F821
    """Étape 2/2 — valide le payout avec le code 2FA reçu par email."""
    payout = await db.affiliate_payouts.find_one({"id": payout_id}, {"_id": 0})
    if not payout:
        raise HTTPException(404, "Payout introuvable")
    batch_id = payout.get("np_batch_id")
    if not batch_id or payout.get("status") != "creating":
        raise HTTPException(400, "Ce payout n'est pas en attente de vérification 2FA.")
    try:
        await _np_verify_payout(batch_id, payload.verification_code)
    except NowPaymentsPayoutError as e:
        raise HTTPException(502, "NOWPayments payout verification failed") from e

    now = datetime.now(timezone.utc).isoformat()
    await db.affiliate_payouts.update_one({"id": payout_id}, {"$set": {
        "status": "processing", "verified_by": admin.get("email"), "verified_at": now,
    }})
    asyncio.create_task(_log_action(admin, "affiliate_payout_verify", f"payout={payout_id} batch={batch_id}", "settings"))
    return {"ok": True, "status": "processing",
            "message": "Payout vérifié et en cours de traitement par NOWPayments."}


async def admin_payout_status(payout_id: str, admin: dict = Depends(get_admin_user)):  # noqa: F821
    """Rafraîchit le statut depuis NOWPayments et synchronise l'état interne."""
    payout = await db.affiliate_payouts.find_one({"id": payout_id}, {"_id": 0})
    if not payout:
        raise HTTPException(404, "Payout introuvable")
    batch_id = payout.get("np_batch_id")
    if not batch_id:
        return {"status": payout.get("status"), "np_status": None}
    try:
        data = await _np_payout_status(batch_id)
    except NowPaymentsPayoutError as e:
        raise HTTPException(502, "NOWPayments payout status unavailable") from e
    np_status = str(data.get("status", "")).lower()
    # Mappe le statut NOWPayments vers l'état interne + finalise si terminé.
    mapped = payout.get("status")
    if np_status in ("finished", "sent", "completed"):
        mapped = "paid"
    elif np_status in ("failed", "rejected"):
        mapped = "failed"
    elif np_status in ("processing", "sending", "waiting", "confirming"):
        mapped = "processing"
    if mapped != payout.get("status"):
        now = datetime.now(timezone.utc).isoformat()
        upd = {"status": mapped, "np_status": np_status, "updated_at": now}
        if mapped == "paid":
            upd["paid_at"] = now
            upd["reference"] = str(data.get("id") or batch_id)
            await db.affiliate_referrals.update_many(
                {"payout_id": payout_id, "status": {"$in": ["pending", "approved"]}},
                {"$set": {"status": "paid", "paid_at": now}},
            )
        await db.affiliate_payouts.update_one({"id": payout_id}, {"$set": upd})
    return {"status": mapped, "np_status": np_status}


# ===== FIRONOVA_PAYOUT_ENDPOINTS_END =====


try:
    from routers.admin_trash import router as admin_trash_router
    from routers.auth_account import router as auth_account_router
    from routers.affiliate import router as affiliate_router
    from routers.email_templates import router as email_templates_router
    from routers.admin_orders import router as admin_orders_router
    from routers.admin_order_ops import router as admin_order_ops_router
    from routers.admin_dispatch_fulfillment import router as admin_dispatch_fulfillment_router
    from routers.admin_misc import router as admin_misc_router
    from routers.catalog_public import router as catalog_public_router
    from routers.admin_commerce import router as admin_commerce_router
    from routers.orders import router as orders_router
    from routers.payments import router as payments_router
    from routers.public import router as public_router
    from routers.reconciliation import router as reconciliation_router
    from routers.seo import router as seo_router
except ImportError:
    from backend.routers.admin_trash import router as admin_trash_router
    from backend.routers.auth_account import router as auth_account_router
    from backend.routers.affiliate import router as affiliate_router
    from backend.routers.email_templates import router as email_templates_router
    from backend.routers.admin_orders import router as admin_orders_router
    from backend.routers.admin_order_ops import router as admin_order_ops_router
    from backend.routers.admin_dispatch_fulfillment import router as admin_dispatch_fulfillment_router
    from backend.routers.admin_misc import router as admin_misc_router
    from backend.routers.catalog_public import router as catalog_public_router
    from backend.routers.admin_commerce import router as admin_commerce_router
    from backend.routers.orders import router as orders_router
    from backend.routers.payments import router as payments_router
    from backend.routers.public import router as public_router
    from backend.routers.reconciliation import router as reconciliation_router
    from backend.routers.seo import router as seo_router

app.include_router(admin_trash_router)
app.include_router(auth_account_router)
app.include_router(affiliate_router)
app.include_router(admin_orders_router)
app.include_router(admin_order_ops_router)
app.include_router(admin_dispatch_fulfillment_router)
app.include_router(admin_misc_router)
app.include_router(email_templates_router)
app.include_router(catalog_public_router)
app.include_router(admin_commerce_router)
app.include_router(orders_router)
app.include_router(payments_router)
app.include_router(public_router)
app.include_router(reconciliation_router)
app.include_router(seo_router)
app.include_router(api)

# Credentialed CORS must use an explicit allowlist. Regex support is opt-in for
# ephemeral preview hosts and should remain unset in production.
_cors_raw = os.environ.get("CORS_ORIGINS", "").strip()
if not _cors_raw:
    _cors_raw = os.environ.get("FRONTEND_URL", "").strip()
if not _cors_raw and not IS_PRODUCTION:
    _cors_raw = "http://localhost:3000,http://localhost:5173"
if not _cors_raw:
    raise RuntimeError("CORS_ORIGINS ou FRONTEND_URL est obligatoire en production")
_cors_origins = [o.strip().rstrip("/") for o in _cors_raw.split(",") if o.strip()]
_cors_origin_regex = os.environ.get("CORS_ORIGIN_REGEX", "").strip() or None
_cors_origin_re = re.compile(_cors_origin_regex, re.IGNORECASE) if _cors_origin_regex else None
if "*" in _cors_origins:
    raise RuntimeError(
        "CORS_ORIGINS doit lister des origines explicites (ex. https://app.fironova.com) — "
        "'*' est interdit avec une auth par cookie (credentialed)."
    )

app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_origin_regex=_cors_origin_regex,
    allow_credentials=True,  # Auth cookie-only : le cookie httpOnly access_token doit traverser
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Accept", "Authorization", "Content-Type", "X-Order-Access-Token"],
)


@app.middleware("http")
async def _shared_mutation_rate_limit(request: Request, call_next):
    if request.method in {"POST", "PUT", "PATCH", "DELETE"} and request.url.path.startswith("/api/"):
        try:
            await _rate_limit(
                "public_mutation",
                _client_ip(request),
                PUBLIC_MUTATION_MAX_PER_MINUTE,
                60,
                "Too many requests. Try again shortly.",
            )
        except HTTPException as exc:
            return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})
    return await call_next(request)


@app.middleware("http")
async def _csrf_origin_guard(request: Request, call_next):
    """Block cross-origin state-mutating requests that carry the session cookie."""
    authz = (request.headers.get("authorization") or "").strip().lower()
    uses_bearer = authz.startswith("bearer ")
    has_auth_cookie = request.cookies.get("access_token") or request.cookies.get("refresh_token")
    if request.method in ("POST", "PUT", "PATCH", "DELETE") and has_auth_cookie and not uses_bearer:
        origin = (request.headers.get("origin") or "").lower().rstrip("/")
        allowed = {o.lower().rstrip("/") for o in _cors_origins}
        matches_regex = bool(_cors_origin_re and _cors_origin_re.fullmatch(origin))
        host = (
            _trusted_forwarded_header(request, "x-forwarded-host")
            or request.headers.get("host")
            or ""
        ).split(",", 1)[0].strip().lower()
        origin_host = origin.split("://", 1)[-1].split("/", 1)[0]

        def _strip_default_port(h: str) -> str:
            if h.endswith(":443"):
                return h[:-4]
            if h.endswith(":80"):
                return h[:-3]
            return h

        same_origin = _strip_default_port(origin_host) == _strip_default_port(host)
        if not origin:
            return Response(status_code=403, content="Origin required", media_type="text/plain")
        if origin not in allowed and not matches_regex and not same_origin:
            return Response(status_code=403, content="Origin not allowed", media_type="text/plain")
    return await call_next(request)


@app.middleware("http")
async def _security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "no-referrer")
    response.headers.setdefault("Permissions-Policy", "camera=(), microphone=(), geolocation=()")
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'none'; base-uri 'none'; form-action 'none'; frame-ancestors 'none'",
    )
    if request.url.path.startswith("/api/"):
        response.headers.setdefault("Cache-Control", "no-store")
        response.headers.setdefault("Pragma", "no-cache")
    if request.url.scheme == "https" or _trusted_forwarded_header(request, "x-forwarded-proto").lower() == "https":
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    return response


@app.get("/robots.txt", include_in_schema=False)
async def robots_txt():
    body = (
        "User-agent: *\n"
        "Allow: /\n"
        f"Sitemap: {SEO_ORIGIN}/api/sitemap.xml\n"
    )
    return Response(content=body, media_type="text/plain")


logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
